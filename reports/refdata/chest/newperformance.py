import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import yfinance as yf
from collections import namedtuple
import os
from openpyxl import Workbook, load_workbook

def flatten_journal_entries(journal_entries):
    data = [entry.to_dict() for entry in journal_entries]
    df = pd.DataFrame(data)

    # Ensure all necessary columns are present
    required_columns = [
        'portfolio', 'investment', 'lotid', 'tax_date', 'ls', 'location', 'financial_account',
        'quantity', 'local', 'book', 'notional', 'oface', 'tranid', 'transaction', 'tradedate',
        'settledate', 'kdbegin', 'kdend', 'ibor_date', 'entry_type', 'feeder', 'running_balances',
        'split_ratio', 'account_key', 'asset_class', 'currency', 'country', 'sector',
        'system_type', 'group1', 'bsgroup', 'performance_category'
    ]
    for column in required_columns:
        if column not in df.columns:
            df[column] = None

    return df
# Utility function to normalize journal entries
def normalize_journal_entries(journal_entries):
    """Normalize journal entries to ensure they are all instances of Journals."""
    normalized_entries = []

    for entry in journal_entries:
        if isinstance(entry, Journals):
            normalized_entries.append(entry)
        elif isinstance(entry, dict):
            normalized_entries.append(Journals(**entry))
        else:
            raise ValueError(f"Unexpected entry type: {type(entry)}. Expected Journals instance or dict.")

    return normalized_entries





def fetch_and_map_groupings(journal_entries, investment_master_path, coa_path):
    # Load the investment master and chart of accounts data
    investment_master_df = pd.read_csv(investment_master_path)
    coa_df = pd.read_csv(coa_path)

    investment_groupings = investment_master_df[['ticker', 'asset_class', 'currency', 'country', 'sector']]
    coa_groupings = coa_df[['SystemName', 'SystemType', 'AccountNum', 'BSGroup', 'PerformanceCategory']]

    for entry in journal_entries:
        investment_row = investment_groupings[investment_groupings['ticker'] == entry.investment]
        coa_row = coa_groupings[coa_groupings['SystemName'] == entry.financial_account]

        if not investment_row.empty:
            entry.asset_class = investment_row.iloc[0]['asset_class']
            entry.currency = investment_row.iloc[0]['currency']
            entry.country = investment_row.iloc[0]['country']
            entry.sector = investment_row.iloc[0]['sector']

        if not coa_row.empty:
            entry.system_type = coa_row.iloc[0]['SystemType']
            entry.group1 = coa_row.iloc[0]['AccountNum']
            entry.bsgroup = coa_row.iloc[0]['BSGroup']
            entry.performance_category = coa_row.iloc[0]['PerformanceCategory']

    return journal_entries

def compute_opening_cash_flows_investments(je_data, level):
    currencies = ['USD', 'AUD', 'GBP', 'EUR', 'JPY']
    valid_accounts = ['Cost']
    group_by_cols = [level, 'ibor_date']
    condition1 = (~je_data['investment'].isin(currencies) & (je_data['book'] > 0))
    condition2 = je_data['financial_account'].isin(valid_accounts)
    opening_flows = je_data[condition1 & condition2]
    opening_flows = opening_flows.groupby(group_by_cols).agg({'local': 'sum', 'book': 'sum'}).reset_index()
    opening_flows.rename(columns={'local': 'Open_CF_Local', 'book': 'Open_CF_Book'}, inplace=True)
    return opening_flows


def compute_cash_flows_currencies(je_data, level):
    currencies = ['USD', 'AUD', 'GBP', 'EUR', 'JPY']
    valid_accounts = ['Cost', 'Payable', 'Receivable', 'DividendsReceivable', 'DividendsPayable']
    group_by_cols = [level, 'ibor_date'] if level != 'ibor_date' else ['ibor_date']
    condition1 = je_data['investment'].isin(currencies)
    condition2 = je_data['financial_account'].isin(valid_accounts)
    currency_flows = je_data[condition1 & condition2]
    currency_flows = currency_flows.groupby(group_by_cols).agg({'local': 'sum', 'book': 'sum'}).reset_index()
    currency_flows.rename(columns={'local': 'Currency_Flows_Local', 'book': 'Currency_Flows_Book'}, inplace=True)
    return currency_flows


def compute_closing_cash_flows_for_investments(je_data, level):
    currencies = ['USD', 'AUD', 'GBP', 'EUR', 'JPY']
    condition1 = je_data['financial_account'].isin(['PriceGainInvestment', 'FXGainInvestment'])
    condition2 = (je_data['financial_account'] == 'Cost') & (je_data['book'] < 0) & (je_data['ls'] == 'l')
    condition3 = (je_data['financial_account'] == 'Cost') & (je_data['book'] > 0) & (je_data['ls'] == 's')
    condition4 = ~je_data['investment'].isin(currencies)
    closing_flows = je_data[(condition1 | condition2 | condition3) & condition4]
    closing_flows_agg = closing_flows.groupby([level, 'ibor_date']).agg({'local': 'sum', 'book': 'sum'}).reset_index()
    closing_flows_agg.rename(columns={'local': 'Close_CF_Local', 'book': 'Close_CF_Book'}, inplace=True)
    return closing_flows_agg


def compute_income(je_data, level):
    income_entries = je_data[(je_data['financial_account'] == 'DividendReceipt') |
                             (je_data['financial_account'] == 'FXGainCurrency') |
                             (je_data['financial_account'] == 'FXGainTradeSettle') |
                             (je_data['financial_account'] == 'DividendExpense')]
    income_je_data = income_entries.groupby([level, 'ibor_date'])[['local', 'book']].sum().reset_index()
    income_je_data.rename(columns={'local': 'Income_Local', 'book': 'Income_Book'}, inplace=True)
    return income_je_data

def compute_daily_twr(journal_entries, level):
    investment_master_path = "C:/Users/hjmne/PycharmProjects/chest/refdata/investment_master.csv"
    investment_master = pd.read_csv(investment_master_path)

    # Flatten journal entries to a DataFrame
    je_data = flatten_journal_entries(journal_entries)

    # Merge with investment master to get additional columns
    je_data = pd.merge(je_data, investment_master, left_on='investment', right_on='ticker', how='left').drop('ticker', axis=1)

    # Rename columns to avoid conflicts
    je_data.rename(columns={
        'investment_x': 'investment',
        'asset_class_x': 'asset_class',
        'currency_x': 'currency',
        'country_x': 'country'
    }, inplace=True)

    # Debugging: Print columns to verify
    print("Columns in je_data after merge and rename:")
    print(je_data.columns)

    je_data['tax_date'] = je_data['tax_date'].astype(str)
    market_values = je_data[je_data['financial_account'] == 'MktVal']
    market_values = market_values.rename(columns={'local': 'EMV_Local', 'book': 'EMV_Book'})

    # Debugging: Print sample data to verify
    print("Sample data from market_values before grouping:")
    print(market_values.head())

    # Check if the necessary columns exist
    if level not in market_values.columns:
        raise KeyError(f"Column '{level}' is missing in the data.")

    market_values = market_values.groupby([level, 'ibor_date']).agg(
        {'EMV_Local': 'sum', 'EMV_Book': 'sum'}).reset_index()
    market_values['ibor_date'] = pd.to_datetime(market_values['ibor_date'])
    market_values = market_values.sort_values(by=[level, 'ibor_date'])
    market_values['BMV_Local'] = market_values.groupby(level)['EMV_Local'].shift(1)
    market_values['BMV_Book'] = market_values.groupby(level)['EMV_Book'].shift(1)

    opening_flows = compute_opening_cash_flows_investments(je_data, level)
    market_values = pd.merge(market_values, opening_flows, on=[level, 'ibor_date'], how='left')

    currency_flows = compute_cash_flows_currencies(je_data, level)
    market_values = pd.merge(market_values, currency_flows, on=[level, 'ibor_date'], how='left')

    closing_flows = compute_closing_cash_flows_for_investments(je_data, level)
    market_values = pd.merge(market_values, closing_flows, on=[level, 'ibor_date'], how='left').fillna(0)

    income_data = compute_income(je_data, level)
    market_values = pd.merge(market_values, income_data, on=[level, 'ibor_date'], how='left').fillna(0)

    market_values['Investment_Changed'] = market_values[level] != market_values[level].shift(1)
    market_values['EMV_Local_Diff'] = np.where(market_values['Investment_Changed'], market_values['EMV_Local'],
                                               market_values['EMV_Local'].diff())
    market_values['EMV_Book_Diff'] = np.where(market_values['Investment_Changed'], market_values['EMV_Book'],
                                              market_values['EMV_Book'].diff())
    market_values['Previous_EMV_Local'] = np.where(market_values['Investment_Changed'], 0,
                                                   market_values['EMV_Local'].shift(1))
    market_values['Previous_EMV_Book'] = np.where(market_values['Investment_Changed'], 0,
                                                  market_values['EMV_Book'].shift(1))

    market_values['EMV_Local_Diff'] = market_values['EMV_Local_Diff'].fillna(0)
    market_values['EMV_Book_Diff'] = market_values['EMV_Book_Diff'].fillna(0)
    market_values['Previous_EMV_Local'] = market_values['Previous_EMV_Local'].fillna(0)

    def calculate_twr(row):
        if row[level] != 'portfolio':
            same_sign_local = (row['Previous_EMV_Local'] >= 0) == (row['Currency_Flows_Local'] >= 0)
            same_sign_book = (row['Previous_EMV_Book'] >= 0) == (row['Currency_Flows_Book'] >= 0)
            denominator_local = (row['Previous_EMV_Local'] + row['Open_CF_Local'] + row['Currency_Flows_Local']) if same_sign_local else (row['Previous_EMV_Local'] + row['Open_CF_Local'])
            denominator_book = (row['Previous_EMV_Book'] + row['Open_CF_Book'] + row['Currency_Flows_Book']) if same_sign_book else (row['Previous_EMV_Book'] + row['Open_CF_Book'])
            numerator_local = (row['EMV_Local'] - row['Previous_EMV_Local'] - row['Open_CF_Local'] - row['Close_CF_Local'] - row['Currency_Flows_Local'] + row['Income_Local'])
            numerator_book = (row['EMV_Book'] - row['Previous_EMV_Book'] - row['Open_CF_Book'] - row['Close_CF_Book'] - row['Currency_Flows_Book'] + row['Income_Book'])
        else:
            numerator_local = (row['EMV_Local'] - row['Previous_EMV_Local'] - row['Open_CF_Local'] - row['Close_CF_Local'] - row['Currency_Flows_Local'])
            denominator_local = row['Previous_EMV_Local']
            numerator_book = (row['EMV_Book'] - row['Previous_EMV_Book'] - row['Open_CF_Book'] - row['Close_CF_Book'] - row['Currency_Flows_Book'])
            denominator_book = row['Previous_EMV_Book']

        twr_local = np.nan if denominator_local == 0 else numerator_local / denominator_local
        twr_book = np.nan if denominator_book == 0 else numerator_book / denominator_book

        return twr_local, twr_book

    market_values[['TWR_Local', 'TWR_Book']] = market_values.apply(calculate_twr, axis=1, result_type='expand')

    market_values.replace([np.inf, -np.inf], np.nan, inplace=True)
    market_values['LocalToDate'] = market_values.groupby(level)['TWR_Local'].transform(lambda x: (1 + x).cumprod())
    market_values['BookToDate'] = market_values.groupby(level)['TWR_Book'].transform(lambda x: (1 + x).cumprod())
    market_values['TWR_Local_Percent'] = market_values['TWR_Local'] * 100
    market_values['TWR_Book_Percent'] = market_values['TWR_Book'] * 100
    market_values['LocalToDate_Percent'] = (market_values['LocalToDate'] - 1) * 100
    market_values['BookToDate_Percent'] = (market_values['BookToDate'] - 1) * 100
    market_values['Category_Flows_Local'] = (
                market_values['Open_CF_Local'] + market_values['Close_CF_Local'] + market_values[
            'Currency_Flows_Local'])
    market_values['Category_Flows_Book'] = (
                market_values['Open_CF_Book'] + market_values['Close_CF_Book'] + market_values['Currency_Flows_Book'])

    return market_values


def fill_blue_row_perf(ws, row_idx):
    fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    for cell in ws[row_idx]:
        cell.fill = fill


def apply_perf_formatting(excel_file_path):
    workbook = openpyxl.load_workbook(excel_file_path)
    for sheet_name in ['Detail', 'Summary']:
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet.freeze_panes = sheet['A2']
            for column in sheet.columns:
                max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0) + 2
                adjusted_width = (max_length + 2) * 1.1
                column_letter = get_column_letter(column[0].column)
                sheet.column_dimensions[column_letter].width = adjusted_width

            subtotal_column_index = 1
            for row in sheet.iter_rows(min_row=2):
                if row[subtotal_column_index - 1].value == 'Subtotal':
                    for cell in row:
                        cell.font = Font(bold=True)

            for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                if i % 2 == 0:
                    for cell in row:
                        cell.fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

            num_format = '#,##0.00'
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, int) or isinstance(cell.value, float):
                        cell.number_format = num_format
            workbook.save(excel_file_path)
    workbook.save(excel_file_path)


# Create Performance Sheets

def create_performance_sheets(journal_entries):
    levels = ['portfolio', 'asset_class', 'analyst', 'country', 'currency', 'investment', 'beta']
    for level in levels:
        performance_metrics = compute_daily_twr(journal_entries, level)

        # Create a new Excel workbook
        workbook = Workbook()
        detail_sheet = workbook.active
        detail_sheet.title = "Detail"

        # Write headers for Detail sheet
        detail_headers = ['Level', 'IBOR Date', 'EMV_Local', 'EMV_Book', 'TWR_Local']
        for col_num, header in enumerate(detail_headers, 1):
            cell = detail_sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)

        # Write performance metrics to Detail sheet
        for row_num, metric in enumerate(performance_metrics.itertuples(), 2):
            detail_sheet.cell(row=row_num, column=1, value=getattr(metric, level))
            detail_sheet.cell(row=row_num, column=2, value=metric.ibor_date)
            detail_sheet.cell(row=row_num, column=3, value=metric.EMV_Local)
            detail_sheet.cell(row=row_num, column=4, value=metric.EMV_Book)
            detail_sheet.cell(row=row_num, column=5, value=metric.TWR_Local)

        # Create Summary sheet
        summary_sheet = workbook.create_sheet(title="Summary")

        # Write headers for Summary sheet
        summary_headers = ['Level', 'Total Local', 'Total Book', 'TWR_Local']
        for col_num, header in enumerate(summary_headers, 1):
            cell = summary_sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)

        # Write summarized performance metrics to Summary sheet
        summary_metrics = performance_metrics.groupby(level).agg({
            'EMV_Local': 'sum',
            'EMV_Book': 'sum',
            'TWR_Local': 'mean'  # Adjust as needed
        }).reset_index()

        for row_num, metric in enumerate(summary_metrics.itertuples(), 2):
            summary_sheet.cell(row=row_num, column=1, value=getattr(metric, level))
            summary_sheet.cell(row=row_num, column=2, value=metric.EMV_Local)
            summary_sheet.cell(row=row_num, column=3, value=metric.EMV_Book)
            summary_sheet.cell(row=row_num, column=4, value=metric.TWR_Local)

        # Save the workbook
        workbook.save(f'performance_{level}.xlsx')

    print("Performance sheets created successfully.")

# def create_performance_sheets(journal_entries):
#     levels = ['investment', 'portfolio', 'asset_class', 'analyst', 'country', 'currency', 'beta']
#     for level in levels:
#         detail_fname = f"C:/Users/hjmne/PycharmProjects/chest/reports/Performance_{level}_Detail.xlsx"
#         summary_fname = f"C:/Users/hjmne/PycharmProjects/chest/reports/Performance_{level}_Summary.xlsx"
#         with pd.ExcelWriter(detail_fname) as detail_writer, pd.ExcelWriter(summary_fname) as summary_writer:
#             twr_result = compute_daily_twr(journal_entries, level)
#             columns_to_drop = ['EMV_Local', 'BMV_Local', 'BMV_Book', 'Open_CF_Local', 'Open_CF_Book',
#                                'Currency_Flows_Local', 'Currency_Flows_Book', 'Close_CF_Local', 'Close_CF_Book',
#                                'Income_Local', 'Investment_Changed', 'EMV_Local_Diff', 'EMV_Book_Diff',
#                                'Previous_EMV_Local', 'Previous_EMV_Book', 'TWR_Local', 'LocalToDate', 'BookToDate',
#                                'TWR_Local_Percent', 'TWR_Book_Percent', 'LocalToDate_Percent', 'Category_Flows_Local']
#             twr_result.drop(columns=columns_to_drop, inplace=True, errors='ignore')
#             twr_result.to_excel(detail_writer, sheet_name='Detail')
#             summary_df = twr_result.drop_duplicates(subset=[level], keep='last')
#             summary_df.to_excel(summary_writer, sheet_name='Summary')
#         apply_perf_formatting(detail_fname)
#         apply_perf_formatting(summary_fname)
#     print("Performance sheets created successfully.")
#

# Comparison Function

def compute_diff(je_data1, je_data2):
    cols_to_include = je_data1.columns[3:]
    return je_data1[cols_to_include].subtract(je_data2[cols_to_include])


def compute_perf_diff(workbook1_path, workbook2_path, sheets_to_compare):
    output_path = 'C:/Users/hjmne/PycharmProjects/chest/repdata/PerformanceReturnsComparison.xlsx'
    if not os.path.exists(output_path):
        wb = Workbook()
        wb.active.title = "Detail"
        wb.save(output_path)

    for sheet_name in sheets_to_compare:
        try:
            je_data1 = pd.read_excel(workbook1_path, sheet_name=sheet_name)
            je_data2 = pd.read_excel(workbook2_path, sheet_name=sheet_name)
            je_data_diff = compute_diff(je_data1, je_data2)
            with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
                book = load_workbook(output_path)
                writer.book = book
                je_data_diff.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            print(f"An error occurred while processing {sheet_name}: {str(e)}")
    print("Function Completed")


# Main Function
# def calculate_and_report_performance(portfolio_name, journal_entries):
#     """Calculate and report performance."""
#     journal_entries_df = flatten_journal_entries(journal_entries)
#     investment_master_path = 'c:/BASE_PATH/refdata/investment_master.csv'
#     coa_path = 'c:/BASE_PATH/refdata/chart_of_accounts.csv'
#     journal_entries_df = fetch_and_map_groupings(journal_entries_df, investment_master_path, coa_path)
#     create_performance_sheets(journal_entries_df)

import os
import pandas as pd

def calculate_and_report_performance(portfolio_name, journal_entries, report_directory):
    """Calculate and report performance."""
    investment_master_path = "C:/BASE_PATH/refdata/investment_master.csv"
    coa_path = "C:/BASE_PATH/refdata/chart_of_accounts.csv"

    # Fetch and map groupings
    journal_entries = fetch_and_map_groupings(journal_entries, investment_master_path, coa_path)

    # Specify the level for grouping
    level = 'portfolio'

    # Compute daily TWR
    performance_metrics = compute_daily_twr(journal_entries, level)

    # Create performance sheets
    create_performance_sheets(performance_metrics, report_directory, portfolio_name)


def create_performance_sheets(performance_metrics, report_directory, portfolio_name):
    """Create performance sheets and save to report directory."""
    detail_file_path = os.path.join(report_directory, f"{portfolio_name}_detail.xlsx")
    summary_file_path = os.path.join(report_directory, f"{portfolio_name}_summary.xlsx")

    # Creating a Pandas Excel writer
    with pd.ExcelWriter(detail_file_path, engine='xlsxwriter') as writer:
        performance_metrics.to_excel(writer, sheet_name='Detail')

    with pd.ExcelWriter(summary_file_path, engine='xlsxwriter') as writer:
        summary = performance_metrics.groupby('level').agg({
            'TWR_Local_Percent': 'mean',
            'TWR_Book_Percent': 'mean',
            'LocalToDate_Percent': 'mean',
            'BookToDate_Percent': 'mean'
        }).reset_index()
        summary.to_excel(writer, sheet_name='Summary')

    print(f"Performance details saved to {detail_file_path}")
    print(f"Performance summary saved to {summary_file_path}")

if __name__ == "__main__":
    Journals = namedtuple('Journals',
                          ['portfolio', 'investment', 'tax_date', 'ls', 'tranid', 'quantity', 'local', 'book',
                           'location', 'financial_account'])
    journal_entries = [
        Journals(portfolio='A', investment='INV1', tax_date='2022-01-01', ls='l', tranid='T1', quantity=100, local=1000,
                 book=900, location='LOC1', financial_account='FA1')]
    investment_master_path = 'path/to/investment_master.csv'
    coa_path = 'path/to/chart_of_accounts.csv'
    journal_entries_df = fetch_and_map_groupings(flatten_journal_entries(journal_entries), investment_master_path,
                                                 coa_path)

    # Call the main function
    calculate_and_report_performance('MyPortfolio', journal_entries_df)

    workbook1_path = 'C:/Users/hjmne/PycharmProjects/chest/repdata/PerformanceReturnsCurrent.xlsx'
    workbook2_path = 'C:/Users/hjmne/PycharmProjects/chest/repdata/PerformanceReturnsPrior.xlsx'
    sheets_to_compare = ['asset_class', 'country', 'currency', 'beta', 'analyst', 'portfolio', 'investment']
    compute_perf_diff(workbook1_path, workbook2_path, sheets_to_compare)
