import report
from openpyxl.styles import PatternFill
import numpy as np
from openpyxl.styles import PatternFill, Font
import openpyxl
import pandas as pd
def calculate_and_report_performance(portfolio_name, journal_entries, view_type):
    """Calculate and report performance."""
#    compute_daily_twr(journal_entries, 'portfolio', portfolio_name)
    create_performance_sheets(journal_entries, view_type)
# Given functions...

def fill_blue_row_perf(ws, row_idx):
    """Fill a specific row in a worksheet with blue color."""
    fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    for cell in ws[row_idx]:
        cell.fill = fill


def check_for_duplicate_columns(df1, df2):
    """
    Check for duplicate columns between two DataFrames and raise an error if any are found.

    Args:
    df1 (pd.DataFrame): The first DataFrame.
    df2 (pd.DataFrame): The second DataFrame.

    Raises:
    ValueError: If duplicate columns are found.
    """
    # Get columns from both DataFrames
    df1_columns = set(df1.columns)
    df2_columns = set(df2.columns)

    # Find the intersection of columns
    duplicate_columns = df1_columns.intersection(df2_columns)

    if duplicate_columns:
        raise ValueError(
            f"Duplicate columns found: {duplicate_columns}. Please ensure there are no duplicate columns before merging.")
    else:
        print("No duplicate columns found. Safe to merge.")


def apply_perf_formatting(excel_file_path):
    workbook = openpyxl.load_workbook(excel_file_path)

    for sheet_name in ['Detail', 'Summary']:
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Example: Freeze the first row
            sheet.freeze_panes = sheet['A2']

            # Example: Set column widths based on the content
            for column in sheet.columns:
                max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0) + 2
                adjusted_width = (max_length + 2) * 1.1
                column_letter = get_column_letter(column[0].column)
                sheet.column_dimensions[column_letter].width = adjusted_width

            subtotal_column_index = 1  # Adjust as per your file

            # Iterate through rows and apply bold font to subtotal rows
            for row in sheet.iter_rows(min_row=2):  # Assuming the first row is headers
                if row[subtotal_column_index - 1].value == 'Subtotal':
                    for cell in row:
                        cell.font = Font(bold=True)

            # Apply zebra striping (assuming 'fill_blue_row' is a predefined function)
            for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                if i % 2 == 0:  # Apply to every even row
                    for cell in row:
                        cell.fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

            num_format = '#,##0.00'
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, int) or isinstance(cell.value, float):
                        cell.number_format = num_format
            # Save the changes to the workbook
            workbook.save(excel_file_path)

            # Apply any other specific formatting needed for performance sheets
            # ...

    workbook.save(excel_file_path)


def style_perf_report(je_data):
    """Style the performance report using specific functions."""
    # Convert the DataFrame to an openpyxl worksheet
    wb = openpyxl.Workbook()
    ws = wb.active = 0
    for r_idx, row in enumerate(je_data.values, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Apply styles
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if i % 2 == 0:  # If the row number is even
            fill_blue_row_perf(ws, i)
        if ws.cell(row=i, column=1).value.startswith('Subtotal'):
            report.underline_row(ws, row)
            report.bold_row(ws, row)

    report.format_numbers(ws)
    report.autosize_columns(ws)

    return ws  # Return the styled worksheet

import pandas as pd

def fetch_and_map_groupings(journal_entries, investment_master_path, coa_path):
    # Load the investment master and chart of accounts data
    investment_master_df = pd.read_csv(investment_master_path)
    coa_df = pd.read_csv(coa_path)

    # Assume that 'investment' and 'SystemName' are the primary keys to map
    investment_groupings = investment_master_df[['investment', 'asset_class', 'currency', 'country', 'sector']]
    coa_groupings = coa_df[['SystemName', 'SystemType', 'Group1', 'BSGroup', 'PerformanceCategory']]

    # Map the groupings to the journal entries DataFrame
    journal_entries = journal_entries.merge(investment_groupings, on='investment', how='left')
    journal_entries = journal_entries.merge(coa_groupings, left_on='financial_account', right_on='SystemName', how='left')

    return journal_entries

# Example usage
# journal_entries_df = pd.DataFrame(journal_entries)  # Convert your journal_entries to a DataFrame
# investment_master_path = 'path/to/investment_master.csv'
# coa_path = 'path/to/chart_of_accounts.csv'

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font


def compute_opening_cash_flows_investments(investment_master, je_data, level):
    currencies = ['USD', 'AUD', 'GBP', 'EUR', 'JPY']
    valid_accounts = ['Cost']

    # Define grouping columns based on the level
    #group_by_cols = [level, 'ibor_date'] if level != 'ibor_date' else ['ibor_date']
    group_by_cols = [level, 'ibor_date']

    # For non-currencies, apply the 'Book' > 0 filter.- For currencies, just check the valid accounts.

    # Conditions for filtering rows
    condition1 = (~je_data['investment'].isin(currencies) & (je_data['book'] > 0))
    condition2 = je_data['financial_account'].isin(valid_accounts)

    # Combine conditions
    opening_flows = je_data[condition1 & condition2]

    opening_flows = opening_flows.groupby(group_by_cols).agg(
        {'local': 'sum', 'book': 'sum'}).reset_index()

    # Rename columns for clarity
    opening_flows.rename(columns={'local': 'Open_CF_Local', 'book': 'Open_CF_Book'}, inplace=True)

    return opening_flows

def compute_cash_flows_currencies(investment_master, je_data, level):
    currencies = ['USD', 'AUD', 'GBP', 'EUR', 'JPY']
    valid_accounts = ['Cost', 'Payable', 'Receivable', 'DividendsReceivable', 'DividendsPayable']

    # Define grouping columns based on the level
    group_by_cols = [level,'ibor_date'] if level != 'ibor_date' else ['ibor_date']

    # For non-currencies, apply the 'Book' > 0 filter. For currencies, just check the valid accounts.

    # Conditions for filtering rows
    condition1 = je_data['investment'].isin(currencies)
    condition2 = je_data['financial_account'].isin(valid_accounts)

    # Combine conditions
    currency_flows = je_data[condition1 & condition2]

    currency_flows = currency_flows.groupby(group_by_cols).agg(
        {'local': 'sum', 'book': 'sum'}).reset_index()

    # Rename columns for clarity
    currency_flows.rename(columns={'local': 'Currency_Flows_Local', 'book': 'Currency_Flows_Book'}, inplace=True)

    return currency_flows

def compute_closing_cash_flows_for_investments(investment_master, je_data, level):
    currencies = ['USD', 'AUD', 'GBP', 'EUR', 'JPY']
    # Condition 1
    condition1 = je_data['financial_account'].isin(['PriceGainInvestment', 'FXGainInvestment'])
    # Condition 2
    condition2 = (je_data['financial_account'] == 'Cost') & (je_data['book'] < 0) & (je_data['ls'] == 'l')
    # Condition 3
    condition3 = (je_data['financial_account'] == 'Cost') & (je_data['book'] > 0) & (je_data['ls'] == 's')
    # Condition 4
    condition4 = ~je_data['investment'].isin(currencies)

    # Combine conditions
    closing_flows = je_data[(condition1 | condition2 | condition3) & condition4]

    # Group by Investment, IBOR Date, and Tax Date
    closing_flows_agg = closing_flows.groupby([level, 'ibor_date']).agg(
        {'local': 'sum', 'book': 'sum'}).reset_index()

    # Rename columns for clarity
    closing_flows_agg.rename(columns={'local': 'Close_CF_Local', 'book': 'Close_CF_Book'}, inplace=True)

    return closing_flows_agg

def compute_income(investment_master, je_data, level):


    print(je_data['financial_account'].unique())

# Must use a bitwise | as Python evaluates the expression in aggregate if OR is used!!!
    income_entries = je_data[(je_data['financial_account'] == 'DividendReceipt') |
                         (je_data['financial_account'] == 'FXGainCurrency') |
                         (je_data['financial_account'] == 'FXGainTradeSettle') |
                         (je_data['financial_account'] == 'DividendExpense')]

    print(income_entries.head())

    # # # Group by both 'Investment' and 'IBOR Date'
    income_je_data = income_entries.groupby([level, 'ibor_date'])[['local', 'book']].sum().reset_index()

    income_je_data.rename(columns={'local': 'Income_Local', 'book': 'Income_Book'}, inplace=True)

    return income_je_data


import numpy as np
#import numpy_financial as npf

def compute_daily_twr(journal_entries, agg_level, level, include_local_currency=True):
    import pandas as pd
    coa_je_data = pd.read_csv("C:/Users/hjmne/PycharmProjects/chest/refdata/chart_of_accounts.csv")
    investment_master = pd.read_csv("C:/Users/hjmne/PycharmProjects/chest/refdata/investment_master.csv")
    for entry in journal_entries:
        if not hasattr(entry, 'financial_account'):
            print("Missing financial_account:", entry)

    def convert_to_dict(entry):
        try:
            return entry._asdict()
        except AttributeError:
            return entry  # or modify as per your requirements

    import pandas as pd
    from collections import namedtuple

    Journals = namedtuple('Journals',
                          ['portfolio', 'investment', 'tax_date', 'ls', 'tranid', 'quantity', 'local', 'book',
                           'location', 'financial_account'])



    def normalize_journal_entries(journal_entries):
        data = [entry.to_dict() for entry in journal_entries]
        return pd.DataFrame(data)

    je_data = normalize_journal_entries(journal_entries)

    # pandas has issue merging where it thinks there are dupe columns, prob because one file is a list
    # renames a bunch of fields to name_x or name_y. this causes some category names not to appear"
    # woraround is rename table 2's vars back to original names but ony the y vars!!!
#    check_for_duplicate_columns(journal_entries, investment_master)
    # Merge je_data with investment_master
    je_data = pd.merge(je_data, investment_master, left_on='investment', right_on='Ticker', how='left').drop(
        'Ticker', axis=1)
    je_data.rename(columns=lambda col: col.replace('_y', '') if col.endswith('_y') else col, inplace=True)


    # given file structures only reset y column names
    # # Rename columns to avoid conflicts
    je_data.rename(columns={
      #  'investment_x': 'investment', #dont
        'Investment_y': 'Investment',  #do
       #  'asset_class_x': 'asset_class',
        'Asset_Class_y': 'Asset_Class',
      #  'currency_x': 'currency',
        'Currency_y': 'Currency',
    #    'country_x': 'country'
        'country_y': 'country'
    }, inplace=True)


    # Filter for 'MarketVal'
    market_values = je_data[je_data['financial_account'] == 'MarketVal']

    # Rename columns
    market_values = market_values.rename(columns={'local': 'EMV_Local', 'book': "EMV_Book"})


    # # Aggregate MarketValues for multiple lots on the same IBOR date
    market_values = market_values.groupby([level, 'ibor_date']).agg({
        'EMV_Local': 'sum',
        'EMV_Book': 'sum'
    }).reset_index()

    # Apply BMV Logic- level + ibor date BMV will need to be adjusted for opening flows below
    # Convert 'ibor_date' column to datetime
    market_values['ibor_date'] = pd.to_datetime(market_values['ibor_date'])
    market_values = market_values.sort_values(by=[level, 'ibor_date'])
    # # Call the function
    # 
    market_values['BMV_Local'] = market_values.groupby(level)['EMV_Local'].shift(1)
    market_values['BMV_Book'] = market_values.groupby(level)['EMV_Book'].shift(1)

    # Call the function
    opening_flows = compute_opening_cash_flows_investments(investment_master, je_data, level)
    # market_values['ibor_date'] = pd.to_datetime(market_values['ibor_date'])
    opening_flows['ibor_date'] = pd.to_datetime(opening_flows['ibor_date'])
    market_values = pd.merge(market_values, opening_flows, on=[level, 'ibor_date'], how='left')


    # market_values['BMV_Local'] -= market_values['Open_CF_Local'].fillna(0)
    # market_values['BMV_Book'] -= market_values['Open_CF_Book'].fillna(0)

    currency_flows = compute_cash_flows_currencies(investment_master, je_data, level)
    market_values['ibor_date'] = pd.to_datetime(market_values['ibor_date'])
    currency_flows['ibor_date'] = pd.to_datetime(currency_flows['ibor_date'])
    market_values = pd.merge(market_values, currency_flows, on=[level,  'ibor_date'], how='left')

    # market_values['BMV_Local'] -= market_values['Currency_Flows_Local'].fillna(0)
    # market_values['BMV_Book'] -= market_values['Currency_Flows_Book'].fillna(0)

    closing_flows = compute_closing_cash_flows_for_investments(investment_master, je_data, level)

    market_values['ibor_date'] = pd.to_datetime(market_values['ibor_date'])
    closing_flows['ibor_date'] = pd.to_datetime(closing_flows['ibor_date'])
    market_values = pd.merge(market_values, closing_flows, on=[level, 'ibor_date'],
                               how='left').fillna(0)

    income_data = compute_income(investment_master, je_data, level)

    # Ensure both 'ibor_date' columns are of datetime64[ns] type
    market_values['ibor_date'] = pd.to_datetime(market_values['ibor_date'])
    income_data['ibor_date'] = pd.to_datetime(income_data['ibor_date'])

    market_values = pd.merge(market_values, income_data, on=[level,'ibor_date'],
                           how='left').fillna(0)

    import pandas as pd
    import numpy as np
    finalized_inputs = market_values
    # Assuming your DataFrame is named finalized_inputs and
    # there is an 'Investment_ID' column to identify changes in investment.

    # Create a column to flag where investment changes
    finalized_inputs['Investment_Changed'] = finalized_inputs[level] != finalized_inputs[
        level].shift(1)

    # Now compute the differences
    # If the investment changed, we take the current 'EMV_Local' or 'EMV_Book'
    # If not, we calculate the normal difference.
    finalized_inputs['EMV_Local_Diff'] = np.where(
        finalized_inputs['Investment_Changed'],
        finalized_inputs['EMV_Local'],
        finalized_inputs['EMV_Local'].diff()
    )

    finalized_inputs['EMV_Book_Diff'] = np.where(
        finalized_inputs['Investment_Changed'],
        finalized_inputs['EMV_Book'],
        finalized_inputs['EMV_Book'].diff()
    )

    # If you need to reach back to the previous day's MVs and set them to 0 when the investment changes,
    # you could use np.where when you select your data for the report.

    # For example, when selecting the 'EMV_Local' for a two-date report:
    finalized_inputs['Previous_EMV_Local'] = np.where(
        finalized_inputs['Investment_Changed'],
        0,  # Set to 0 if the investment changed
        finalized_inputs['EMV_Local'].shift(1)  # Otherwise, use the previous day's MV
    )
    finalized_inputs['Previous_EMV_Book'] = np.where(
        finalized_inputs['Investment_Changed'],
        0,  # Set to 0 if the investment changed
        finalized_inputs['EMV_Book'].shift(1)  # Otherwise, use the previous day's MV
    )

    # Now the Previous_EMV_Local is set to 0 for days where Investment_Changed is True,
    # and it's the previous day's MV otherwise.

    # Don't forget to fill NaN values that were generated by shift() or diff() methods
    finalized_inputs['EMV_Local_Diff'] = finalized_inputs['EMV_Local_Diff'].fillna(0)
    finalized_inputs['EMV_Book_Diff'] = finalized_inputs['EMV_Book_Diff'].fillna(0)
    finalized_inputs['Previous_EMV_Local'] = finalized_inputs['Previous_EMV_Local'].fillna(0)



    def calculate_twr(row):
        # Define the TWR calculation based on the level
        print(f"Processing row: {row}")  # This will print the entire row's information
        if row[level] != 'portfolio':
            # Determine if the signs of the previous market value and currency flows are the same
            same_sign_local = (row['Previous_EMV_Local'] >= 0) == (row['Currency_Flows_Local'] >= 0)
            same_sign_book = (row['Previous_EMV_Book'] >= 0) == (row['Currency_Flows_Book'] >= 0)

            # Update denominator based on currency flow signs
            if same_sign_local:
                denominator_local = (row['Previous_EMV_Local'] + row['Open_CF_Local'] + row['Currency_Flows_Local'])
            else:
                denominator_local = (row['Previous_EMV_Local'] + row['Open_CF_Local'])

            if same_sign_book:
                denominator_book = (row['Previous_EMV_Book'] + row['Open_CF_Book'] + row['Currency_Flows_Book'])
            else:
                denominator_book = (row['Previous_EMV_Book'] + row['Open_CF_Book'])

            # Update numerator
            numerator_local = (
                        row['EMV_Local'] - row['Previous_EMV_Local'] - row['Open_CF_Local'] - row['Close_CF_Local'] -
                        row['Currency_Flows_Local'] + row['Income_Local'])
            numerator_book = (row['EMV_Book'] - row['Previous_EMV_Book'] - row['Open_CF_Book'] - row['Close_CF_Book'] -
                              row['Currency_Flows_Book'] + row['Income_Book'])
        else:  # Portfolio level calculation
            # Update numerator and denominator for portfolio level calculation
            numerator_local = (
                    row['EMV_Local'] - row['Previous_EMV_Local'] - row['Open_CF_Local'] - row['Close_CF_Local'] -
                    row['Currency_Flows_Local'])
            denominator_local = (row['Previous_EMV_Local'])

            numerator_book = (row['EMV_Book'] - row['Previous_EMV_Book'] - row['Open_CF_Book'] - row['Close_CF_Book'] -
                              row['Currency_Flows_Book'])
            denominator_book = (row['Previous_EMV_Book'])

        # Handle potential division by zero for Local
        twr_local = np.nan if denominator_local == 0 else numerator_local / denominator_local
        # Handle potential division by zero for Book
        twr_book = np.nan if denominator_book == 0 else numerator_book / denominator_book

        return twr_local, twr_book

        # Apply the function and expand the results into two new columns

    finalized_inputs[['TWR_Local', 'TWR_Book']] = finalized_inputs.apply(calculate_twr, axis=1, result_type='expand')
    import numpy as np


    import pandas as pd
    import numpy as np
    import report

    # Drop rows with NaN or infinite values
    finalized_inputs.replace([np.inf, -np.inf], np.nan, inplace=True)
    # finalized_inputs.dropna(subset=['TWR_Local', 'TWR_Book'], inplace=True)

    # styled_ws = style_perf_report(finalized_inputs)  # Replace finalized_inputs with your DataFrame name
    # wb = styled_ws.parent  # Get the workbook object to save it
    # wb.save('styled_performance_report.xlsx')
    # Save to Excel file

    # 2. Chain link the TWRs to calculate the cumulative performance
    finalized_inputs['LocalToDate'] = finalized_inputs.groupby(level)['TWR_Local'].transform(
        lambda x: (1 + x).cumprod())
    finalized_inputs['BookToDate'] = finalized_inputs.groupby(level)['TWR_Book'].transform(
        lambda x: (1 + x).cumprod())

    # 3. Convert the TWR and chain-linked values back to percentage format for reporting
    finalized_inputs['TWR_Local_Percent'] = finalized_inputs['TWR_Local'] * 100
    finalized_inputs['TWR_Book_Percent'] = finalized_inputs['TWR_Book'] * 100
    finalized_inputs['LocalToDate_Percent'] = (finalized_inputs['LocalToDate'] - 1) * 100
    finalized_inputs['BookToDate_Percent'] = (finalized_inputs['BookToDate'] - 1) * 100

    finalized_inputs['Category_Flows_Local'] = (finalized_inputs['Open_CF_Local'] + finalized_inputs['Close_CF_Local'] + \
                                         finalized_inputs['Currency_Flows_Local'])
    finalized_inputs['Category_Flows_Book'] = finalized_inputs['Open_CF_Book'] + finalized_inputs['Close_CF_Book'] + \
                                        finalized_inputs['Currency_Flows_Book']



    # cols_to_keep = [level, 'ibor_date', 'EMV_Book', 'Category_Flows_Book', 'Income_Book',
    #                 'TWR_Book_Percent', 'BookToDate_Percent', 'Investment_Changed']
    # finalized_inputs = finalized_inputs[cols_to_keep]

    import pandas as pd
    import yfinance as yf

    # Fetch data for each index
    sp500 = yf.download('^GSPC', start='2022-01-01', end='2022-12-31')
    # Calculate daily returns in percentage for each index
    sp500['S&P 500'] = sp500['Adj Close'].pct_change() * 100
    sp500.reset_index(inplace=True)
    # Merge each index's data into your DataFrame
    finalized_inputs = pd.merge(finalized_inputs, sp500[['Date', 'S&P 500']],
                                left_on='ibor_date', right_on='Date', how='left')
    finalized_inputs.drop('Date', axis=1, inplace=True)

    nasdaq = yf.download('^IXIC', start='2022-01-01', end='2022-12-31')
    # Calculate daily returns in percentage for each index
    nasdaq['NASDAQ'] = nasdaq['Adj Close'].pct_change() * 100
    nasdaq.reset_index(inplace=True)
    # Merge each index's data into your DataFrame
    finalized_inputs = pd.merge(finalized_inputs, nasdaq[['Date', 'NASDAQ']],
                                left_on='ibor_date', right_on='Date', how='left')
    finalized_inputs.drop('Date', axis=1, inplace=True)
    
    russell2000 = yf.download('^RUT', start='2022-01-01', end='2022-12-31')
    # Calculate daily returns in percentage for each index
    russell2000['Russell 2000'] = russell2000['Adj Close'].pct_change() * 100
    russell2000.reset_index(inplace=True)
    # Merge each index's data into your DataFrame
    finalized_inputs = pd.merge(finalized_inputs, russell2000[['Date', 'Russell 2000']],
                                left_on='ibor_date', right_on='Date', how='left')
    finalized_inputs.drop('Date', axis=1, inplace=True)

    ief = yf.download('IEF', start='2022-01-01', end='2022-12-31')
    # Calculate daily returns in percentage for each index
    ief['IEF30YrTreas'] = ief['Adj Close'].pct_change() * 100
    ief.reset_index(inplace=True)
    # Merge each index's data into your DataFrame
    finalized_inputs = pd.merge(finalized_inputs, ief[['Date', 'IEF30YrTreas']],
                                left_on='ibor_date', right_on='Date', how='left')
    finalized_inputs.drop('Date', axis=1, inplace=True)

    import pandas as pd

    # Assuming 'df' is your original DataFrame and it has a column 'Investment_Changed' indicating the change

    # We find the indices where the investment changes, skipping the first record
    change_indices = finalized_inputs.index[finalized_inputs['Investment_Changed'] == True].tolist()[1:]  # Skip the first record

    # Add the last index of the DataFrame if it's not already in the list
    if finalized_inputs.index[-1] not in change_indices:
        change_indices.append(finalized_inputs.index[-1])

    # We create the summary DataFrame by selecting the rows just before the change
    # This will include the last record before a change, and the very last record of the dataset
    summary_records = [finalized_inputs.iloc[i - 1] for i in change_indices if i > 0]

    # Convert the list of Series objects into a DataFrame
    summary_finalized_inputs = pd.DataFrame(summary_records)

    # Reset index for the new summary DataFrame
    summary_finalized_inputs.reset_index(drop=True, inplace=True)

    # If you want to include only the last record of each asset_class,
    # you might want to drop duplicates keeping the last occurrence
    summary_finalized_inputs = summary_finalized_inputs.drop_duplicates(subset=[level], keep='last')

    # Now 'summary_df' contains the summary information for each category just before an investment change

    return finalized_inputs, summary_finalized_inputs

    # fnamechoice = "C:/Users/hjmne/PycharmProjects/chest/repdata/daily_twr_dump.xlsx"
    # finalized_inputs.to_excel(fnamechoice, index=False)


import pandas as pd
from itertools import product


import pandas as pd


def create_summary(journal_entries, level, type):
    # Ensure 'journal_entries' DataFrame is sorted by category and date
    journal_entries.sort_values(by=[level, 'ibor_date'], inplace=True)

    # Find the last index of each category
    last_indices = journal_entries.drop_duplicates(subset=[level], keep='last').index

    # Select the rows that are at these last indices
    summary = journal_entries.loc[last_indices]

    return summary

    import pandas as pd

def create_performance_sheets(journal_entries, view_type):
    levels = ['investment', 'portfolio', 'Asset_Class', 'Analyst', 'Country', 'Currency', 'Beta']
    more_levels = ['asset_class', 'portfolio', 'investment']

    for level in levels:
        detail_fname = f"C:/Users/hjmne/PycharmProjects/chest/reports/Performance_{level}_Detail" +view_type+".xlsx"
        summary_fname = f"C:/Users/hjmne/PycharmProjects/chest/reports/Performance_{level}_Summary"+view_type+".xlsx"

        with pd.ExcelWriter(detail_fname) as detail_writer, pd.ExcelWriter(summary_fname) as summary_writer:
            twr_result, _ = compute_daily_twr(journal_entries, level, level, False)

            if twr_result.empty:
                print(f"Warning: twr_result is empty for level {level}. Skipping.")
                continue

            columns_to_drop = ['EMV_Local', 'BMV_Local', 'BMV_Book', 'Open_CF_Local',
                               'Open_CF_Book', 'Currency_Flows_Local', 'Currency_Flows_Book',
                               'Close_CF_Local', 'Close_CF_Book', 'Income_Local',
                               'Investment_Changed', 'EMV_Local_Diff', 'EMV_Book_Diff',
                               'Previous_EMV_Local', 'Previous_EMV_Book', 'TWR_Local',
                               'LocalToDate', 'BookToDate',
                               'TWR_Local_Percent', 'TWR_Book_Percent', 'LocalToDate_Percent',
                               'Category_Flows_Local']

            twr_result.drop(columns=columns_to_drop, inplace=True, errors='ignore')

            twr_result.to_excel(detail_writer,  sheet_name='Detail')
            summary_df = create_summary(twr_result,  level, view_type)
            summary_df.to_excel(summary_writer, sheet_name='Summary')

        apply_perf_formatting(detail_fname)
        apply_perf_formatting(summary_fname)


    print("Performance sheets created successfully.")

import os
import pandas as pd
from openpyxl import Workbook

def compute_diff(je_data1, je_data2):
    # Exclude the first three columns
    cols_to_include = je_data1.columns[3:]

    return je_data1[cols_to_include].subtract(je_data2[cols_to_include])

from openpyxl import load_workbook

def compute_perf_diff(workbook1_path, workbook2_path, sheets_to_compare):
    print("Function Started")
    output_path = 'C:/Users/hjmne/PycharmProjects/chest/repdata/PerformanceReturnsComparison.xlsx'

    # List of sheet names to compare
    sheets_to_compare = ['asset_class', 'country', 'currency', 'beta', 'analyst', 'portfolio', 'investment']

    if not os.path.exists(output_path):
        print("Output file does not exist. Creating a new workbook.")
        wb = Workbook()
        wb.active.title = "Detail"
        wb.save(output_path)

    for sheet_name in sheets_to_compare:
        try:
            print(f"Processing sheet: {sheet_name}")
            print("Available sheets in workbook1:", pd.ExcelFile(workbook1_path).sheet_names)
            print("Available sheets in workbook2:", pd.ExcelFile(workbook2_path).sheet_names)

            # Load sheets from the two Excel files
            je_data1 = pd.read_excel(workbook1_path, sheet_name=sheet_name)
            je_data2 = pd.read_excel(workbook2_path, sheet_name=sheet_name)

            print(f"Size of je_data1 from sheet {sheet_name}: {je_data1.shape}")
            print(f"Size of je_data2 from sheet {sheet_name}: {je_data2.shape}")

            # Compute the difference
            je_data_diff = compute_diff(je_data1, je_data2)
            print(f"Size of je_data_diff from sheet {sheet_name}: {je_data_diff.shape}")

            # # Load the workbook for writing and remove the sheet if it exists
            # book = load_workbook(output_path)
            # if sheet_name in book.sheetnames:
            #     book.remove(book[sheet_name])

            book = load_workbook('your_file.xlsx')
            all_hidden = all(sheet.sheet_state == 'hidden' for sheet in book)
            if all_hidden:
                book[book.sheetnames[0]].sheet_state = 'visible'

            book.save('your_file.xlsx')

            # Save the difference to a new sheet in the Excel file
            with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
                writer.book = book
                je_data_diff.to_excel(writer, sheet_name=sheet_name, index=False)
                print(f'Difference for {sheet_name} saved to {output_path}')

        except Exception as e:
            print(f"An error occurred while processing {sheet_name}: {str(e)}")

    print("Function Completed")

if __name__ == "__main__":
    workbook1_path = 'C:/Users/hjmne/PycharmProjects/chest/repdata/PerformanceReturnsCurrent.xlsx'
    workbook2_path = 'C:/Users/hjmne/PycharmProjects/chest/repdata/PerformanceReturnsPrior.xlsx'
    sheets_to_compare = ['asset_class', 'country', 'currency', 'beta', 'analyst', 'portfolio', 'investment']
    compute_perf_diff(workbook1_path, workbook2_path, sheets_to_compare)

