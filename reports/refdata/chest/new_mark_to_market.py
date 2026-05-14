y
import logging
import json
import bookkeeping
from bookkeeping import Journals, BookkeepingSpace

# Set up logging
logging.basicConfig(level=logging.INFO)

# Helper function to parse datetime strings

def new_get_data_and_format(date):
    price_file = "C:/Users/hjmne/PycharmProjects/chest/refdata/price_master.csv"
    fx_file = "C:/Users/hjmne/PycharmProjects/chest/refdata/fx_master.csv"
    price_data = bookkeeping.load_price_data(price_file)
    fx_data = bookkeeping.load_fx_data(fx_file)
    formatted_date = "{}/{}/{}".format(date.month, date.day, date.year)

    return price_data, fx_data, formatted_date
def combine_journal_entries(all_journal_entries, all_marked_journal_entries):
    combined_entries = [je.__dict__ if isinstance(je, Journals) else je for je in all_journal_entries]
    combined_entries += [je.__dict__ if isinstance(je, Journals) else je for je in all_marked_journal_entries]
    return combined_entries

def new_prepare_gl_data_for_reporting(period_start, period_cutoff, fund, mark_each_day, derive_mktval, mark_tax_lots=False):
    new_combine_je_files(fund)
    journals = new_combined_file_python()
    repository = BookkeepingSpace()

    if isinstance(period_cutoff, str):
        period_cutoff = parse_datetime(period_cutoff)

    space1, all_journal_entries, marked_journal_entries = new_build_sub_ledger_from_journals(
        journals, period_start, period_cutoff, mark_each_day, derive_mktval
    )

    logging.info("All Journal Entries in Repository:\n{}".format([je for je in all_journal_entries][:5]))
    #performance.calculate_and_report_performance(fund, journals)
    return space1, all_journal_entries, marked_journal_entries

# Helper Functions
def fetch_all_files(directory):
    all_files = []
    for foldername, subfolders, filenames in os.walk(directory):
        for filename in filenames:
            if filename.endswith('.json') and 'bookkeeping' not in filename.lower():
                full_path = os.path.join(foldername, filename)
                all_files.append(full_path)
    logging.info(f"Files fetched: {len(all_files)}")  # Print the number of files fetched
    return all_files

def new_combine_je_files(fund):
    files_to_combine = fetch_all_files(f'C:/Users/hjmne/PycharmProjects/chest/funds/{fund}/periods')

    combined_journals = []
    for path in files_to_combine:
        try:
            with open(path, 'r') as f:
                try:
                    journals = json.load(f)
                except json.decoder.JSONDecodeError:
                    journals = []
                combined_journals.extend(journals)
        except FileNotFoundError:
            continue

    with open('combined_journals.json', 'w') as f:
        json.dump(combined_journals, f, default=new_extended_serializer)

    logging.info(f"Combined journals: {len(combined_journals)}")  # Print the number of combined journals
    return combined_journals

def new_combined_file_python():
    filename = "C:/Users/hjmne/PycharmProjects/chest/combined_journals.json"
    with open(filename, 'r') as f:
        data = json.load(f)

    allowed_keys = ['portfolio', 'investment', 'tax_date', 'ls', 'location', 'financial_account', 'quantity', 'local',
                    'book', 'tranid', 'transaction', 'feeder', 'tradedate', 'settledate', 'kdbegin', 'kdend',
                    'ibor_date', 'entry_type', 'running_balances', 'split_ratio']

    filtered_data = []
    for entry in data:
        new_entry = {k: entry[k] for k in allowed_keys if k in entry}
        if 'tax date' in entry:
            tax_date_value = entry['tax date']
            if isinstance(tax_date_value, str):
                new_entry['tax_date'] = parse_datetime(tax_date_value)
            elif isinstance(tax_date_value, int) and tax_date_value == 0:
                new_entry['tax_date'] = 0
            else:
                new_entry['tax_date'] = tax_date_value
        filtered_data.append(new_entry)

    datetime_fields = ['tax_date', 'tradedate', 'settledate', 'kdbegin', 'kdend', 'ibor_date']
    for entry in filtered_data:
        entry['tranid'] = int(entry.get('tranid') or 0)
        for field in datetime_fields:
            if field in entry and isinstance(entry[field], str):
                entry[field] = parse_datetime(entry[field])
            elif field == 'tax_date' and isinstance(entry[field], int) and entry[field] == 0:
                entry[field] = 0
        entry['tranid'] = int(entry.get('tranid', 0))

    journal_entries = [Journals(**entry) for entry in filtered_data]
    logging.info(f"Journal entries created: {len(journal_entries)}")  # Print the number of journal entries created
    return journal_entries

def filter_bookkeeping_space_for_assets_liabilities(bookkeeping_space):
    asset_liability_info = bookkeeping_space.get_all_asset_liability_bookkeeping_info()
    filtered_space = [
        info for info in asset_liability_info
        if info[5] in ('Cost', 'Payable', 'Receivable', 'SpotFxReceivable', 'SpotFxPayable', 'ExpensesPayable')
    ]
    return filtered_space

def new_extended_serializer(obj):
    if isinstance(obj, datetime):
        return obj.isoformat()
    elif isinstance(obj, Journals):
        return {
            "portfolio": obj.portfolio,
            "investment": obj.investment,
            "tax_date": obj.tax_date.isoformat() if isinstance(obj.tax_date, datetime) else obj.tax_date,
            "ls": obj.ls,
            "location": obj.location,
            "financial_account": obj.financial_account,
            "quantity": obj.quantity,
            "local": obj.local,
            "book": obj.book,
            "tranid": obj.tranid,
            "transaction": obj.transaction,
            "tradedate": obj.tradedate.isoformat() if isinstance(obj.tradedate, datetime) else obj.tradedate,
            "settledate": obj.settledate.isoformat() if isinstance(obj.settledate, datetime) else obj.settledate,
            "kdbegin": obj.kdbegin.isoformat() if isinstance(obj.kdbegin, datetime) else obj.kdbegin,
            "kdend": obj.kdend.isoformat() if isinstance(obj.kdend, datetime) else obj.kdend,
            "ibor_date": obj.ibor_date.isoformat() if isinstance(obj.ibor_date, datetime) else obj.ibor_date,
            "entry_type": obj.entry_type,
            "feeder": obj.feeder,
            "running_balances": obj.running_balances,
            "split_ratio": obj.split_ratio,
            "account_key": obj.account_key
        }
    raise TypeError(f"Object of type {obj.__class__.__name__} is not JSON serializable")

def new_get_all_lots_for_marking(sub_ledger):
    all_lots = []
    asset_liability_info = sub_ledger.get_all_asset_liability_bookkeeping_info()
    logging.info(f"Asset Liability Info: {asset_liability_info[:5]}")  # Log the first 5 items for verification

    if not asset_liability_info:  # Check if the list is empty
        logging.info("No asset/liability information available.")
        return all_lots

    for info in asset_liability_info:
        logging.info(f"Processing info: {info}")
        # Assuming info is a list in the correct order of fields
        if info[5] in ('Cost', 'Payable', 'Receivable', 'SpotFxReceivable', 'SpotFxPayable', 'ExpensesPayable'):
            all_lots.append([
                info[0],  # portfolio
                info[1],  # investment
                info[2],  # lotid
                info[3],  # tax_date
                info[4],  # ls
                info[5],  # location
                info[7],  # quantity
                info[8],  # local
                info[9],   # book
                info[10]   #notional
            ])
    logging.info(f"Extracted Lots: {all_lots[:5]}")  # Log the first 5 extracted lots for verification
    return all_lots

def new_calculate_marks(tax_lots, sub_ledger, price_data, fx_data, date):
    mark_date = parse_datetime(date)
    if tax_lots.empty:
        logging.info("The input DataFrame is empty. No processing will be done.")
        return []

    required_columns = {'portfolio', 'investment', 'lotid','tax_date', 'ls', 'location', 'quantity', 'local', 'book'}
    if not required_columns.issubset(tax_lots.columns):
        raise KeyError(f"DataFrame is missing one or more required columns: {required_columns - set(tax_lots.columns)}")

    mark_records = []
    current_investment = None
    subspace = None

    for _, row in tax_lots.iterrows():
        if row['investment'] != current_investment:
            current_investment = row['investment']
            subspace = sub_ledger.asset_liability_repository.get_position_space(current_investment)

        record_key = (row['portfolio'], row['investment'],  row['lotid'], row['tax_date'], row['ls'], row['location'])
        investment_type = subspace.get_information_field("AIF", "investment_type") if subspace else None
        pricing_factor = float(subspace.get_information_field("AIF", "pricing_factor") if subspace else 1)

        ticker = row['investment']
        quantity = row['quantity']
        local = row['local']
        book = row['book']
        notional = row['notional']


        if isinstance(mark_date, datetime):
            formatted_date = f"{mark_date.month}/{mark_date.day}/{mark_date.year}"
        else:
            formatted_date = mark_date.strftime('%Y-%m-%d')

        if not pricing_factor:
            pricing_factor = 1

        data_for_date = price_data.get(formatted_date, {})
        price_data_filtered = data_for_date.get(ticker, {})
        price = price_data_filtered.get('price', 6.78787) * float(pricing_factor)  # Set price to default if None
        currency = price_data_filtered.get('currency', "USD")  # Set currency to "USD" if None

        fx_rate = fx_data.get(formatted_date, {}).get(currency, 1)

        logging.debug(f"Date: {formatted_date}, Investment: {row['investment']}, Price: {price}, Currency: {currency}, FX Rate: {fx_rate}")

        mkt_val_local = price * row['quantity'] - row['notional']
        mkt_val_book = mkt_val_local * fx_rate
        pgain_local = mkt_val_local - row['local'] if investment_type != "FUTURE" else mkt_val_local
        pgain_book = pgain_local * fx_rate
        totgain_book = mkt_val_book - row['book'] if investment_type != "FUTURE" else 0

        record_to_add = (
            row['quantity'], row['local'], row['book'], mkt_val_local, mkt_val_book, pgain_local, pgain_book, totgain_book,
            0, mark_date, investment_type)  # fx_gain is set to 0 and not calculated here
        mark_records.append((record_key, record_to_add))

        logging.debug(f"Calculated Mark Record: {record_to_add}")

    logging.info("Calculated Mark Records:\n{}".format(mark_records))
    return mark_records

def fetch_unrealized_gl(subspace, price_stat_key, fx_stat_key):
    try:
        price_local_query = subspace.query_balance(
            account_type=price_stat_key[-1],
            portfolio=price_stat_key[0],
            investment=price_stat_key[1],
            lotid = price_stat_key[2],
            tax_date=price_stat_key[3],
            ls=price_stat_key[4],
            location=price_stat_key[5]
        )
        price_book_query = subspace.query_balance(
            account_type=price_stat_key[-1],
            portfolio=price_stat_key[0],
            investment=price_stat_key[1],
            lotid=price_stat_key[2],
            tax_date=price_stat_key[3],
            ls=price_stat_key[4],
            location=price_stat_key[5]
        )
        fx_book_query = subspace.query_balance(
            account_type=fx_stat_key[-1],
            portfolio=fx_stat_key[0],
            investment=fx_stat_key[1],
            lotid=price_stat_key[2],
            tax_date=fx_stat_key[3],
            ls=fx_stat_key[4],
            location=fx_stat_key[5]
        )

        price_local = price_local_query[0][1] if price_local_query else 0
        price_book = price_book_query[0][2] if price_book_query else 0
        fx_book = fx_book_query[0][2] if fx_*book_query else 0

        return price_local, price_book, fx_book
    except Exception as e:
        logging.error(f"Error fetching unrealized GL balances: {e}")
        return 0, 0, 0

def new_post_accounting_marks(records_to_mark, date, derive_mktval, space):
    mark_date = parse_datetime(date.strftime('%Y-%m-%d %H:%M:%S'))  # Ensure date is a string
    if records_to_mark is None:
        return []

    journal_entries = []

    for record in records_to_mark:
        account_key, (
            quantity, local, book, mvlocal, mvbook, pgain_local, pgain_book, totgain_book, _, date,
            investment_type) = record
        curr_fx = mvbook / mvlocal if mvlocal != 0 else 1

        # Construct the full account keys with the financial accounts
        price_stat_key = account_key + ('UnrealGLPriceStat',)
        fx_stat_key = account_key + ('UnrealFXStat',)

        # Query the correct subspace
        subspace = space.asset_liability_repository.investment_spaces_library[account_key[1]]

        # Fetch balances for UnrealGLPriceStat and UnrealFXStat
        prev_unrealized_price_local, prev_unrealized_price_book, prev_unrealized_fx_book = fetch_unrealized_gl(
            subspace, price_stat_key, fx_stat_key)

        # Calculate the changes in unrealized gains/losses
        chg_unrealized_price_local = mvlocal - local - prev_unrealized_price_local
        chg_unrealized_price_book = chg_unrealized_price_local * curr_fx
        chg_unrealized_fx_gain = totgain_book - pgain_book - prev_unrealized_fx_book

        # Apply the deltas to the previous unrealized balances to get the current ones
        curr_unrealized_price_local = prev_unrealized_price_local + chg_unrealized_price_local
        curr_unrealized_price_book = prev_unrealized_price_book + chg_unrealized_price_book
        curr_unrealized_fx_book = prev_unrealized_fx_book + chg_unrealized_fx_gain

        net_local = chg_unrealized_price_local
        net_book = chg_unrealized_price_book
        net_book_fx = chg_unrealized_fx_gain

        net_local_stat = curr_unrealized_price_local
        net_book_stat = curr_unrealized_price_book
        net_book_fx_stat = curr_unrealized_fx_book

        tdate = date
        # Set tradedate to end of the day
        tdate = date.replace(hour=23, minute=00, second=00)

        if net_local or net_book:
            markA = {
                "portfolio": account_key[0],
                "investment": account_key[1],
                "lotid": account_key[2],
                "tax_date": account_key[3],
                "ls": account_key[4],
                "location": account_key[5],
                "financial_account": "UnrealGLPrice",
                "quantity": 0,
                "local": net_local,
                "book": net_book,
                "notional": None,
                "oface": None,
                "tranid": None,
                "transaction": "AcctMark",
                "tradedate": tdate,
                "settledate": tdate,
                "kdbegin": tdate,
                "kdend": tdate,
                "ibor_date": tdate,  # Assigning ibor_date to tdate
                "entry_type": "Asset/Liability"
            }
            journal_entries.append(markA)

            try:
                journal_entry = Journals(**markA)
                space.post_journal_entry(journal_entry)  # Post the journal entry to the space
                logging.info(f"Posted journal entry: {markA}")
            except AttributeError as e:
                logging.error(f"Error in posting MarkA: {e}, {markA}")
                raise

            markB = {
                "portfolio": account_key[0],
                "investment": account_key[1],
                "lotid": account_key[2],
                "tax_date": account_key[3],
                "ls": account_key[4],
                "location": account_key[5],
                "financial_account": "UnrealPriceRevExp",
                "quantity": 0,
                "local": -net_local,
                "book": -net_book,
                "notional": None,
                "oface": None,
                "tranid": None,
                "transaction": "AcctMark",
                "tradedate": tdate,
                "settledate": tdate,
                "kdbegin": tdate,
                "kdend": tdate,
                "ibor_date": tdate,  # Assigning ibor_date to tdate
                "entry_type": "Revenue/Expense/Capital"
            }
            journal_entries.append(markB)

            try:
                journal_entry = Journals(**markB)
                space.post_journal_entry(journal_entry)  # Post the journal entry to the space
                logging.info(f"Posted journal entry: {markB}")
            except AttributeError as e:
                logging.error(f"Error in posting MarkB: {e}, {markB}")
                raise

        if chg_unrealized_fx_gain:
            markFX = {
                "portfolio": account_key[0],
                "investment": account_key[1],
                "lotid": account_key[2],
                "tax_date": account_key[3],
                "ls": account_key[4],
                "location": account_key[5],
                "financial_account": "UnrealGLFX",
                "quantity": 0,
                "local": 0,
                "book": net_book_fx,
                "notional": None,
                "oface": None,
                "tranid": None,
                "transaction": "AcctMark",
                "tradedate": tdate,
                "settledate": tdate,
                "kdbegin": tdate,
                "kdend": tdate,
                "ibor_date": tdate,  # Assigning ibor_date to tdate
                "entry_type": "Asset/Liability"
            }
            journal_entries.append(markFX)

            try:
                journal_entry = Journals(**markFX)
                space.post_journal_entry(journal_entry)  # Post the journal entry to the space
                logging.info(f"Posted journal entry: {markFX}")
            except AttributeError as e:
                logging.error(f"Error in posting MarkFX: {e}, {markFX}")
                raise

            markFXOffset = {
                "portfolio": account_key[0],
                "investment": account_key[1],
                "lotid": account_key[2],
                "tax_date": account_key[3],
                "ls": account_key[4],
                "location": account_key[5],
                "financial_account": "UnrealFXRevExp",
                "quantity": 0,
                "local": 0,
                "book": -net_book_fx,
                "notional": None,
                "oface": None,
                "tranid": None,
                "transaction": "AcctMark",
                "tradedate": tdate,
                "settledate": tdate,
                "kdbegin": tdate,
                "kdend": tdate,
                "ibor_date": tdate,  # Assigning ibor_date to tdate
                "entry_type": "Revenue/Expense/Capital"
            }
            journal_entries.append(markFXOffset)

            try:
                journal_entry = Journals(**markFXOffset)
                space.post_journal_entry(journal_entry)  # Post the journal entry to the space
                logging.info(f"Posted journal entry: {markFXOffset}")
            except AttributeError as e:
                logging.error(f"Error in posting MarkFXOffset: {e}, {markFXOffset}")
                raise

        if net_book and derive_mktval:
            markP = {
                "portfolio": account_key[0],
                "investment": account_key[1],
                "lotid": account_key[2],
                "tax_date": account_key[3],
                "ls": account_key[4],
                "location": account_key[5],
                "financial_account": "MktVal",
                "quantity": quantity,
                "local": mvlocal,
                "book": mvbook,
                "notional": None,
                "oface": None,
                "tranid": None,
                "transaction": "PerfMark",
                "tradedate": tdate,
                "settledate": tdate,
                "kdbegin": tdate,
                "kdend": tdate,
                "ibor_date": tdate,  # Assigning ibor_date to tdate
                "entry_type": "Asset/Liability"
            }
            journal_entries.append(markP)

            try:
                journal_entry = Journals(**markP)
                space.post_journal_entry(journal_entry)  # Post the journal entry to the space
                logging.info(f"Posted journal entry: {markP}")
            except AttributeError as e:
                logging.error(f"Error in posting MarkP: {e}, {markP}")
                raise

            markPRE = {
                "portfolio": account_key[0],
                "investment": account_key[1],
                "lotid": account_key[2],
                "tax_date": account_key[3],
                "ls": account_key[4],
                "location": account_key[5],
                "financial_account": "MktValRE",
                "quantity": 0,
                "local": -mvlocal,
                "book": -mvbook,
                "notional": None,
                "oface": None,
                "tranid": None,
                "transaction": "AcctMark",
                "tradedate": tdate,
                "settledate": tdate,
                "kdbegin": tdate,
                "kdend": tdate,
                "ibor_date": tdate,  # Assigning ibor_date to tdate
                "entry_type": "Revenue/Expense/Capital"
            }
            journal_entries.append(markPRE)

            try:
                journal_entry = Journals(**markPRE)
                space.post_journal_entry(journal_entry)  # Post the journal entry to the space
                logging.info(f"Posted journal entry: {markPRE}")
            except AttributeError as e:
                logging.error(f"Error in posting MarkPRE: {e}, {markPRE}")
                raise

        if net_local_stat or net_book_stat:
            markStatPrice = {
                "portfolio": account_key[0],
                "investment": account_key[1],
                "lotid": account_key[2],
                "tax_date": tdate,  # Use current timestamp for tax_date
                "ls": account_key[4],
                "location": account_key[5],
                "financial_account": "UnrealGLPriceStat",
                "quantity": quantity,
                "local": curr_unrealized_price_local,  # Use current unrealized price local
                "book": curr_unrealized_price_book,    # Use current unrealized price book
                "notional": None,
                "oface": None,
                "tranid": None,
                "transaction": "StatMark",
                "tradedate": tdate,
                "settledate": tdate,
                "kdbegin": tdate,
                "kdend": tdate,
                "ibor_date": tdate,  # Assigning ibor_date to tdate
                "entry_type": "Asset/Liability"
            }
            journal_entries.append(markStatPrice)

            try:
                journal_entry = Journals(**markStatPrice)
                space.post_journal_entry(journal_entry)  # Post the journal entry to the space
                logging.info(f"Posted journal entry: {markStatPrice}")
            except AttributeError as e:
                logging.error(f"Error in posting MarkStatPrice: {e}, {markStatPrice}")
                raise

        if net_book_stat:
            markStatFX = {
                "portfolio": account_key[0],
                "investment": account_key[1],
                "lotid": account_key[2],
                "tax_date": tdate,  # Use current timestamp for tax_date
                "ls": account_key[4],
                "location": account_key[5],
                "financial_account": "UnrealFXGLStat",
                "quantity": quantity,
                "local": 0,
                "book": curr_unrealized_fx_book,  # Use current unrealized fx book
                "notional": None,
                "oface": None,
                "tranid": None,
                "transaction": "StatMark",
                "tradedate": tdate,
                "settledate": tdate,
                "kdbegin": tdate,
                "kdend": tdate,
                "ibor_date": tdate,  # Assigning ibor_date to tdate
                "entry_type": "Asset/Liability"
            }
            journal_entries.append(markStatFX)

            try:
                journal_entry = Journals(**markStatFX)
                space.post_journal_entry(journal_entry)  # Post the journal entry to the space
                logging.info(f"Posted journal entry: {markStatFX}")
            except AttributeError as e:
                logging.error(f"Error in posting MarkStatFX: {e}, {markStatFX}")
                raise

    return journal_entries

def new_mark_positions(space1, date, derive_mktval):
    current_unrealized_gl = {"price_local": {}, "price_book": {}, "fx_book": {}}  # Store current period's unrealized GL numbers
    journal_entries = []

    logging.info(f"Marking positions for {date}")

    tax_lots = new_get_all_lots_for_marking(space1)
    if tax_lots:
        logging.info(f"Number of tax lots for marking: {len(tax_lots)}")
        f1, f2, *rest = new_get_data_and_format(date)
        lot_df = pd.DataFrame(tax_lots, columns=['portfolio', 'investment', 'lotid', 'tax_date', 'ls', 'location', 'quantity', 'local', 'book', 'notional'])
        marked_records = new_calculate_marks(lot_df, space1, f1, f2, date.strftime('%Y-%m-%d'))
        journal_entries.extend(new_post_accounting_marks(marked_records, date, derive_mktval, space1))

        for entry in journal_entries:
            if entry['financial_account'] == 'UnrealGLPrice':
                key = (entry['portfolio'], entry['investment'], entry['lotid'], entry['tax_date'], entry['ls'], entry['location'])
                current_unrealized_gl["price_local"][key] = entry['local']
                current_unrealized_gl["price_book"][key] = entry['book']
            elif entry['financial_account'] == 'UnrealFXRevExp':
                key = (entry['portfolio'], entry['investment'], entry['lotid'], entry['tax_date'], entry['ls'], entry['location'])
                current_unrealized_gl["fx_book"][key] = entry['book']

    logging.info(f"Finished marking positions for {date}")

    return current_unrealized_gl, journal_entries

def new_build_sub_ledger_from_journals(journals, start_date, end_date, mark_each_day, derive_mktval):
    space1 = BookkeepingSpace()
   # period_start = start_date - BDay(1)
    rollup_date = start_date
    all_journal_entries = []
    all_marked_journal_entries = []

    journals.sort(key=lambda je: je.ibor_date)

    # Initialize rollup_date to the earliest ibor_date if it's after the start_date
    if journals and rollup_date < journals[0].ibor_date:
        rollup_date = journals[0].ibor_date - BDay(1)
        rollup_date = rollup_date.replace(hour=23, minute=59, second=59)

    # Process journals up to period start - 1 BD
    while rollup_date < start_date:
        logging.info(f"Processing rollup date: {rollup_date}")

        for je in journals:
            if je.ibor_date <= rollup_date and je not in all_journal_entries:
                space1.post_journal_entry(je)
                all_journal_entries.append(je)

        rollup_date += BDay(1)

    # Perform marking at the period start
    logging.info(f"Marking positions at period start: {start_date}")
    tax_lots = new_get_all_lots_for_marking(space1)
    if tax_lots:
        current_unrealized_gl, marked_entries = new_mark_positions(space1, start_date, derive_mktval)
        all_marked_journal_entries.extend(marked_entries)

    # Process journals from period start to end_date
    rollup_date = start_date
    while rollup_date <= end_date:
        logging.info(f"Processing rollup date: {rollup_date}")

        for je in journals:
            if je.ibor_date <= rollup_date and je not in all_journal_entries:
                space1.post_journal_entry(je)
                all_journal_entries.append(je)

        if mark_each_day or rollup_date == end_date:
            tax_lots = new_get_all_lots_for_marking(space1)
            if tax_lots:
                current_unrealized_gl, marked_entries = new_mark_positions(space1, rollup_date, derive_mktval)
                all_marked_journal_entries.extend(marked_entries)

        rollup_date += BDay(1)
        if not mark_each_day and rollup_date < end_date:
            rollup_date = end_date

    return space1, all_journal_entries, all_marked_journal_entries

from pandas.tseries.offsets import BDay



def fetch_regular_journal_entries(start_period, end_period, space1):
    journal_entries = space1.journal_entries
    regular_journal_entries = [je.__dict__ for je in journal_entries if start_period <= je.tradedate <= end_period]
    return regular_journal_entries

def transform_journal_entry(entry):
    transformed_entry = {
        'portfolio': entry.get('portfolio'),
        'investment': entry.get('investment'),
        'lotid': entry.get('lotid'),
        'tax_date': entry.get('tax_date'),
        'ls': entry.get('ls'),
        'tranid': entry.get('tranid'),
        'quantity': entry.get('quantity'),
        'local': entry.get('local'),
        'book': entry.get('book'),
        'location': entry.get('location'),
        'financial_account': entry.get('financial_account')
    }
    return Journals(**transformed_entry)

def transform_journal_entries(journal_entries):
    transformed_entries = [transform_journal_entry(entry) if isinstance(entry, dict) else entry for entry in journal_entries]
    return transformed_entries

account_mapping = {
    'Cost': (0, 1, 2),
    'UnrealGLPrice': (0, 3, 4),
    'UnrealGLFX': (0, 'skip', 5),
    'PriceGainInvestment': (0, 6, 7),
    'FXGainInvestment': (0, 6, 7),
    'DividendReceipt': (0, 8, 9),
    'ContributedCost': (10, 11, 12),
    'Receivable': (0, 1, 2),
    'Payable': (0, 1, 2),
    'DividendsReceivable': (0, 1, 2),
    'FXGainTradeSettle': (0, 8, 9),
    'FXGainCurrency': (0, 8, 9),
}

def assign_values(entry, mapping):
    assigned_values = [0] * 13  # Initialize a list with 13 columns set to 0
    values = [entry['quantity'], entry['local'], entry['book']]

    for i, map_index in enumerate(mapping):
        if map_index != 'skip':
            assigned_values[map_index] += values[i]

    return assigned_values

from datetime import datetime
import pandas as pd
from pandas.tseries.offsets import BDay
import logging
import openpyxl
from openpyxl.styles import PatternFill, Font
import os

def generate_comprehensive_report_and_pivot(start_period, end_period, fund, output_file, derive_mktval, mark_each_day=False):
    start_period = pd.to_datetime(start_period)
    end_period = pd.to_datetime(end_period)
    start_period_adjusted = start_period - BDay(1)
    start_period_adjusted = start_period_adjusted.replace(hour=23, minute=59, second=59)

    space1, all_journal_entries, marked_journal_entries = new_prepare_gl_data_for_reporting(
        start_period_adjusted, end_period, fund, mark_each_day, derive_mktval
    )

    combined_entries = combine_journal_entries(all_journal_entries, marked_journal_entries)
    combined_journal_entries_df = pd.DataFrame(combined_entries)

    required_columns = {'investment', 'portfolio', 'tax_date', 'ls', 'location', 'quantity', 'local', 'book'}
    if not required_columns.issubset(combined_journal_entries_df.columns):
        raise KeyError(f"DataFrame is missing one or more required columns: {required_columns - set(combined_journal_entries_df.columns)}")

    combined_journal_entries_df = combined_journal_entries_df.sort_values(by=['investment', 'ibor_date', 'financial_account'])
    combined_journal_entries_df.to_csv('journal_entries_report.csv', index=False)

    try:
        logging.info("DataFrame info:")
        logging.info(combined_journal_entries_df.info())
        logging.info("First few rows of the DataFrame:")
        logging.info(combined_journal_entries_df.head())

        grouping_columns = ['portfolio', 'investment', 'lotid', 'ibor_date', 'ls', 'location', 'financial_account', 'transaction']
        for column in grouping_columns:
            unique_values = combined_journal_entries_df[column].unique()
            logging.info(f"Unique values in {column}: {unique_values}")

        logging.info("Checking for missing values in grouping columns:")
        logging.info(combined_journal_entries_df[grouping_columns].isnull().sum())

        combined_journal_entries_df = combined_journal_entries_df[
            ~combined_journal_entries_df['financial_account'].isin(['UnrealPriceRevExp', 'UnrealFXRevExp',
            'UnrealFXGLStat', 'UnrealGLPriceStat'])]

        combined_journal_entries_df['quantity'] = combined_journal_entries_df['quantity'].astype(float)
        combined_journal_entries_df['local'] = combined_journal_entries_df['local'].astype(float)
        combined_journal_entries_df['book'] = combined_journal_entries_df['book'].astype(float)

        combined_journal_entries_df['index'] = combined_journal_entries_df.index
        grouping_columns.append('index')
        grouped = combined_journal_entries_df.groupby(grouping_columns)
        logging.info(f"Number of groups formed: {len(grouped)}")

        rows = []

        for name, group in grouped:
            portfolio, investment, lotid, ibor_date, ls, location, financial_account, transaction, index = name
            logging.info(f"Processing group: {name} with {len(group)} records")

            row = {
                'portfolio': portfolio,
                'investment': investment,
                'lotid': lotid,
                'ibor_date': ibor_date,
                'ls': ls,
                'location': location,
                'financial_account': financial_account,
                'transaction': transaction,
                'quantity': 0,
                'local': 0,
                'book': 0,
                'unrealgllocal': 0,
                'unrealglbook': 0,
                'unrealfxbook': 0,
                'realizedlocal': 0,
                'realizedbook': 0,
                'incomelocal': 0,
                'incomebook': 0,
                'capitalshares': 0,
                'capitallocal': 0,
                'capitalbook': 0,
                'tranid': group['tranid'].iloc[0],
                'entry_type': group['entry_type'].iloc[0]
            }

            # Add logging before the loop starts
            logging.info(f"Starting to process entries for group: {name}")

            try:
                for _, entry in group.iterrows():
                    logging.info(f"Processing entry: {entry.to_dict()}")

                    if financial_account in account_mapping:
                        assigned_values = assign_values(entry, account_mapping[financial_account])
                        row['quantity'] += assigned_values[0]
                        row['local'] += assigned_values[1]
                        row['book'] += assigned_values[2]
                        row['unrealgllocal'] += assigned_values[3]
                        row['unrealglbook'] += assigned_values[4]
                        row['unrealfxbook'] += assigned_values[5]
                        row['realizedlocal'] += assigned_values[6]
                        row['realizedbook'] += assigned_values[7]
                        row['incomelocal'] += assigned_values[8]
                        row['incomebook'] += assigned_values[9]
                        row['capitalshares'] += assigned_values[10]
                        row['capitallocal'] += assigned_values[11]
                        row['capitalbook'] += assigned_values[12]
                    else:
                        row['local'] += entry['local']
                        row['book'] += entry['book']
            except Exception as e:
                logging.error(f"Error processing entries for group {name}: {e}")

            logging.info(f"Finished processing entries for group: {name}")
            logging.info(f"Row to be appended: {row}")
            rows.append(row)

        logging.info("Finished processing all groups.")
        logging.info("Creating DataFrame from rows.")

        pivot_table_df = pd.DataFrame(rows)

        # Debugging: Check DataFrame before saving
        logging.info("Pivot table DataFrame:")
        logging.info(pivot_table_df)

        book_sum = pivot_table_df['book'].sum()
        unrealglbook_sum = pivot_table_df['unrealglbook'].sum()
        unrealfxbook_sum = -pivot_table_df['unrealfxbook'].sum()
        realizedbook_sum = -pivot_table_df['realizedbook'].sum()
        incomebook_sum = -pivot_table_df['incomebook'].sum()
        capitalbook_sum = -pivot_table_df['capitalbook'].sum()
        mkt_val_book = book_sum + unrealglbook_sum + unrealfxbook_sum
        comprehensive_sum = capitalbook_sum + realizedbook_sum + incomebook_sum + unrealfxbook_sum + unrealglbook_sum

        summary_data = pd.DataFrame([{
            'portfolio': 'Summary Book Value',
            'investment': '',
            'lotid': '',
            'ibor_date': '',
            'ls': '',
            'location': '',
            'financial_account': '',
            'transaction': '',
            'quantity': '',
            'local': '',
            'book': book_sum,
            'unrealgllocal': '',
            'unrealglbook': unrealglbook_sum,
            'unrealfxbook': unrealfxbook_sum,
            'realizedlocal': '',
            'realizedbook': realizedbook_sum,
            'incomelocal': '',
            'incomebook': incomebook_sum,
            'capitalshares': '',
            'capitallocal': '',
            'capitalbook': capitalbook_sum,
            'tranid': '',
            'entry_type': ''
        }, {
            'portfolio': 'Summary MarketVal Book Derived from Book Cost & Unrealized',
            'investment': '',
            'lotid': '',
            'ibor_date': '',
            'ls': '',
            'location': '',
            'financial_account': '',
            'transaction': '',
            'quantity': '',
            'local': '',
            'book': mkt_val_book,
            'unrealgllocal': '',
            'unrealglbook': '',
            'unrealfxbook': '',
            'realizedlocal': '',
            'realizedbook': '',
            'incomelocal': '',
            'incomebook': '',
            'capitalshares': '',
            'capitallocal': '',
            'capitalbook': '',
            'tranid': '',
            'entry_type': 'Summary MktValBook Derived from Capital & Net Earnings'
        }, {
            'portfolio': 'Summary',
            'investment': '',
            'lotid': '',
            'ibor_date': '',
            'ls': '',
            'location': '',
            'financial_account': '',
            'transaction': '',
            'quantity': '',
            'local': '',
            'book': comprehensive_sum,
            'unrealgllocal': '',
            'unrealglbook': '',
            'unrealfxbook': '',
            'realizedlocal': '',
            'realizedbook': '',
            'incomelocal': '',
            'incomebook': '',
            'capitalshares': '',
            'capitallocal': '',
            'capitalbook': '',
            'tranid': '',
            'entry_type': 'Comprehensive Sum'
        }])

        pivot_table_df = pd.concat([pivot_table_df, summary_data], ignore_index=True)

        logging.info("Saving pivot table to Excel file...")
        pivot_table_df.to_excel(output_file, index=False)

        wb = openpyxl.load_workbook(output_file)
        ws = wb.active

        fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        bold_font = Font(bold=True)

        summary_rows = ['Summary Book Value', 'Summary MarketVal Book Derived from Book Cost & Unrealized',
                        'Comprehensive Sum']

        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'

                if i % 2 == 0:
                    cell.fill = fill

                if cell.value in summary_rows:
                    for summary_cell in row:
                        summary_cell.font = bold_font

        wb.save(output_file)
        logging.info(f"Pivot table saved with formatting. Number of records in pivot table: {len(pivot_table_df)}")

        #return pivot_table_df

    except Exception as e:
        logging.error(f"An error occurred: {e}")
        return None

# Example usage
start_period_str = '2022-01-12'
end_period_str = '2022-03-31'
fund = "XYZMutualFund"
mark_each_day = False
derive_mktval = False

output_file = 'BASE_PATH/reports/investment_pivot_report.xlsx'
report = generate_comprehensive_report_and_pivot(start_period_str, end_period_str, fund, output_file, derive_mktval, mark_each_day)

