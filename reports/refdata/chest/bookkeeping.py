
import bond_calc
import utilities
import mark_to_market
from collections import OrderedDict
import logging
import datetime
from collections import OrderedDict



import pandas as pd
from collections import OrderedDict



# In EventScheduler class
class EventScheduler:
    def __init__(self, sub_ledger):
        self.events = OrderedDict()
        self.sub_ledger = sub_ledger
        self.event_type_precedence = {
            'buy_equity': 1075, 'sell_equity': 1111, 'short_equity': 1111, 'cover_equity': 1111,
            'buy_bond': 1075, 'sell_bond': 1111, 'short_bond': 1111, 'cover_bond': 1111,
            'buy_future': 1075, 'sell_future': 1111, 'short_future': 1111, 'cover_future': 1111,
            'open_swap': 1075, 'reset_swap': 1111, 'spot_fx': 1111, 'forward_fx': 1111,
            'deposit_currency': 1090, 'withdraw_currency': 1090, 'split_equity': 2111,
            'dividend_equity': 2122, 'option_exercise_open': 1076, 'option_exercise_close': 1077,
            'option_assign_open': 1076, 'option_assign_close': 1077, 'mark_event': 9000,
            'perf_mark': 9000, 'allocate': 9500, 'settle_bond_flows_in': 1050, 'settle_bond_flows_out': 1050,
            'settle_single_flow_in': 1050, 'settle_single_flow_out': 1050, 'settle_multiple_flows_in_out': 1050,
            'settle_pay_rec_by_tranid': 1050, 'accrue_interest': 1060
        }


    def schedule_event(self, tradedate, event_function, *args, event_type=None):
        if not callable(event_function):
            raise ValueError("event_function must be callable")

        event_id = len(self.events) + 1  # Generate a unique identifier for the event
        self.events[event_id] = {
            'tradedate': tradedate,
            'event_function': event_function,
            'args': args,
            'event_type': event_type
        }

    def sort_events(self, reverse=False):
        sorted_items = sorted(
            self.events.items(),
            key=lambda x: (self.event_type_precedence.get(x[1]['event_type'], float('inf')), x[1]['tradedate']),
            reverse=reverse
        )
        self.events = OrderedDict(sorted_items)

    def run_next_event(self):
        while self.events:
            try:
                next_event_key, next_event_value = self.events.popitem(last=False)
                current_date = next_event_value['tradedate']
            #    print(f"Processing event {next_event_key} on date {current_date}")

                event_function = next_event_value['event_function']
                event_args = next_event_value['args']

                if not callable(event_function):
                    raise TypeError(f"Expected a callable event_function, but got {type(event_function).__name__}")

                event_function(*event_args)
           #     print(f"Processed event function {event_function} with args {event_args}")

            except Exception as e:
                logging.error(f"Error processing event {next_event_key}: {e}")
                print(f"Error processing event {next_event_key}: {e}")
                print(f"Event arguments: {event_args}")  # Print the event arguments

                # Handle error appropriately (e.g., continue, stop, retry)

        return


def accrue_interest(sub_ledger, af, portfolio, investment, current_date, fx_rates_df):
    investment_cache = {}

    # Cache investment types and subspaces to minimize redundant accesses
    for investment in sub_ledger.asset_liability_repository.investment_spaces_library.keys():
        if investment not in investment_cache:
            subspace = sub_ledger.asset_liability_repository.get_position_space(investment)
            investment_type = subspace.get_information_field("AIF", "Investment_Type")
            if investment_type:
                investment_type = investment_type.strip()
                investment_cache[investment] = (investment_type, subspace)
            else:
                #print(f"Investment type for {investment} is None, skipping this investment.")
                continue

    # Process only BOND investments
    for investment, (investment_type, subspace) in investment_cache.items():
        if investment_type == "BOND":
            print(f"Processing BOND investment: {investment}")

            net_positions = af.calculate_net_positions(portfolio=portfolio, investment=investment, date=current_date)
            print(f"Net Positions for {investment}: {net_positions}")

            for location, positions in net_positions.items():
                for position_type, qty in positions.items():
                    ls = 'l' if 'long' in position_type else 's'
                    if ls == 's':
                        qty = -qty

                    issue_date_str = subspace.get_information_field('AIF', 'issue_date')
                    first_coupon_date_str = subspace.get_information_field('AIF', 'first_coupon_date')
                    next_to_last_coupon_date_str = subspace.get_information_field('AIF', 'next_to_last_coupon_date')
                    maturity_date_str = subspace.get_information_field('AIF', 'maturity_date')
                    coupon_rate = float(subspace.get_information_field('AIF', 'coupon_rate'))
                    day_count_convention = 'actual/365'
                    payment_frequency = 'semi-annual'
                    semi_split = "C"

                    issue_date = datetime.strptime(issue_date_str, '%m/%d/%Y')
                    first_coupon_date = datetime.strptime(first_coupon_date_str, '%m/%d/%Y')
                    next_to_last_coupon_date = datetime.strptime(next_to_last_coupon_date_str, '%m/%d/%Y')
                    maturity_date = datetime.strptime(maturity_date_str, '%m/%d/%Y')
                    valuation_date = current_date

                    accrued_interest, days_in_period, days_of_accrual = bond_calc.calculate_accrued_interest(
                        issue_date, first_coupon_date, day_count_convention, payment_frequency,
                        next_to_last_coupon_date, maturity_date, valuation_date, coupon_rate, semi_split
                    )

                    coupon = qty * accrued_interest

                    if coupon > 0:
                        faal = "InterestReceivable"
                        faie = "InterestReceipt"
                    else:
                        faal = "InterestPayable"
                        faie = "InterestExpense"

                    payment_currency = subspace.get_information_field('AIF', 'payment_currency')
                    fx_rate = 1  # Example placeholder
                    coupon_in_book_terms = coupon * fx_rate

                    bcoup = Journals(portfolio, payment_currency, 0, ls, location, faal, coupon_in_book_terms,
                                     coupon_in_book_terms, coupon_in_book_terms, 0,
                                     "Accrual", valuation_date, valuation_date, valuation_date,
                                     valuation_date, valuation_date, "Asset/Liability")
                    sub_ledger.post_journal_entry(bcoup)

                    bcoupRE = Journals(portfolio, investment, 0, ls, location, faie, 0, -coupon_in_book_terms,
                                       -coupon_in_book_terms, 0,
                                       "Accrual", valuation_date, valuation_date, valuation_date, valuation_date,
                                       valuation_date,
                                       "Revenue/Expense/Capital")
                    sub_ledger.post_journal_entry(bcoupRE)
        else:
            print(f"Skipping non-BOND investment: {investment}, Type: {investment_type}")

class Event:
    def __init__(self, transaction=None, method=None, tradedate=None, settledate=None,
                 actualsettledate=None, ibor_date=None, portfolio=None, investment=None,
                 tranid=None, location=None, strategy=None, quantity=None, total_amount=None,
                 total_amount_base=None, fx_rate=None, accounting_impact_fields=None, data=None):
        self.transaction = transaction
        self.method = method
        self.tradedate = tradedate
        self.settledate = settledate
        self.actualsettledate = actualsettledate
        self.ibor_date = ibor_date
        self.portfolio = portfolio
        self.investment = investment
        self.tranid = tranid
        self.location = location
        self.strategy = strategy
        self.quantity = quantity
        self.total_amount = total_amount
        self.total_amount_base = total_amount_base
        self.fx_rate = fx_rate
        self.accounting_impact_fields = accounting_impact_fields
        self.data = data
        self.future_events = []
        self.trade_events = []



class SpaceManager:
    def __init__(self):
        # Initialize separate spaces for sub_ledger and general_ledger
        self.spaces = {
            'sub_ledger': {
                'data': BookkeepingSpace()  # The main space for posting journals
            },
            'general_ledger': {
                'data': BookkeepingSpace()  # Separate space for GL posting and aggregation
            },

            # ... other spaces can be initialized here ...
        }

    def get_space(self, space_name):
        """
        Retrieve a space by its name.
        :param space_name: The name of the space to retrieve.
        :return: The requested space instance.
        """
        if space_name in self.spaces:
            return self.spaces[space_name]['data']
        else:
            raise ValueError(f"Space '{space_name}' not found in registered spaces.")

    def clear_space(self, space_name):
        """
        Clears the data in the specified space by using the reset method.
        :param space_name: The name of the space to clear.
        """
        if space_name in self.spaces:
            space_instance = self.spaces[space_name]['data']
            space_instance.reset_all()  # Use the centralized reset method
        else:
            raise ValueError(f"Space '{space_name}' not found in registered spaces.")

    def register_space(self, space_name, space_instance):
        """
        Register a new space under the given name.
        :param space_name: The name for the space to register.
        :param space_instance: The instance of the space to register.
        """
        self.spaces[space_name] = {'data': space_instance}

    def query_sub_ledger(self, space_type, criteria):
        if space_type not in self.sub_ledger:
            raise ValueError(f"Invalid space_type: {space_type}")

        results = [e for e in self.sub_ledger[space_type].entries if
                   all(criteria[key] == e[0][key] for key in criteria)]
        return results

    def query_journal_entries(self, criteria):
        return [e for e in self.journal_entries if all(getattr(e, key) == criteria[key] for key in criteria)]

import pandas as pd

import pandas as pd

from datetime import datetime, timedelta

class AdministrativeFacility:
    def __init__(self):
        self.records = {}  # Manages records by transaction ID
        self.position_cache = {}  # Caching mechanism

    def __iter__(self):
        # This allows the object to be iterable over its records
        return iter(self.records.values())

    def add_record(self, tranid, portfolio, investment, position, position_effect, location, currency, qty,
                   currency_amount, status):
        new_record = {
            'tranid': tranid,
            'portfolio': portfolio,
            'investment': investment,
            'position': position,
            'position_effect': position_effect,
            'location': location,
            'currency': currency,
            'qty': qty,
            'currency_amount': currency_amount,
            'status': status
        }
        self.records[tranid] = new_record
        print(f"Record for TranID {tranid} added to SMF.")

    def update_record_status(self, tranid, new_status, portfolio):
        record = self.records.get(tranid, {})
        if record and record.get('portfolio') == portfolio:
            self.records[tranid]['status'] = new_status
            print(f"Status of record {tranid} updated to {new_status}.")
        else:
            print(f"Record not found for TranID {tranid} with portfolio {portfolio}.")

    def query_records(self, portfolio, **filters):
        return [record for record in self if record.get('portfolio', '').strip().lower() == portfolio.lower() and
                all(record.get(k, '').strip().lower() == v.lower() for k, v in filters.items())]

    def calculate_net_positions(self, portfolio, investment, date, status='Settled'):
        cache_key = (portfolio, investment, date)
        if cache_key in self.position_cache:
            return self.position_cache[cache_key]

        net_positions = {}
        for record in self.query_records(portfolio, investment=investment, status=status):
            location = record['location']
            position_key = f"{record['position']}_{record['position_effect']}"
            if location not in net_positions:
                net_positions[location] = {}
            if position_key not in net_positions[location]:
                net_positions[location][position_key] = 0
            effect_multiplier = 1 if record['position_effect'] == 'open' else -1
            net_positions[location][position_key] += record['qty'] * effect_multiplier

        self.position_cache[cache_key] = net_positions
        return net_positions

    def cleanup_records(self, retention_period_days=365):
        """ Clean up records that have been fully settled and net to zero, and are older than the retention period. """
        today = datetime.now()
        keys_to_remove = []
        for tranid, record in self.records.items():
            record_date = datetime.strptime(record['settled_date'], '%Y-%m-%d')
            if (today - record_date).days > retention_period_days and self.is_fully_settled_and_netted(tranid):
                keys_to_remove.append(tranid)

        for tranid in keys_to_remove:
            del self.records[tranid]
            print(f"Record {tranid} removed from SMF due to cleanup policy.")

    def is_fully_settled_and_netted(self, tranid):
        """ Determine if a transaction is fully settled and netted to zero. """
        record = self.records[tranid]
        # Assuming there's a way to check if the position is netted to zero (simplified here)
        return record['status'] == 'Settled' and record['net_position'] == 0

from datetime import datetime

from datetime import datetime
from typing import Optional, Tuple
from datetime import datetime
from typing import Optional, Tuple
from datetime import datetime
from typing import Optional, Tuple


class Journals:
    sequence_counter = 0  # Static counter to track the sequence

    def __init__(self, portfolio: str = "", investment: str = "", lotid: int = 0, tax_date: Optional[datetime] = None,
                 ls: str = "", location: str = "", financial_account: str = "",
                 quantity: float = 0.0, local: float = 0.0, book: float = 0.0,
                 notional: float = 0.0, oface: Optional[float] = None, tranid: Optional[int] = 0, transaction: str = "",
                 tradedate: Optional[datetime] = None, settledate: Optional[datetime] = None,
                 kdbegin: Optional[datetime] = None, kdend: Optional[datetime] = None,
                 ibor_date: Optional[datetime] = None, entry_type: str = "", feeder: str = "",
                 running_balances: Tuple[float, float, float] = (0.0, 0.0, 0.0),
                 split_ratio: float = 1.0, account_key: Optional[Tuple[str, str, datetime, str, str]] = None,
                 period: str = "", journal_type: str = "", sequence_number: Optional[int] = None):

        # Assign values to attributes
        self.portfolio = portfolio
        self.investment = investment
        self.lotid = lotid
        self.tax_date = datetime.fromisoformat(tax_date) if isinstance(tax_date, str) else tax_date
        self.ls = ls
        self.location = location
        self.financial_account = financial_account
        self.quantity = quantity
        self.local = local
        self.book = book
        self.notional = notional
        self.oface = oface
        self.tranid = tranid
        self.transaction = transaction
        self.tradedate = datetime.fromisoformat(tradedate) if isinstance(tradedate, str) else tradedate
        self.settledate = datetime.fromisoformat(settledate) if isinstance(settledate, str) else settledate
        self.kdbegin = datetime.fromisoformat(kdbegin) if isinstance(kdbegin, str) else kdbegin
        self.kdend = datetime.fromisoformat(kdend) if isinstance(kdend, str) else kdend
        self.ibor_date = datetime.fromisoformat(ibor_date) if isinstance(ibor_date, str) else ibor_date
        self.entry_type = entry_type
        self.feeder = feeder
        self.running_balances = running_balances
        self.split_ratio = split_ratio
        self.account_key = account_key or (
            self.portfolio, self.investment, self.lotid, self.tax_date, self.ls, self.location, self.financial_account)

        # Metadata attributes
        self.period = period
        self.journal_type = journal_type

        # Assign a sequence number if not provided
        if sequence_number is None:
            self.sequence_number = Journals.sequence_counter
            Journals.sequence_counter += 1
        else:
            self.sequence_number = sequence_number

        # Print statements for debugging
       # print(f"Journal entry created: {self}")


    def __str__(self):
        return f"JournalEntry(portfolio={self.portfolio}, investment={self.investment}, lotid={self.lotid}, " \
               f"tax_date={self.tax_date}, ls={self.ls}, location={self.location}, " \
               f"financial_account={self.financial_account}, quantity={self.quantity}, " \
               f"local={self.local}, book={self.book}, notional={self.notional}, oface={self.oface})"

    @classmethod
    def fetch_market_values(cls, journal_entries, edate, portfolio, investment, lotid, financial_account="MarketVal"):
        """
        Fetch market values from the journal entries based on the given criteria.
        """
        # Truncate time from edate
        edate = edate.date() if isinstance(edate, datetime) else edate

        for entry in journal_entries:
            # Truncate time from entry.tradedate
            entry_date = entry.tradedate.date() if isinstance(entry.tradedate, datetime) else entry.tradedate

            if (entry.financial_account == financial_account and
                    entry.investment == investment and
                    entry.lotid == lotid and
                    entry_date == edate and
                    entry.entry_type == "Revenue/Expense/Capital"):
                # Return the market values if a match is found
                return entry.local, entry.book

        # Return None if no match is found
        return None, None

    def to_dict(self):
        def format_date(date_value):
            if isinstance(date_value, str):
                try:
                    date_value = datetime.fromisoformat(date_value)
                except ValueError:
                    return date_value
            return date_value.strftime("%Y-%m-%dT%H:%M:%S") if isinstance(date_value, datetime) else date_value

        return {
            # Other fields...
            "portfolio": self.portfolio,
            "investment": self.investment,
            "lotid": self.lotid,
            "tax_date": self.tax_date.strftime("%Y-%m-%dT%H:%M:%S") if self.tax_date else None,
            "ls": self.ls,
            "location": self.location,
            "financial_account": self.financial_account,
            "quantity": self.quantity,
            "local": self.local,
            "book": self.book,
            "notional": self.notional,
            "oface": self.oface,
            "tranid": self.tranid,
            "transaction": self.transaction,
            "tradedate": self.tradedate.strftime("%Y-%m-%dT%H:%M:%S") if self.tradedate else None,
            "settledate": self.settledate.strftime("%Y-%m-%dT%H:%M:%S") if self.settledate else None,
            "kdbegin": self.kdbegin.strftime("%Y-%m-%dT%H:%M:%S") if self.kdbegin else None,
            "kdend": self.kdend.strftime("%Y-%m-%dT%H:%M:%S") if self.kdend else None,
            "ibor_date": self.ibor_date.strftime("%Y-%m-%dT%H:%M:%S") if self.ibor_date else None,
            "entry_type": self.entry_type,
            "feeder": self.feeder,
            "running_balances": self.running_balances,
            "split_ratio": self.split_ratio,
            "account_key": self.account_key,
            # Uncomment and handle default values if these fields are needed:
            # "asset_class": self.asset_class if hasattr(self, 'asset_class') else None,
            # "currency": self.currency if hasattr(self, 'currency') else None,
            # "country": self.country if hasattr(self, 'country') else None,
            # "sector": self.sector if hasattr(self, 'sector') else None,
            # "system_type": self.system_type if hasattr(self, 'system_type') else None,
            # "group1": self.group1 if hasattr(self, 'group1') else None,
            # "bsgroup": self.bsgroup if hasattr(self, 'bsgroup') else None,
            # "performance_category": self.performance_category if hasattr(self, 'performance_category') else None,
        }

    @classmethod
    def from_dict(cls, data):
        # Normalize tax_date, default to January 1, 1970, if empty or invalid
        if 'tax_date' in data and data['tax_date']:
            tax_date = datetime.strptime(data['tax_date'], "%Y-%m-%dT%H:%M:%S").replace(hour=0, minute=0, second=0)
        else:
            tax_date = datetime(1970, 1, 1, 0, 0, 0)

        return cls(
            portfolio=data['portfolio'],
            investment=data['investment'],
            lotid=data['lotid'],
            tax_date=tax_date,
            ls=data['ls'],
            location=data['location'],
            financial_account=data['financial_account'],
            transaction=data['transaction'],
            quantity=data['quantity'],
            local=data['local'],
            book=data['book'],
            notional=data['notional'],
            oface=data['oface'],
            tradedate=datetime.strptime(data['tradedate'], "%Y-%m-%dT%H:%M:%S"),
            settledate=datetime.strptime(data['settledate'], "%Y-%m-%dT%H:%M:%S"),
            ibor_date=datetime.strptime(data['ibor_date'], "%Y-%m-%dT%H:%M:%S"),
            kdbegin=data['kdbegin'],
            kdend=data['kdend'],
            entry_type=data['entry_type'],
            feeder=data['feeder'],
            running_balances=data['running_balances'],
            split_ratio=data['split_ratio'],
            account_key=data['account_key']
        )
@staticmethod
def from_json(entry):
    if 'tax date' in entry:
        entry['tax_date'] = entry.pop('tax date')

    entry.pop('lines', None)

    date_fields = ['ibor_date', 'kdbegin', 'kdend', 'tradedate', 'settledate', 'tax_date']
    for date_field in date_fields:
        if date_field in entry:
            if isinstance(entry[date_field], str):
                entry[date_field] = entry[date_field].replace('T', ' ')
                try:
                    entry[date_field] = datetime.strptime(entry[date_field], "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    pass
            elif isinstance(entry[date_field], int):
                pass

    if 'running_balances' in entry:
        entry['running_balances'] = tuple(entry['running_balances'])
    if 'account_key' in entry:
        entry['account_key'] = tuple(entry['account_key'])

    return Journals(**entry)

from typing import Optional, Tuple, List


import pandas as pd
from typing import Optional, Tuple
from datetime import datetime






class StatisticalRepository:
    def __init__(self):
        # Use a dictionary to store data keyed by a tuple of relevant identifiers
        self.data = {}

    def add_entry(self, mark_date, portfolio, investment, lotid, tax_date, stat_account_name, ls, location, stat_repo, local_value,
                  book_value):
        # Store data based on the key that includes ls and location
        key = (mark_date, portfolio, investment, lotid, tax_date, stat_account_name, ls, location)
        self.data[key] = {'local': local_value, 'book': book_value}

    def clear(self):
        """Clears all data in the repository."""
        self.data.clear()

    def get_entry(self, mark_date, portfolio, investment, lotid, tax_date, stat_account_name, ls, location, stat_repo):
        # Implementation to retrieve data based on all these dimensions
        key = (mark_date, portfolio, investment, lotid, tax_date, stat_account_name, ls, location)
        if key in self.data:
            return self.data[key]['local'], self.data[key]['book']
        else:
            return None, None

    def get_entries_in_mark_date_range(
            self,
            start_mark_date: datetime,
            end_mark_date: datetime,
            portfolio: str,
            investment: str,
            lot_id: str,
            tax_date: datetime,
            key: str
    ) -> List[Tuple[datetime, float, float]]:
        # Collect all matching entries
        results = []
        for entry_key, values in self.data.items():
            # Unpack the key tuple
            entry_portfolio, entry_investment, entry_lot_id, entry_tax_date, entry_key, entry_mark_date = entry_key
            if (entry_portfolio == portfolio and
                    entry_investment == investment and
                    entry_lot_id == lot_id and
                    entry_tax_date == pd.to_datetime(tax_date) and
                    entry_key == key and
                    entry_mark_date and
                    start_mark_date <= entry_mark_date <= end_mark_date):
                results.append((entry_mark_date, values['value_local'], values['value_book']))

        return results

    def clear(self):
        """Clears all data in the repository."""
        self.data.clear()  # This removes all entries from the dictionary


class QuantityTracker:
    def __init__(self):
        self.lot_quantities = {}

    def update_quantity(self, lot_id, quantity_change):
        """ Update the quantity for a given lot_id """
        if lot_id in self.lot_quantities:
            self.lot_quantities[lot_id] += quantity_change
        else:
            self.lot_quantities[lot_id] = quantity_change

    def get_quantity(self, lot_id):
        """ Retrieve the current quantity for a given lot_id """
        return self.lot_quantities.get(lot_id, 0)


class RevenueExpenseCapitalRepository:
    def __init__(self):
        self.entries = []
        self.entries.clear()
        self.investment_spaces_library = {}

    def add_entry(self, entry):
        self.entries.append(entry)

    def __iter__(self):
        return iter(self.entries)

    def reset(self):
        """Reset the repository by clearing all entries and subspaces."""
        self.entries.clear()
        self.investment_spaces_library.clear()  # Clear subspaces if they exist

    def get_position_space(self, investment):
        if investment not in self.investment_spaces_library:
            self.investment_spaces_library[investment] = AssetLiabilitySubspace()
        return self.investment_spaces_library[investment]

    def query_balance(self, tranid, account_type, investment):
        subspace = self.get_position_space(investment)
        return subspace.query_balance(tranid, account_type)

    def post_journal_entry(self, je, lot_portfolio, lot_investment, lot_lotid, lot_tax_date, lot_ls, lot_location, lot_financial_account,
            net_unrealized_price_local, net_unrealized_price_book):
        """
        If the journal entry with the same key exists, update it.
        Otherwise, append the new journal entry.
        """
        for entry in self.entries:
            if (entry.portfolio == lot_portfolio and
                    entry.investment == lot_investment and
                    entry.lotid == lot_lotid and
                    entry.tax_date == lot_tax_date and
                    entry.ls == lot_ls and
                    entry.location == lot_location and
                    entry.financial_account == lot_financial_account):
                # Update the existing entry
                entry.local = net_unrealized_price_local
                entry.book = net_unrealized_price_book
                return

        # If not found, append the new entry
        self.entries.append(je)

    def find_most_recent_entry(entries, portfolio, investment, lotid, tax_date, ls, location, financial_account):
        """
        Find the most recent entry in the list of entries that matches the given parameters.

        Args:
            entries (list): List of journal entry objects.
            portfolio (str): The portfolio identifier.
            investment (str): The investment identifier.
            lotid (str): The lot identifier.
            tax_date (str): The tax date.
            ls (str): The LS (long/short) identifier.
            location (str): The location identifier.
            financial_account (str): The financial account identifier.

        Returns:
            tuple: A tuple containing the local and book values of the most recent matching entry,
                   or (None, None) if no matching entry is found.
        """
        # Iterate over the list in reverse order to find the most recent entry
        for entry in reversed(entries):
            if (entry.portfolio == portfolio and entry.investment == investment and
                    entry.lotid == lotid and entry.ls == ls and entry.location == location and
                    entry.financial_account == financial_account):
                # Extract the local and book values
                local_value = entry.local
                book_value = entry.book

                # Print debug information
                print(
                    f"Record found: Investment: {investment}, Lot ID: {lotid}, Local Value: {local_value}, Book Value: {book_value}")

                return local_value, book_value

        # Print debug information if no record is found
        print(f"No matching record found for Investment: {investment}, Lot ID: {lotid}")

        # Return None if no matching entry is found
        return None, None

    def get_journal_value(self, portfolio, investment, lotid, tax_date, ls, location, financial_account):
        # Check if the investment sub-space exists
        if investment not in self.investment_spaces_library:
            return None

        subspace = self.investment_spaces_library[investment]

        # Retrieve the entry if it exists
        entry_key = (portfolio, investment, lotid, tax_date, ls, location, financial_account)
        if entry_key in subspace.entries:
            return subspace.entries[entry_key]
        else:
            return None



import pandas as pd
from datetime import datetime
import csv
#PERIOD_END_DATE = '12/30/2022'
  # Define the period-end date globally


class AIF:
    def __init__(self, **kwargs):
        self.fields = {}
        for key, value in kwargs.items():
            self.fields[key] = value

    def update_field(self, key, value):
        self.fields[key] = value

    def get_field(self, key):
        return self.fields.get(key, None)

class CIF:
    def __init__(self, **kwargs):
        self.fields = {}
        for key, value in kwargs.items():
            self.fields[key] = value

    def update_field(self, key, value):
        self.fields[key] = value

    def get_field(self, key):
        return self.fields.get(key, None)

class RIF:
    def __init__(self, **kwargs):
        self.fields = {}
        for key, value in kwargs.items():
            self.fields[key] = value

    def update_field(self, key, value):
        self.fields[key] = value

    def get_field(self, key):
        return self.fields.get(key, None)
class AssetLiabilitySubspace:
    _init_count = 0  # Debugging counter

    def __init__(self):
        AssetLiabilitySubspace._init_count += 1
        self.investments = {}  # Dictionary to store investment-related data
        self.information_fields_library = {}  # Dictionary to store additional information fields
        self.entries = {}  # Dictionary to store journal entries
        # print(f"AssetLiabilitySubspace initialized: {id(self)}. Total instances: {AssetLiabilitySubspace._init_count}")

    def update_investment_space(self, investment, field_type, attribute, value):
        if investment not in self.investments:
            self.investments[investment] = {}
        if field_type not in self.investments[investment]:
            self.investments[investment][field_type] = {}
        self.investments[investment][field_type][attribute] = value
        # print(f"Updated investment: {investment}, field_type: {field_type}, attribute: {attribute} set to {value}")

    def reset(self):
        """Reset the subspace by clearing all entries and information fields."""
        self.entries.clear()
        self.information_fields_library.clear()
        print("Subspace has been reset.")

    def load_aif_data(self, aif_data):
        """Load AIF data into the subspace's information fields."""
        for key, value in aif_data.items():
            self.information_fields_library[key] = value
        # print(f"Loaded AIF data: {aif_data}")

    def get_information_field(self, field_type, attribute):
        """Retrieve specific information field value."""
        return self.information_fields_library.get(attribute)

    def post_journal_entry_to_subspace(self, je):

        if not isinstance(je.tax_date, datetime):
            raise TypeError(f"Expected je.tax_date to be a datetime object, but got {type(je.tax_date)} instead.")

        """Post a journal entry to the subspace."""
        key = (je.portfolio, je.investment, je.lotid, je.tax_date, je.ls, je.location, je.financial_account)
        if key in self.entries:
            old_values = self.entries[key]
            updated_values = (
                old_values[0] + je.quantity,
                old_values[1] + je.local,
                old_values[2] + je.book,
                old_values[3] + je.notional if old_values[3] is not None else je.notional,
                old_values[4] + je.oface if old_values[4] is not None else je.oface
            )
            if abs(updated_values[0]) < .01 and abs(updated_values[1]) < .01 and abs(updated_values[2]) < .01:
                del self.entries[key]
            else:
                self.entries[key] = updated_values
        else:
            self.entries[key] = (je.quantity, je.local, je.book, je.notional, je.oface)
        # print(f"Posted journal entry: {je}")
class AssetLiabilityRepository:
    def __init__(self):
        self.investment_spaces_library = {}  # Dictionary to store investment subspaces
        self.entries = []  # List to store journal entries
        # print(f"AssetLiabilityRepository initialized: {id(self)}")

    def reset_subspaces(self):
        """Reset all subspaces by clearing their entries."""
        for subspace in self.investment_spaces_library.values():
            subspace.reset()  # Call reset method on each subspace
        self.investment_spaces_library.clear()  # Clear the dictionary holding the subspaces

    def reset(self):
        """Reset the repository by clearing all entries and subspaces."""
        self.entries.clear()  # Clear entries directly in the repository
        self.reset_subspaces()  # Reset subspaces using the internal method

    def add_entry(self, entry):
        """Add a journal entry to the repository."""
        self.entries.append(entry)

    def get_journal_entries(self):
        """Retrieve all journal entries."""
        return self.entries


    def fetch_aif_data(self, investment):
        """Fetch AIF data for a specific investment (implementation to be filled)."""
        return fetch_investment_master(investment)

    def fetch_bond_data(self, investment):
        """Fetch bond data for a specific investment (implementation to be filled)."""
        return fetch_bond_info(investment)

    def get_position_space(self, investment):
        """Get or create an investment subspace."""
        if investment not in self.investment_spaces_library:
            # Create a new subspace for the investment
            self.investment_spaces_library[investment] = AssetLiabilitySubspace()
            aif_data = self.fetch_aif_data(investment)
            if aif_data:
                self.investment_spaces_library[investment].load_aif_data(aif_data)
        return self.investment_spaces_library[investment]

    def get_information_field(self, investment, field_type, attribute):
        """Retrieve specific information field value from a subspace."""
        subspace = self.get_position_space(investment)
        return subspace.get_information_field(field_type, attribute)

    def combined_assets_liabilities(self):
        """Aggregate and return all entries from all subspaces."""
        aggregated_entries = []
        for key, subspace in self.investment_spaces_library.items():
            aggregated_entries.extend(subspace.entries.items())
        return aggregated_entries

    def update_investment(self, investment, field_type, attribute, value):
        """Update an investment attribute in a subspace."""
        subspace = self.get_position_space(investment)
        subspace.update_investment_space(investment, field_type, attribute, value)

    def reload_aif_data_for_all_subspaces(self):
        """Reload AIF data for all subspaces."""
        for investment, subspace in self.investment_spaces_library.items():
            aif_data = self.fetch_aif_data(investment)
            if aif_data:
                subspace.load_aif_data(aif_data)
        print("AIF data has been reloaded for all subspaces.")


class BookkeepingSpace:
    def __init__(self):
        # Initialize your repositories and other attributes here
        self.asset_liability_repository = AssetLiabilityRepository()
        self.investments = {}
        self.stat_repo = StatisticalRepository()
        self.asset_liability_sub_spaces = {}
        self.journal_entries = []
        self.initialized = True  # Mark as initialized
        print(f"BookkeepingSpace initialized: {id(self)}")
    def get_revenue_expense_space(self):
        """Retrieve revenue and expense entries in a consistent format."""
        formatted_entries = []
        for entry in self.revenue_expense_repository.entries:
            # Directly construct the key and values based on the known structure
            # Assuming 'entry' has attributes or ways to access:
            # portfolio, investment, tax_lot_num, ls, location, financial_account for the key
            # and quantity, local, book for the values
            key = (
                entry.portfolio, entry.investment, entry.lotid, entry.tax_date, entry.ls, entry.location,
                entry.financial_account)
            values = (entry.quantity, entry.local, entry.book)

            # Append the structured entry to the formatted_entries list
            formatted_entry = (key, values)
            formatted_entries.append(formatted_entry)

        return formatted_entries

    def get_position_space(self, investment):
        """Proxy method to simplify access to investment spaces."""
        return self.asset_liability_repository.get_position_space(investment)

    def post_journal_entry(self, je):
        """
        Post a journal entry to the appropriate subspace.

        :param je: The journal entry to post.
        """
        # Store the journal entry in the local list for bookkeeping
        self.journal_entries.append(je)

        # Always route the journal entry to the appropriate subspace for Asset/Liability entries
        if je.entry_type == 'Asset/Liability':
            space = self.asset_liability_repository.get_position_space(je.investment)
            space.post_journal_entry_to_subspace(je)

        # For Revenue/Expense/Capital entries, route to the revenue/expense repository
        elif je.entry_type == 'Revenue/Expense/Capital':
            self.revenue_expense_repository.add_entry(je)

        # If the entry type doesn't match known categories, raise an error
        else:
            raise ValueError(f"Invalid entry type: {je.entry_type}")

        # Print statements for debugging
      #  print(f"Journal entry posted: {je}")

    def get_or_create_asset_liability_sub_space(self, investment):
        if investment not in self.asset_liability_sub_spaces:
            self.asset_liability_sub_spaces[investment] = AssetLiabilitySubspace()
      #      print(f"Created AssetLiabilitySubSpace for investment: {investment}")
        else:
            print(f"Retrieved existing AssetLiabilitySubSpace for investment: {investment}")
        return self.asset_liability_sub_spaces[investment]

    def combined_space_list(self):
        # Get all asset/liability accounts
        asset_liability_accounts = self.all_asset_liability_bookkeeping_accounts_info()

        # Initialize the list for combined accounts
        combined_space_list = []

        # Add all asset/liability accounts to the list
        combined_space_list.extend(asset_liability_accounts)  # Assuming asset_liability_accounts is a list or iterable

        # Get revenue/expense/capital accounts
        revenue_expense_space = self.get_revenue_expense_space()

        # Filter and add only "MarketValue" type accounts
        for account in revenue_expense_space:  # Assuming revenue_expense_space is iterable
            if account.get('Type') == 'MarketValue':  # Replace 'Type' with the actual key for account type
                combined_space_list.append(account)

        # Return the final combined list
        return combined_space_list

    @classmethod
    def get_instance(cls):
        if cls._instance is None:
            cls._instance = cls()
        return cls._instance

    @classmethod
    def create_new_instance(cls):
        cls._instance = cls()
        return cls._instance


    def combine_journal_entries(self):
        # Gather all journal entries from both repositories
        revenue_expense_entries = self.revenue_expense_repository.entries
        statistical_entries = self.statistical_repository.entries
        asset_liability_entries = self.statistical_repository.entries
        # Convert both sets of entries into DataFrames
        df_revenue_expense = pd.DataFrame([je.__dict__ for je in revenue_expense_entries])
        df_statistical = pd.DataFrame([je.__dict__ for je in statistical_entries])

        # Combine the two DataFrames
        combined_df = pd.concat([df_revenue_expense, df_statistical], ignore_index=True)

        return combined_df

    def __init__(self):
        if not hasattr(self, 'initialized'):
            self.initialized = True
            self.journal_entries = []  # List to store journal entries
            self.subspaces = {}  # Dictionary to store subspaces for different investments
            self.revenue_expense_repository = RevenueExpenseCapitalRepository()  # Repository for revenue and expenses
            self.asset_liability_repository = AssetLiabilityRepository()  # Repository for assets and liabilities
            self.statistical_repository = StatisticalRepository()
            self.asset_liability_sub_spaces = AssetLiabilitySubspace()
            self.statistical_entries = []  # List to store statistical entries
            self.all_assets_liabilities_accounts_list = []  # List to store all asset and liability accounts
            self.all_bookkeeping_accounts_list = []  # List to store all bookkeeping accounts
            self.existing_account_keys = set()  # Set to track existing account keys
            self.check_duplicates = False  # Flag to enable/disable duplicate check
            self.entries = []  # list to store jes
            self.asset_liability_entries = []  # Initialize the attribute
            self.revenue_expense_entries = []  # Initialize the attribute
            self.statistical_entries = []  # Initialize the attribute
            self.balances = {}
            self.investment_spaces_library = {}
            self.journal_entries = []  # Centralized storage if needed

    def reset_all(self):
        """Reset the entire bookkeeping space, including AIF and other subspaces."""
        # Reset journal entries
        self.journal_entries.clear()

        # Reset asset and liability repository
        self.asset_liability_repository.reset()

        # Reload AIF data for all subspaces
        self.asset_liability_repository.reload_aif_data_for_all_subspaces()

        # Reset other repositories
        self.revenue_expense_repository.reset()
        self.statistical_repository.clear()

        # Check if self.asset_liability_sub_spaces is correctly initialized as a dictionary
        if isinstance(self.asset_liability_sub_spaces, dict):
            # Iterate through each subspace and reset
            for subspace in self.asset_liability_sub_spaces.values():
                if hasattr(subspace, 'reset'):
                    subspace.reset()
                else:
                    print(f"Warning: Subspace {subspace} does not have a reset method.")

            # Clear the dictionary of subspaces
            self.asset_liability_sub_spaces.clear()
        else:
            print("Warning: asset_liability_sub_spaces is not a dictionary. Resetting may not work as intended.")

        # Ensure investment subspaces in asset liability repository are also reset
        for subspace in self.asset_liability_repository.investment_spaces_library.values():
            if hasattr(subspace, 'reset'):
                subspace.reset()
            else:
                print(f"Warning: Investment subspace {subspace} does not have a reset method.")

        # Clear the dictionary of investment subspaces
        self.asset_liability_repository.investment_spaces_library.clear()

        # Reset other lists and sets
        self.all_assets_liabilities_accounts_list.clear()
        self.all_bookkeeping_accounts_list.clear()
        self.existing_account_keys.clear()
        self.investment_spaces_library.clear()
        self.statistical_entries.clear()

        # Optionally clear any GL-specific spaces if distinct from sub-ledger
        if hasattr(self, 'general_ledger'):
            self.general_ledger.clear()

        print("All bookkeeping spaces, subspaces, and related repositories have been reset.")


    def update_investment_space(self, investment, field_type, attribute, value):
        if investment not in self.asset_liability_repository.investment_spaces_library:
            self.asset_liability_repository.investment_spaces_library[investment] = AssetLiabilitySubspace()
        subspace = self.asset_liability_repository.investment_spaces_library[investment]
        if field_type not in subspace.information_fields_library:
            subspace.information_fields_library[field_type] = {}
        subspace.information_fields_library[field_type][attribute] = value

    def reload_aif_data(self):
        """Reload AIF data using the existing method in AssetLiabilityRepository."""
        # Assuming the repository's load method directly populates the necessary AIF data
        self.asset_liability_repository.load_aif_data(self.aif_data)
        print("AIF data has been reloaded into AssetLiabilityRepository.")

    def update_journal_value(self, portfolio, investment, lotid, tax_date,  ls, location, financial_account, local_value,
                             book_value):
        # Check if the investment sub-space exists
        if investment not in self.investment_spaces_library:
            self.investment_spaces_library[investment] = AssetLiabilitySubspace()

        subspace = self.investment_spaces_library[investment]

        # Check if the entry exists, if not, create it
        entry_key = (portfolio, investment, lotid, tax_date, ls, location, financial_account)
        if entry_key not in subspace.entries:
            subspace.entries[entry_key] = {
                'local': local_value,
                'book': book_value
            }
        else:
            # Update the existing entry
            subspace.entries[entry_key]['local'] = local_value
            subspace.entries[entry_key]['book'] = book_value

    def get_journal_value(self, portfolio, investment, lotid, ls, location, financial_account):
        # Check if the investment sub-space exists
        if investment not in self.investment_spaces_library:
            return None

    def add_journal_value(self, entry):
        """
        Appends a new journal entry to the list of entries.
        """
        self.entries.append(entry)

    def fetch_investment_master_data(self, investment):
        return fetch_investment_master(investment)

    def fetch_bond_info_data(self, investment):
        return fetch_bond_info(investment)

    def get_position_space(self, investment):
        return self.asset_liability_repository.get_position_space(investment)

    def get_information_field(self, investment, field_type, attribute):
        return self.asset_liability_repository.get_information_field(investment, field_type, attribute)

    def fetch_investment_master_data(self, investment):
        return utilities.fetch_investment_master(investment)

    def fetch_bond_info_data(self, investment):
        return utilities.fetch_bond_info(investment)

    def get_investment_attribute(self, field_type, investment, attribute):
        investment_space = self.asset_liability_repository.get_position_space(investment)
        if not investment_space:
            raise ValueError(f"Investment space for {investment} not found")

        if field_type == "AIF":
            return investment_space.get_information_field('AIF', attribute)
        elif field_type == "CIF":
            return investment_space.get_information_field('CIF', attribute)
        elif field_type == "RIF":
            return investment_space.get_information_field('RIF', attribute)
        else:
            raise ValueError(f"Unknown field type: {field_type}")

    def get_statistical_repository(self):
        return self.statistical_repository

    def get_all_entries(self):
        return self.asset_liability_repository.entries + self.revenue_expense_repository.entries + self.statistical_entries

    def query_balance_by_tranid(self, tranid, account_type, investment):
        results = []
        for entry in self.asset_liability_entries:
            if entry.tranid == tranid and entry.financial_account == account_type and entry.investment == investment:
                results.append(entry)
        return results

    def query_futures_balance(self, tranid, account_type, investment):
        results = []
        subspace = self.asset_liability_repository.get_position_space(investment)
        if subspace:
            print(f"Querying subspace for investment: {investment}")
          #  print(f"Subspace entries: {subspace.entries}")
            for entry in subspace.entries:
                portfolio, inv, lotid, tax_date, ls, location, financial_account, quantity, local, book, notional, oface = entry
                print(
                    f"Checking entry: portfolio={portfolio}, inv={inv}, lotid={lotid}, financial_account={financial_account}")
                if lotid == tranid and financial_account == account_type and inv == investment:
                    print(f"Match found: {entry}")
                    results.append(entry)
        else:
            print(f"No subspace found for investment: {investment}")
        return results

    def get_all_entries(self):
        return self.asset_liability_entries + self.revenue_expense_entries

    @property
    def sub_ledger(self):
        return self.get_all_entries()

    def unrealized_balances(self):
        balances = []
        for info in self.get_all_asset_liability_bookkeeping_info():
            if info[5] in ('Cost', 'Payable', 'Receivable', 'SpotFxReceivable', 'SpotFxPayable', 'ExpensesPayable'):
                balances.append([
                    info[0],  # portfolio
                    info[1],  # investment
                    info[2],  # lotid
                    info[3],  # tax_date
                    info[4],  # ls
                    info[5],  # location
                    info[7],  # quantity
                    info[8],  # local
                    info[9]  # book
                ])
        return balances

    def query_unrealized_balances(self, price_stat_key, fx_stat_key):
        subspace = self.asset_liability_repository.get_position_space(price_stat_key[1])

        price_local_query = subspace.query_balance(
            account_type=price_stat_key[5],
            portfolio=price_stat_key[0],
            investment=price_stat_key[1],
            tax_date=price_stat_key[2],
            ls=price_stat_key[3],
            location=price_stat_key[4]
        )
        price_book_query = subspace.query_balance(
            account_type=price_stat_key[5],
            portfolio=price_stat_key[0],
            investment=price_stat_key[1],
            tax_date=price_stat_key[2],
            ls=price_stat_key[3],
            location=price_stat_key[4]
        )
        fx_book_query = subspace.query_balance(
            account_type=fx_stat_key[5],
            portfolio=fx_stat_key[0],
            investment=fx_stat_key[1],
            tax_date=fx_stat_key[2],
            ls=fx_stat_key[3],
            location=fx_stat_key[4]
        )

        price_local = price_local_query[0][1] if price_local_query else 0
        price_book = price_book_query[0][2] if price_book_query else 0
        fx_book = fx_book_query[0][2] if fx_book_query else 0

        return price_local, price_book, fx_book
    def query_balance(self, tranid, account_type, investment):
        return self.asset_liability_repository.query_balance(tranid, account_type, investment)
    def enable_duplicate_check(self):
        self.check_duplicates = True

    def disable_duplicate_check(self):
        self.check_duplicates = False

    def store_to_parquet(self, filename):
        df = pd.DataFrame(self.get_all_asset_liability_bookkeeping_info())
        df.to_parquet(filename)

    @classmethod
    def load_from_parquet(cls, filename):
        df = pd.read_parquet(filename)
        instance = cls()
        instance.load_asset_liability_bookkeeping_info(df.to_dict(orient='records'))
        return instance

    def combined_assets_liabilities(self):
        aggregated_entries = []
        for subspace in self.asset_liability_repository.investment_spaces_library.values():
            aggregated_entries.extend(subspace.entries.items())
        return aggregated_entries

    def all_asset_liability_bookkeeping_accounts_info(self):
        combined_asset_liability_entries = self.combined_assets_liabilities()
        all_asset_liability_bookkeeping_accounts_info = []

        for entry in combined_asset_liability_entries:
            key, values = entry[0], entry[1]
            portfolio, investment, tax_lot_num, ls, location, financial_account = key
            quantity, local, book = values
            booksp_row = [portfolio, investment, tax_lot_num, ls, location, financial_account, quantity, local, book]
            all_asset_liability_bookkeeping_accounts_info.append(booksp_row)

        return all_asset_liability_bookkeeping_accounts_info


    def get_combined_space(self):
        asset_liability_space = self.combined_assets_liabilities()
        revenue_expense_space = self.get_revenue_expense_space()
        combined_space = asset_liability_space + revenue_expense_space
        return combined_space

    #
    # def combined_assets_liabilities(self):
    #     return self.asset_liability_repository.combined_assets_liabilities()

    def new_get_all_lots_for_marking(sub_ledger):
        all_lots = []
        asset_liability_info = sub_ledger.get_all_asset_liability_bookkeeping_info()
        logging.info(f"Asset Liability Info: {asset_liability_info[:5]}")  # Log the first 5 items for verification

        if not asset_liability_info:  # Check if the list is empty
            logging.info("No asset/liability information available.")
            return all_lots

        for info in asset_liability_info:
            logging.info(f"Processing info: {info}")
            if isinstance(info, dict):  # Ensure each item is a dictionary
                if info.get('financial_account') in (
                'Cost', 'Payable', 'Receivable', 'SpotFxReceivable', 'SpotFxPayable', 'ExpensesPayable'):
                    all_lots.append((
                                    info.get('portfolio'), info.get('investment'), info.get('lotid'),info.get('tax_date'), info.get('ls'),
                                    info.get('location'), info.get('quantity'), info.get('local'), info.get('book'),
                                    info.get('notional'), info.get('oface')))
        logging.info(f"Extracted Lots: {all_lots[:5]}")  # Log the first 5 extracted lots for verification
        return all_lots

    def get_all_asset_liability_bookkeeping_info(self):
        combined_asset_liability_entries = self.combined_assets_liabilities()
        all_asset_liability_bookkeeping_accounts_info = []

        for entry in combined_asset_liability_entries:
            key, values = entry[0], entry[1]
            portfolio, investment, lotid, tax_date, ls, location, financial_account = key
            quantity, local, book, notional, oface = values
            booksp_row = [portfolio, investment, lotid, tax_date, ls, location, financial_account, quantity, local,
                          book, notional, oface]
            all_asset_liability_bookkeeping_accounts_info.append(booksp_row)

        return all_asset_liability_bookkeeping_accounts_info

    def all_asset_liability_bookkeeping_accounts_info(self):
        combined_asset_liability_entries = self.combined_assets_liabilities()
        all_asset_liability_bookkeeping_accounts_info = []

        for entry in combined_asset_liability_entries:
            key, values = entry[0], entry[1]
            portfolio, investment, lotid, tax_lot_num, ls, location, financial_account = key
            quantity, local, book, notional, oface = values
            booksp_row = [portfolio, investment, lotid, tax_lot_num, ls, location, financial_account, quantity, local, book,
                          notional, oface]
            all_asset_liability_bookkeeping_accounts_info.append(booksp_row)

        return all_asset_liability_bookkeeping_accounts_info

    def update_all_bookkeeping_accounts_list(self):
        # This method updates the all_bookkeeping_accounts_list based on the latest information
        self.all_bookkeeping_accounts_list = self.get_all_asset_liability_bookkeeping_info()
        print(f"Updated all_bookkeeping_accounts_list: {self.all_bookkeeping_accounts_list}")

    # Ensure you call this method to update the list
    def ensure_all_bookkeeping_accounts_list_updated(self):
        if not self.all_bookkeeping_accounts_list:
            self.update_all_bookkeeping_accounts_list()

    # def get_combined_space(self):
    #     asset_liability_space = self.combined_assets_liabilities()
    #     revenue_expense_space = self.get_revenue_expense_space()
    #
    #     combined_space = asset_liability_space + revenue_expense_space
    #     return combined_space


    def serialize_journal_entries(self, journal_entries, fund):
        """
        Serialize journal entries to a CSV file.

        Args:
            journal_entries (list): List of journal entry objects.
            file_path (str): Path to the output CSV file.
        """
        # Define CSV field names based on journal entry attributes
        fieldnames = [
            'Portfolio', 'Investment', 'Lot ID', 'Tax Date', 'LS', 'Location', 'Financial Account',
            'Quantity', 'Local', 'Book', 'Notional', 'OFace', 'Tran ID', 'Transaction', 'Trade Date',
            'Settle Date', 'KDBegin', 'KDEnd', 'IBOR Date', 'Entry Type', 'Feeder',
            'Running Balances', 'Split Ratio', 'Account Key'
        ]
        file_path = 'C:/BASE_PATH/refdata/jeoutput/'+fund+'.csv'
        # Open the CSV file in write mode
        with open(file_path, 'w', newline='') as csvfile:
            # Create a CSV writer object
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

            # Write the CSV header
            writer.writeheader()

            # Serialize each journal entry and write it to the CSV file
            for entry in journal_entries:
                serialized_entry = {
                    'Portfolio': entry.portfolio,
                    'Investment': entry.investment,
                    'LotID': entry.lotid,
                    'Tax Date': entry.tax_date,
                    'LS': entry.ls,
                    'Location': entry.location,
                    'Financial Account': entry.financial_account,
                    'Quantity': entry.quantity,
                    'Local': entry.local,
                    'Book': entry.book,
                    'Notional': entry.notional,
                    'OFace': entry.oface,
                    'Tran ID': entry.tranid,
                    'Transaction': entry.transaction,
                    'Trade Date': entry.tradedate,
                    'Settle Date': entry.settledate,
                    'KDBegin': entry.kdbegin,
                    'KDEnd': entry.kdend,
                    'IBOR Date': entry.ibor_date,
                    'Entry Type': entry.entry_type,
                    'Feeder': entry.feeder,
                    'Running Balances': entry.running_balances,
                    'Split Ratio': entry.split_ratio,
                    'Account Key': entry.account_key
                }

                writer.writerow(serialized_entry)

        print(f"Serialized journal entries saved to {file_path}")

    def serialize_journal_entries_to_excel(journal_entries, file_path, max_rows_per_sheet=1048575):
        from openpyxl import Workbook
        """
        Serialize journal entries to an Excel file.

        Args:
            journal_entries (list): List of journal entry objects.
            file_path (str): Path to the output Excel file.
            max_rows_per_sheet (int): Maximum number of rows per sheet.
        """
        # Create a new workbook
        wb = Workbook()

        # Split data into multiple sheets
        num_sheets = (len(journal_entries) // max_rows_per_sheet) + 1
        for sheet_idx in range(num_sheets):
            ws = wb.create_sheet(title=f"Sheet_{sheet_idx + 1}")

            # Set headers
            headers = [
                "Portfolio", "Investment", "Lot ID", "Tax Date", "LS", "Location", "Financial Account",
                "Quantity", "Local", "Book", "Notional", "Oface", "Tran ID", "Transaction", "Trade Date",
                "Settle Date", "KDBegin", "KDEnd", "IBOR Date", "Entry Type", "Feeder", "Running Balances",
                "Split Ratio", "Account Key"
            ]
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)
                ws.cell(row=1, column=col_idx).font = Font(bold=True)

            # Write data to the sheet
            start_idx = sheet_idx * max_rows_per_sheet
            end_idx = min((sheet_idx + 1) * max_rows_per_sheet, len(journal_entries))
            for row_idx, entry in enumerate(journal_entries[start_idx:end_idx], start=2):
                for col_idx, attr in enumerate(["portfolio", "investment", "lotid", "tax_date", "ls", "location",
                                                "financial_account", "quantity", "local", "book", "notional","oface",
                                                "tranid", "transaction", "tradedate", "settledate", "kdbegin", "kdend",
                                                "ibor_date", "entry_type", "feeder", "running_balances",
                                                "split_ratio", "account_key"], start=1):
                    value = getattr(entry, attr)
                    if isinstance(value, (list, tuple)):
                        value = ', '.join(map(str, value))
                    ws.cell(row=row_idx, column=col_idx, value=value)

        # Save the workbook to the specified file path
        wb.save(file_path)

        print(f"Serialized journal entries saved to {file_path}")

    def get_revenue_expense_space(self):
        return []

    @staticmethod
    def convert_bookkeeping_objects_to_df(bookkeeping_objects):
        data = [{
            'portfolio': obj.account_key[0],
            'investment': obj.account_key[1],
            'lotid': obj.account_key[2],
            'tax_date': obj.account_key[3],
            'ls': obj.account_key[4],
            'location': obj.account_key[5],
            'financial_account': obj.account_key[6],
            'quantity': obj.quantity,
            'local': obj.local,
            'book': obj.book,
            'notional': obj.notional,
            'oface': obj.oface,
        } for obj in bookkeeping_objects]
        return pd.DataFrame(data)

    def get_revenue_expense_space(self):
        return []

    @staticmethod
    def convert_bookkeeping_objects_to_df(bookkeeping_objects):
        data = [{
            'portfolio': obj.account_key[0],
            'investment': obj.account_key[1],
            'lotid': obj.account_key[2],
            'tax_date': obj.account_key[3],
            'ls': obj.account_key[4],
            'location': obj.account_key[5],
            'financial_account': obj.account_key[6],
            'quantity': obj.quantity,
            'local': obj.local,
            'book': obj.book,
            'notional': obj.notional,
            'oface': obj.oface,
        } for obj in bookkeeping_objects]
        return pd.DataFrame(data)

    # def combined_space(self):
    #     # Similarly, ensure key-value pairs are maintained in this aggregation
    #     combined_entries = []
    #     # Combine asset and liability entries with keys
    #     combined_entries.extend(self.combined_assets_liabilities())
    #     # Include revenue/expense and statistical entries, ensuring keys are preserved
    #     # This example assumes you adjust for key-value pairing as needed
    #     combined_entries.extend(self.revenue_expense_repository.entries)  # Hypothetical example
    #     combined_entries.extend(
    #         [(key, value) for key, value in enumerate(self.statistical_entries)])  # Example for list-based structure
    #     return combined_entries

    def get_revenue_expense_entries(self):
        """Directly access revenue and expense entries."""
        return self.revenue_expense_repository.entries
    def print_sub_ledger(self):
        for key, value in self.bs:
            quantity, local, book = value
            print(f"Key: {key} Quantity: {quantity} Local: {local} Book: {book}")

    def sum_bs_quantitys(self):
        total_quantity = sum(entry[1][0] for entry in self.bs)
        return total_quantity

    def sum_bs_book(self):
        total_bv = sum(entry[1][2] for entry in self.bs)
        return total_bv

    def get_all_journal_entries(self):
        # Ensure both attributes are lists before concatenation
        al_entries = self.asset_liability_repository.entries if isinstance(self.asset_liability_repository.entries,
                                                                           list) else list(
            self.asset_liability_repository.entries.values())
        re_entries = self.revenue_expense_repository.entries if isinstance(self.revenue_expense_repository.entries,
                                                                           list) else list(
            self.revenue_expense_repository.entries.values())
        return al_entries + re_entries

    def add_entry(self, entry):
        # Based on entry type, add to the appropriate space
        # Assuming entry has an attribute 'type'
        if entry.type == 'Asset/Liability':
            self.asset_liability_repository.add_entry(entry)
        elif entry.type == 'Revenue/Expense':
            self.revenue_expense_repository.add_entry(entry)

    def build_general_ledger_from_journals(self, journals, period_end):
        import time
        bs_start_time = time.time()

        # # Initialize or get the GeneralLedger space
        # general_ledger_space = self  # Assuming 'self' is an instance of the GeneralLedger class

        space_manager = SpaceManager()
        general_ledger = space_manager.get_space('general_ledger')  # Reset sub_ledger before processing

        space_manager.clear_space('general_ledger')  # Reset sub_ledger before processing
        # Filter and process JEs for General Ledger
        for idx, je in enumerate(journals):
            if je.ibor_date <= period_end:
                general_ledger.post_journal_entry(je)  # Posting to general ledger space

                # Optional: Print status every N entries
                if idx % 10 == 0:
                    print(f"Processed {idx} journal entries for General Ledger...")
            else:
                break  # Exit loop when period_end is exceeded

        print("Finished processing all journal entries for General Ledger!")
        bs_end_time = time.time()
        fetch_time = bs_end_time - bs_start_time

        print("\nElapsed time - General Ledger build from data: {:.6f}".format(fetch_time))

        # Return the state of the General Ledger space (optional)
        return general_ledger

    @property
    def sub_ledger(self):
        return self.asset_liability_entries + self.revenue_expense_entries

    def add_journal_entry(self, entry):
        self.journal_entries.append(entry)

    def process_stock_split(self, investment, split_ratio: float):
        split_quantities = {}
        for je in self.asset_liability_entries:
            if je[0][1] != investment:
                continue  # Skip entries that do not match the specified investment
            pos_type = je[0][3]
            additional_quantity = je[1][0] * split_ratio - je[1][0]
            split_quantities[pos_type] = additional_quantity

        return split_quantities

    def aggregate_entries(self):
        # Combine both asset/liability and revenue/expense entries into a single list
        all_entries = self.asset_liability_entries + self.revenue_expense_entries
        return all_entries

    def lot_iterator_by_custodian(self, investment, entry_type, new_shares_ratio):
        if entry_type == "dividends":
            entries = self.asset_liability_entries
        elif entry_type == "stock_splits":
            entries = self.asset_liability_entries  # Or you can use another appropriate attribute
        else:
            raise ValueError("Invalid entry_type provided")

        for entry in entries:
            key, value = entry
            # Assuming that key is in the format (portfolio, investment, lot_id, pos_type, location, financial_account)
            portfolio, inv, lotid, lot_id, pos_type, location, fa = key

            # Perform additional filtering or processing based on the specific entry_type if needed
            # For example, check the investment and new_shares_ratio

            # If the entry meets the desired conditions, yield the relevant lot information
            yield lot_id, value[0] * new_shares_ratio, location

    def query_aggregate_balances(self, query):
        query_parts = query.split('.')
        dimensions = query_parts[:-1]
        metric = query_parts[-1]
        aggregate_dict = self.aggregates
        for dimension in dimensions:
            if dimension not in aggregate_dict:
                return None
            aggregate_dict = aggregate_dict[dimension]

        if metric not in aggregate_dict:
            return None
        return aggregate_dict[metric]


    class Investment:
        def __init__(self, type, name, legs=None):
            self.type = type
            self.name = name
            self.legs = legs or []

        def get_legs_by_exposure(self, exposure_type):
            return [leg for leg in self.legs if leg.exposure_type == exposure_type]

        def report_exposure(self):
            report = {}
            for exposure_type in ExposureType:
                legs = self.get_legs_by_exposure(exposure_type)
                total_notional = sum(leg.notional_amount for leg in legs)
                total_local = sum(leg.local for leg in legs)
                total_book = sum(leg.book for leg in legs)
                report[exposure_type.value] = {
                    "Total Notional": total_notional,
                    "Total Local": total_local,
                    "Total Book": total_book
                }
            return report

    from enum import Enum

    class ExposureType(Enum):
        EQUITY = "Equity"
        FIXED_INCOME = "Fixed Income"
        COMMODITY = "Commodity"
        CURRENCY = "Currency"

class CurrencyReclassifier:
    def __init__(self, sub_ledger):
        self.sub_ledger = sub_ledger

    def get_combined_cash_quantity(self, investment):
        account_types = ['Cost', 'Receivable', 'Payable']
        return self._get_combined_quantity(investment, account_types)

    def get_separated_cash_quantities(self, investment):
        cost_quantity = self._get_quantity(investment, 'Cost')
        receivable_quantity = self._get_quantity(investment, 'Receivable')
        payable_quantity = self._get_quantity(investment, 'Payable')
        return cost_quantity, receivable_quantity, payable_quantity

    def _get_combined_quantity(self, investment, account_types):
        combined_quantity = 0
        for record in self.sub_ledger.get_journal_entries():
            if record['investment'] == investment and record['financial_account'] in account_types:
                combined_quantity += record['quantity']
        return combined_quantity

    def _get_quantity(self, investment, account_type):
        quantity = 0
        for record in self.sub_ledger.get_journal_entries():
            if record['investment'] == investment and record['financial_account'] == account_type:
                quantity += record['quantity']
        return quantity

    class Leg:
        def __init__(self, leg_name, notional_amount, type, quantity, local, book, notional, oface, exposure_type, investment,
                     day_count=None, accrual_rate=None, reset_date=None):
            self.leg_name = leg_name
            self.notional_amount = notional_amount
            self.type = type
            self.quantity = quantity
            self.local = local
            self.book = book
            self.notional = notional
            self.oface = oface
            self.exposure_type = exposure_type
            self.investment = investment
            self.day_count = day_count
            self.accrual_rate = accrual_rate
            self.reset_date = reset_date

def store_journals(journals):
    fieldnames = ['portfolio', 'investment', 'tax_lot_num', 'ls', 'location', 'financial_account']

    with open("journal_entries.csv", 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(fieldnames)

        for journal in journals:
            journal_tuple = (
                journal.portfolio,
                journal.investment,
                journal.lotid,
                journal.tax_lot_num,
                journal.ls,
                journal.location,
                journal.financial_account
            )
            writer.writerow(journal_tuple)

def update_running_balances( jes, level_preference):
    from collections import defaultdict
    running_totals = defaultdict(lambda: [0, 0, 0])  # Default to (0, 0, 0) for (quantity, local, book)

    # We sort the journal entries to ensure we're processing them in the order they occurred.
    # If they're already in the desired order, this step can be skipped.
    sorted_jes = sorted(jes, key=lambda je: (
    je.ibor_date, je.tranid is None, je.tranid, je.investment,  je.ls, je.location, je.financial_account))
   # sorted_jes = sorted(jes, key=lambda je: je.tranid)  # sort option.

    for je in sorted_jes:
        key = tuple(getattr(je, attr) for attr in level_preference)

        # Update running totals
        running_totals[key][0] += je.quantity
        running_totals[key][1] += je.local
        running_totals[key][2] += je.book

        # Assign the current running totals to this journal entry
        je.running_balances = tuple(running_totals[key])

def parse_journal_entries(file_path):
    with open(file_path, "r") as file:
        lines = file.readlines()

    journal_entries = []
    entry_data = {}
    for line in lines:
        line = line.strip()
        if line.startswith("Entry Count"):
            if entry_data:
                journal_entry = Journals(**entry_data)
                journal_entries.append(journal_entry)
            entry_data = {}
        elif line:
            key, value = line.split(":", 1)
            key = key.strip()
            value = value.strip()
            entry_data[key] = value

    if entry_data:
        journal_entry = Journals(**entry_data)
        journal_entries.append(journal_entry)

    return journal_entries

from datetime import datetime

def check_date_format(date_str, format):
    try:
        datetime.strptime(date_str, format)
        return True
    except ValueError:
        return False

def add_leading_zero(df):
    df.index = df.index.astype(str)  # Ensures that the index is of type string
    df.index = df.index.map(lambda x: x if len(x.split('/')[0]) == 2 or x.split('/')[0] == "0" else '0'+x)
    return df

def format_date(date_string):
    date_obj = datetime.strptime(date_string, '%m/%d/%Y')
    return date_obj.strftime('%m/%d/%Y').lstrip("0").replace("/0", "/")

from dateutil.parser import parse

def parse_date(date_string):
    try:
        return parse(date_string).date()
    except ValueError:
        raise ValueError(f"Unable to parse date from string: {date_string}")

def load_price_data(price_file):
    price_data = {}
    with open(price_file, 'r') as file:
        reader = csv.DictReader(file)
        for row in reader:
            date = parse_date(row['date']).isoformat() # Using the ISO format
            date_obj = datetime.strptime(row['date'], '%m/%d/%Y')  # Parsing date
            date = date_obj.strftime('%m/%d/%Y').lstrip("0").replace("/0", "/")
            ticker = row['ticker']
            currency = row['currency']
            price = float(row['price'])

            if date not in price_data:
                price_data[date] = {}

            # Create a nested dictionary for each ticker containing 'price' and 'currency' fields
            price_data[date][ticker] = {
                'price': price,
                'currency': currency
            }

    return price_data
def load_fx_data(fx_file):
    fx_data = {}
    with open(fx_file, 'r') as file:
        reader = csv.DictReader(file)
        for row in reader:
            date = parse_date(row['date']).isoformat() # Using the ISO format
            date_obj = datetime.strptime(row['date'], '%m/%d/%Y')  # Parsing date
            date = date_obj.strftime('%m/%d/%Y').lstrip("0").replace("/0", "/")
        # Formatting date
            currency = row['currency']
            if date not in fx_data:
                fx_data[date] = {}
            fx_data[date][currency] = float(row['price'])
    return fx_data

import csv

def load_coa_from_csv():
    filename = "C:/Users/hjmne/PycharmProjects/chest/refdata/chart_of_accounts.csv"
    with open(filename, mode='r') as file:
        reader = csv.DictReader(file, delimiter='\t')
        return [row for row in reader]

def get_system_type(coa, system_name):
    for account in coa:
        if account["SystemName"] == system_name:
            return account["SystemType"]
    return None  # return None if no matching SystemName is found

def validate_bookkeeping_entry(entry):
    expected_keys = ['portfolio', 'investment', 'lotid', 'tax_lot', 'ls', 'location', 'financial_account', 'quantity', 'local', 'book', 'notional', 'oface']

    # Check if the entry has all the expected keys
    if not all(key in entry for key in expected_keys):
        return False

    # Check data types of the entry
    if not isinstance(entry['portfolio'], str) or \
            not isinstance(entry['investment'], str) or \
            not isinstance(entry['lotid'], str) or \
            not isinstance(entry['tax_lot'], int) or \
            not isinstance(entry['ls'], str) or \
            not isinstance(entry['location'], str) or \
            not isinstance(entry['financial_account'], str) or \
            not isinstance(entry['quantity'], int) or \
            not isinstance(entry['local'], float) or \
            not isinstance(entry['book'], float) or \
            not isinstance(entry['notional'], float) or \
            not isinstance(entry['oface'], float):
            return False

    return True

def query_income_activity(journal_entries):
    # Dictionary to hold roll-up data
    rollup_data = {}

    for je in journal_entries:
        financial_account = je.financial_account
        investment = je.investment
        feeder = je.feeder

        if financial_account in ["Income", "DividendReceipt", "PriceGainInvestment", "FXGainInvestment",
                                 "InterestReceipt", "ContributedCost", "FXGainInvestment", "PerfFee", "MgmtFee",
                                 "FXGainLossInvestment", "UnrealGLRevExp", "FXGainTradeSettle"]:
            key = (investment, financial_account, feeder)
            rollup_data.setdefault(key, {"local": 0, "book": 0})

            rollup_data[key]["local"] += je.local  # Assuming 'local' is an attribute of Journals
            rollup_data[key]["book"] += je.book    # Assuming 'book' is an attribute of Journals

    summary_list = []
    for key, values in rollup_data.items():
        investment, account, feeder = key
        summary_list.append({
            "investment": investment,
            "account": account,
            "feeder" : feeder,
            "local": values["local"],
            "book": values["book"]
        })

    return summary_list

def query_balance_sheet_activity(journal_entries):
    # Dictionary to hold roll-up data
    rollup_data = {}

    for je in journal_entries:
        financial_account = je.financial_account
        investment = je.investment
        feeder = je.feeder

        key = (investment, financial_account, feeder)
        rollup_data.setdefault(key, {"quantity": 0, "local": 0, "book": 0})

        rollup_data[key]["quantity"] += je.quantity  # Assuming 'local' is an attribute of Journals
        rollup_data[key]["local"] += je.local  # Assuming 'local' is an attribute of Journals
        rollup_data[key]["book"] += je.book    # Assuming 'book' is an attribute of Journals
        rollup_data[key]["notional"] += je.notional  # Assuming 'book' is an attribute of Journals
        rollup_data[key]["oface"] += je.oface  # Assuming 'book' is an attribute of Journals

    summary_list = []
    for key, values in rollup_data.items():
        investment, account, feeder = key
        summary_list.append({
            "investment": investment,
            "account": account,
            "feeder" : feeder,
            "quantity": values["quantity"],
            "local": values["local"],
            "book": values["book"],
            "notional": values["notional"],
            "oface": values["oface"],
        })

    return summary_list


import os
def parse_chart_of_accounts(filename):
    """Parses the chart of accounts and returns a dictionary."""
    chart_dict = {}
    if not isinstance(filename, str):
        raise TypeError(f"Expected a string for filename, but got {type(filename)} with value {filename}")

    with open(filename, 'r') as file:
        # Skipping the header row
        next(file)

        for line in file:
            parts = line.strip().split('\t')  # Using tab as the delimiter

            # Ensure there are 6 elements, fill in missing with empty string

            while len(parts) < 6:
                parts.append('')

            system_name, space_pref, user_description, gl_account_num, group1, group2 = parts
            chart_dict[system_name] = {
                "SpacePref": space_pref,
                "UserDescription": user_description,
                "GLAccountNum": gl_account_num,
                "Group1": group1,
                "Group2": group2
            }
    return chart_dict


def fetch_mark_records_by_account_key(account_key, all_entries):
    # First, filter the entries based on the given account_key
    filtered_entries = [entry for entry in all_entries if entry.account_key == account_key]

    # Then, filter the filtered_entries by type_name
    asset_records = [entry for entry in filtered_entries if entry.type_name == "UnrealGLAsset"]
    revexp_records = [entry for entry in filtered_entries if entry.type_name == "UnrealGLRevExp"]

    return asset_records, revexp_records
