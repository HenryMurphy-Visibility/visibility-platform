import datetime
import csv
import time
import os
from collections import OrderedDict
import psutil


class EventScheduler:
    def __init__(self):
        self.events = OrderedDict()
        self.event_type_precedence = {
            'buy_equity': 1075, 'sell_equity': 1111, 'short_equity': 1111, 'cover_equity': 1111,
            'buy_bond': 1075, 'sell_bond': 1111, 'short_bond': 1111, 'cover_bond': 1111,
            'buy_future': 1075, 'sell_future': 1111, 'short_future': 1111, 'cover_future': 1111,
            'open_swap': 1075, 'reset_swap': 1111, 'spot_fx': 1111, 'forward_fx': 1111,
            'deposit_currency': 1090, 'withdraw_currency': 1090, 'split_equity': 2111,
            'dividend_equity': 2122, 'option_exercise_open': 1076, 'option_exercise_close': 1077,
            'option_assign_open': 1076, 'option_assign_close': 1077, 'mark-to-market': 9000,
            'perf_mark': 9000, 'allocate': 9500, 'settle_bond_flows_in': 1050, 'settle_bond_flows_out': 1050,
            'settle_single_flow_in': 1050, 'settle_single_flow_out': 1050, 'settle_multiple_flows_in_out': 1050,
            'settle_pay_rec_by_tranid': 1050
        }
        # with open('log_data.csv', 'w', newline='') as file:
        #     writer = csv.writer(file)
        #     writer.writerow(["Increment", "Elapsed Time", "Memory Usage", "Events Per Second"])

    def schedule_event(self, tradedate, event_function, *args, event_type=None):
        if not callable(event_function):
            raise ValueError("event_function must be callable")

        event_id = len(self.events) + 1  # Generate a unique identifier for the event
        self.events[event_id] = {
            'tradedate': tradedate,
            'event_function': event_function,
            'args': args,
            'event_type': event_type
        }

        # Debug statement to print the scheduled event function
      #  print(f"Scheduled event: {event_type} with function {event_function.__name__} on {tradedate}")


    def sort_events(self, reverse=False):
        sorted_items = sorted(
            self.events.items(),
            key=lambda x: (self.event_type_precedence.get(x[1]['event_type'], float('inf')), x[1]['tradedate']),
            reverse=reverse
        )
        self.events = OrderedDict(sorted_items)

    def get_memory_usage(self):
        process = psutil.Process()
        memory_info = process.memory_info()
        return memory_info.rss

    def run_next_event(self, testcount, counter_start_time):
        previous_elapsed_time = 0
        while self.events:
            next_event_key, next_event_value = self.events.popitem(last=False)
            event_function = next_event_value['event_function']
            event_args = next_event_value['args']

            if not callable(event_function):
                raise TypeError(f"Expected a callable event_function, but got {type(event_function).__name__}")

            event_function(*event_args)
            testcount += 1
            if testcount % 1000 == 0:
                counter_time = time.time()
                elapsed_time = counter_time - counter_start_time + 2
                memory_usage = self.get_memory_usage()
                denom = elapsed_time - previous_elapsed_time
                if denom == 0:
                    denom = 1
                events_per_second = 1000 / denom if testcount != 1000 else 0

                # Print summary metrics
                # print(f"Processed {testcount} events")
                # print(f"Elapsed time: {elapsed_time:.6f} seconds")
                # print(f"Memory usage: {memory_usage / (1024 ** 2):.2f} MB")  # Convert bytes to MB
                # print(f"Events per second: {events_per_second:.2f}")

                with open('log_data.csv', 'a', newline='') as file:
                    writer = csv.writer(file)
                    writer.writerow([testcount, f"{elapsed_time:.6f}", memory_usage, f"{events_per_second:.2f}"])

                previous_elapsed_time = elapsed_time

        return testcount

class Event:
    def __init__(self, transaction=None, method=None, tradedate=None, settledate=None,
                 actualsettledate=None, ibor_date=None, portfolio=None, investment=None,
                 tranid=None, location=None, strategy=None, quantity=None, total_amount=None,
                 total_amount_base=None, fx_rate=None, accounting_impact_fields=None, data=None):
        self.transaction = transaction
        self.method = method
        self.tradedate = tradedate
        self.settledate = settledate
        self.actualsettledate = actualsettledate
        self.ibor_date = ibor_date
        self.portfolio = portfolio
        self.investment = investment
        self.tranid = tranid
        self.location = location
        self.strategy = strategy
        self.quantity = quantity
        self.total_amount = total_amount
        self.total_amount_base = total_amount_base
        self.fx_rate = fx_rate
        self.accounting_impact_fields = accounting_impact_fields
        self.data = data
        self.future_events = []
        self.trade_events = []

# Example usage

    def generate_chart(csv_file_path):
        import pandas as pd
        data = pd.read_csv(csv_file_path)
        data = pd.read_csv(csv_file_path, header=None)
        data.columns = ['Increment', 'Elapsed Time', 'Memory Usage', 'Events Per Second']

        # Convert 'Events Per Second' to numeric, handling non-numeric values if necessary
        data['Events Per Second'] = pd.to_numeric(data['Events Per Second'], errors='coerce')


        data.at[1, 'Events Per Second'] = data.iloc[2]['Events Per Second']

        plt.figure(figsize=(14, 7))
        plt.plot(data['Increment'], data['Events Per Second'], marker='o', linestyle='-', color='b')
        plt.title('Events Processed per Second per 1000 Increment')
        plt.xlabel('Increment (x1000)')
        plt.ylabel('Events Processed per Second')
        plt.grid(True)
        plt.savefig('events_per_second_chart.png')
        plt.show()

   # generate_chart('log_data.csv')

    # def __init__(self):
    #     self.events = []  # Use list and pop
    # def schedule_event(self, tradedate, TOpenPrecedence, counter, event_function, *args, event_type=None):
    #     event_key = (tradedate, TOpenPrecedence, counter)
    #     heapq.heappush(self.events, (event_key, (event_function, args, event_type)))
    #
    #
    # def run_next_event(self):
    #     if self.events:
    #         next_event_key, next_event_value = heapq.heappop(self.events)
    #         event_function, event_args, event_type = next_event_value
    #         event_function(*event_args)
    #

class SpaceManager:
    def __init__(self):
        self.spaces = {
            'investment_accounting_space': {
                'data': BookkeepingSpace()  # Or however you instantiate this space
            },
            # ... other spaces ...
        }
    def clear_space(self, space_name):
        if space_name in self.spaces:
            space_instance = self.spaces[space_name]['data']
            space_instance.journal_entries = []
            space_instance.asset_liability_entries = []
            space_instance.revenue_expense_entries = []
            # ... clear other attributes similarly ...

    def get_space(self, space_name):
        # Return the space if it exists
        if space_name in self.spaces:
            return self.spaces[space_name]['data']

        # Otherwise, create a new space, register it, and return it
        else:
            new_space = BookkeepingSpace()  # or whatever the constructor for the space is
            self.register_space(space_name, new_space)
            return new_space

    def register_space(self, space_name, space_instance):
        self.spaces[space_name] = {'data': space_instance}

    def clear_space(self, space_name):
        if space_name in self.spaces:
            space_instance = self.spaces[space_name]['data']
            space_instance.journal_entries = []
            space_instance.asset_liability_entries = []
            space_instance.revenue_expense_entries = []
            # ... clear other attributes similarly ...
        else:
            raise ValueError(f"'{space_name}' not found in registered spaces.")

    def create_child_space(parent_space, data):
        child_space_id = uuid.uuid4()
        child_space = {
            "id": child_space_id,
            "parent_id": parent_space["id"],
            "data": data
        }
        return child_space

    def create_space(data):
        space_id = uuid.uuid4()
        space = {
            "id": space_id,
            "data": data
        }
        return space

    def interact_spaces(self, source_space_name, target_space_name, criteria):
        source_space = self.spaces[source_space_name]
        target_space = self.spaces[target_space_name]
        # ... define logic ...

    def change_space(self, from_space_name, to_space_name):
        # Assuming you want to transfer the data from one space to another:
        self.spaces[to_space_name].journal_entries.extend(self.spaces[from_space_name].journal_entries)
        self.spaces[from_space_name].journal_entries.clear()

    def query_sub_ledger(self, space_type, criteria):
        if space_type not in self.sub_ledger:
            raise ValueError(f"Invalid space_type: {space_type}")

        results = [e for e in self.sub_ledger[space_type].entries if
                   all(criteria[key] == e[0][key] for key in criteria)]
        return results

    def query_journal_entries(self, criteria):
        return [e for e in self.journal_entries if all(getattr(e, key) == criteria[key] for key in criteria)]

import pandas as pd

import pandas as pd

class AdministrativeFacility:
    def __init__(self):
        self.records = {}  # Manages records by transaction ID


    def __iter__(self):
        # This allows the object to be iterable over its records
        return iter(self.records.values())

    def add_record(self, tranid, portfolio, investment, position, position_effect, location, currency, qty,
                   currency_amount, status):
        new_record = {
            'tranid': tranid,
            'portfolio': portfolio,
            'investment': investment,
            'position': position,
            'position_effect': position_effect,
            'location': location,
            'currency': currency,
            'qty': qty,
            'currency_amount': currency_amount,
            'status': status
        }
        self.records[tranid] = new_record
        print(f"Record for TranID {tranid} added to SMF.")

    def update_record_status(self, tranid, new_status, portfolio):
        record = self.records.get(tranid, {})
        if record and record.get('portfolio') == portfolio:
            self.records[tranid]['status'] = new_status
            print(f"Status of record {tranid} updated to {new_status}.")
        else:
            print(f"Record not found for TranID {tranid} with portfolio {portfolio}.")

    def query_records(self, portfolio, **filters):
        return [record for record in self if record.get('portfolio', '').strip().lower() == portfolio.lower() and
                all(record.get(k, '').strip().lower() == v.lower() for k, v in filters.items())]

    def calculate_net_positions(self, portfolio, investment, status='Settled'):
        net_positions = {}
        for record in self.query_records(portfolio, investment=investment, status=status):
            location = record['location']
            position_key = f"{record['position']}_{record['position_effect']}"
            if location not in net_positions:
                net_positions[location] = {}
            if position_key not in net_positions[location]:
                net_positions[location][position_key] = 0
            effect_multiplier = 1 if record['position_effect'] == 'open' else -1
            net_positions[location][position_key] += record['qty'] * effect_multiplier

        return net_positions

    def cleanup_records(self, retention_period_days=365):
        """ Clean up records that have been fully settled and net to zero, and are older than the retention period. """
        today = datetime.now()
        keys_to_remove = []
        for tranid, record in self.records.items():
            record_date = datetime.strptime(record['settled_date'], '%Y-%m-%d')
            if (today - record_date).days > retention_period_days and self.is_fully_settled_and_netted(tranid):
                keys_to_remove.append(tranid)

        for tranid in keys_to_remove:
            del self.records[tranid]
            print(f"Record {tranid} removed from SMF due to cleanup policy.")

    def is_fully_settled_and_netted(self, tranid):
        """ Determine if a transaction is fully settled and netted to zero. """
        record = self.records[tranid]
        # Assuming there's a way to check if the position is netted to zero (simplified here)
        return record['status'] == 'Settled' and record['net_position'] == 0


class Journals:
    def __init__(self, portfolio: str = "", investment: str = "", tax_date: datetime = None,
                 ls: str = "", location: str = "", financial_account: str = "",
                 quantity: float = 0.0, local: float = 0.0, book: float = 0.0,
                 tranid: int = 0, transaction: str = "", tradedate: datetime = None,
                 settledate: datetime = None, kdbegin: datetime = None, kdend: datetime = None,
                 ibor_date: datetime = None, entry_type: str = "", feeder: str = "",
                 running_balances: tuple = (0.0, 0.0, 0.0),
                 split_ratio: float = 1.0, account_key: tuple = None,
                 ):
    #     self.entries = [] # Assuming entries is a list of tuples or a similar structure
    # def __getitem__(self, key):
    #     return self.entries[key]

        # if Journals.chart_of_accounts is None:
        #     raise ValueError("Chart of Accounts not loaded.")
        if transaction == "StockSplit" or transaction == "StockDividend":
            ibor_date = tradedate

        self.portfolio = portfolio
        self.investment = investment
        self.tax_date = tax_date
        self.ls = ls
        self.location = location
        self.financial_account = financial_account
        self.quantity = quantity
        self.local = local
        self.book = book
        self.tranid = tranid
        self.transaction = transaction
        self.tradedate = tradedate
        self.settledate = settledate
        self.kdbegin = kdbegin
        self.kdend = kdend
        self.ibor_date = ibor_date
        self.entry_type = entry_type
        self.feeder = feeder
        self.running_balances = running_balances
        self.lines = []
        self.split_ratio = split_ratio
        self.account_key = account_key or (
            self.portfolio, self.investment, self.tax_date, self.ls, self.location)

    def get_tax_lot_for_report(self):
        if isinstance(self.tax_date, int):
            return 'NA'
        else:
            return str(self.tax_date)

    from datetime import datetime

    @staticmethod
    def from_json(entry):
        # Replace 'tax date' key with 'tax_date'
        if 'tax date' in entry:
            entry['tax_date'] = entry.pop('tax date')

        # Remove the 'lines' key if it exists
        entry.pop('lines', None)

        # Handling date fields
        date_fields = ['ibor_date', 'kdbegin', 'kdend', 'tradedate', 'settledate', 'tax_date']
        for date_field in date_fields:
            if date_field in entry:
                if isinstance(entry[date_field], str):
                    # Replace 'T' with a space in the date string
                    entry[date_field] = entry[date_field].replace('T', ' ')
                    try:
                        # Try to convert to datetime object
                        entry[date_field] = datetime.strptime(entry[date_field], "%Y-%m-%d %H:%M:%S")
                    except ValueError:
                        # If conversion fails, assume it's an integer and leave it as is
                        pass
                elif isinstance(entry[date_field], int):
                    # If it's already an integer, leave it as is
                    pass

        # Convert list to tuple for running_balances and account_key, if needed
        if 'running_balances' in entry:
            entry['running_balances'] = tuple(entry['running_balances'])
        if 'account_key' in entry:
            entry['account_key'] = tuple(entry['account_key'])

        return Journals(**entry)


    def to_dict(self):
        return {
            "portfolio": self.portfolio,
            "investment": self.investment,
            "tax date" : self.tax_date,
            "ls" : self.ls,
            "location" : self.location,
            "financial_account" : self.financial_account,
            "quantity" : self.quantity,
            "local" : self.local,
            "book" : self.book,
            "tranid" : self.tranid,
            "transaction" : self.transaction,
            "tradedate" : self.tradedate,
            "settledate" : self.settledate,
            "kdbegin" : self.kdbegin,
            "kdend" : self.kdend,
            "ibor_date" : self.ibor_date,
            "entry_type" : self.entry_type, # "Asset/Liability" or "Revenue/Expense/Capital" or Stat
            "feeder" : self.feeder,
            "running_balances" : self.running_balances,
            "lines" : self.lines,
            "split_ratio" : self.split_ratio

        # ... and so on for every other attribute
        }


    def items(self):
        return {
            "portfolio": self.portfolio,
            "investment": self.investment,
            "tax_date": self.tax_date,
            "ls": self.ls,
            "location": self.location,
            "financial_account": self.financial_account,
            "quantity": self.quantity,
            "local": self.local,
            "book": self.book,
            "tranid": self.tranid,
            "transaction": self.transaction,
            "tradedate": self.tradedate,
            "settledate": self.settledate,
            "kdbegin": self.kdbegin,
            "kdend": self.kdend,
            "ibor_date": self.ibor_date,
            "entry_type": self.entry_type,
            "feeder": self.feeder,
            "running_balances": self.running_balances,
            "lines": self.lines,
            "split_ratio": self.split_ratio,
            "account_key": self.account_key
        }


       # self.tax_date = tax_date if tax_date is not None else tradedate
    def __repr__(self):
        return f"Journals(portfolio={self.portfolio}, investment={self.investment}, tax_date={self.tax_date},ls={self.ls}," \
               f"tranid={self.tranid},quantity={self.quantity},local={self.local}," \
               f"book={self.book},location={self.location}, financial_account={self.financial_account})"  # add other attributes as needed

    def append_if_valid_date(entry, journal_list, start_period):
        """Appends the entry to the journal list if the ibor_date is greater than or equal to the start_period."""
        if entry.ibor_date >= start_period:
            journal_list.append(entry)

class AIFEntry:
    def __init__(self, aif_id, type, subtype, data):
        self.aif_id = aif_id
        self.type = type
        self.subtype = subtype
        self.data = data

class AIFEvent:
    def __init__(self, aif_id, type, subtype, data, trade_date=None):
        self.aif_id = aif_id
        self.type = type # e.g. portgolio, investment, location, financial_account
        self.subtype = subtype # e.g. bond info, account paranters, investment paramters
        self.data = data
        self.trade_date = trade_date  # Optional: For backdating or bitemporal handling

class StatisticalRepository:
    def __init__(self):
        self.aifs = {}  # Key: AifId, Value: AIFEntry

    def add_aif(self, aif_id, type, subtype, data):
        aif_entry = AIFEntry(aif_id, type, subtype, data)
        self.aifs[aif_id] = aif_entry

    def get_aif(self, aif_id):
        return self.aifs.get(aif_id)

    def remove_aif(self, aif_id):
        if aif_id in self.aifs:
            del self.aifs[aif_id]

    def find_aifs_by_type_subtype(self, type, subtype=None):
        return [aif for aif in self.aifs.values() if aif.type == type and (subtype is None or aif.subtype == subtype)]

    # def post_aif_event(statistical_repository: StatisticalRepository, aif_event: AIFEvent):
    #     statistical_repository.add_aif(aif_event.aif_id, aif_event.type, aif_event.subtype, aif_event.data)
    #
# Example usage
# if __name__ == "__main__":
#     # Initialize the statistical repository
#     stat_repo = StatisticalRepository()
#
#     # Create an AIFEvent
#     aif_event = AIFEvent("AIF001", "Investment", "Futures Valuation Info", {"contract_size": 1000, "tick_size": 0.01})
#
#     # Post the AIFEvent to the statistical repository
#     post_aif_event(stat_repo, aif_event)
#
#     # Retrieve and print the AIF to verify
#     retrieved_aif = stat_repo.get_aif("AIF001")
#     print(f"AIF ID: {retrieved_aif.aif_id}, Type: {retrieved_aif.type}, Subtype: {retrieved_aif.subtype}, Data: {retrieved_aif.data}")

import pandas as pd

class RevenueExpenseCapitalRepository:
    def __init__(self):
        self.entries = []
        self.investment_spaces_library = {}

    def add_entry(self, entry):
        self.entries.append(entry)

    def __iter__(self):
        return iter(self.entries)

    def get_position_space(self, investment):
        if investment not in self.investment_spaces_library:
            self.investment_spaces_library[investment] = AssetLiabilitySubspace()
        return self.investment_spaces_library[investment]

    def post_journal_entry(self, je):
        space = self.get_position_space(je.investment)
        space.post_journal_entry_to_subspace(je)

    def query_balance(self, tranid, account_type, investment):
        subspace = self.get_position_space(investment)
        return subspace.query_balance(tranid, account_type)

class AssetLiabilityRepository:
    def __init__(self):
        self.investment_spaces_library = {}
        self.entries = {}

    def get_position_space(self, investment):
        if investment not in self.investment_spaces_library:
            self.investment_spaces_library[investment] = AssetLiabilitySubspace()
        return self.investment_spaces_library[investment]

    def get_investment_attribute(self, field_type, investment, attribute):
        subspace = self.get_position_space(investment)
        return subspace.get_information_field(field_type, attribute)

    def ensure_aif_lookups(self, je):
        # Ensure AIF lookups are performed
        investment = je.investment
        subspace = self.get_position_space(investment)
        if subspace:
            subspace.load_investment_info(investment)
            logging.debug(f"Performed AIF lookups for investment {investment}")

    def reset_subspaces(self):
        for subspace in self.investment_spaces_library.values():
            subspace.reset_entries()

    def post_journal_entry(self, je):
        if je.entry_type == 'Asset/Liability':
            space = self.get_position_space(je.investment)
            space.post_journal_entry_to_subspace(je)

    def query_balance(self, tranid, account_type, investment):
        subspace = self.get_position_space(investment)
        return subspace.query_balance(tranid, account_type)

class AssetLiabilitySubspace:
    def __init__(self):
        self.entries = {}

    def reset_entries(self):
        self.entries = {}

    def post_journal_entry_to_subspace(self, je):
        key = (je.portfolio, je.investment, je.tax_date, je.ls, je.location, je.financial_account)
        if key in self.entries:
            old_values = self.entries[key]
            updated_values = (old_values[0] + je.quantity, old_values[1] + je.local, old_values[2] + je.book)
            if abs(updated_values[0]) < .01 and abs(updated_values[1]) < .01 and abs(updated_values[2]) < .01:
                del self.entries[key]
            else:
                self.entries[key] = updated_values
        else:
            self.entries[key] = (je.quantity, je.local, je.book)

    def query_balance(self, tranid, account_type, investment=None):
        results = []
        for key, values in self.entries.items():
            portfolio, investment_key, tax_date, ls, location, financial_account = key
            quantity, local, book = values
            if financial_account == account_type and abs(local) > 0.01:
                results.append((quantity, local, book))
        return results


class BookkeepingSpace:
    _instance = None

    def __new__(cls, *args, **kwargs):
        if not cls._instance:
            cls._instance = super().__new__(cls, *args, **kwargs)
        return cls._instance

    def __init__(self, check_duplicates=False):
        if not hasattr(self, 'initialized'):
            self.initialized = True
            self.journal_entries = []
            self.subspaces = {}
            self.revenue_expense_repository = RevenueExpenseCapitalRepository()
            self.asset_liability_repository = AssetLiabilityRepository()
            self.statistical_entries = []
            self.all_assets_liabilities_accounts_list = []
            self.all_bookkeeping_accounts_list = []
            self.existing_account_keys = set()
            self.check_duplicates = check_duplicates
            self.entries = {}

    @classmethod
    def get_instance(cls):
        if cls._instance is None:
            cls._instance = cls()
        return cls._instance

    @classmethod
    def create_new_instance(cls, *args, **kwargs):
        new_instance = super().__new__(cls)
        cls.__init__(new_instance, *args, **kwargs)
        return new_instance

    def reset_investment_subspaces(self):
        self.asset_liability_repository.reset_subspaces()

    def reset_all(self):
        self.revenue_expense_repository.entries = []
        self.asset_liability_repository.reset_subspaces()
        self.statistical_entries = []
        self.journal_entries = []
        self.all_assets_liabilities_accounts_list = []
        self.all_bookkeeping_accounts_list = []
        self.existing_account_keys.clear()

    def query_balance(self, tranid, account_type, investment):
        return self.asset_liability_repository.query_balance(tranid, account_type, investment)

    def add_account_entry(self, entry, list_type):
        account_key = entry.get("account_key")
        if account_key in self.existing_account_keys:
            raise ValueError(f"Duplicate entry for account key: {account_key}")
        else:
            self.existing_account_keys.add(account_key)
            if list_type == "assets_liabilities":
                self.all_assets_liabilities_accounts_list.append(entry)
            elif list_type == "bookkeeping":
                self.all_bookkeeping_accounts_list.append(entry)

        if self.check_duplicates and account_key in self.unique_account_keys:
            raise ValueError(f"Duplicate entry for account key: {account_key}")
        else:
            self.all_accounts_list.append(entry)
            if self.check_duplicates:
                self.unique_account_keys.add(account_key)

    def reset_accounts(self):
        self.all_assets_liabilities_accounts_list = []
        self.all_bookkeeping_accounts_list = []
        self.existing_account_keys.clear()

    def enable_duplicate_check(self):
        self.check_duplicates = True

    def disable_duplicate_check(self):
        self.check_duplicates = False

    def combined_assets_liabilities(self):
        aggregated_entries = []
        for subspace in self.asset_liability_repository.investment_spaces_library.values():
            aggregated_entries.extend(subspace.entries.items())
        return aggregated_entries

    def get_all_asset_liability_bookkeeping_info(self):
        combined_asset_liability_entries = self.combined_assets_liabilities()
        all_asset_liability_bookkeeping_accounts_info = []

        for entry in combined_asset_liability_entries:
            key, values = entry[0], entry[1]
            portfolio, investment, tax_lot_num, ls, location, financial_account = key
            quantity, local, book = values
            booksp_row = [portfolio, investment, tax_lot_num, ls, location, financial_account, quantity, local,
                          book]
            all_asset_liability_bookkeeping_accounts_info.append(booksp_row)

        return all_asset_liability_bookkeeping_accounts_info

    def all_asset_liability_bookkeeping_accounts_info(self):
        combined_asset_liability_entries = self.combined_assets_liabilities()
        all_asset_liability_bookkeeping_accounts_info = []

        for entry in combined_asset_liability_entries:
            key, values = entry[0], entry[1]
            portfolio, investment, tax_lot_num, ls, location, financial_account = key
            quantity, local, book = values
            booksp_row = [portfolio, investment, tax_lot_num, ls, location, financial_account, quantity, local, book]
            all_asset_liability_bookkeeping_accounts_info.append(booksp_row)

        return all_asset_liability_bookkeeping_accounts_info


    def get_combined_space(self):
        asset_liability_space = self.combined_assets_liabilities()
        revenue_expense_space = self.get_revenue_expense_space()
        combined_space = asset_liability_space + revenue_expense_space
        return combined_space

    def serialize_journal_entries(self, journal_entries, fund):
        """
        Serialize journal entries to a CSV file.

        Args:
            journal_entries (list): List of journal entry objects.
            file_path (str): Path to the output CSV file.
        """
        # Define CSV field names based on journal entry attributes
        fieldnames = [
            'Portfolio', 'Investment', 'Tax Date', 'LS', 'Location', 'Financial Account',
            'Quantity', 'Local', 'Book', 'Tran ID', 'Transaction', 'Trade Date',
            'Settle Date', 'KDBegin', 'KDEnd', 'IBOR Date', 'Entry Type', 'Feeder',
            'Running Balances', 'Split Ratio', 'Account Key'
        ]
        file_path = 'C:/BASE_PATH/refdata/jeoutput/'+fund+'.csv'
        # Open the CSV file in write mode
        with open(file_path, 'w', newline='') as csvfile:
            # Create a CSV writer object
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

            # Write the CSV header
            writer.writeheader()

            # Serialize each journal entry and write it to the CSV file
            for entry in journal_entries:
                serialized_entry = {
                    'Portfolio': entry.portfolio,
                    'Investment': entry.investment,
                    'Tax Date': entry.tax_date,
                    'LS': entry.ls,
                    'Location': entry.location,
                    'Financial Account': entry.financial_account,
                    'Quantity': entry.quantity,
                    'Local': entry.local,
                    'Book': entry.book,
                    'Tran ID': entry.tranid,
                    'Transaction': entry.transaction,
                    'Trade Date': entry.tradedate,
                    'Settle Date': entry.settledate,
                    'KDBegin': entry.kdbegin,
                    'KDEnd': entry.kdend,
                    'IBOR Date': entry.ibor_date,
                    'Entry Type': entry.entry_type,
                    'Feeder': entry.feeder,
                    'Running Balances': entry.running_balances,
                    'Split Ratio': entry.split_ratio,
                    'Account Key': entry.account_key
                }

                writer.writerow(serialized_entry)

        print(f"Serialized journal entries saved to {file_path}")

    def serialize_journal_entries_to_excel(journal_entries, file_path, max_rows_per_sheet=1048575):
        from openpyxl import Workbook
        """
        Serialize journal entries to an Excel file.

        Args:
            journal_entries (list): List of journal entry objects.
            file_path (str): Path to the output Excel file.
            max_rows_per_sheet (int): Maximum number of rows per sheet.
        """
        # Create a new workbook
        wb = Workbook()

        # Split data into multiple sheets
        num_sheets = (len(journal_entries) // max_rows_per_sheet) + 1
        for sheet_idx in range(num_sheets):
            ws = wb.create_sheet(title=f"Sheet_{sheet_idx + 1}")

            # Set headers
            headers = ["Portfolio", "Investment", "Tax Date", "LS", "Location", "Financial Account",
                       "Quantity", "Local", "Book", "Tran ID", "Transaction", "Trade Date", "Settle Date",
                       "KDBegin", "KDEnd", "IBOR Date", "Entry Type", "Feeder", "Running Balances",
                       "Split Ratio", "Account Key"]
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)
                ws.cell(row=1, column=col_idx).font = Font(bold=True)

            # Write data to the sheet
            start_idx = sheet_idx * max_rows_per_sheet
            end_idx = min((sheet_idx + 1) * max_rows_per_sheet, len(journal_entries))
            for row_idx, entry in enumerate(journal_entries[start_idx:end_idx], start=2):
                for col_idx, attr in enumerate(["portfolio", "investment", "tax_date", "ls", "location",
                                                "financial_account", "quantity", "local", "book", "tranid",
                                                "transaction", "tradedate", "settledate", "kdbegin", "kdend",
                                                "ibor_date", "entry_type", "feeder", "running_balances",
                                                "split_ratio", "account_key"], start=1):
                    value = getattr(entry, attr)
                    if isinstance(value, (list, tuple)):
                        value = ', '.join(map(str, value))
                    ws.cell(row=row_idx, column=col_idx, value=value)

        # Save the workbook to the specified file path
        wb.save(file_path)

        print(f"Serialized journal entries saved to {file_path}")

    def get_revenue_expense_space(self):
        return []

    @staticmethod
    def convert_bookkeeping_objects_to_df(bookkeeping_objects):
        data = [{
            'portfolio': obj.account_key[0],
            'investment': obj.account_key[1],
            'tax_date': obj.account_key[2],
            'ls': obj.account_key[3],
            'location': obj.account_key[4],
            'financial_account': obj.account_key[5],
            'quantity': obj.quantity,
            'local': obj.local,
            'book': obj.book,
        } for obj in bookkeeping_objects]
        return pd.DataFrame(data)

    def get_revenue_expense_space(self):
        return []

    @staticmethod
    def convert_bookkeeping_objects_to_df(bookkeeping_objects):
        data = [{
            'portfolio': obj.account_key[0],
            'investment': obj.account_key[1],
            'tax_date': obj.account_key[2],
            'ls': obj.account_key[3],
            'location': obj.account_key[4],
            'financial_account': obj.account_key[5],
            'quantity': obj.quantity,
            'local': obj.local,
            'book': obj.book,
        } for obj in bookkeeping_objects]
        return pd.DataFrame(data)

    # def combined_space(self):
    #     # Similarly, ensure key-value pairs are maintained in this aggregation
    #     combined_entries = []
    #     # Combine asset and liability entries with keys
    #     combined_entries.extend(self.combined_assets_liabilities())
    #     # Include revenue/expense and statistical entries, ensuring keys are preserved
    #     # This example assumes you adjust for key-value pairing as needed
    #     combined_entries.extend(self.revenue_expense_repository.entries)  # Hypothetical example
    #     combined_entries.extend(
    #         [(key, value) for key, value in enumerate(self.statistical_entries)])  # Example for list-based structure
    #     return combined_entries

    def get_revenue_expense_entries(self):
        """Directly access revenue and expense entries."""
        return self.revenue_expense_repository.entries

    def get_revenue_expense_space(self):
        """Retrieve revenue and expense entries in a consistent format."""
        formatted_entries = []
        for entry in self.revenue_expense_repository.entries:
            # Directly construct the key and values based on the known structure
            # Assuming 'entry' has attributes or ways to access:
            # portfolio, investment, tax_lot_num, ls, location, financial_account for the key
            # and quantity, local, book for the values
            key = (
            entry.portfolio, entry.investment, entry.tax_date, entry.ls, entry.location, entry.financial_account)
            values = (entry.quantity, entry.local, entry.book)

            # Append the structured entry to the formatted_entries list
            formatted_entry = (key, values)
            formatted_entries.append(formatted_entry)

        return formatted_entries

    def get_position_space(self, investment):
        """Proxy method to simplify access to investment spaces."""
        return self.asset_liability_repository.get_position_space(investment)

    def post_journal_entry(self, je):
        # Append to journal_entries for audit trail
        self.journal_entries.append(je)

        if je.entry_type == 'Asset/Liability':
            # Direct Asset/Liability entries to their specific investment subspace
            space = self.asset_liability_repository.get_position_space(je.investment)
            # Here, we assume space.post_journal_entry(je) correctly updates or appends entries within the subspace
            space.post_journal_entry_to_subspace(je)  # This method needs to handle the logic as per your system's requirements
        elif je.entry_type == 'Revenue/Expense/Capital':
            # Handle Revenue/Expense/Capital entries via the repository
            self.revenue_expense_repository.add_entry(je)
        elif je.entry_type == 'Statistical':
            # Append Statistical entries to a list or manage via a dedicated repository if implemented
            self.statistical_entries.append(je)
        else:
            raise ValueError("Invalid entry type")

    # Threshold to remove any negligible values
    #     threshold = 0.0001
    #     if abs(new_quantity) < threshold and abs(new_local) < threshold and abs(new_book) < threshold:
    #         entries.remove((key, (new_quantity, new_local, new_book)))
    def print_sub_ledger(self):
        for key, value in self.bs:
            quantity, local, book = value
            print(f"Key: {key} Quantity: {quantity} Local: {local} Book: {book}")

    def sum_bs_quantitys(self):
        total_quantity = sum(entry[1][0] for entry in self.bs)
        return total_quantity

    def sum_bs_book(self):
        total_bv = sum(entry[1][2] for entry in self.bs)
        return total_bv

    def get_all_journal_entries(self):
        # Return combined journal entries from both spaces
        return self.asset_liability_repository.entries + self.revenue_expense_repository.entries

    def add_entry(self, entry):
        # Based on entry type, add to the appropriate space
        # Assuming entry has an attribute 'type'
        if entry.type == 'Asset/Liability':
            self.asset_liability_repository.add_entry(entry)
        elif entry.type == 'Revenue/Expense':
            self.revenue_expense_repository.add_entry(entry)

    def build_sub_ledger_from_journals(self, journals, period_end):
        import time
        bs_start_time = time.time()
        space1 = BookkeepingSpace.create_new_instance()
        # space2 = Bookkeeping()

        # Filter and process JEs for space1
        for idx, je in enumerate(journals):
            if je.ibor_date <= period_end:
                space1.post_journal_entry(je)

                #     # Optional: Print status every N entries
                #     if idx % 10 == 0:
                #         print(f"Processed {idx} journal entries for space1...")
                # else:
                #     break  # Exit loop when period_start is reached
                #
                # # Save space1's state
                # space1_state = space1.__dict__.copy()
                #
                # # Continue processing JEs for space2
                # for je in journals[idx:]:  # Start from the current index
                #     if je.ibor_date <= period_end:
                #         space2.post_journal_entry(je)

                # Optional: Print status every N entries
                if idx % 10 == 0:
                    print(f"Processed {idx} journal entries for space2...")
            else:
                break  # Exit loop when period_end is exceeded

        print("Finished processing all journal entries!")
        bs_end_time = time.time()
        fetch_time = bs_end_time - bs_start_time

        print("\nElapsed time- BS build from data: {:.6f}".format(fetch_time))

        # Return the states of space1 and space2
        return space1

    @property
    def sub_ledger(self):
        return self.asset_liability_entries + self.revenue_expense_entries

    def add_journal_entry(self, entry):
        self.journal_entries.append(entry)

    def process_stock_split(self, investment, split_ratio: float):
        split_quantities = {}
        for je in self.asset_liability_entries:
            if je[0][1] != investment:
                continue  # Skip entries that do not match the specified investment
            pos_type = je[0][3]
            additional_quantity = je[1][0] * split_ratio - je[1][0]
            split_quantities[pos_type] = additional_quantity

        return split_quantities

    def aggregate_entries(self):
        # Combine both asset/liability and revenue/expense entries into a single list
        all_entries = self.asset_liability_entries + self.revenue_expense_entries
        return all_entries

    def lot_iterator_by_custodian(self, investment, entry_type, new_shares_ratio):
        if entry_type == "dividends":
            entries = self.asset_liability_entries
        elif entry_type == "stock_splits":
            entries = self.asset_liability_entries  # Or you can use another appropriate attribute
        else:
            raise ValueError("Invalid entry_type provided")

        for entry in entries:
            key, value = entry
            # Assuming that key is in the format (portfolio, investment, lot_id, pos_type, location, financial_account)
            portfolio, inv, lot_id, pos_type, location, fa = key

            # Perform additional filtering or processing based on the specific entry_type if needed
            # For example, check the investment and new_shares_ratio

            # If the entry meets the desired conditions, yield the relevant lot information
            yield lot_id, value[0] * new_shares_ratio, location

    def query_aggregate_balances(self, query):
        query_parts = query.split('.')
        dimensions = query_parts[:-1]
        metric = query_parts[-1]
        aggregate_dict = self.aggregates
        for dimension in dimensions:
            if dimension not in aggregate_dict:
                return None
            aggregate_dict = aggregate_dict[dimension]

        if metric not in aggregate_dict:
            return None
        return aggregate_dict[metric]


    class Investment:
        def __init__(self, type, name, legs=None):
            self.type = type
            self.name = name
            self.legs = legs or []

        def get_legs_by_exposure(self, exposure_type):
            return [leg for leg in self.legs if leg.exposure_type == exposure_type]

        def report_exposure(self):
            report = {}
            for exposure_type in ExposureType:
                legs = self.get_legs_by_exposure(exposure_type)
                total_notional = sum(leg.notional_amount for leg in legs)
                total_local = sum(leg.local for leg in legs)
                total_book = sum(leg.book for leg in legs)
                report[exposure_type.value] = {
                    "Total Notional": total_notional,
                    "Total Local": total_local,
                    "Total Book": total_book
                }
            return report

    from enum import Enum

    class ExposureType(Enum):
        EQUITY = "Equity"
        FIXED_INCOME = "Fixed Income"
        COMMODITY = "Commodity"
        CURRENCY = "Currency"

    class Leg:
        def __init__(self, leg_name, notional_amount, type, quantity, local, book, exposure_type, investment,
                     day_count=None, accrual_rate=None, reset_date=None):
            self.leg_name = leg_name
            self.notional_amount = notional_amount
            self.type = type
            self.quantity = quantity
            self.local = local
            self.book = book
            self.exposure_type = exposure_type
            self.investment = investment
            self.day_count = day_count
            self.accrual_rate = accrual_rate
            self.reset_date = reset_date

def store_journals(journals):
    fieldnames = ['portfolio', 'investment', 'tax_lot_num', 'ls', 'location', 'financial_account']

    with open("journal_entries.csv", 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(fieldnames)

        for journal in journals:
            journal_tuple = (
                journal.portfolio,
                journal.investment,
                journal.tax_lot_num,
                journal.ls,
                journal.location,
                journal.financial_account
            )
            writer.writerow(journal_tuple)

def update_running_balances( jes, level_preference):
    from collections import defaultdict
    running_totals = defaultdict(lambda: [0, 0, 0])  # Default to (0, 0, 0) for (quantity, local, book)

    # We sort the journal entries to ensure we're processing them in the order they occurred.
    # If they're already in the desired order, this step can be skipped.
    sorted_jes = sorted(jes, key=lambda je:  (str(je.ibor_date),je.tranid is None, je.tranid, je.investment, je.ls,je.location, je.financial_account))
   # sorted_jes = sorted(jes, key=lambda je: je.tranid)  # sort option.

    for je in sorted_jes:
        key = tuple(getattr(je, attr) for attr in level_preference)

        # Update running totals
        running_totals[key][0] += je.quantity
        running_totals[key][1] += je.local
        running_totals[key][2] += je.book

        # Assign the current running totals to this journal entry
        je.running_balances = tuple(running_totals[key])

def parse_journal_entries(file_path):
    with open(file_path, "r") as file:
        lines = file.readlines()

    journal_entries = []
    entry_data = {}
    for line in lines:
        line = line.strip()
        if line.startswith("Entry Count"):
            if entry_data:
                journal_entry = Journals(**entry_data)
                journal_entries.append(journal_entry)
            entry_data = {}
        elif line:
            key, value = line.split(":", 1)
            key = key.strip()
            value = value.strip()
            entry_data[key] = value

    if entry_data:
        journal_entry = Journals(**entry_data)
        journal_entries.append(journal_entry)

    return journal_entries



import csv

import os
def close_a_period(period_name, old_je):
    # Generate the timestamp
    timestamp = datetime.now().strftime('%Y_%m_%d_%H_%M_%S')

    # Create a CSV file with period name and timestamp in the description
    file_name = f"{period_name}_timestamp_{timestamp}.csv"
    file_name = file_name.replace(" ", "_")  # Replace whitespace with underscore
    file_path = f"C:/Users/hjmne/PycharmProjects/chest/PeriodFiles/{file_name}"

    # Check if the file already exists
    if os.path.isfile(file_path):
        print("File already exists.")
        return

    # Prepare the data to be written
    data = []
   #- data.append(['Timestamp', timestamp])
    data.append(['portfolio', 'investment', 'tax_lot_num', 'ls', 'location', 'financial_account',
                 'quantity', 'local', 'book', 'tranid', 'transaction', 'tradedate', 'settledate',
                 'kdbegin', 'kdend','ibor_date', 'running_balances', 'entry_count'])

    for record in old_je:
        data.append([
            record.portfolio,
            record.investment,
            record.tax_lot_num,
            record.ls,
            record.location,
            record.financial_account,
            record.quantity,
            record.local,
            record.book,
            record.tranid,
            record.transaction,
            record.tradedate,
            record.settledate,
            record.kdbegin,
            record.kdend,
            record.ibor_date,
            record.running_balances,
        ])

    # Write the data to the file
    with open(file_path, 'w', newline='') as file:
        writer = csv.writer(file)
        writer.writerows(data)

from datetime import datetime

def check_date_format(date_str, format):
    try:
        datetime.strptime(date_str, format)
        return True
    except ValueError:
        return False

def add_leading_zero(df):
    df.index = df.index.astype(str)  # Ensures that the index is of type string
    df.index = df.index.map(lambda x: x if len(x.split('/')[0]) == 2 or x.split('/')[0] == "0" else '0'+x)
    return df

def format_date(date_string):
    date_obj = datetime.strptime(date_string, '%m/%d/%Y')
    return date_obj.strftime('%m/%d/%Y').lstrip("0").replace("/0", "/")

from dateutil.parser import parse

def parse_date(date_string):
    try:
        return parse(date_string).date()
    except ValueError:
        raise ValueError(f"Unable to parse date from string: {date_string}")

def load_price_data(price_file):
    price_data = {}
    with open(price_file, 'r') as file:
        reader = csv.DictReader(file)
        for row in reader:
            date = parse_date(row['date']).isoformat() # Using the ISO format
            date_obj = datetime.strptime(row['date'], '%m/%d/%Y')  # Parsing date
            date = date_obj.strftime('%m/%d/%Y').lstrip("0").replace("/0", "/")
            ticker = row['ticker']
            currency = row['currency']
            price = float(row['price'])

            if date not in price_data:
                price_data[date] = {}

            # Create a nested dictionary for each ticker containing 'price' and 'currency' fields
            price_data[date][ticker] = {
                'price': price,
                'currency': currency
            }

    return price_data
def load_fx_data(fx_file):
    fx_data = {}
    with open(fx_file, 'r') as file:
        reader = csv.DictReader(file)
        for row in reader:
            date = parse_date(row['date']).isoformat() # Using the ISO format
            date_obj = datetime.strptime(row['date'], '%m/%d/%Y')  # Parsing date
            date = date_obj.strftime('%m/%d/%Y').lstrip("0").replace("/0", "/")
        # Formatting date
            currency = row['currency']
            if date not in fx_data:
                fx_data[date] = {}
            fx_data[date][currency] = float(row['price'])
    return fx_data

import csv

def load_coa_from_csv():
    filename = "C:/Users/hjmne/PycharmProjects/chest/refdata/chart_of_accounts.csv"
    with open(filename, mode='r') as file:
        reader = csv.DictReader(file, delimiter='\t')
        return [row for row in reader]

def get_system_type(coa, system_name):
    for account in coa:
        if account["SystemName"] == system_name:
            return account["SystemType"]
    return None  # return None if no matching SystemName is found

def validate_bookkeeping_entry(entry):
    expected_keys = ['portfolio', 'investment', 'tax_lot', 'ls', 'location', 'financial_account', 'quantity', 'local', 'book']

    # Check if the entry has all the expected keys
    if not all(key in entry for key in expected_keys):
        return False

    # Check data types of the entry
    if not isinstance(entry['portfolio'], str) or \
            not isinstance(entry['investment'], str) or \
            not isinstance(entry['tax_lot'], int) or \
            not isinstance(entry['ls'], str) or \
            not isinstance(entry['location'], str) or \
            not isinstance(entry['financial_account'], str) or \
            not isinstance(entry['quantity'], int) or \
            not isinstance(entry['local'], float) or \
            not isinstance(entry['book'], float):
        return False

    return True

def query_income_activity(journal_entries):
    # Dictionary to hold roll-up data
    rollup_data = {}

    for je in journal_entries:
        financial_account = je.financial_account
        investment = je.investment
        feeder = je.feeder

        if financial_account in ["Income", "DividendReceipt", "PriceGainInvestment", "FXGainInvestment",
                                 "InterestReceipt", "ContributedCost", "FXGainInvestment", "PerfFee", "MgmtFee",
                                 "FXGainLossInvestment", "UnrealGLRevExp", "FXGainTradeSettle"]:
            key = (investment, financial_account, feeder)
            rollup_data.setdefault(key, {"local": 0, "book": 0})

            rollup_data[key]["local"] += je.local  # Assuming 'local' is an attribute of Journals
            rollup_data[key]["book"] += je.book    # Assuming 'book' is an attribute of Journals

    summary_list = []
    for key, values in rollup_data.items():
        investment, account, feeder = key
        summary_list.append({
            "investment": investment,
            "account": account,
            "feeder" : feeder,
            "local": values["local"],
            "book": values["book"]
        })

    return summary_list

def query_balance_sheet_activity(journal_entries):
    # Dictionary to hold roll-up data
    rollup_data = {}

    for je in journal_entries:
        financial_account = je.financial_account
        investment = je.investment
        feeder = je.feeder


        key = (investment, financial_account, feeder)
        rollup_data.setdefault(key, {"quantity": 0, "local": 0, "book": 0})

        rollup_data[key]["quantity"] += je.quantity  # Assuming 'local' is an attribute of Journals
        rollup_data[key]["local"] += je.local  # Assuming 'local' is an attribute of Journals
        rollup_data[key]["book"] += je.book    # Assuming 'book' is an attribute of Journals

    summary_list = []
    for key, values in rollup_data.items():
        investment, account, feeder = key
        summary_list.append({
            "investment": investment,
            "account": account,
            "feeder" : feeder,
            "quantity": values["quantity"],
            "local": values["local"],
            "book": values["book"]
        })

    return summary_list


import os
def parse_chart_of_accounts(filename):
    """Parses the chart of accounts and returns a dictionary."""
    chart_dict = {}
    if not isinstance(filename, str):
        raise TypeError(f"Expected a string for filename, but got {type(filename)} with value {filename}")

    with open(filename, 'r') as file:
        # Skipping the header row
        next(file)

        for line in file:
            parts = line.strip().split('\t')  # Using tab as the delimiter

            # Ensure there are 6 elements, fill in missing with empty string

            while len(parts) < 6:
                parts.append('')

            system_name, space_pref, user_description, gl_account_num, group1, group2 = parts
            chart_dict[system_name] = {
                "SpacePref": space_pref,
                "UserDescription": user_description,
                "GLAccountNum": gl_account_num,
                "Group1": group1,
                "Group2": group2
            }
    return chart_dict


def fetch_mark_records_by_account_key(account_key, all_entries):
    # First, filter the entries based on the given account_key
    filtered_entries = [entry for entry in all_entries if entry.account_key == account_key]

    # Then, filter the filtered_entries by type_name
    asset_records = [entry for entry in filtered_entries if entry.type_name == "UnrealGLAsset"]
    revexp_records = [entry for entry in filtered_entries if entry.type_name == "UnrealGLRevExp"]

    return asset_records, revexp_records


