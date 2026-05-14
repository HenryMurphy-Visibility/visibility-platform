import cProfile
import time
import openpyxl
import loaddata
from shared_data import current_period_data
import gui
import performance
#import generic_domain
import report
import heapq
import utilities
import pivottables
from utilities import kivy_process_inputs
journal_entries = []
import gc
events = []  # This is a heap-based priority queue
#import swaps_domain
import equity_domain
import  bond_domain
import futures_domain
import global_domain
import closed_period
from bookkeeping import BookkeepingSpace,   Journals, EventScheduler, Event, AssetLiabilityRepository, SettlementChores
import bookkeeping
from collections import OrderedDict
import currency_domain
import pandas as pd
event = Event()
import os
from typing import List, Tuple
from utilities import parse_date
from utilities import get_fx_rate, get_price

sub_ledger = BookkeepingSpace()
utilities.load_investment_master_to_aif(sub_ledger, 'c:/BASE_PATH/refdata/investment_master.csv')
utilities.load_bond_info_to_aif(sub_ledger, 'c:/BASE_PATH/refdata/bond_info.csv')


#bookkeeping_space = BookkeepingSpace(check_duplicates=False)

smf = SettlementChores()

# Enable duplicate checking for debugging or specific operations
#bookkeeping_space.enable_duplicate_check()
asset_liability_repository = AssetLiabilityRepository()
on = 0
tranid = 0
import validate
journal_entries = []

events = []  # This is a heap-based priority queue

if not isinstance(sub_ledger, BookkeepingSpace):
    raise TypeError("sub_ledger is not an instance of BookkeepingSpace")

markjes = journal_entries

def schedule_event(date, precedence, function, args):
    # Combine date and precedence into a tuple
    # The heapq module always makes a min-heap, so earliest dates come first
    heapq.heappush(events, ((date, precedence), function, args))
#
# def schedule_event(date, function, args):
#     #Theheapq module always makes a min-heap, so earliest dates come first
#     heapq.heappush(events, (date, function, args))

def run_events_until(end_date):
    while events and events[0][0] <= end_date:
    #    date, function, args = heapq.heappop(events
        date, function, args = heapq.heappop(events)
        function(*args)

def check_journal_balances(journal_entries, threshold=0.01):
    balances = {}
    accts_to_skip = ["PriceGainStatOffset", "FXGainStatOffset", "UnrealGLFX", "UnrealGLPrice", "MarketVal"]

    for entry in journal_entries:
        tranid = entry.tranid
        book = entry.book
        fin_acct = entry.financial_account
        transaction = entry.transaction

        if tranid not in balances:
            balances[tranid] = 0

        # # Skip the accounts that should be excluded
        if fin_acct in accts_to_skip:
            continue

        balances[tranid] += book

    out_of_balance_transactions = {}

    for tranid, balance in balances.items():
        if abs(balance) > threshold:
            out_of_balance_transactions[tranid] = balance

    return out_of_balance_transactions

def map_trade_amount(trade_type, amount):
    if trade_type == 'buy':
        return abs(amount)
    elif trade_type == 'sell':
        return -abs(amount)
    else:
        raise ValueError("Invalid trade type: {}".format(trade_type))

#
# def check_same_sign(num1, num2):
#     if (num1 >= 0 and num2 >= 0) or (num1 <= 0 and num2 <= 0):
#         return True
#     else:
#         return False

def write_closed_period_parameters(current_period_data):

    import json

    # Extract the parameters from the data dictionary
    current_period_start_str = current_period_data.get("current_period_start")
    current_period_cutoff_str = current_period_data.get("current_period_cutoff")
    current_period_knowledge_str = current_period_data.get("current_period_knowledge")
    prior_period_start_str = current_period_data.get("prior_period_start")
    prior_period_cutoff_str = current_period_data.get("prior_period_cutoff")
    prior_period_knowledge_str = current_period_data.get("prior_period_knowledge")
    period_name = current_period_data.get("period_name")
    prior_period_name = current_period_data.get("prior_period_name")
    selected_fund = current_period_data.get("selected_fund")

    # Create a dictionary to store the dates
    dates_dict = {
        "current_period_start": current_period_start_str,
        "current_period_cutoff": current_period_cutoff_str,
        "current_period_knowledge": current_period_knowledge_str,
        "prior_period_start": prior_period_start_str,
        "prior_period_cutoff": prior_period_cutoff_str,
        "prior_period_knowledge": prior_period_knowledge_str,
        "period_name": period_name,
        "prior_period_name": prior_period_name,
        "selected_fund": selected_fund
    }

    # Define the text file path to save the parameters
    txt_filename = "C:/Users/hjmne/PycharmProjects/chest/configs/inputs.txt"

    # Open the text file in append mode and write the dictionary as JSON
    with open(txt_filename, 'a') as txt_file:
        # Convert dictionary to JSON and write it as a string, then add a newline
        txt_file.write(json.dumps(dates_dict) + "\n")

    print(f"Parameters for period {period_name} appended to {txt_filename}")

import pandas as pd


from bookkeeping import AssetLiabilityRepository

investment_master_path = 'c:/BASE_PATH/refdata/investment_master.csv'
bond_info_path = 'c:/BASE_PATH/refdata/bond_info.csv'
repository = sub_ledger.asset_liability_sub_spaces

utilities.load_investment_master_to_aif(repository, investment_master_path)
utilities.load_bond_info_to_aif(repository, bond_info_path)

def save_to_journals(journals_sheet, journal_entries):
    # Clear existing data in the Journals sheet
    journals_sheet.delete_rows(1, journals_sheet.max_row)

    # Write the updated journal entries to the Journals sheet
    for je in journal_entries:
        je_row_data = [je.portfolio, je.investment, je.tax_lot_num, je.ls, je.location, je.financial_account,
                    je.quantity, je.local, je.book, je.tranid, je.transaction, je.ibor_date, je.running_balances[0],
                    je.running_balances[1], je.running_balances[2]]
        journals_sheet.append(je_row_data)

def clear_sheet(sheet):
    sheet.delete_rows(1, sheet.max_row)
def save_to_accounting_space(accounting_space_sheet, sub_ledger):
    # Clear existing data in the AccountingSpace sheet
    accounting_space_sheet.delete_rows(1, accounting_space_sheet.max_row)

    # Write the updated bookkeeping space to the AccountingSpace sheet
    for key, (quantity, local, book) in sub_ledger.bs.items():
        portfolio, investment, tax_date, location, financial_account = key
        booksp_row = [portfolio, investment, tax_date,location, financial_account, quantity, local, book]
        accounting_space_sheet.append(booksp_row)

if not isinstance(sub_ledger, BookkeepingSpace):
                raise TypeError("sub_ledger is not an instance of BookkeepingSpace")

def build_period_mark_list(event_record_args, investments_to_mark):
    tradedate = parse_date(event_record_args['tradedate'], 'tradedate')
    investment = event_record_args['investment']
    portfolio = event_record_args['portfolio']
    price = 0  # Set to 0 initially
    fx_rate = 0  # Set to 0 initially

    if tradedate not in investments_to_mark:
        investments_to_mark[tradedate] = set()

    investments_to_mark[tradedate].add((portfolio, investment, price, fx_rate))

    return tradedate
# In create_mark_events_from_list functiono
def fetch_investment_prices(investment):
    # Placeholder function to fetch investment prices data
    # Replace this with actual implementation
    # Example return value: a DataFrame with 'date' and 'price' columns
    return pd.DataFrame({
        'date': pd.to_datetime(['2023-01-01', '2023-01-02']),
        'price': [100, 102]
    })

def fetch_fx_rates(currency):
    # Placeholder function to fetch FX rates data
    # Replace this with actual implementation
    # Example return value: a DataFrame with 'date' and 'fx_rate' columns
    return pd.DataFrame({
        'date': pd.to_datetime(['2023-01-01', '2023-01-02']),
        'fx_rate': [1.1, 1.2]
    })


def process_events(space_manager, events_sheet,  fund, process_start_date, period_start,
                   period_cutoff, knowledge_cutoff,  journal_entries, sub_ledger, general_ledger,
                    tdate_fx, scheduler, stat_repo, price_data,
                   fx_data, mark_daily, aggregate_marks, include_marks):
    space_manager.clear_space('sub_ledger')
    space_manager.clear_space('general_ledger')

    last_processed_date = period_start
    investments_to_mark = {}

    if not isinstance(sub_ledger, BookkeepingSpace):
        raise TypeError("sub_ledger is not an instance of BookkeepingSpace")

    import time
    start_time = time.time()
    marked = False

    retrieved_events_records = []
    TPrecedence = 1075
    TOpenPrecedence = 3111
    TOptionPrecedence = 3275
    TClosePrecedence = 3211
    SPrecedence = 1111
    CSplitPrecedence = 2111
    CDivPrecedence = 2122
    markPrecedence = 9000
    allocatePrecedence = 9500

    counter = 1

    f_time = time.time()
    fetch_time = f_time - start_time

    import csv
    from collections import OrderedDict

  #  print("Reading in for fund " + fund)
    start_time = time.time()

    csv_file_path = 'C:/Users/hjmne/PycharmProjects/chest/refdata/pooltest/' + fund + '.csv'
    retrieved_events_records = []

    with open(csv_file_path, mode='r', encoding='utf-8') as csvfile:
        csv_reader = csv.DictReader(csvfile)
        for row in csv_reader:
            retrieved_events_records.append(row)

 #   print("Finished the retrieval")
    from datetime import timedelta

    def is_business_day(date):
        return date.weekday() < 5  # Monday to Friday are business days

    # Generate a list of business days between two dates
    def generate_business_days(start_date, end_date):
        current_date = start_date
        business_days = []
        while current_date <= end_date:
            if is_business_day(current_date):
                business_days.append(current_date)
            current_date += timedelta(days=1)
        return business_days

    # Build a list of candidates
    def build_candidates(retrieved_events_records, period_end):
        candidates = {}
        for event_record in retrieved_events_records:
            tradedate = parse_date(event_record['tradedate'], 'tradedate')
            if tradedate > period_end:
                continue

            investment = event_record['investment']
            portfolio = event_record['portfolio']
            if not investment:
                continue

            key = (portfolio, investment)
            if key not in candidates:
                candidates[key] = tradedate
        #    print(f"Added to candidate list: {portfolio}, {investment}, {tradedate}")

        return candidates

    def create_mark_events(scheduler, candidates, period_start, period_end, sub_ledger, price_data, fx_data,
                           market="US"):
        business_days = generate_business_days(period_start, period_end)
        marked_currencies = set()  # Track which currencies have been marked each day

        for (portfolio, investment), tradedate in candidates.items():
            # Retrieve the currency for the investment using the original method
            currency = sub_ledger.get_investment_attribute('AIF', investment, 'Currency')

            if currency is None:
                print(f"Warning: Currency for investment {investment} not found. Skipping.")
                continue

            for business_day in business_days:
                if business_day >= tradedate:
                    # Process marks for the investment itself
                    if (investment, business_day) not in marked_currencies:
                        price = get_price(investment, business_day, price_data)
                        try:
                            fx_rate = get_fx_rate(currency, business_day, fx_data)  # FX rate lookup
                        except ValueError as e:
                            print(f"Error: {e}")
                            continue  # Skip this event if FX rate is not found

                        # Schedule the mark event for the investment
                        scheduler.schedule_event(
                            business_day,
                            global_domain.mark_event,
                            portfolio,
                            investment,
                            business_day,
                            stat_repo,
                            sub_ledger,
                            price,
                            fx_rate
                        )

                        # Mark the investment as processed for this day
                        marked_currencies.add((investment, business_day))

                    # Process marks for the associated currency (only if it hasn't been marked yet for this day)
                    if (currency, business_day) not in marked_currencies:
                        try:
                            fx_rate = get_fx_rate(currency, business_day, fx_data)
                        except ValueError as e:
                            print(f"Error: {e}")
                            continue  # Skip this event if FX rate is not found

                        # Schedule the mark event for the currency
                        scheduler.schedule_event(
                            business_day,
                            global_domain.mark_event,
                            portfolio,
                            currency,
                            business_day,
                            stat_repo,
                            sub_ledger,
                            1,  # Currency price is typically 1
                            fx_rate
                        )

                        # Mark the currency as processed for this day
                        marked_currencies.add((currency, business_day))

        print(f"Scheduled marks for candidates: {candidates}")

    # Main processing loop
    for event_record_args in retrieved_events_records:
        portfolio = event_record_args['portfolio']
        method = event_record_args['method']
        tradedate = event_record_args['tradedate']
        settledate = event_record_args['settledate']
        kdbegin = event_record_args['kdbegin']
        kdend = event_record_args['kdend']
        investment = event_record_args['investment']
        payment_currency = event_record_args['payment_currency']
        tdate_fx = float(event_record_args['tdate_fx'].strip()) if event_record_args['tdate_fx'].strip() else 0
        location = event_record_args['location']
        strategy = event_record_args['strategy']
        quantity = float(event_record_args['quantity'].strip()) if event_record_args['quantity'].strip() else 0
        price = float(event_record_args['price'].strip()) if event_record_args['price'].strip() else 0
        notional = float(event_record_args['notional'].strip()) if event_record_args['notional'].strip() else None
        original_face = float(event_record_args['original_face'].strip()) if event_record_args[
            'original_face'].strip() else None
        local = float(event_record_args['total_amount'].strip()) if event_record_args['total_amount'].strip() else 0
        book = float(event_record_args['total_amount_base'].strip()) if event_record_args[
            'total_amount_base'].strip() else 0
        tranid = event_record_args['tranid']
        transaction = event_record_args['transaction']
        accrued_local = float(event_record_args['accrued_local'].strip()) if event_record_args[
            'accrued_local'].strip() else 0
        accrued_book = float(event_record_args['accrued_book'].strip()) if event_record_args[
            'accrued_book'].strip() else 0
        new_shares = float(event_record_args['new_shares'].strip()) if event_record_args[
            'new_shares'].strip() else 0
        old_shares = float(event_record_args['old_shares'].strip()) if event_record_args[
            'old_shares'].strip() else 0
        per_share = float(event_record_args['per_share'].strip()) if event_record_args['per_share'].strip() else 0
        legin = event_record_args['legin']
        legout = event_record_args['legout']
        allocation_entities = event_record_args['allocation_entities']
        allocation_percents = event_record_args['allocation_percents']
        financial_account = event_record_args['financial_account']
        buy_currency = event_record_args['buy_currency']
        sell_currency = event_record_args['sell_currency']
        if event_record_args['buy_amt']:
            buy_amt = float(event_record_args['buy_amt'])
        else:
            buy_amt = 0.0
        if event_record_args['sell_amt']:
            sell_amt = float(event_record_args['sell_amt'])
        else:
            sell_amt = 0.0
        feeder = event_record_args['feeder']
        put_call = event_record_args['put_call']
        strike = float(event_record_args['strike'].strip()) if event_record_args['strike'].strip() else 0
        underlying = event_record_args['underlying']
        if isinstance(tradedate, str):
            tradedate = parse_date(tradedate, 'tradedate')
        if isinstance(settledate, str):
            settledate = parse_date(settledate, 'settledate')
        if isinstance(kdbegin, str):
            kdbegin = parse_date(kdbegin, 'kdbegin')
        if isinstance(kdend, str):
            kdend = parse_date(kdend, 'kdend')

        if tradedate > period_cutoff or \
                (kdend is not None and kdend < knowledge_cutoff) or \
                (kdbegin is not None and kdbegin > knowledge_cutoff):
            continue

        if tranid == 0:
            continue

        if method == "buy_equity":
            scheduler.schedule_event(tradedate, equity_domain.buy_equity, portfolio, investment,
                                     location, quantity, local, book, journal_entries, sub_ledger,
                                     tranid,
                                     transaction, tradedate, settledate, kdbegin, kdend, payment_currency, tdate_fx)
            if tradedate != settledate and settledate <= period_cutoff:
                scheduler.schedule_event(settledate, currency_domain.settle_single_flow_out, portfolio,
                                         payment_currency, location, quantity, local, book, journal_entries,
                                         sub_ledger, tranid, "Settlement", tradedate, settledate,
                                         kdbegin, kdend, fx_data)
        if method == "buy_bond":
            scheduler.schedule_event(tradedate,  bond_domain.buy_bond, portfolio,
                                     investment,
                                     location, quantity, local, book, journal_entries, sub_ledger, tranid,
                                     transaction, tradedate, settledate, kdbegin, kdend, payment_currency, smf, accrued_local, accrued_book)

            if tradedate != settledate and settledate <= period_cutoff:
                scheduler.schedule_event(settledate, currency_domain.settle_bond_flow_out,
                                         portfolio, payment_currency, investment, location, quantity, local, book, journal_entries,
                                         sub_ledger, tranid, "Settlement", tradedate, settledate, kdbegin,
                                         kdend, smf, accrued_local, accrued_book, fx_data)

                # Schedule another event to update SMF record status
                scheduler.schedule_event(settledate, bond_domain.schedule_update_smf_record_status, smf,
                                         tranid, "Settled", portfolio)
        if method == "buy_future":
            scheduler.schedule_event(tradedate, futures_domain.buy_future, portfolio, investment,
                                     location, quantity, local, book, journal_entries, sub_ledger,tranid,
                                     transaction, tradedate, settledate, kdbegin, kdend, payment_currency,
                                      tdate_fx, notional, price)

            if tradedate != settledate and settledate < period_cutoff and local !=0:
                scheduler.schedule_event(settledate, currency_domain.settle_single_flow_out, portfolio,
                                         payment_currency, location, quantity, local, book, journal_entries,
                                         sub_ledger, tranid, "Settlement", tradedate, settledate, kdbegin,
                                         kdend, fx_data)
        elif method == "sell_equity":
            closing_method = "FIFO"
            scheduler.schedule_event(tradedate, equity_domain.sell_equity, portfolio, investment,
                                     location, quantity, local, book, closing_method, journal_entries,
                                     sub_ledger, tranid, transaction, tradedate, settledate, kdbegin, kdend,
                                     payment_currency, tdate_fx)

            if tradedate != settledate and settledate < period_cutoff:
                scheduler.schedule_event(settledate, currency_domain.settle_single_flow_in, portfolio,
                                         payment_currency, location, quantity, local, book, journal_entries,
                                         sub_ledger, tranid, "Settlement", tradedate, settledate, kdbegin,
                                         kdend, fx_data)

        elif method == "sell_bond":
            closing_method = "FIFO"
            scheduler.schedule_event(tradedate, bond_domain.sell_bond, portfolio, investment,
                                     location, quantity, local, book, closing_method, journal_entries,
                                     sub_ledger, tranid, transaction, tradedate, settledate, kdbegin, kdend,
                                     payment_currency, smf, accrued_local,
                                     accrued_book)

            if tradedate != settledate and settledate <= period_cutoff:
                scheduler.schedule_event(settledate, currency_domain.settle_bond_flow_in, portfolio,
                                         payment_currency, investment, location, quantity, local, book, journal_entries,
                                         sub_ledger, tranid, "Settlement", tradedate, settledate, kdbegin,
                                         kdend,  smf, accrued_local, accrued_book)
                #Schedule another event to update SMF record status
                scheduler.schedule_event(settledate,
                                         bond_domain.schedule_update_smf_record_status, smf,
                                         tranid, "Settled", portfolio)
        elif method == "sell_future":
            closing_method = "FIFO"
            scheduler.schedule_event(tradedate, futures_domain.sell_future, portfolio, investment,
                                     location, quantity, local, book, closing_method, journal_entries,
                                     sub_ledger, tranid, transaction, tradedate, settledate, kdbegin, kdend,
                                     payment_currency, tdate_fx, notional, price)

            if tradedate != settledate and settledate <= period_cutoff:
                # Schedule settlement event
                scheduler.schedule_event(
                    settledate, currency_domain.settle_pay_rec_by_tranid, portfolio, investment, location,  quantity, local, book, journal_entries,
                    sub_ledger, tranid, "FutureSettlement", tradedate, settledate, kdbegin,
                        kdend, payment_currency, smf, fx_data)


        elif method == "short_equity":
            scheduler.schedule_event(tradedate,  equity_domain.short_equity, portfolio, investment,
                                     location, quantity, local, book, journal_entries, sub_ledger, tranid,
                                     transaction, tradedate, settledate, kdbegin, kdend, payment_currency,
                                      tdate_fx)

            if tradedate != settledate and settledate <= period_cutoff:
                scheduler.schedule_event(settledate,  currency_domain.settle_single_flow_in, portfolio,
                                         payment_currency, location, quantity, local, book, journal_entries,
                                         sub_ledger, tranid, "Settlement", tradedate, settledate, kdbegin,
                                         kdend, fx_data)

        if method == "short_bond":
            scheduler.schedule_event(tradedate, bond_domain.short_bond, portfolio, investment,
                                     location, quantity, local, book, journal_entries, sub_ledger, tranid,
                                     transaction, tradedate, settledate, kdbegin, kdend, payment_currency,
                                      smf, accrued_local, accrued_book)

            def settle_bond_flow_in(portfolio, payment_currency, investment, location, quantity, local,
                                    book, journal_entries, sub_ledger, tranid, transaction, tradedate,
                                    settledate,
                                    kdbegin, kdend, smf, accrued_local, accrued_book, fx_data):

                # Schedule another event to update SMF record status
                scheduler.schedule_event(settledate,
                                         bond_domain.schedule_update_smf_record_status, smf,
                                         tranid, "Settled", portfolio)
        if method == "short_future":
            scheduler.schedule_event(tradedate, futures_domain.buy_future, portfolio, investment,
                                     location, quantity, local, book, journal_entries, sub_ledger, tranid,
                                     transaction, tradedate, settledate, kdbegin, kdend, payment_currency,
                                     tdate_fx, notional, price)

            if tradedate != settledate and settledate <= period_cutoff and local != 0:
                scheduler.schedule_event(settledate, currency_domain.settle_single_flow_out, portfolio,
                                         payment_currency, location, quantity, local, book, journal_entries,
                                         sub_ledger, tranid, "Settlement", tradedate, settledate, kdbegin,
                                         kdend, fx_data)
        elif method == "cover_equity":
            closing_method = "FIFO"
            scheduler.schedule_event(tradedate, equity_domain.cover_equity, portfolio, investment,
                                     location, quantity, local, book, closing_method, journal_entries,
                                      sub_ledger, tranid, transaction, tradedate, settledate, kdbegin, kdend,
                                     payment_currency, tdate_fx )

            if tradedate != settledate and settledate <= period_cutoff:
                scheduler.schedule_event(settledate, currency_domain.settle_single_flow_out, portfolio,
                                         payment_currency, location, quantity, local, book, journal_entries,
                                         sub_ledger, tranid, "Settlement", tradedate, settledate, kdbegin,
                                         kdend, fx_data)
        elif method == "cover_bond":
            closing_method = "FIFO"
            scheduler.schedule_event(tradedate, bond_domain.cover_bond, portfolio, investment,
                                     location, quantity, local, book, closing_method, journal_entries,
                                     sub_ledger, tranid, transaction, tradedate, settledate, kdbegin, kdend,
                                     payment_currency, smf, accrued_local,
                                     accrued_book)

            if tradedate != settledate and settledate <= period_cutoff:
                scheduler.schedule_event(settledate, currency_domain.settle_bond_flow_out,
                                         portfolio,
                                         payment_currency, investment, location, quantity, local, book, journal_entries,
                                         sub_ledger, tranid, "Settlement", tradedate, settledate, kdbegin,
                                         kdend, smf, accrued_local,
                                         accrued_book, fx_data)
                # Schedule another event to update SMF record status
                scheduler.schedule_event(settledate,
                                         bond_domain.schedule_update_smf_record_status, smf,
                                         tranid, "Settled", portfolio)

        elif method == "cover_future":
            closing_method = "FIFO"
            scheduler.schedule_event(tradedate, futures_domain.sell_future, portfolio, investment,
                                     location, quantity, local, book, closing_method, journal_entries,
                                     sub_ledger, tranid, transaction, tradedate, settledate, kdbegin, kdend,
                                     payment_currency, tdate_fx, notional, price)

            if tradedate != settledate and settledate <= period_cutoff:
                # Schedule settlement event
                scheduler.schedule_event(
                    settledate, currency_domain.settle_pay_rec_by_tranid, portfolio, investment, location, quantity,
                    local, book, journal_entries,
                    sub_ledger, tranid, "FutureSettlement", tradedate, settledate, kdbegin,
                    kdend, payment_currency, smf, fx_data)


        elif method == "dividend_equity":
            scheduler.schedule_event(tradedate, equity_domain.dividend_equity, portfolio, investment,
                                     journal_entries,sub_ledger, tranid, transaction, tradedate, settledate, kdbegin, kdend,
                                     payment_currency, per_share, period_start)

            if tradedate != settledate and settledate <= period_cutoff:
                financial_account_in = "DividendsReceivable"
                financial_account_out = "DividendsPayable"
                scheduler.schedule_event(settledate, currency_domain.settle_multiple_flows_in_out, portfolio,
                                        payment_currency, investment, financial_account_in, financial_account_out, journal_entries,
                                         sub_ledger, tranid, "Settlement", tradedate, settledate, kdbegin,
                                         kdend, smf, fx_data)
                # portfolio, payment_currency, financial_account_in, financial_account_out,
                # journal_entries, sub_ledger, tranid, transaction, tradedate, settledate,
                # kdbegin, kdend):

        elif method == "bond_coupon":
            scheduler.schedule_event(tradedate, bond_domain.bond_coupon, portfolio, investment,
                                     journal_entries,sub_ledger, tranid, transaction, tradedate, settledate, kdbegin, kdend,
                                     payment_currency, per_share,  smf)

            if tradedate != settledate and settledate <= period_cutoff:
                financial_account_in = "InterestReceivable"
                financial_account_out = "InterestPayable"
                scheduler.schedule_event(settledate,  currency_domain.settle_multiple_flows_in_out, portfolio,
                                        payment_currency, investment, financial_account_in, financial_account_out, journal_entries,
                                         sub_ledger, tranid, "Settlement", tradedate, settledate, kdbegin,
                                         kdend, smf, fx_data)

        elif method == "split_equity":
            scheduler.schedule_event(tradedate, equity_domain.split_equity, portfolio, investment,
                                     journal_entries, sub_ledger, tranid, transaction, tradedate, settledate,
                                     kdbegin, kdend, new_shares, old_shares)



        elif method == "deposit_currency":
            scheduler.schedule_event(tradedate, currency_domain.deposit_currency, portfolio,
                                     payment_currency, location, quantity, local, book, journal_entries,
                                     sub_ledger, tranid, transaction, tradedate, settledate, kdbegin, kdend, fx_data)

        elif method == "swap_open":
            scheduler.schedule_event(tradedate, swaps_domain.swap_open, portfolio, investment,
                                     location, quantity, local, book, journal_entries, sub_ledger, tranid,
                                     transaction, tradedate, settledate, kdbegin, kdend, payment_currency, period_start,  smf, legin, legout)

        elif method == "withdraw_currency":
            scheduler.schedule_event(tradedate, currency_domain.withdraw_currency, portfolio,
                                     payment_currency, location, local, local, book, journal_entries,
                                     sub_ledger, tranid, transaction, tradedate, settledate, kdbegin, kdend,
                                     period_start, tdate_fx, fx_data)
        elif method == "perfmark":
            if create_performance:
                scheduler.schedule_event(tradedate, global_domain.performance_mark,
                                         tradedate, sub_ledger,
                                          tranid)

        elif method == "allocate":
            scheduler.schedule_event(tradedate, global_domain.allocate, portfolio,
                                 investment, location, quantity, local, book,
                                 tranid, transaction, tradedate, settledate, kdbegin, kdend, period_start,
                                 period_cutoff,  smf, allocation_entities, allocation_percents)

        elif method == "expense":
            scheduler.schedule_event(tradedate, currency_domain.expense, portfolio,
                        payment_currency, location, quantity, local, book, financial_account, journal_entries,
                        sub_ledger, tranid, transaction, tradedate, settledate, kdbegin, kdend, period_start, investment_accounting_space)

        elif method == "spot_fx":
            scheduler.schedule_event(tradedate, currency_domain.spotfx, portfolio, investment, location,
                                     quantity, local, book, journal_entries, sub_ledger, tranid,transaction, tradedate,
                                     settledate, kdbegin, kdend, buy_currency, sell_currency, buy_amt,
                                     sell_amt)
            if settledate <= period_cutoff:
                if tradedate == settledate: #temp
                    SPrecedence = 8000
                scheduler.schedule_event(settledate, currency_domain.settle_single_flow_out, portfolio,
                                         sell_currency, location, sell_amt, sell_amt, buy_amt, journal_entries,
                                         sub_ledger, tranid, "SpotSettlement", tradedate, settledate, kdbegin,
                                         kdend, smf, fx_data)
                scheduler.schedule_event(settledate, currency_domain.settle_single_flow_in, portfolio,
                                         buy_currency, location, buy_amt, buy_amt, buy_amt, journal_entries,
                                         sub_ledger, tranid, "SpotSettlement", tradedate, settledate, kdbegin,
                                         kdend, smf, fx_data)
        elif method == "exercise_option_open":
            closing_method = "FIFO"
            scheduler.schedule_event(tradedate, equity_domain.exercise_option_open, portfolio,
                                     investment,
                                     location, quantity, local, book, closing_method, journal_entries,
                                     sub_ledger, tranid, transaction, tradedate, settledate, kdbegin, kdend,
                                     payment_currency,put_call, strike, underlying, tdate_fx)

            if tradedate != settledate and settledate <= period_cutoff:
                financial_account_in = "Receivable"
                financial_account_out = "Payable"
                scheduler.schedule_event(settledate, currency_domain.settle_multiple_flows_in_out, portfolio,
                                        payment_currency, investment, financial_account_in, financial_account_out, journal_entries,
                                         sub_ledger, tranid, "Settlement", tradedate, settledate, kdbegin,
                                         kdend,  smf, fx_data)
        elif method == "exercise_option_close":
            closing_method = "FIFO"
            scheduler.schedule_event(tradedate, equity_domain.exercise_option_close,
                                     portfolio,
                                     investment,
                                     location, quantity, local, book, closing_method, journal_entries,
                                     sub_ledger, tranid, transaction, tradedate, settledate, kdbegin,
                                     kdend,payment_currency,put_call, strike, underlying, tdate_fx)

            if tradedate != settledate and settledate <= period_cutoff:
                financial_account_in = "Receivable"
                financial_account_out = "Payable"
                scheduler.schedule_event(settledate,
                                         currency_domain.settle_multiple_flows_in_out, portfolio,
                                         payment_currency, investment, financial_account_in, financial_account_out,
                                         journal_entries,
                                         sub_ledger, tranid, "Settlement", tradedate, settledate,
                                         kdbegin,
                                         kdend, smf, fx_data)
        elif method == "assign_option_open":
            closing_method = "FIFO"
            scheduler.schedule_event(tradedate, equity_domain.assign_option_open,
                                     portfolio,
                                     investment,
                                     location, quantity, local, book, closing_method, journal_entries,
                                     sub_ledger, tranid, transaction, tradedate, settledate, kdbegin,
                                     kdend,payment_currency,put_call, strike, underlying, tdate_fx)

            if tradedate != settledate and settledate <= period_cutoff:
                financial_account_in = "Receivable"
                financial_account_out = "Payable"
                scheduler.schedule_event(settledate,
                                         currency_domain.settle_multiple_flows_in_out, portfolio,
                                         payment_currency, investment, financial_account_in, financial_account_out,
                                         journal_entries,
                                         sub_ledger, tranid, "Settlement", tradedate, settledate,
                                         kdbegin,
                                         kdend,  smf, fx_data)
        elif method == "assign_option_close":
            closing_method = "FIFO"
            scheduler.schedule_event(tradedate, equity_domain.assign_option_close,
                                     portfolio,
                                     investment,
                                     location, quantity, local, book, closing_method, journal_entries,
                                     sub_ledger, tranid, transaction, tradedate, settledate, kdbegin,
                                     kdend,payment_currency, put_call, strike, underlying, tdate_fx)

            if tradedate != settledate and settledate <= period_cutoff:
                financial_account_in = "Receivable"
                financial_account_out = "Payable"
                scheduler.schedule_event(settledate,
                                         currency_domain.settle_multiple_flows_in_out, portfolio,
                                         payment_currency, investment, financial_account_in, financial_account_out,
                                         journal_entries,
                                         sub_ledger, tranid, "Settlement", tradedate, settledate,
                                         kdbegin,
                                         kdend,  smf, fx_data)

            # def spotfx(portfolio, investment, location, qty, local, book, journal_entries, sub_ledger, tranid,
            #            transaction, tradedate, settledate, kdbegin, kdend, buy_currency, sell_currency,
            #            buy_amt, sell_amt, period_start, investment_accounting_space):

        counter += 1
        if counter == 1:
            schedule_time = s_time - f_time
            # print("\nElapsed time- Reading the file: {:.6f}".format(schedule_time))

        if counter % 1000 == 0:
            s_time = time.time()
            schedule_time = s_time - f_time
            # print("\nElapsed time- Scheduling Events:"+str(counter)+" {:.6f}".format(schedule_time))

    counter_start_time = time.time()
    process_begin = time.time()
    testcount =0

   # sort must be in ascending orde  r for processing to be sequenced correctly!!!
   # scheduler.sort_events()
   #  investments_to_mark = build_cumulative_mark_list(retrieved_events_records, period_cutoff)
   #
   #  create_mark_events_from_list(scheduler, investments_to_mark, last_processed_date, journal_entries,
   #                               sub_ledger, price_data, fx_data)

    if include_marks:
        candidates = build_candidates(retrieved_events_records, period_cutoff)
        create_mark_events(scheduler, candidates, period_start, period_cutoff,
                           sub_ledger, price_data, fx_data, mark_daily)
    print("Scheduling complete for portfolio"+ portfolio)
    # Sort events based on precedence and tradedate
    scheduler.sort_events()

    # Process events
    while scheduler.events:
        try:
            totcount = scheduler.run_next_event()
        except KeyError as e:
            print(f"KeyError processing event: {e}")
        except TypeError as e:
            print(f"TypeError processing event: {e}")

    print("Processing complete for portfolio"+ portfolio)
    after_process_time = time.time()

    elapsed_time = after_process_time - process_begin # TAKE OUT STARTUP TIME FOR KEY SETUP?4 = {dict: 4} {'args': ('MyBondPortfolio', 'FUTURE', 'Goldman', 5.0, 0.0, 0.0, [], <bookkeeping.BookkeepingSpace object at 0x0000023E19CDBED0>, '3', 'Settlement', datetime.datetime(2022, 1, 4, 0, 0), datetime.datetime(2022, 1, 10, 0, 0), datetime.datetime(2022, 1, 4, 0,... View
  #  print("\nElapsed raw transactions processing time after scheduling: {:.6f}".format(elapsed_time))
    # if elapsed_time != 0:
    #     print("\nTransactions per second:", totcount/(elapsed_time )) # TAKE OUT EVENT KEY SETUP


    combined_asset_liability_entries_for_marking = sub_ledger.combined_assets_liabilities()
    all_asset_liability_for_marking = []
    # Assuming combined_entries is structured correctly as expected
    for entry in combined_asset_liability_entries_for_marking:
        key, values = entry  # Unpack key and values
        # Unpack the key tuple
        portfolio, investment, lotid, tax_lot_num, ls, location, financial_account = key

        # Unpack the values tuple, now including notional, original_face, and settlement_status
        quantity, local, book, notional, oface = values

        # Prepare the row for output, including new fields
        booksp_row = [portfolio, investment, lotid, tax_lot_num, ls, location, financial_account,
                      quantity, local, book, notional, oface]

        all_asset_liability_for_marking.append(booksp_row)

    # MARK-TO-MARKET
    # rever se any previous marks and mark again yielding deltas for the period
    #
    # from mark_to_market import accounting_mark_logic
    # accounting_mark_logic(period_cutoff, sub_ledger, tranid)
    # from mark_to_market import mark_to_market, post_accounting_marks
    # marked_results = mark_to_market(sub_ledger, PERIOD_END_DATE)
    # post_accounting_marks(marked_results, period_cutoff, sub_ledger, tpranid=None)



    #  bookkeeping.store_journals(journal_entries)
    journal_entries_dict = {}

    for entry in sub_ledger.journal_entries:
        key = (entry.portfolio, entry.investment, entry.lotid, entry.tax_date, entry.ls, entry.location, entry.financial_account)
        if key in journal_entries_dict:
            journal_entries_dict[key].append(entry)
        else:
            journal_entries_dict[key] = [entry]

   # print(isinstance(investment_accounting_space, Bookkeeping))
   # Example usage
    combined_space_custom = sub_ledger.get_combined_space()
    # Assuming combined_entries is structured correctly as expected
    all_bookkeeping_accounts_list = []
    for entry in combined_space_custom:
        key, values = entry[0], entry[1]  # Extracting key and values from each entry
        # Unpacking key into individual variables
        portfolio, investment, lotid,  tax_date, ls, location, financial_account = key

        # Using unpack_values to safely extract data from values
        quantity, local, book = utilities.unpack_values(values)

        # Create a bookkeeping row with all the data
        booksp_row = [
            portfolio, investment, lotid,  tax_date, ls, location, financial_account,
            quantity, local, book
        ]

        # Append this row to the list of all bookkeeping accounts
        all_bookkeeping_accounts_list.append(booksp_row)

    combined_asset_liability_entries = []
    combined_asset_liability_entries = sub_ledger.combined_assets_liabilities()
    all_asset_liability_accounts_list = []
    # Assuming combined_entries is structured correctly as expected
    for entry in combined_asset_liability_entries:
        key, values = entry[0], entry[1]  # If combined_entries contains entries as ((key), (values))
        # Or directly, if combined_entries is indeed structured as (key, values)
        portfolio, investment, lotid, tax_date, ls, location, financial_account = key
        # Using unpack_values to safely extract data from values
        quantity, local, book = utilities.unpack_values(values)

        booksp_row = [portfolio, investment, lotid,  tax_date, ls, location, financial_account, quantity, local, book]
        all_asset_liability_accounts_list.append(booksp_row)
    #
    # # for key, v in combined_entries:
    #     portfolio, investment, tax_lot_num, ls, location, financial_account = key
    #     quantity, local, book = v[0], v[1], v[2]
    #     booksp_row = [portfolio, investment, tax_lot_num, ls, location, financial_account, quantity, local, book]
    #     sub_ledger_list.append(booksp_row)
    #
  #  sub_ledger.serialize_journal_entries(investment_accounting_space.journal_entries, fund)

    #VERIFY JOURNALS ARE BALANCED
    out_of_balance = check_journal_balances(journal_entries, threshold=0.01)
    print(out_of_balance)

    return
    # review any out of balance transactions
    out_of_balance = check_journal_balances(sub_ledger.journal_entries, threshold=0.01)
    print(out_of_balance)  # Output: {1: 50.0}

    return
    # Verify that the bookkeeping space is balanced
    # total_bv = sub_ledger.sum_bs_book()
    # print("\nBalance Check: {:.6f}".format(total_bv))
    #
    # if total_bv > -.01 and total_bv < .01:
    #     print("Bookkeeping space is balanced.")
    # else:
    #     print("Bookkeeping space is not balanced by ", total_bv)


   # prepare data for reporting

    # # Example of generating a report
    # asset_liability_repo = asset_liability_repository  # Assuming bookkeeping is your Bookkeeping instance
    # all_assets_and_liabilites = asset_liability_repo.aggregate_entries_for_reporting()
    # # sub_ledger = BookkeepingSpace()
    # if timeset == "current" and create_performance and not close_period:
    #     filen = 'AccountingResultsCurrentPeriod'+fund
    #     fname = 'C:/Users/hjmne/PycharmProjects/chest/reports/PerformanceSummaries.xlsx'
    #     performance.create_performance_sheets(investment_accounting_space.journal_entries)
    #
    # if timeset == "current" and compare_to_prior_period:
    #     filen = 'AccountingResultsCurrentPeriod'+fund
    #     report.create_accounting_reports(investment_accounting_space.journal_entries, filen,
    #                                      all_asset_liability_accounts_list, period_cutoff)


           # filen = '--Managerial--'+fund
           # report.journals_style_reports(investment_accounting_space.journal_entries,
           #                           sub_ledger,
           #                           period_start, period_cutoff, filen)
           #
         #  filen = '--Managerial--'+portfolio
         #   filen = f'--Managerial--'+fund
         #   report.position_report_by_sector(all_asset_liability_accounts_list, period_cutoff,knowledge_cutoff,filen)

           #filen = '--Managerial--'+fund
           # report.realized_gains_losses(sub_ledger.journal_entries,
           #                               sub_ledger, period_start,
           #                               period_cutoff,
           #                               filen)
           # filen = '--Managerial--'+fund
           # report.total_income_earned(sub_ledger.journal_entries,
           #                             sub_ledger, period_start, period_cutoff,
           #                             filen)
           # #
           # filen = '--Managerial--'+fund
           # report.journals_by_tranid(journal_entries,
           #                            sub_ledger, period_start, period_cutoff,
           #                            filen)

           # filen = '--Managerial--'+fund
           # report.journals_by_posting_order(journal_entries,
           #                            sub_ledger, period_start, period_cutoff,
           #                            filen)

           #
           # filen = '--Managerial--'+fund
           # report.valuation_style_reports(all_asset_liability_accounts_list, period_cutoff, filen)
           # #
           #
           #
           # filen = '--Managerial--'+fund
           # report.cost_basis_balance_sheet(all_bookkeeping_accounts_list, period_cutoff, filen)

           #
           # filen = '--Managerial--'+fund
           # report.position_report_by_asset_class(all_asset_liability_accounts_list, period_cutoff, filen)

#-----------------------------General Ledger Reporting-----------------------------

     #      import testreports
       #    space = testreports.prepare_gl_data_for_reporting(period_cutoff, fund)
       #    gl_space, jes, al, alc = space # jes-journals, al - assets/liabilities, alc - all bookkkeeping acounts

           # filen = 'General Ledger'+fund
           # report.position_report_by_asset_class(al, period_cutoff, filen)


           # filen = '--General Ledger--'+fund
           # report.cost_basis_balance_sheet(alc, period_cutoff, filen)

           # filen = '--General Ledger--'+fund
           # report.prior_period_adjustments(jes, al, period_start, period_cutoff, filen)
           #
           # filen = "--General Ledger--"+fund
           # report.journals_by_tranid(jes, al,
           #                           period_start, period_cutoff,
           #                            filen)
           #
           # filen = '--General Ledger--'+fund
           # report.position_report_by_sector(al, period_cutoff, knowledge_cutoff,  filen)
           #
           #
           # filen = '--General Ledger--'
           # report.journals_style_reports(jes, al,
           #                               period_start, period_cutoff, filen)

    # # # BASE
    # if timeset == "base" and not close_period and compare_to_prior_period: # get prior period journals up to knowledge cutoff
    #         filen = 'AccountingResultsPriorPeriod'+fund
    #         report.create_accounting_reports(investment_accounting_space.journal_entries, filen,
    #                                          all_asset_liability_accounts_list, period_cutoff)
    #
    #         report.diff_two_excel_files("AccountingResultsPriorPeriod"+fund+".xlsx", "AccountingResultsCurrentPeriod"+fund+".xlsx")
    #



    # -----------------------CLOSE PERIOD----------------------------------
    # 1. Create Pickle file of journals for prior period adjusted for Prior Period KNOWLEDGE
    #
    # with pior period journals that have already been stored because period was already closed
    # if timeset == "current" and period_iteration == "First":  # SAVE JES TO PICKLE
    #     import pickle
    #     prior_period_journals_not_adjusted_py = investment_accounting_space.journal_entries
    #     filtered_journals_py = [entry for entry in prior_period_journals_not_adjusted_py
    #                             if (entry.ibor_date <= prior_period_cutoff)]
    #
    #     import pickle
    #     start_time = time.time()
    #     # Saving prior period journals to pickle
    #     prior_period_journals_not_adjusted_file_path_pickle = "C:/Users/hjmne/PycharmProjects/chest/funds/"+fund+"/periods/prior_period_journals_not_adjusted_pkl.pkl"
    #     with open(prior_period_journals_not_adjusted_file_path_pickle, "wb") as file:
    #         pickle.dump(filtered_journals_py, file)
    #     pickle_duration = time.time() - start_time
    #     print(f"Pickle Time: {pickle_duration}, pickle Time: {pickle_duration}")
    #
    #  #2. prior period journals adjusted for current knowledge
    # if timeset == "current" and period_iteration == "Second":
    #     import pickle
    #     prior_period_journals_adjusted_py = investment_accounting_space.journal_entries
    #     filtered_journals_py = [entry for entry in prior_period_journals_adjusted_py
    #                             if (entry.ibor_date <= prior_period_cutoff)]
    #
    #     import pickle
    #     start_time = time.time()
    #     # Saving prior period journals to pickle
    #     prior_period_journals_adjusted_file_path_pickle = "C:/Users/hjmne/PycharmProjects/chest/funds/"+fund+"/periods/prior_period_journals_adjusted_pkl.pkl"
    #     with open(prior_period_journals_adjusted_file_path_pickle, "wb") as file:
    #         pickle.dump(filtered_journals_py, file)
    #     pickle_duration = time.time() - start_time
    #     print(f"Pickle Time: {pickle_duration}, pickle Time: {pickle_duration}")
    #
    #     # Third iteration
    #     # 1. get pickle files call adjustment routine, post adjustments
    #     # 2. filter for current jes amd post them
    #
    # if timeset == "current" and period_iteration == "Third":
    #     import json
    #     import os
    #     import pickle
    #     from datetime import datetime, timedelta
    #
    #     period_journals = investment_accounting_space.journal_entries
    #
    #     current_period_journals_py = [entry for entry in period_journals
    #                             if (entry.ibor_date >= period_start and entry.ibor_date <= period_cutoff)]
    #
    #
    #     # Pickle file containing prior period adjustsments based on CURRENT KNOWLEDGE - Ref 2.
    #     # Load file and convert to Python format
    #     prior_period_journals_file_path_not_adjusted_pkl = "C:/Users/hjmne/PycharmProjects/chest/funds/"+fund+"/periods/prior_period_journals_not_adjusted_pkl.pkl"
    #
    #     # Check if file exists
    #     if os.path.exists(prior_period_journals_file_path_not_adjusted_pkl):
    #         with open(prior_period_journals_file_path_not_adjusted_pkl, "rb") as file:
    #             prior_period_journals_not_adjusted_py = pickle.load(file)
    #
    #             # Pickle file containing prior period adjustsments based on CURRENT KNOWLEDGE - Ref 2.
    #             # Load file and convert to Python format
    #
    #     prior_period_journals_file_path_adjusted_pkl = "C:/Users/hjmne/PycharmProjects/chest/funds/"+fund+"/periods/prior_period_journals_adjusted_pkl.pkl"
    #
    #         # Check if file exists
    #     if os.path.exists(prior_period_journals_file_path_adjusted_pkl):
    #         with open(prior_period_journals_file_path_adjusted_pkl, "rb") as file:
    #             prior_period_journals_adjusted_py = pickle.load(file)
    #
    #     if len(prior_period_journals_not_adjusted_py) != 0:
    #         adjusting_journals = closed_period.create_adjustment_records(prior_period_journals_not_adjusted_py,
    #                                                                      prior_period_journals_adjusted_py,
    #                                                                      )
    #         # Store the adjustments into a json file, store in current period directory
    #         closed_period.store_adjusting_journals(period_start, adjusting_journals, fund, period_name,
    #                                                    (prior_period_cutoff + timedelta(seconds=1)))
    #     else:
    #         print("File not found:", prior_period_journals_file_path_not_adjusted_pkl)
    #         journals2 = []
    #
    #     # Convert the filtered journal entries to dictionaries
    #     journal_entries_py = [entry.to_dict() for entry in current_period_journals_py]
    #     # Convert the journal entries to dictionaries if they're not already
    #
    #     # Serialize to JSON
    #     # # Optionally, if you need to serialize these objects back to JSON
    #     json_data_jsn = json.dumps(journal_entries_py,
    #                                default=closed_period.extended_serializer)
    #
    #     # write journals into the current period file directory
    #     current_journals_file_path = "C:/Users/hjmne/PycharmProjects/chest/funds/"+fund+"/periods/" + period_name + "/journals.json"
    #     # Optionally, write to a file
    #     with open(current_journals_file_path, 'w') as file:
    #         file.write(json_data_jsn)
    #
    #     # Here we must combine all period files, convert them into python format once combined, and
    #     # then feed them into the process that creates a GL temporal view
    #     # os.remove(prior_period_journals_adjusted_file_path_pickle)
    #     # combine the json files
    #     closed_period.combine_je_files(fund)
    #     # convert to pythin
    #     journals = closed_period.combined_file_python()  # Assuming this returns a list of Journals objects
    #
    #     # feed into bookkeeping space builder to create the GL view
    #     # Example usage
    #     repository = BookkeepingSpace()
    #     space1 = repository.build_sub_ledger_from_journals(journals, prior_period_cutoff)
    #     asset_liability_data = space1.combined_assets_liabilities()
    #     je_data = space1.journal_entries
    #
# import csv
# import datetime
# #import logging
# import os
# import psutil
# import time
# from multiprocessing import Process
# from sharding import add_shard
# import bookkeeping
#
# def distribute_portfolios(shards, portfolio_list, base_path):
#     portfolios_to_process = [os.path.join(base_path, f"{portfolio}.csv") for portfolio in portfolio_list]
#
#     # Distribute portfolios among the shards
#     for i, portfolio in enumerate(portfolios_to_process):
#         shard = shards[i % len(shards)]
#         shard.add_portfolio(portfolio)
#
#
# from datetime import datetime
# import pandas as pd
# import logging
# import os
# import csv
# import time
# import bookkeeping
#
#
# import cProfile
# import pstats
# import io
# import functools
#
# # Profiling decorator
# def profile_function(func):
#     @functools.wraps(func)
#     def wrapper_profile_function(*args, **kwargs):
#         pr = cProfile.Profile()
#         pr.enable()
#         result = func(*args, **kwargs)
#         pr.disable()
#         s = io.StringIO()
#         ps = pstats.Stats(pr, stream=s).sort_stats(pstats.SortKey.CUMULATIVE)
#         ps.print_stats()
#         print(s.getvalue())
#         return result
#     return wrapper_profile_function
# @profile_function
# def prepare_parameters_and_process_events(current_period_data, portfolio_file):
#     logging.info(f"Processing domain-specific events for portfolio file: {portfolio_file}")
#
#     # Extract the portfolio name from the file path
#     portfolio_name = os.path.splitext(os.path.basename(portfolio_file))[0]
#     logging.info(f"Extracted portfolio name: {portfolio_name}")
#
#     # Construct the correct file path for the portfolio
#     portfolio_file = os.path.join('C:/BASE_PATH/refdata/pooltest', f'{portfolio_name}.csv')
#     temp_file = portfolio_file + '.tmp'
#
#     # Read and write portfolio file
#     with open(portfolio_file, mode='r', encoding='utf-8') as infile, open(temp_file, mode='w', encoding='utf-8',
#                                                                           newline='') as outfile:
#         reader = csv.reader(infile)
#         writer = csv.writer(outfile)
#         header = next(reader)
#         writer.writerow(header)
#         for row in reader:
#             row[0] = portfolio_name  # Assuming the first column is the portfolio column
#             writer.writerow(row)
#
#     os.replace(temp_file, portfolio_file)
#
#     selected_report = current_period_data.get("selected_report", "Default Report Type")
#
#     manager = bookkeeping.SpaceManager()
#     investment_accounting_space = manager.get_space('investment_accounting_space')
#     fund_structures_space = manager.get_space('fund_structures_space')
#
#     coa_path = "C:/Users/hjmne/PycharmProjects/chest/refdata/chart_of_accounts.csv"
#     coa = bookkeeping.load_coa_from_csv()  # Assuming this function loads the chart of accounts
#
#     date_fields = ["current_period_start", "current_period_cutoff", "current_period_knowledge",
#                    "prior_period_start", "prior_period_cutoff", "prior_period_knowledge"]
#
#     # Convert date strings to datetime objects once
#     for field in date_fields:
#         if not current_period_data.get(field):
#             logging.error(f"Error: {field} data not provided.")
#             return
#
#     current_period_start = datetime.strptime(current_period_data["current_period_start"], "%Y-%m-%d:%H:%M:%S")
#     current_period_cutoff = datetime.strptime(current_period_data["current_period_cutoff"], "%Y-%m-%d:%H:%M:%S")
#     current_period_knowledge = datetime.strptime(current_period_data["current_period_knowledge"], "%Y-%m-%d:%H:%M:%S")
#     prior_period_start = datetime.strptime(current_period_data["prior_period_start"], "%Y-%m-%d:%H:%M:%S")
#     prior_period_cutoff = datetime.strptime(current_period_data["prior_period_cutoff"], "%Y-%m-%d:%H:%M:%S")
#     prior_period_knowledge = datetime.strptime(current_period_data["prior_period_knowledge"], "%Y-%m-%d:%H:%M:%S")
#
#     numport = 1
#     process_current = "Yes"
#     process_base = "Yes"
#     selected_fund = current_period_data.get("selected_fund", "MyPortfolio")
#     if selected_fund == "Select a Fund":
#         selected_fund = "MyPortfolio"
#     close_period = False
#     create_performance = False
#     prior_period_name = "foo"
#     events_sheet = ""
#     tdate_fx = 1
#
#     logging.info(f"Processing portfolio: {portfolio_name}")
#     processing_start_time = time.time()
#
#     scheduler = bookkeeping.EventScheduler()
#
#     if process_current == "Yes" and not close_period:
#         sub_ledger_list = []
#         sub_ledger.reset_all()
#         sub_ledger.reset_investment_subspaces()
#         fund_structures_space.journal_entries = []
#         fund_structures_space.asset_liability_entries = []
#         fund_structures_space.revenue_expense_entries = []
#         fund_structures_space.journal_entries = []
#         composite_journal_lists = []
#
#         process_start_date = current_period_start
#         if not isinstance(sub_ledger, bookkeeping.BookkeepingSpace):
#             raise TypeError("sub_ledger is not an instance of BookkeepingSpace")
#
#         process_events(events_sheet, "current", 1, portfolio_name, process_start_date, current_period_start,
#                        current_period_cutoff, current_period_knowledge, portfolio_name,
#                        journal_entries, sub_ledger,
#                        investment_accounting_space, coa, tdate_fx, create_performance,
#                        scheduler, price_data, fx_data)
#
#
# def process_shard(shard, current_period_data):
#     for portfolio_file in shard.portfolios:
#         prepare_parameters_and_process_events(current_period_data, portfolio_file)
#
# def start_system(current_period_data, portfolio_list_file):
#     logging.info("Starting system with data: %s", current_period_data)
#
#     # Define the initial shards
#     shards = []
#     logging.info("Adding shards")
#     num_cores = psutil.cpu_count(logical=False)
#     for i in range(num_cores):
#         add_shard(shards, name=f"pool{i + 1}", current_period_data=current_period_data, process_domain_events=prepare_parameters_and_process_events, cpu_core=i)
#
#     # Load the portfolio list from the CSV file
#     portfolio_list = []
#     with open(portfolio_list_file, mode='r', encoding='utf-8') as file:
#         reader = csv.reader(file)
#         portfolio_list = [row[0] for row in reader if row[0] != 'PortfolioName']
#
#     base_path = os.path.dirname(portfolio_list_file)
#
#     # Distribute portfolios among the shards
#     distribute_portfolios(shards, portfolio_list, base_path)
#
#     # Start each shard in a separate process
#     logging.info("Starting shard processes")
#     processes = []
#     for shard in shards:
#         p = Process(target=process_shard, args=(shard, current_period_data))
#         p.start()
#         processes.append(p)
#
#     # Ensure processes are properly joined
#     for p in processes:
#         p.join()
#
#     logging.info("System processing completed.")
#
#     # Print processing times for each portfolio
#     for shard in shards:
#         for i, processing_time in enumerate(shard.processing_times):
#             logging.info(f"Shard {shard.name} processed portfolio {i + 1} in {processing_time} seconds")
