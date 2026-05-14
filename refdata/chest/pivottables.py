import report
import jinja2
# from openpyxl import load_workbook
# from openpyxl.utils import get_column_letter
#
# def create_hyperlink_to_sheet(sheet, target_sheet_name, cell, text):
#     sheet[f'{cell}'] = text
#     sheet[f'{cell}'].hyperlink = f"#{target_sheet_name}!A1"
#     sheet[f'{cell}'].style = "Hyperlink"
#
# wb = load_workbook('workbook.xlsx')
# sheet1 = wb['Sheet1']
# sheet2 = wb['Sheet2']
#
# # Add a hyperlink in Sheet1 A6 to Sheet2 A1
# create_hyperlink_to_sheet(sheet1, 'Sheet2', 'A6', 'Go to Sheet2')
#
# # Add a hyperlink in Sheet2 A6 to Sheet1 A1
# create_hyperlink_to_sheet(sheet2, 'Sheet1', 'A6', 'Go to Sheet1')
#
# # Save the workbook
# wb.save('workbook_with_links.xlsx')

import pandas as pd
from openpyxl import load_workbook

import performance


# Function to create hyperlinks to all other sheets in the workbook
def add_hyperlinks_to_all_sheets(filename):
    wb = load_workbook(filename)

#     for sheet_name in wb.sheetnames:
#         sheet = wb[sheet_name]
#
#         # Starting row for hyperlinks
#         row = 1
#
#         for target_sheet_name in wb.sheetnames:
#             if target_sheet_name != sheet_name:
#                 cell = f'A{row}'
#                 hyperlink_text = f'Go to {target_sheet_name}'
#                 sheet[cell] = hyperlink_text
#                 sheet[cell].hyperlink = f"#{target_sheet_name}!A1"
#                 sheet[cell].style = "Hyperlink"
#
#                 # Increment row for next hyperlink
#                 row += 1
#
#     # Save the workbook with hyperlinks
#     wb.save(filename)
#
#
# # # Create sample data and write to Excel file
# # df = pd.DataFrame({'Data': range(10)})
# # with pd.ExcelWriter('workbook.xlsx', engine='openpyxl') as writer:
# #     for i in range(1, 4):  # Creating 3 sheets as an example
# #         df.to_excel(writer, sheet_name=f'Sheet{i}', index=False)
#
# # Add hyperlinks to all sheets in the workbook
# add_hyperlinks_to_all_sheets('workbook.xlsx')

import openpyxl

# for _, event_row in events_df.iterrows():
#     if event_row['method'] == 'SomeSpecificMethod':
#         # Display in a format optimized for this method
#         group1 = ['portfolio', 'method', 'source']
#         group2 = ['tradedate', 'settledate', 'location']
#         print(event_row[group1].to_string(index=False))
#         print(event_row[group2].to_string(index=False))
#     else:
#         # Default display format
#         group1 = ['portfolio', 'method', 'source', 'tradedate', 'settledate']
#         group2 = ['kdbegin', 'kdend', 'investment', 'payment_currency', 'location']
#         print(event_row[group1].to_string(index=False))
#         print(event_row[group2].to_string(index=False))
#
#     print("-------------------------")



def display_event_with_journals(events_df, journals_df, filename):
    # Open the file for writing
    with open(filename, 'w') as file:

        # Ensure both dataframes are sorted by the transaction ID
        events_df = events_df.sort_values(by='tranid')
        journals_df = journals_df.sort_values(by='Tran ID')

        # Group journals by their transaction ID for easier lookup
        grouped_journals = journals_df.groupby('Tran ID')

        # Helper function to format group data
        def format_group_data(row, group):
            return "\n".join([f"{col}: {row[col]}" for col in group])

        # For each event, display the event details and then its corresponding journals
        for _, event_row in events_df.iterrows():
            trades = ['buy', 'sell', 'short', 'cover']
            # Split columns into groups for multi-line display
            base1_terms = ['portfolio', 'method', 'source', 'tradedate', 'settledate','kdbegin', 'kdend']
            base2_terms = ['investment', 'payment_currency', 'location', 'strategy', 'tranid', 'transaction']
            trade_terms = ['quantity', 'total_amount', 'total_amount_base']




            # ... Add more groups as needed

            # Print each group of columns in a separate line
            print(format_group_data(event_row, base1_terms))
            print(format_group_data(event_row, base2_terms))
            print(format_group_data(event_row, trade_terms))
            if event_row['method'] == 'dividend':
                div_terms = ['per_share']
                print(format_group_data(event_row, div_terms))
            if event_row['method'] == 'split':
                split_terms = ['old_shares', 'new_shares']
                print(format_group_data(event_row, split_terms))
            if event_row['method'] == 'spotfx':
                fx_terms = ['buy_currency', 'sell_currency', 'buy_amt', 'sell_amt']
                print(format_group_data(event_row, fx_terms))
            if event_row['method'] == 'allocate':
                alloc_terms = ['allocation_entities', 'allocation_percents']
                print(format_group_data(event_row, alloc_terms))




            # ... Print more groups as needed

            print("-------------------------")

            tranid = event_row['tranid']
            file.write("Event Record for TranID: " + str(tranid) + "\n")
            file.write(format_group_data(event_row, base1_terms) + "\n")
            file.write(format_group_data(event_row, base2_terms) + "\n")
            file.write(format_group_data(event_row, trade_terms) + "\n")
            if event_row['method'] == 'dividend':
                file.write(format_group_data(event_row, div_terms) + "\n")
            if event_row['method'] == 'split':
                file.write(format_group_data(event_row, split_terms) + "\n")
            if event_row['method'] == 'spotfx':
                file.write(format_group_data(event_row, fx_terms) + "\n")
            if event_row['method'] == 'allocate':
                file.write(format_group_data(event_row, alloc_terms) + "\n")

            # ... Write more groups as needed

            file.write("\nAssociated Journal Entries:\n")

            # Check if there are journals associated with this event
            if tranid in grouped_journals.groups:
                je_lines = grouped_journals.get_group(tranid)
                file.write(je_lines.to_string() + "\n")
            else:
                file.write("No associated journal entries for this event.\n")
            file.write("\n-------------------------\n")

def build_pivot_tables(period_end):

    # Load the Excel file into a DataFrame
    import pandas as pd
    events_df = pd.read_excel("C:/Users/hjmne/PycharmProjects/chest/refdata/portfolio1.xlsx")
    events_df["tradedate"] = pd.to_datetime(events_df["tradedate"], format="%m/%d/%Y:%H:%M:%S")
    events_df = events_df[events_df["tradedate"] <= period_end]
    #events_df['Record Type'] = 'Event'

    # Load the 'AccountingJournals' sheet
    accounting_journals_df = pd.read_excel("C:/Users/hjmne/PycharmProjects/chest/repdata/CurrentPeriodResultsforFund1.xlsx",
                                           sheet_name='AccountingJournals')

    events_df = events_df.sort_values(by='tranid')
    accounting_journals_df = accounting_journals_df.sort_values(by='Tran ID')
#    grouped_journals = accounting_journals_df.groupby('Tran ID')

    display_event_with_journals(events_df, accounting_journals_df, "C:/Users/hjmne/PycharmProjects/chest/repdata/pivotdatasets/eventsandjournals.txt")



    valuation_df = pd.read_excel("C:/Users/hjmne/PycharmProjects/chest/repdata/pivotdatasets/valuation_report.xlsx")

    import pandas as pd

    # List of sheets to skip
    sheets_to_skip = ['Parameters']

    # Load the workbook
    file_path = "C:/Users/hjmne/PycharmProjects/chest/repdata/performance_returns.xlsx"
    xls = pd.ExcelFile(file_path)

    # Function to clean and transform the data
    def clean_transform_save(sheet_name):
        # Load the sheet
        df = pd.read_excel(xls, sheet_name=sheet_name)

        # Remove unnecessary columns
        cols_to_keep = [sheet_name, 'ibor_date', 'BMV_Book', 'EMV_Book', 'Open_CF_Book', 'Close_CF_Book',
                        'TWR_Book', 'BookToDate_Percent']
        df = df[cols_to_keep]
        df['Capital Flows'] = df['Open_CF_Book'] + df['Close_CF_Book']

        # Reorder columns
        cols_to_keep = [sheet_name, 'ibor_date', 'BMV_Book', 'EMV_Book', 'Capital Flows',
                        'TWR_Book', 'BookToDate_Percent']

        df = df[cols_to_keep]

        # Rename columns
        df = df.rename(columns={
            'ibor_date': 'Date',
            'TWR_Book': 'TWR',
            'BookToDate_Percent': 'Cumulative TWR'
        })

        # Style and format
        def currency_format(x):
            return "${:,.2f}".format(x)

        df.style.format({'EMV_Book': currency_format, 'BMV_Book': currency_format})

        # Save to a new Excel file
        output_path = f'C:/Users/hjmne/PycharmProjects/chest/repdata/pivotdatasets/{sheet_name}_twr.xlsx'
        df.to_excel(output_path, engine='openpyxl', index=False)
        print(f"Saved: {output_path}")

    # Loop through each sheet in the workbook
    for sheet_name in xls.sheet_names:
        if sheet_name not in sheets_to_skip:
            clean_transform_save(sheet_name)

    import pandas as pd

    # Specify the file path and sheet name
    # file_path = 'C:/Users/hjmne/PycharmProjects/chest/repdata/performance_returns.xlsx'
    # sheet_name = 'portfolio'
    #
    # # Read the Excel sheet into a DataFrame
    # df = pd.read_excel(file_path, sheet_name=sheet_name)

    # # Display the first few rows of the DataFrame
    # print("Original DataFrame:")
    # print(df.head())
    #
    # # Create a pivot table (modify this part according to your specific needs)
    # # As an example, let's assume you have 'Category' and 'Amount' columns in your DataFrame,
    # # and you want to create a pivot table to sum the amounts for each category.
    # df = df.pivot(index='ibor_date', columns='BMV_Book', values='Value')
    # # EMV_Local
    # # EMV_Book
    # # BMV_Local
    # # BMV_Book
    # # Open_CF_Local
    # # Open_CF_Book
    # # Close_CF_Local
    # # Close_CF_Book
    # # Income
    # # Local
    # # Income
    # # Book
    # # TWR_Local
    # # TWR_Book
    # # LocalToDate
    # # BookToDate
    # # TWR_Local_Percent
    # # TWR_Book_Percent
    # # LocalToDate_Percent
    # # BookToDate_Percent
    #
    # # Display the pivot table
    # print("\nPivot Table:")
    # print(pivot_table)
    #
    # assetclass_ir_df = pd.read_excel("C:/Users/hjmne/PycharmProjects/chest/repdata/performance_returns.xlsx",
    #                                 sheet_name='asset_class')
    #
    # investment_ir_df = pd.read_excel("C:/Users/hjmne/PycharmProjects/chest/repdata/performance_returns.xlsx",
    #                                 sheet_name='investment')
    #
    # #    ir_df = performance.create_performance_sheets(journals, je_df)
    #
    # print ("here")
    #
    # import pandas as pd

#
    #     'Beginning MV': pivot_table['Beginning MV'].iloc[0],
    #     'Ending MV': pivot_table['Ending MV'].iloc[-2],
    #     'Capital Adds/Subs': pivot_table['Capital Adds/Subs'].sum(),  # or however you want to compute this
    #     'Income': pivot_table['Income'].sum(),  # or however you want to compute this
    #     'Period ROR': 'NA',
    #     'Cumulative ROR': pivot_table['Cumulative ROR'].iloc[-2]
    # }
    #
    # pivot_table.loc['Summary'] = pd.Series(summary_values)
#
#     print ("here")
#
#
# #    pivot_table.loc['Summary'] = pd.Series(summary_values)
#
#     pivot_table.to_excel('C:/Users/hjmne/PycharmProjects/chest/repdata/portfolio_pt.xlsx', engine='openpyxl')

