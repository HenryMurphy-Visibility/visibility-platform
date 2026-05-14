import pandas as pd
import logging
from pandas.tseries.offsets import BDay
import openpyxl
from openpyxl.styles import PatternFill, Font

# Configure logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
file_handler = logging.FileHandler('processing.log')
file_handler.setLevel(logging.DEBUG)
file_handler.setFormatter(logging.Formatter('%(asctime)s - %(asctime)s - %(levelname)s - %(message)s'))
logging.getLogger().addHandler(file_handler)

# Provided account_mapping dictionary
account_mapping = {
    'Cost': (0, 1, 2),
    'UnrealPriceGL': (0, 3, 4),
    'UnrealFXGL': (0, 'skip', 5),
    'PriceGainStatOffset': (0, 3, 4),
    'FXGainStatOffset': (0, 'skip', 5),
    'PriceGainInvestment': (0, 6, 7),
    'FXGainInvestment': (0, 6, 7),
    'DividendReceipt': (0, 8, 9),
    'UnearnedIncome': (0, 8, 9),
    'ContributedCost': (10, 11, 12),
    'Receivable': (0, 1, 2),
    'Payable': (0, 1, 2),
    'DividendsReceivable': (0, 1, 2),
    'FXGainTradeSettle': (0, 8, 9),
    'FXGainCurrency': (0, 8, 9),
}

def assign_values(entry, mapping):
    assigned_values = [0] * 13  # Initialize a list with 13 columns set to 0
    values = [entry['quantity'], entry['local'], entry['book']]

    for i, map_index in enumerate(mapping):
        if map_index != 'skip':
            assigned_values[map_index] += values[i]

    return assigned_values

# Function to convert journal entries to DataFrame
def convert_journal_entries_to_df(journal_entries):
    if isinstance(journal_entries[0], dict):
        data = [
            [
                je['portfolio'],
                je['transaction'],
                je['investment'],
                je['ibor_date'],
                je['tradedate'],
                je['settledate'],
                je['lotid'],
                je['tranid'],
                je['tax_date'],
                je['ls'],
                je['location'],
                je['financial_account'],
                je['quantity'],
                je['local'],
                je['book'],
                je['notional'],
                je['oface']
            ]
            for je in journal_entries
        ]
    else:
        data = [
            [
                je.portfolio,
                je.transaction,
                je.investment,
                je.ibor_date,
                je.tradedate,
                je.settledate,
                je.lotid,
                je.tranid,
                je.tax_date,
                je.ls,
                je.location,
                je.financial_account,
                je.quantity,
                je.local,
                je.book,
                je.notional,
                je.oface
            ]
            for je in journal_entries
        ]

    columns = [
        'portfolio',
        'transaction',
        'investment',
        "ibor_date",
        "tradedate",
        "settledate",
        'lotid',
        'tranid',
        'tax_date',
        'ls',
        'location',
        'financial_account',
        "quantity",
        'local',
        'book',
        'notional',
        'oface',
    ]

    df = pd.DataFrame(data, columns=columns)
    logging.info(f"Converted journal entries to DataFrame:\n{df.head()}")
    return df

# Function to clean dates
def clean_dates(df, date_columns):
    for column in date_columns:
        df[column] = pd.to_datetime(df[column], errors='coerce')
    return df

# Main function to generate the report and pivot table


import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from pandas.tseries.offsets import BDay
import logging

import pandas as pd
import logging
from pandas.tseries.offsets import BDay
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

def clean_dates(df, date_columns):
    for col in date_columns:
        df[col] = pd.to_datetime(df[col], errors='coerce')
    return df
import pandas as pd
import logging
from pandas.tseries.offsets import BDay
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

import pandas as pd

def clean_dates(df, date_columns):
    """
    Converts specified columns to datetime format, handling errors by coercing them.

    Args:
    - df (pd.DataFrame): The DataFrame to clean.
    - date_columns (list of str): The list of column names to convert to datetime.

    Returns:
    - pd.DataFrame: The DataFrame with specified columns converted to datetime.
    """
    for col in date_columns:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')
        else:
            print(f"Warning: Column {col} not found in DataFrame. Skipping conversion.")
    return df


def generate_comprehensive_report_and_pivot(space_manager, journal_entries, ledger_choice, start_period, end_period, fund,
                                            output_file, derive_mktval, fx_data):
    start_period = pd.to_datetime(start_period)
    end_period = pd.to_datetime(end_period)
    start_period_adjusted = start_period - BDay(1)
    start_period_adjusted = start_period_adjusted.replace(hour=23, minute=59, second=59)

    # Convert the journal entries list to a DataFrame
    combined_journal_entries_df = pd.DataFrame([je.__dict__ for je in journal_entries])

    if 'portfolio' in combined_journal_entries_df.columns:
        combined_journal_entries_df = combined_journal_entries_df.drop(columns=['portfolio'])

    combined_journal_entries_df = clean_dates(combined_journal_entries_df, ['ibor_date', 'tradedate', 'settledate', 'tax_date'])
    combined_journal_entries_df = combined_journal_entries_df.sort_values(by=['investment', 'lotid', 'ibor_date', 'financial_account'])

    # Filter out MarketVal financial account
    combined_journal_entries_df = combined_journal_entries_df[
        (combined_journal_entries_df['financial_account'] != 'MarketVal') |
        (
                (combined_journal_entries_df['local'] == 0) &
                (combined_journal_entries_df['book'] == 0)
        )
    ]

    combined_journal_entries_df.to_csv('journal_entries_report.csv', index=False)

    try:
        logging.info("DataFrame info:")
        logging.info(combined_journal_entries_df.info())
        logging.info("First few rows of the DataFrame:")
        logging.info(combined_journal_entries_df.head())

        # Combine DataFrames
        combined_df = combined_journal_entries_df
        combined_df = combined_df.sort_values(by=['investment', 'ibor_date', 'financial_account'])

        rows = []

        for name, group in combined_df.groupby(
                ['transaction', 'investment', 'lotid', 'ibor_date', 'ls', 'location', 'financial_account']):
            transaction, investment, lotid, ibor_date, ls, location, financial_account = name
            logging.info(f"Processing group: {name} with {len(group)} records")

            row = {
                'transaction': transaction,
                'investment': investment,
                'lotid': lotid,
                'ibor_date': ibor_date,
                'ls': ls,
                'location': location,
                'financial_account': financial_account,
                'quantity': 0,
                'local': 0,
                'book': 0,
                'unrealgllocal': 0,
                'unrealglbook': 0,
                'unrealfxbook': 0,
                'realizedlocal': 0,
                'realizedbook': 0,
                'incomelocal': 0,
                'incomebook': 0,
                'capitalshares': 0,
                'capitallocal': 0,
                'capitalbook': 0,
                'tranid': None,
                'entry_type': None
            }

            if 'tranid' in group.columns:
                row['tranid'] = group['tranid'].iloc[0]
            else:
                logging.error(f"'tranid' column missing in group: {name}")
                continue

            if 'entry_type' in group.columns:
                row['entry_type'] = group['entry_type'].iloc[0]
            else:
                row['entry_type'] = 'N/A'

            logging.info(f"Starting to process entries for group: {name}")
            combined_df = combined_df.sort_values(by=['investment', 'ibor_date', 'financial_account'])

            for _, entry in group.iterrows():
                logging.info(f"Processing entry: {entry.to_dict()}")

                if entry['financial_account'] in account_mapping:
                    assigned_values = assign_values(entry, account_mapping[entry['financial_account']])
                    row['quantity'] += assigned_values[0]
                    row['local'] += assigned_values[1]
                    row['book'] += assigned_values[2]
                    row['unrealgllocal'] += assigned_values[3]
                    row['unrealglbook'] += assigned_values[4]
                    row['unrealfxbook'] += assigned_values[5]
                    row['realizedlocal'] += assigned_values[6]
                    row['realizedbook'] += assigned_values[7]
                    row['incomelocal'] += assigned_values[8]
                    row['incomebook'] += assigned_values[9]
                    row['capitalshares'] += assigned_values[10]
                    row['capitallocal'] += assigned_values[11]
                    row['capitalbook'] += assigned_values[12]
                else:
                    row['local'] += entry['local']
                    row['book'] += entry['book']

            logging.info(f"Finished processing entries for group: {name}")
            logging.info(f"Row to be appended: {row}")
            rows.append(row)

        logging.info("Finished processing all groups.")
        logging.info("Creating DataFrame from rows.")

        pivot_table_df = pd.DataFrame(rows)
        pivot_table_df = pivot_table_df.sort_values(by=['investment', 'ibor_date', 'transaction'])  # Final sort

        logging.info("Pivot table DataFrame:")
        logging.info(pivot_table_df)

        book_sum = pivot_table_df['book'].sum()
        unrealglbook_sum = pivot_table_df['unrealglbook'].sum()
        unrealfxbook_sum = pivot_table_df['unrealfxbook'].sum()
        realizedbook_sum = -pivot_table_df['realizedbook'].sum()
        incomebook_sum = -pivot_table_df['incomebook'].sum()
        capitalbook_sum = -pivot_table_df['capitalbook'].sum()
        mkt_val_book = book_sum + unrealglbook_sum + unrealfxbook_sum
        comprehensive_sum = capitalbook_sum + realizedbook_sum + incomebook_sum + unrealfxbook_sum + unrealglbook_sum

        summary_data = pd.DataFrame([{
            'portfolio': 'Summary Book Value',
            'investment': '',
            'transaction': '',
            'lotid': '',
            'ibor_date': '',
            'ls': '',
            'location': '',
            'financial_account': '',
            'quantity': '',
            'local': '',
            'book': book_sum,
            'unrealgllocal': '',
            'unrealglbook': unrealglbook_sum,
            'unrealfxbook': unrealfxbook_sum,
            'realizedlocal': '',
            'realizedbook': realizedbook_sum,
            'incomelocal': '',
            'incomebook': incomebook_sum,
            'capitalshares': '',
            'capitallocal': '',
            'capitalbook': capitalbook_sum,
            'tranid': '',
            'entry_type': ''
        }, {
            'portfolio': 'Summary MarketVal Book Derived from Book Cost & Unrealized',
            'investment': '',
            'transaction': '',
            'lotid': '',
            'ibor_date': '',
            'ls': '',
            'location': '',
            'financial_account': '',
            'quantity': '',
            'local': '',
            'book': mkt_val_book,
            'unrealgllocal': '',
            'unrealglbook': '',
            'unrealfxbook': '',
            'realizedlocal': '',
            'realizedbook': '',
            'incomelocal': '',
            'incomebook': '',
            'capitalshares': '',
            'capitallocal': '',
            'capitalbook': '',
            'tranid': '',
            'entry_type': 'Summary MktValBook Derived from Capital & Net Earnings'
        }, {
            'portfolio': 'Summary',
            'investment': '',
            'transaction': '',
            'lotid': '',
            'ibor_date': '',
            'ls': '',
            'location': '',
            'financial_account': '',
            'quantity': '',
            'local': '',
            'book': comprehensive_sum,
            'unrealgllocal': '',
            'unrealglbook': '',
            'unrealfxbook': '',
            'realizedlocal': '',
            'realizedbook': '',
            'incomelocal': '',
            'incomebook': '',
            'capitalshares': '',
            'capitallocal': '',
            'capitalbook': '',
            'tranid': '',
            'entry_type': 'Comprehensive Sum'
        }])

        pivot_table_df = pd.concat([pivot_table_df, summary_data], ignore_index=True)

        logging.info("Saving pivot table to Excel file...")
        pivot_table_df.to_excel(output_file, index=False, engine='openpyxl')

        wb = load_workbook(output_file)
        ws = wb.active

        # Define the blue fill pattern
        blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

        # Apply blue fill to rows where 'transaction' column equals 'Valuation'
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            if row[0].value == 'Valuation':  # Assuming 'transaction' is the 1st column (index 0)
                for cell in row:
                    cell.fill = blue_fill

            # Apply number formatting and font sizes as previously defined
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
                cell.font = Font(size=11)  # Set normal font size

        # Apply bold font for summary rows
        summary_rows = ['Summary Book Value', 'Summary MarketVal Book Derived from Book Cost & Unrealized', 'Comprehensive Sum']
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            if row[0].value in summary_rows:
                for cell in row:
                    cell.font = Font(bold=True)

        # Freeze the top row for column titles
        ws.freeze_panes = 'A2'

        # Auto-size the columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(output_file)
    except Exception as e:
        logging.error(f"An error occurred: {e}")
    finally:
        if 'pivot_table_df' in locals():
            logging.info(f"Pivot table saved with formatting. Number of records in pivot table: {len(pivot_table_df)}")
        else:
            logging.info("Pivot table was not created due to an error.")
