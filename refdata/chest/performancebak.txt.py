import report
from openpyxl.styles import PatternFill

from openpyxl.styles import PatternFill, Font
import openpyxl
import pandas as pd


# Given functions...

def fill_blue_row_perf(ws, row_idx):
    """Fill a specific row in a worksheet with blue color."""
    fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    for cell in ws[row_idx]:
        cell.fill = fill


def style_perf_report(merge_mv_data):
    """Style the performance report using specific functions."""
    # Convert the DataFrame to an openpyxl worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    for r_idx, row in enumerate(merge_mv_data.values, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Apply styles
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if i % 2 == 0:  # If the row number is even
            fill_blue_row_perf(ws, i)
        if ws.cell(row=i, column=1).value.startswith('Subtotal'):
            report.underline_row(ws, row)
            report.bold_row(ws, row)

    report.format_numbers(ws)
    report.autosize_columns(ws)

    return ws  # Return the styled worksheet


def compute_opening_cash_flows_investments(merge_mv_datainv, merge_mv_data, level):
    currencies = ['USD', 'AUD', 'GBP', 'EUR', 'JPY']
    valid_accounts = ['Cost', 'Payable', 'Receivable']

    # Define grouping columns based on the level
    group_by_cols = [level, 'ibor_date'] if level != 'ibor_date' else ['ibor_date']

    # For non-currencies, apply the 'Book' > 0 filter. For currencies, just check the valid accounts.

    # Conditions for filtering rows
    condition1 = (~merge_mv_data['investment'].isin(currencies) & (merge_mv_data['book'] > 0))
    condition2 = merge_mv_data['financial_account'].isin(valid_accounts)

    # Combine conditions
    opening_flows = merge_mv_data[condition1 & condition2]

    opening_flows = opening_flows.groupby(group_by_cols).agg(
        {'local': 'sum', 'book': 'sum'}).reset_index()

    # Rename columns for clarity
    opening_flows.rename(columns={'local': 'Open_CF_Local', 'book': 'Open_CF_Book'}, inplace=True)

    return opening_flows

def compute_cash_flows_currencies(merge_mv_datainv, merge_mv_data, level):
    currencies = ['USD', 'AUD', 'GBP', 'EUR', 'JPY']
    valid_accounts = ['Cost', 'Payable', 'Receivable', 'FXGainTradeSettle', 'FXCurrency']

    # Define grouping columns based on the level
    group_by_cols = [level, 'ibor_date'] if level != 'ibor_date' else ['ibor_date']

    # For non-currencies, apply the 'Book' > 0 filter. For currencies, just check the valid accounts.

    # Conditions for filtering rows
    condition1 = merge_mv_data['investment'].isin(currencies)
    condition2 = merge_mv_data['financial_account'].isin(valid_accounts)

    # Combine conditions
    currency_flows = merge_mv_data[condition1 & condition2]

    currency_flows = currency_flows.groupby(group_by_cols).agg(
        {'local': 'sum', 'book': 'sum'}).reset_index()

    # Rename columns for clarity
    currency_flows.rename(columns={'local': 'Currency_Flows_Local', 'book': 'Currency_Flows_Book'}, inplace=True)

    return currency_flows

def compute_closing_cash_flows_for_investments(merge_mv_datainv, merge_mv_data, level):
    currencies = ['USD', 'AUD', 'GBP', 'EUR', 'JPY']
    # Condition 1
    condition1 = merge_mv_data['financial_account'].isin(['PriceGainInvestment', 'FXGainInvestment'])
    # Condition 2
    condition2 = (merge_mv_data['financial_account'] == 'Cost') & (merge_mv_data['book'] < 0) & (merge_mv_data['ls'] == 'l')
    # Condition 3
    condition3 = (merge_mv_data['financial_account'] == 'Cost') & (merge_mv_data['book'] > 0) & (merge_mv_data['ls'] == 's')
    # Condition 4
    condition4 = ~merge_mv_data['investment'].isin(currencies)

    # Combine conditions
    closing_flows = merge_mv_data[(condition1 | condition2 | condition3) & condition4]

    # Group by Investment, IBOR Date, and Tax Date
    closing_flows_agg = closing_flows.groupby([level, 'ibor_date']).agg(
        {'local': 'sum', 'book': 'sum'}).reset_index()

    # Rename columns for clarity
    closing_flows_agg.rename(columns={'local': 'Close_CF_Local', 'book': 'Close_CF_Book'}, inplace=True)

    return closing_flows_agg

def compute_income(merge_mv_datainv, merge_mv_data, level):

    # valid_accounts = ['DividendReceipt', 'FXGainTradeSettle']
    # income_entries = merge_mv_data['financial_account'].isin(valid_accounts)

    income_entries = merge_mv_data[(merge_mv_data['financial_account'] == 'FXGainTradeSettle')]

    # Group by both 'Investment' and 'IBOR Date'
    income_merge_mv_data = income_entries.groupby([level, 'ibor_date'])[['local', 'book']].sum().reset_index()

    income_merge_mv_data.rename(columns={'local': 'Income Local', 'book': 'Income Book'}, inplace=True)

    return income_merge_mv_data


import numpy as np
#import numpy_financial as npf



def compute_daily_twr(journal_entries, agg_level, level='portfolio', include_local_currency=True):
    import pandas as pd
    coa_merge_mv_data = pd.read_csv("C:/Users/hjmne/PycharmProjects/chest/refdata/chart_of_accounts.csv")
    merge_mv_datainv = pd.read_csv("C:/Users/hjmne/PycharmProjects/chest/refdata/investment_master.csv")
    for entry in journal_entries:
        if not hasattr(entry, 'financial_account'):
            print("Missing financial_account:", entry)

    from collections import namedtuple

    # Define a named tuple for Journals

    import pandas as pd

    # Convert namedtuples to dict within DataFrame
    def convert_to_dict(entry):
        try:
            return entry._asdict()
        except AttributeError:
            return entry  # or modify as per your requirements

    import pandas as pd
    from collections import namedtuple

    Journals = namedtuple('Journals',
                          ['portfolio', 'investment', 'tax_date', 'ls', 'tranid', 'quantity', 'local', 'book',
                           'location', 'financial_account'])



    def normalize_journal_entries(journal_entries):
        data = [entry.to_dict() for entry in journal_entries]
        return pd.DataFrame(data)

    merge_mv_data = normalize_journal_entries(journal_entries)
    # Merge merge_mv_data with merge_mv_datainv to get the asset_class for each investment
    merge_mv_data = pd.merge(merge_mv_data, merge_mv_datainv[['ticker', agg_level]], left_on='investment', right_on='ticker', how='left').drop(
        'ticker', axis=1)


    merge_mv_data['tax date'] = merge_mv_data['tax date'].astype(str)
    print(merge_mv_data.head())

    # Filter for 'MktVal'
    mktval_data_merge_mv_data = merge_mv_data[merge_mv_data['financial_account'] == 'MktVal']

    # Rename columns
    mktval_data_merge_mv_data = mktval_data_merge_mv_data.rename(columns={'local': 'EMV_Local', 'book': "EMV_Book"})


    # Aggregate MktValues for multiple lots on the same IBOR date
    agg_mktval_data = mktval_data_merge_mv_data.groupby([level, 'ibor_date']).agg({
        'EMV_Local': 'sum',
        'EMV_Book': 'sum'
    }).reset_index()

    # Apply BMV Logic- level + ibor date BMV will need to be adjusted for opening flows below
    # Convert 'ibor_date' column to datetime
    agg_mktval_data['ibor_date'] = pd.to_datetime(agg_mktval_data['ibor_date'])
    agg_mktval_data = agg_mktval_data.sort_values(by=[level, 'ibor_date'])
    # Call the function

    agg_mktval_data['BMV_Local'] = agg_mktval_data.groupby(level)['EMV_Local'].shift(1)
    agg_mktval_data['BMV_Book'] = agg_mktval_data.groupby(level)['EMV_Book'].shift(1)

    # Call the function
    opening_flows = compute_opening_cash_flows_investments(merge_mv_datainv, merge_mv_data, level)
    agg_mktval_data['ibor_date'] = pd.to_datetime(agg_mktval_data['ibor_date'])
    opening_flows['ibor_date'] = pd.to_datetime(opening_flows['ibor_date'])
    agg_mktval_data = pd.merge(agg_mktval_data, opening_flows, on=[level, 'ibor_date'], how='left')

    # agg_mktval_data['BMV_Local'] -= agg_mktval_data['Open_CF_Local'].fillna(0)
    # agg_mktval_data['BMV_Book'] -= agg_mktval_data['Open_CF_Book'].fillna(0)

    currency_flows = compute_cash_flows_currencies(merge_mv_datainv, merge_mv_data, level)
    agg_mktval_data['ibor_date'] = pd.to_datetime(agg_mktval_data['ibor_date'])
    currency_flows['ibor_date'] = pd.to_datetime(currency_flows['ibor_date'])
    agg_mktval_data = pd.merge(agg_mktval_data, currency_flows, on=[level, 'ibor_date'], how='left')

    # agg_mktval_data['BMV_Local'] -= agg_mktval_data['Currency_Flows_Local'].fillna(0)
    # agg_mktval_data['BMV_Book'] -= agg_mktval_data['Currency_Flows_Book'].fillna(0)

    closing_flows = compute_closing_cash_flows_for_investments(merge_mv_datainv, merge_mv_data, level)

    agg_mktval_data['ibor_date'] = pd.to_datetime(agg_mktval_data['ibor_date'])
    closing_flows['ibor_date'] = pd.to_datetime(closing_flows['ibor_date'])
    agg_mktval_data = pd.merge(agg_mktval_data, closing_flows, on=[level, 'ibor_date'],
                               how='left').fillna(0)

    income_data = compute_income(merge_mv_datainv, merge_mv_data, level)
    # Ensure both 'ibor_date' columns are of datetime64[ns] type
    agg_mktval_data['ibor_date'] = pd.to_datetime(agg_mktval_data['ibor_date'])
    income_data['ibor_date'] = pd.to_datetime(income_data['ibor_date'])

    agg_mktval_data = pd.merge(agg_mktval_data, income_data, on=[level, 'ibor_date'],
                           how='left').fillna(0)

    merged_mv_data = agg_mktval_data.groupby([level, 'ibor_date']).sum().reset_index()

    # Calculate the End Market Value differences
    merged_mv_data['EMV_Local_Diff'] = merged_mv_data['EMV_Local'].diff().fillna(0)
    merged_mv_data['EMV_Book_Diff'] = merged_mv_data['EMV_Book'].diff().fillna(0)

    merged_mv_data['TWR_Local'] = (
            (merged_mv_data['EMV_Local'].diff() -
             merged_mv_data['Open_CF_Local'] - merged_mv_data['Close_CF_Local'] -
             merged_mv_data['Currency_Flows_Local'] + merged_mv_data['Income Local']) /
            (merged_mv_data['EMV_Local'].shift(1) + merged_mv_data['Open_CF_Local'])
    )

    merged_mv_data['TWR_Book'] = (
            (merged_mv_data['EMV_Book'].diff() -
             merged_mv_data['Open_CF_Book'] - merged_mv_data['Close_CF_Book'] -
             merged_mv_data['Currency_Flows_Book'] + merged_mv_data['Income Book']) /
            (merged_mv_data['EMV_Book'].shift(1) + merged_mv_data['Open_CF_Book'])
    )

    import pandas as pd
    import numpy as np
    import report

    # Drop rows with NaN or infinite values
    merged_mv_data.replace([np.inf, -np.inf], np.nan, inplace=True)
    # merged_mv_data.dropna(subset=['TWR_Local', 'TWR_Book'], inplace=True)

    # styled_ws = style_perf_report(merged_mv_data)  # Replace merged_mv_data with your DataFrame name
    # wb = styled_ws.parent  # Get the workbook object to save it
    # wb.save('styled_performance_report.xlsx')
    # Save to Excel file

    # 2. Chain link the TWRs to calculate the cumulative performance
    merged_mv_data['LocalToDate'] = merged_mv_data.groupby(level)['TWR_Local'].transform(
        lambda x: (1 + x).cumprod())
    merged_mv_data['BookToDate'] = merged_mv_data.groupby(level)['TWR_Book'].transform(
        lambda x: (1 + x).cumprod())

    # 3. Convert the TWR and chain-linked values back to percentage format for reporting
    merged_mv_data['TWR_Local_Percent'] = merged_mv_data['TWR_Local'] * 100
    merged_mv_data['TWR_Book_Percent'] = merged_mv_data['TWR_Book'] * 100
    merged_mv_data['LocalToDate_Percent'] = (merged_mv_data['LocalToDate'] - 1) * 100
    merged_mv_data['BookToDate_Percent'] = (merged_mv_data['BookToDate'] - 1) * 100

    # merge_mv_data['Capital Flows'] = merge_mv_data['Open_CF_Book'] + merge_mv_data['Close_CF_Book'] + merge_mv_data['Currency_Flows_Book']`
    #
    # cols_to_keep = [level, 'ibor_date', 'BMV_Book', 'Capital Flows', 'Income Book',
    #                 'TWR_Book', 'BookToDate_Percent']
 #   merge_mv_data = merge_mv_data[cols_to_keep]
    
    return merged_mv_data
    # fnamechoice = "C:/Users/hjmne/PycharmProjects/chest/repdata/daily_twr_dump.xlsx"
    # merged_mv_data.to_excel(fnamechoice, index=False)


import pandas as pd
from itertools import product

def create_performance_sheets(journal_entries,fname):
    import pandas as pd
    from itertools import product

    # Assuming you have an Excel writer set up
    with pd.ExcelWriter(fname) as writer:
        levels = ['asset_class', 'country', 'currency', 'beta', 'analyst']
        local_currency_flags = [False, False, False, False, False]

        params_merge_mv_data = pd.DataFrame({"Levels": levels, "Include Local Currency": local_currency_flags})
        params_merge_mv_data.to_excel(writer, sheet_name="Parameters")

        twr_results_merge_mv_data = []
        # portfolio_column = je_merge_mv_data['Portfolio']  # Assuming 'portfolio' is a column in your journal_entries DataFrame

        for level, include_local_currency in zip(levels, local_currency_flags):
            twr_result = compute_daily_twr(journal_entries, level, level, include_local_currency)
            sheet_name = f"{level}"

            twr_result.to_excel(writer, sheet_name=sheet_name)
            twr_results_merge_mv_data.append(twr_result)

        more_levels = ['portfolio', 'asset_class', 'investment']
        more_local_currency_flags = [False, False, False]

        for level, include_local_currency in zip(more_levels, more_local_currency_flags):
            twr_result = compute_daily_twr(journal_entries, "asset_class", level, include_local_currency)
            sheet_name = f"{level}"

        # if level != 'Portfolio':
        #     twr_result.insert(0, 'Portfolio', portfolio_column)

            twr_result.to_excel(writer, sheet_name=sheet_name)
            twr_results_merge_mv_data.append(twr_result)

    return twr_results_merge_mv_data


import os
import pandas as pd
from openpyxl import Workbook

def compute_diff(merge_mv_data1, merge_mv_data2):
    # Exclude the first three columns
    cols_to_include = merge_mv_data1.columns[3:]

    return merge_mv_data1[cols_to_include].subtract(merge_mv_data2[cols_to_include])

from openpyxl import load_workbook

def compute_perf_diff(workbook1_path, workbook2_path, sheets_to_compare):
    print("Function Started")
    output_path = 'C:/Users/hjmne/PycharmProjects/chest/repdata/PerformanceReturnsComparison.xlsx'

    # List of sheet names to compare
    sheets_to_compare = ['asset_class', 'country', 'currency', 'beta', 'analyst', 'portfolio', 'investment']

    if not os.path.exists(output_path):
        print("Output file does not exist. Creating a new workbook.")
        wb = Workbook()
        wb.active.title = "Sheet1"
        wb.save(output_path)

    for sheet_name in sheets_to_compare:
        try:
            print(f"Processing sheet: {sheet_name}")
            print("Available sheets in workbook1:", pd.ExcelFile(workbook1_path).sheet_names)
            print("Available sheets in workbook2:", pd.ExcelFile(workbook2_path).sheet_names)

            # Load sheets from the two Excel files
            merge_mv_data1 = pd.read_excel(workbook1_path, sheet_name=sheet_name)
            merge_mv_data2 = pd.read_excel(workbook2_path, sheet_name=sheet_name)

            print(f"Size of merge_mv_data1 from sheet {sheet_name}: {merge_mv_data1.shape}")
            print(f"Size of merge_mv_data2 from sheet {sheet_name}: {merge_mv_data2.shape}")

            # Compute the difference
            merge_mv_data_diff = compute_diff(merge_mv_data1, merge_mv_data2)
            print(f"Size of merge_mv_data_diff from sheet {sheet_name}: {merge_mv_data_diff.shape}")

            # Load the workbook for writing and remove the sheet if it exists
            book = load_workbook(output_path)
            if sheet_name in book.sheetnames:
                book.remove(book[sheet_name])

            # Save the difference to a new sheet in the Excel file
            with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
                writer.book = book
                merge_mv_data_diff.to_excel(writer, sheet_name=sheet_name, index=False)
                print(f'Difference for {sheet_name} saved to {output_path}')

        except Exception as e:
            print(f"An error occurred while processing {sheet_name}: {str(e)}")

    print("Function Completed")

if __name__ == "__main__":
    workbook1_path = 'C:/Users/hjmne/PycharmProjects/chest/repdata/PerformanceReturnsCurrent.xlsx'
    workbook2_path = 'C:/Users/hjmne/PycharmProjects/chest/repdata/PerformanceReturnsPrior.xlsx'
    sheets_to_compare = ['asset_class', 'country', 'currency', 'beta', 'analyst', 'portfolio', 'investment']
    compute_perf_diff(workbook1_path, workbook2_path, sheets_to_compare)

