import pandas as pd


def fetch_data(bookkeeping_list):
    """
    Load base data from bookkeeping list.
    Each entry contains information for each tax lot position.
    """
    bookkeeping_space_list = []
    for entry in bookkeeping_list:
        portfolio, investment, lotid, tax_date, ls, location, financial_account, quantity, local, book, notional, oface = entry
        bookkeeping_space_list.append({
            'Portfolio': portfolio,
            'Investment': investment,
            'Lot_ID': lotid,
            'Tax_Date': tax_date,
            'Ls': ls,
            'Location': location,    # Ensure 'Location' is correctly added here
            'Financial_Account': financial_account,
            'Quantity': quantity,
            'Local Cost': local,
            'Book Cost': book,
            'Notional': notional,
            'Original Face': oface
        })
    return pd.DataFrame(bookkeeping_space_list)

def calculate_valuation(df, edate, price_data, fx_data, ledger_space):
    """
    Calculate market values and unrealized gains/losses for each net position entry.
    """
    # Initialize new columns for valuation results in df
    df['Market Value Local'] = 0
    df['Market Value Book'] = 0
    df['Price GL Local'] = 0
    df['Price GL Book'] = 0
    df['FX GL Book'] = 0

    # Calculate values for each row
    for idx, row in df.iterrows():
        investment = row['Investment']
        quantity = row['Quantity']
        local_cost = row['Local Cost']
        book_cost = row['Book Cost']

        # Get price and FX rate for valuation
        price = price_data.get(investment)
        currency = ledger_space.get_attribute_field(investment, 'AIF', 'Currency')
        fx_rate = fx_data.get(currency)

        if price is None or fx_rate is None:
            print(f"Missing data for {investment}: Price or FX rate is unavailable.")
            continue

        # Calculate market values
        pricing_factor = ledger_space.get_attribute_field(investment, 'AIF', 'Pricing_Factor') or 1
        mktval_local = quantity * price * pricing_factor
        mktval_book = mktval_local * fx_rate

        # Calculate unrealized gains/losses
        price_gl_local = mktval_local - local_cost      # Price gain/loss in local currency
        price_gl_book = price_gl_local * fx_rate        # Price gain/loss in book currency
        fx_gl_book = mktval_book - (book_cost + price_gl_book)  # FX gain/loss in book currency

        # Assign calculated values to the DataFrame
        df.at[idx, 'Market Value Local'] = mktval_local
        df.at[idx, 'Market Value Book'] = mktval_book
        df.at[idx, 'Price GL Local'] = price_gl_local
        df.at[idx, 'Price GL Book'] = price_gl_book
        df.at[idx, 'FX GL Book'] = fx_gl_book

    # Debugging: Confirm columns after adding valuations
    print("Columns after valuation calculation:", df.columns)
    return df

def calculate_subtotals(df):
    """
    Calculate subtotals for each net position (by Investment, Ls, and Location if present).
    """
    # Confirm the existence of required columns for grouping
    grouping_columns = ['Investment']
    if 'Ls' in df.columns:
        grouping_columns.append('Ls')
    if 'Location' in df.columns:
        grouping_columns.append('Location')

    # Debugging: Print grouping columns to confirm they are available
    print("Grouping by columns:", grouping_columns)

    subtotal_df = df.groupby(grouping_columns).agg({
        'Quantity': 'sum',
        'Local Cost': 'sum',
        'Book Cost': 'sum',
        'Market Value Local': 'sum',
        'Market Value Book': 'sum',
        'Price GL Local': 'sum',
        'Price GL Book': 'sum',
        'FX GL Book': 'sum'
    }).reset_index()

    subtotal_df['Portfolio'] = 'Subtotal'
    subtotal_df['Lot_ID'] = ''
    subtotal_df['Tax_Date'] = ''
    subtotal_df['Financial_Account'] = ''

    return subtotal_df

def calculate_grand_totals(df):
    """
    Calculate grand totals for the entire dataset.
    """
    grand_totals = df.agg({
        'Quantity': 'sum',
        'Local Cost': 'sum',
        'Book Cost': 'sum',
        'Market Value Local': 'sum',
        'Market Value Book': 'sum',
        'Price GL Local': 'sum',
        'Price GL Book': 'sum',
        'FX GL Book': 'sum'
    }).to_dict()

    grand_totals.update({
        'Portfolio': 'Grand Totals',
        'Investment': '',
        'Lot_ID': '',
        'Tax_Date': '',
        'Ls': '',
        'Location': '',
        'Financial_Account': ''
    })
    return pd.DataFrame([grand_totals])

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

def format_and_output(df, fname):
    """
    Format and save the final report to an Excel file with alternating row colors, bold subtotal/grand total rows,
    and number formatting with commas.
    """
    output_filepath = f{BASE_PATH}/reports/{fname}"
    df.to_excel(output_filepath, index=False, sheet_name='Report')

    # Load the workbook and the specific sheet to apply formatting
    wb = openpyxl.load_workbook(output_filepath)
    ws = wb['Report']

    # Define styles for alternating rows and bold totals
    light_blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    bold_font = Font(bold=True)

    # Identify which rows are subtotal and grand total rows for bolding
    subtotal_keyword = 'Subtotal'
    grand_total_keyword = 'Grand Totals'

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        row_num = row[0].row

        # Alternating row colors
        fill = light_blue_fill if row_num % 2 == 0 else white_fill
        for cell in row:
            cell.fill = fill

        # Bold formatting for subtotal and grand total rows
        if row[0].value == subtotal_keyword or row[0].value == grand_total_keyword:
            for cell in row:
                cell.font = bold_font

        # Number formatting for columns based on header
        for cell in row:
            col_header = ws.cell(row=1, column=cell.column).value
            if col_header in ["Market Value Book", "Market Value Local", "Price GL Book", "Price GL Local", "FX GL Book", "Local Cost", "Book Cost"]:
                cell.number_format = "#,##0.00"  # Comma format with two decimal places
            elif col_header == "Percent of Portfolio":
                cell.number_format = "0.00%"  # Percentage format

    # Adjust column width to fit content
    for column in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column if cell.value)
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Save the workbook with formatting
    wb.save(output_filepath)
    print(f"Report saved with formatting at {output_filepath}")

def filter_records(df, column, conditions, operator="OR", mode="include"):
    """
    Filter records in a DataFrame based on multiple conditions with an inclusion or exclusion mode.

    Parameters:
    - df: The DataFrame to operate on.
    - column: The column to apply the filter conditions to.
    - conditions: A list of values for filtering.
    - operator: Logical operator to combine conditions ("AND" or "OR").
    - mode: Mode of filtering ('include' or 'exclude').

    Returns:
    - DataFrame after filtering.
    """
    if operator.upper() == "OR":
        if mode.lower() == "include":
            return df[df[column].isin(conditions)]
        elif mode.lower() == "exclude":
            return df[~df[column].isin(conditions)]
    elif operator.upper() == "AND":
        mask = pd.Series([True] * len(df)) if mode.lower() == "include" else pd.Series([False] * len(df))
        for condition in conditions:
            if mode.lower() == "include":
                mask = mask & (df[column] == condition)
            elif mode.lower() == "exclude":
                mask = mask | (df[column] != condition)
        return df[mask]
    else:
        raise ValueError("Operator must be 'AND' or 'OR'")

def calculate_percentage_of_portfolio(df, grand_totals, column_to_percent):
    """
    Calculate the percentage of the portfolio each investment represents.

    Parameters:
    - df: pandas DataFrame containing the investment data.
    - grand_totals: DataFrame or Series containing the grand totals.
    - column_to_percent: The name of the column to calculate percentages for.

    Returns:
    - DataFrame with a new column for percentage of the portfolio.
    """
    # Retrieve the grand total for the specified column
    grand_total_value = grand_totals[column_to_percent].iloc[0] if isinstance(grand_totals, pd.DataFrame) else grand_totals[column_to_percent]

    # Avoid division by zero by checking if grand_total_value is non-zero
    if grand_total_value == 0:
        print(f"Warning: Grand total for '{column_to_percent}' is zero. Percent calculation skipped.")
        df['Percent of Portfolio'] = 0
    else:
        # Calculate percentage of portfolio for each row
        df['Percent of Portfolio'] = (df[column_to_percent] / grand_total_value) * 100

    return df


##################################################################################################
def tax_lot_appraisal_new(bookkeeping_list, ledger_space, stat_repo, edate, fname, price_data, fx_data):
    if not fname.endswith('.xlsx'):
        fname += '.xlsx'

    # 1. Fetch base data
    df = fetch_data(bookkeeping_list)

    # 2. Apply filters if necessary (for example, filtering on 'Financial_Account')
    # Example placeholder: df = filter_records(df, 'Financial_Account', ['Exclude1', 'Exclude2'], operator="OR", mode="exclude")

    # Debugging: Print initial columns to confirm required columns like 'Quantity', 'Local Cost', 'Book Cost' are present
    print("Initial columns after data fetch:", df.columns)

    # 3. Calculate valuations - this should add columns like 'Market Value Local', 'Market Value Book', etc.
    df = calculate_valuation(df, edate, price_data, fx_data, ledger_space)

    # Debugging: Print columns after valuation calculation to confirm all valuation columns were added
    print("Columns after valuation calculation:", df.columns)

    # 4. Calculate subtotals by grouping by Investment, Ls, and Location
    try:
        subtotal_df = calculate_subtotals(df)
        print("Subtotal columns:", subtotal_df.columns)
    except KeyError as e:
        print(f"KeyError during subtotal calculation: {e}")
        return

    # 5. Concatenate the original and subtotal DataFrames
    df_with_subtotals = pd.concat([df, subtotal_df], ignore_index=True)

    # 6. Calculate grand totals and append to DataFrame
    try:
        grand_totals = calculate_grand_totals(df_with_subtotals)
        print("Grand totals:", grand_totals)
    except KeyError as e:
        print(f"KeyError during grand total calculation: {e}")
        return

    df_with_totals = pd.concat([df_with_subtotals, grand_totals], ignore_index=True)

    # 7. Calculate percentage of portfolio for Market Value Book, if needed
    df_with_totals = calculate_percentage_of_portfolio(df_with_totals, grand_totals, 'Market Value Book')

    # 8. Format and output the report
    format_and_output(df_with_totals, fname)

def position_report_new(bookkeeping_list, ledger_space, stat_repo, edate, fname, price_data, fx_data):
    if not fname.endswith('.xlsx'):
        fname += '.xlsx'

    # 1. Fetch base data
    df = fetch_data(bookkeeping_list)
    print("Columns after fetching data:", df.columns)  # Debugging

    # Ensure required columns are present before grouping
    required_columns = ['Book Cost', 'Local Cost']
    for col in required_columns:
        if col not in df.columns:
            df[col] = 0  # Initialize missing columns

    # 2. Group by Investment and aggregate
    df = df.groupby('Investment', as_index=False).agg({
        'Quantity': 'sum',
        'Local Cost': 'sum',
        'Book Cost': 'sum'
    })

    print("Columns after grouping:", df.columns)  # Debugging

    # 3. Calculate valuations - this should add columns like 'Market Value Local', 'Market Value Book', etc.
    df = calculate_valuation(df, edate, price_data, fx_data, ledger_space)
    print("Columns after valuation calculation:", df.columns)  # Debugging

    # 4. Calculate subtotals by net position
    subtotal_df = calculate_subtotals(df)
    print("Subtotal DataFrame columns:", subtotal_df.columns)  # Debugging

    # 5. Concatenate the original and subtotal DataFrames
    df_with_subtotals = pd.concat([df, subtotal_df], ignore_index=True)

    # 6. Calculate grand totals and append to DataFrame
    grand_totals = calculate_grand_totals(df_with_subtotals)
    print("Grand totals DataFrame:", grand_totals)  # Debugging
    df_with_totals = pd.concat([df_with_subtotals, grand_totals], ignore_index=True)

    # 7. Calculate percentage of portfolio for Market Value Book
    df_with_totals = calculate_percentage_of_portfolio(df_with_totals, grand_totals, 'Market Value Book')

    # 8. Format and output the report to an Excel file
    format_and_output(df_with_totals, fname)
