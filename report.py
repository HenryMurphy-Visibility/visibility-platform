import mark_to_market
#import xlwings
import main
import bookkeeping

from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
import utilities
#from VisibilityProcessing import MASTER_QUERY_SPACES


def underline_row(ws, row):
    u_font = Font(underline="single")
    for cell in row:
        cell.font = u_font

def bold_row(ws, row):
    b_font = Font(bold=True)
    for cell in row:
        cell.font = b_font

def fill_blue_row(row):
    fill = PatternFill(start_color="ADD8E6",
                   end_color="ADD8E6",
                   fill_type="solid")
    for cell in row:
        cell.fill = fill

def format_numbers(ws):
    num_format = '#,##0.00'
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, int) or isinstance(cell.value, float):
                cell.number_format = num_format
def autosize_columns(worksheet):
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[openpyxl.utils.cell.get_column_letter(column[0].column)].width = adjusted_width


def fill_light_green_row(row):
    fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    for cell in row:
        cell.fill = fill

def add_subtotal(df):
    numeric_columns = ['Quantity', 'Local', 'Book', 'Mkt Val Local', 'Mkt Val Book', 'PGain Local', 'PGain Book', 'TotGain Book', 'Fx Gain']
    # add or remove columns to suit your actual dataframe

    out = pd.DataFrame()
    for key, values in df.groupby(['Portfolio', 'Investment', 'LS', 'Location', 'Financial Account']):
        # use .loc to select the numeric columns you want to sum up
        subtotal = values.loc[:, numeric_columns].sum()
        subtotal_line = pd.Series([f'Subtotal', f'{key[1]}', f'{key[2]}', f'{key[3]}', f'{key[4]}'],
                                  index=['Portfolio', 'Investment', 'LS', 'Location', 'Financial Account'])
        subtotal = pd.concat([subtotal, subtotal_line])
        values = pd.concat([values, subtotal.T.to_frame().T])
        out = pd.concat([out, values])
    return out
import pandas as pd

def capitalize_after_underscore(name):
    parts = name.split('_')
    return '_'.join(part.capitalize() for part in parts)

def capitalize_after_underscore(col):
    # Split the column name by underscores, capitalize each part, and then join them back with underscores
    parts = col.split('_')
    # Capitalize the first letter of each part, keeping other letters in the part unchanged
    capitalized_parts = [part[0].upper() + part[1:] if part else '' for part in parts]
    # Join the parts back together with underscores
    return '_'.join(capitalized_parts)

def standardize_columns(df):
    df.columns = [capitalize_after_underscore(col) for col in df.columns]
    return df

def standardize_key_values(df, key_columns):
    for col in key_columns:
        df[col] = df[col].astype(str).apply(capitalize_after_underscore)
    return df

def fetch_and_map_groupings(reporting_space, investment_master_path, coa_path):
    if isinstance(investment_master_path, list):
        investment_master_df = pd.DataFrame(investment_master_path)
    else:
        investment_master_df = pd.read_csv(investment_master_path)

    coa_df = pd.read_csv(coa_path)

    reporting_space = standardize_columns(reporting_space)
    investment_master_df = standardize_columns(investment_master_df)
    coa_df = standardize_columns(coa_df)
#
  #  reporting_space = standardize_key_values(reporting_space, ['investment', 'financial_account'])
   # investment_master_df = standardize_key_values(investment_master_df, ['Ticker'])
   # coa_df = standardize_key_values(coa_df, ['System_Name'])

    investment_groupings = investment_master_df[['Ticker', 'Asset_Class', 'Currency', 'Country', 'Sector', 'Industry','Analyst']]
    coa_groupings = coa_df[['System_Name', 'System_Type', 'BS_Group_Name', 'BS_Group', 'Performance_Category']]

    reporting_space = reporting_space.merge(investment_groupings, left_on='Investment', right_on='Ticker', how='left')
    reporting_space = reporting_space.merge(coa_groupings, left_on='Financial Account', right_on='System_Name', how='left')

    return reporting_space



def flatten_records(df, account_links):
    flattened_records = []
    grouped = df.groupby(['Portfolio', 'Investment', 'Tax Date', 'LS', 'Location'])

    for name, group in grouped:
        print(f"Processing group: {name}")  # Debugging print
        for primary_account, linked_accounts in account_links.items():
            print(f"  Primary account: {primary_account}")  # Debugging print
            primary_records = group[group['Financial Account'] == primary_account]
            secondary_records = group[group['Financial Account'].isin(linked_accounts)]

            if primary_records.empty:
                print(f"  No primary records found for account: {primary_account}")  # Debugging print
                continue

            if secondary_records.empty:
                print(f"  No secondary records found for linked accounts: {linked_accounts}")  # Debugging print

            for _, primary_row in primary_records.iterrows():
                secondary_data = {sec_acc: 0.0 for sec_acc in linked_accounts}  # Initialize secondary values to zero

                print(f"Primary record: {primary_row['Financial Account']} with Quantity: {primary_row['Quantity']}")

                for _, secondary_row in secondary_records.iterrows():
                    secondary_account = secondary_row['Financial Account']
                    if secondary_account in linked_accounts:
                        secondary_data[secondary_account] += secondary_row[
                            'Local']  # or 'Book', depending on the field needed

                        print(f"  Linked secondary record: {secondary_account} with Local: {secondary_row['Local']}")

                flattened_record = {
                    'Portfolio': primary_row['Portfolio'],
                    'Investment': primary_row['Investment'],
                    'Lot ID': primary_row['Lot ID'],
                    'Tax Date':  primary_row['Tax Date'],
                    'LS': primary_row['LS'],
                    'Location': primary_row['Location'],
                    'Primary Account': primary_row['Financial Account'],
                    'Primary Quantity': primary_row['Quantity'],
                    'Primary Local': primary_row['Local'],
                    'Primary Book': primary_row['Book'],
                    # 'Notional': primary_row['Notional'],
                    # 'OFace': primary_row['OFace'],
                }
                for linked_acc in linked_accounts:
                    flattened_record[linked_acc] = secondary_data[linked_acc]
                    print(f"  Aggregated {linked_acc}: {secondary_data[linked_acc]}")

                flattened_records.append(flattened_record)

    return flattened_records


import os
import time
import shutil
#import xlwings as xw
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# Ensure all Excel instances are closed
def close_all_excel_instances():
    for app in xw.apps:
        app.quit()
    print("All Excel instances closed.")


def consolidate_and_format_workbook(portfolio):
    destination_file_path = f"C:/users/hjmne/pycharmprojects/chest/gui_data/{portfolio}_consolidated.xlsx"

    # Close Excel and clear temporary files if any
    close_all_excel_instances()
    temp_file_path = f"{destination_file_path}.temp"
    if os.path.exists(temp_file_path):
        os.remove(temp_file_path)
        print(f"Removed temp file: {temp_file_path}")

    retry_count = 0
    max_retries = 3
    app = None

    while retry_count < max_retries:
        try:
            app = xw.App(visible=False)
            wb = app.books.add()

            # Simulated data
            data_sources = {
                "Investment Master": pd.DataFrame({
                    "Investment": ["Stock A", "Bond B", "Stock C"],
                    "Quantity": [100, 200, 150],
                    "Price": [10.5, 105.0, 15.5]
                }),
                "Prices": pd.DataFrame({
                    "Date": ["2023-01-01", "2023-01-02", "2023-01-03"],
                    "Stock A": [10.5, 10.7, 10.6],
                    "Bond B": [105.0, 105.2, 104.9]
                }),
                "FX Rates": pd.DataFrame({
                    "Date": ["2023-01-01", "2023-01-02", "2023-01-03"],
                    "USD/EUR": [0.85, 0.86, 0.87]
                }),
                "Bond Info": pd.DataFrame({
                    "Bond Name": ["Bond B", "Bond C"],
                    "Maturity Date": ["2030-01-01", "2035-01-01"],
                    "Yield": [2.5, 3.0]
                }),
                "Events": pd.DataFrame({
                    "Event Date": ["2023-01-01", "2023-01-02"],
                    "Event Type": ["Dividend", "Interest"],
                    "Amount": [1000, 2000]
                })
            }

            # Adding sheets with simulated data
            for sheet_name, data in data_sources.items():
                ws = wb.sheets.add(sheet_name)
                ws.range("A1").value = data
                print(f"Mock data added to '{sheet_name}' sheet.")

            # Save temporarily for formatting and close
            wb.save(temp_file_path)
            wb.close()
            print(f"Workbook saved to temp file: {temp_file_path}")

            format_workbook(temp_file_path)  # Apply formatting
            shutil.move(temp_file_path, destination_file_path)
            print(f"Formatted workbook saved as: {destination_file_path}")
            break  # Exit loop on success

        except Exception as e:
            retry_count += 1
            print(f"Attempt {retry_count}: Error during consolidation - {e}")
            time.sleep(2)  # Short pause before retrying

        finally:
            if app:
                app.quit()
                print("Excel application closed.")

    if retry_count == max_retries:
        print("Failed to consolidate and format workbook after multiple attempts.")


def format_workbook(file_path):
    """Apply formatting to the workbook using openpyxl."""
    try:
        wb = load_workbook(file_path)
        color_fill_1 = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        color_fill_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            max_col_width = {}

            for r_idx, row in enumerate(ws.iter_rows(), start=1):
                for c_idx, cell in enumerate(row, start=1):
                    if r_idx == 1:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="center")
                    else:
                        cell.fill = color_fill_1 if r_idx % 2 == 0 else color_fill_2
                    max_col_width[c_idx] = max(max_col_width.get(c_idx, 0), len(str(cell.value) or ""))

            for col, width in max_col_width.items():
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].width = width + 2

        wb.save(file_path)
        wb.close()
        print("Workbook formatting applied and saved.")

    except Exception as e:
        print(f"Error formatting workbook: {e}")



def convert_to_dict(entry):
    try:
        return entry._asdict()
    except AttributeError:
        return entry  # Modify this as per your requirements

def normalize_journal_entries(journal_entries):
    data = [convert_to_dict(entry) for entry in journal_entries]
    return pd.DataFrame(data)

# # Example usage
# journal_entries = [
#     Journals(portfolio='P1', investment='Inv1', tax_date='2023-01-01', ls='LS1', tranid='T1',
#              quantity=10, local=100, book='B1', location='Loc1', financial_account='FA1'),
#     # ... add more Journals entries here
# ]
#
# je_data = normalize_journal_entries(journal_entries)
# You would then merge je_data with investment_master as needed
# Example (assuming investment_master and other variables are defined):
# je_data = pd.merge(je_data, investment_master[['ticker', 'agg_level']], left_on='investment', right_on='ticker', how='left')

# Function to filter JE dataset for gain/loss events
import pandas as pd

def replace_column_name(df, old_name, new_name):
    """
    Replaces the name of a column in a pandas DataFrame.

    :param df: pandas DataFrame in which the column name needs to be replaced.
    :param old_name: The current name of the column.
    :param new_name: The new name for the column.
    :return: DataFrame with the column name replaced.
    """
    if old_name in df.columns:
        df = df.rename(columns={old_name: new_name})
    else:
        print(f"Column '{old_name}' does not exist in the DataFrame.")
    return df





def filter_for_gain_loss_events(df, column_name):
    # Check if the column exists
    if column_name not in df.columns:
        raise KeyError(f"Column '{column_name}' not found in DataFrame")

    # Proceed with your logic assuming the column exists
    merged_df = df.merge(coa_df[['System_Name', column_name]], on=column_name, how='left')
    return merged_df

def filter_for_gain_loss_events(je_df, column_name):
    coa_df = pd.read_csv("C:/Users/hjmne/PycharmProjects/chest/refdata/chart_of_accounts.csv")

    def filter_for_gain_loss_events(df, column_name):
        # Check if the column exists
        if column_name not in df.columns:
            raise KeyError(f"Column '{column_name}' not found in DataFrame")


    # Merge je_df with coa_df on 'Financial Account' and 'SystemName'
    merged_df = je_df.merge(coa_df[['System_Name', column_name]],
                            left_on='Financial Account',
                            right_on='System_Name',
                            how='left')

    # Filter groups where 'SummaryReport' equals 'RealizedIncome'
    gain_loss_groups = merged_df.groupby(['IBOR Date', 'Tran ID']).filter(
        lambda x: (x['SummaryReport'] == 'RealizedIncome').any()
    )

    return gain_loss_groups

# # Usage example:
# # Assuming je_df is your journal entries DataFrame and coa_df is your Chart of Accounts DataFrame
# filtered_je_df = filter_for_gain_loss_events(df)

def filter_records(df, column, conditions, operator="OR", mode="include"):
    """
    Filter records in a DataFrame based on multiple conditions with an inclusion or exclusion mode.

    Parameters:
    - df: The DataFrame to operate on.
    - column: The column to apply the filter conditions to.
    - conditions: A list of values for filtering.
    - operator: Logical operator to combine conditions ("AND" or "OR").
    - mode: Mode of filtering ('include' or 'exclude').

    Returns:
    - DataFrame after filtering.
    """
    if operator.upper() == "OR":
        if mode.lower() == "include":
            return df[df[column].isin(conditions)]
        elif mode.lower() == "exclude":
            return df[~df[column].isin(conditions)]
    elif operator.upper() == "AND":
        mask = pd.Series([True] * len(df)) if mode.lower() == "include" else pd.Series([False] * len(df))
        for condition in conditions:
            if mode.lower() == "include":
                mask = mask & (df[column] == condition)
            elif mode.lower() == "exclude":
                mask = mask | (df[column] != condition)
        return df[mask]
    else:
        raise ValueError("Operator must be 'AND' or 'OR'")

# Example usage
# Assuming df is your DataFrame
# To include rows where 'Financial Account' matches any of the conditions
# included_df_or = filter_records(df, 'Financial Account', ['Cost', 'Revenue'], 'OR', 'include')
#
# # To exclude rows where 'Financial Account' matches any of the conditions
# excluded_df_or = filter_records(df, 'Financial Account', ['Cost', 'Revenue'], 'OR', 'exclude')

import pandas as pd

def fetch_bookkeeping_space_data(bookkeeping_space):
    """
    Create a DataFrame from bookkeeping space entries.

    Parameters:
    - bookkeeping_space: A collection of bookkeeping entries, where each entry is a tuple
      consisting of a key (another tuple) and a value (tuple of quantity, local, and book).

    Returns:
    - A pandas DataFrame representing the bookkeeping data.
    """
    bookkeeping_space_list = []
    for key, (quantity, local, book) in bookkeeping_space:
        portfolio, investment, tax_lot_num, ls, location, financial_account = key
        booksp_row = [portfolio, investment, tax_lot_num, ls, location, financial_account, quantity, local, book]
        bookkeeping_space_list.append(booksp_row)

    df = pd.DataFrame(bookkeeping_space_list, columns=[
        'Portfolio', 'Investment', 'Tax Lot Num', 'LS', 'Location', 'Financial Account', 'Quantity', 'Local', 'Book'
    ])

    return df

# Usage example
# # Assuming you have a 'bookkeeping_space' variable defined as per your structure
# df_bookkeeping = fetch_bookkeeping_space_data(bookkeeping_space)

def merge_and_drop_duplicates(df1, df2, key):

    # Merge the dataframes on the specified key
    merged_df = pd.merge(df1, df2, on=key, how='inner')

    # Drop duplicate columns
    merged_df = merged_df.loc[:,~merged_df.columns.duplicated()]

    return merged_df

# Example usage:
# merged_df = merge_and_drop_duplicates('path_to_file1.csv', 'path_to_file2.csv', 'Portfolio')
# print(merged_df.head())  # This will print the first 5 rows of the merged dataframe without duplicates

def create_combined_column(df, inv_master_path, coa_path, inv_column=None, coa_column=None, new_column_name="Combined"):
    """
    Create a new column in the DataFrame by fetching a column from investment_master or chart_of_accounts,
    or by joining one column from each.

    Parameters:
    - df: The DataFrame to which the new column will be added.
    - inv_master_path: Path to the investment_master CSV file.
    - coa_path: Path to the chart_of_accounts CSV file.
    - inv_column: The name of the column to fetch from investment_master.
    - coa_column: The name of the column to fetch from chart_of_accounts.
    - new_column_name: The name of the new column to be added to the DataFrame.

    Returns:
    - The DataFrame with the new combined column added.
    """
    # Load the investment_master and chart_of_accounts data
    inv_master_df = pd.read_csv(inv_master_path)
    coa_df = pd.read_csv(coa_path)

    # Check which columns to fetch and combine
    if inv_column and not coa_column:
        # Fetch the column from investment_master
        combined_data = inv_master_df.set_index('Investment')[inv_column]
    elif coa_column and not inv_column:
        # Fetch the column from chart_of_accounts
        combined_data = coa_df.set_index('SystemName')[coa_column]
    elif inv_column and coa_column:
        # Fetch and join the columns from both investment_master and chart_of_accounts
        combined_data = inv_master_df.set_index('Investment')[inv_column] + ' ' + coa_df.set_index('SystemName')[coa_column]
    else:
        raise ValueError("At least one of inv_column or coa_column must be provided.")

    # Map the combined data to the DataFrame using the 'Investment' column
    df[new_column_name] = df['Investment'].map(combined_data)

    return df

# Example usage:
# df = pd.DataFrame({'Investment': ['Investment1', 'Investment2'], 'OtherColumn': [123, 456]})
# df = create_combined_column(df,
#                             inv_master_path='path_to_investment_master.csv',
#                             coa_path='path_to_chart_of_accounts.csv',
#                             inv_column='InvColumn',
#                             coa_column='CoaColumn',
#                             new_column_name='NewCombinedColumn')


def drop_n_columns(df, n):
    """
    Drop the last N columns from a DataFrame.

    Parameters:
    - df: pandas DataFrame from which to drop columns.
    - n: The number of last columns to drop.

    Returns:
    - DataFrame with the last N columns dropped.
    """
    # Check if the DataFrame has at least n columns
    if n > len(df.columns):
        raise ValueError("The DataFrame doesn't have enough columns to drop.")

    # Drop the last n columns
    df = df.drop(df.columns[-n:], axis=1)
    return df

def drop_columns_by_name(df, columns_to_drop):
    """
    Drop specified columns from a DataFrame by name.

    Parameters:
    - df: pandas DataFrame from which to drop columns.
    - columns_to_drop: List of column names to drop.

    Returns:
    - DataFrame with specified columns dropped.
    """
    df = df.drop(columns_to_drop, axis=1, inplace = True)
    return df


def move_column_in_dataframe(df, column, position):
    """
    Move a column in a DataFrame to a specific position.

    Parameters:
    - df: pandas DataFrame.
    - column: The name of the column to move.
    - position: The position index to move the column to.

    Returns:
    - DataFrame with the column moved to the new position.
    """
    # Get a list of columns in the DataFrame
    columns = list(df.columns)

    # Remove the column to move from its current position
    columns.remove(column)

    # Insert the column at the desired position
    columns.insert(position, column)

    # Return the DataFrame with columns in the new order
    return df[columns]


# Assuming df_final_report is your DataFrame and 'Percent of Portfolio' is the column you want to move
# Move 'Percent of Portfolio' to be the second column (position index 1)
# df_final_report = move_column_in_dataframe(df_final_report, 'Percent of Portfolio', 1)


def calculate_percentage_of_portfolio(df, grand_totals, column_to_percent):
    """
    Calculate the percentage of the portfolio each investment represents.

    Parameters:
    - df: pandas DataFrame containing the investment data.
    - grand_totals: DataFrame or Series containing the grand totals.
    - column_to_percent: The name of the column to calculate percentages for.

    Returns:
    - DataFrame with a new column for percentage of the portfolio.
    """
    # Get the grand total value for the specified column
    grand_total_value = grand_totals[column_to_percent].iloc[0]

    # Calculate the percentage of the grand total for each investment
    df['Percent of Portfolio'] = (df[column_to_percent] / grand_total_value) * 100

    # If you want to merge this back into the original dataframe, you can return it directly
    return df


# Use the subroutine after calculating subtotals and before finalizing the report
# Assume df_subtotals is your dataframe after inserting subtotals
# Assume grand_totals is your precalculated grand totals dataframe
# Assume 'Book' is the column you're interested in for percentage calculation
#df_final_report = calculate_percentage_of_portfolio(df_subtotals, grand_totals, 'Book')


# Now df_final_report will have a new column 'Percent of Portfolio' with the calculated values


def fetch_from_inv_master_and_merge(df, inv_master_path, inv_column, df_key='Investment'):
    """
    Fetch a single column from the investment master and merge with the DataFrame.

    Parameters:
    - df: DataFrame to merge with.
    - inv_master_path: Path to the investment_master CSV file.
    - inv_column: The name of the column to fetch from investment_master.
    - df_key: The name of the column in df to use as the primary key for merging.

    Returns:
    - Merged DataFrame.
    """
    inv_master_df = pd.read_csv(inv_master_path)
    merged_df = df.merge(inv_master_df[[df_key, inv_column]], on=df_key, how='left')
    return merged_df

def fetch_from_coa_master_and_merge(df, coa_path, coa_column, df_key='SystemName'):
    """
    Fetch a single column from the chart of accounts master and merge with the DataFrame.

    Parameters:
    - df: DataFrame to merge with.
    - coa_path: Path to the chart_of_accounts CSV file.
    - coa_column: The name of the column to fetch from chart_of_accounts.
    - df_key: The name of the column in df to use as the primary key for merging.

    Returns:
    - Merged DataFrame.



    """
    coa_df = pd.read_csv(coa_path)
    df = df.reset_index()
    coa_df = coa_df.reset_index()

    merged_df = df.merge(coa_df[[coa_column, df_key]],
                              left_on='Financial Account', # left being df
                              right_on='SystemName', #right beinf coa_df
                              how='left')

    # If you don't want to keep the 'SystemName' column after merging, you can drop it
    merged_df.drop('SystemName', axis=1, inplace=True)
    return merged_df


from collections import defaultdict
import pandas as pd

from collections import defaultdict
import pandas as pd


def update_running_balances(journal_entries, group_keys):
    """
    Update each journal entry with running balances for quantity, local, and book.

    Parameters:
    - journal_entries: List of `Journals` objects.
    - group_keys: List of fields to group by, e.g., ['portfolio', 'investment', 'ls', 'location', 'financial_account'].
    """
    # Convert journal entries to a DataFrame for easy grouping and sorting
    df = pd.DataFrame([
        {
            'portfolio': je.portfolio,
            'investment': je.investment,
            'ls': je.ls,
            'location': je.location,
            'financial_account': je.financial_account,
            'ibor_date': je.ibor_date,
            'quantity': je.quantity,
            'local': je.local,
            'book': je.book,
            'entry': je  # Store reference to original Journals object
        }
        for je in journal_entries
    ])

    # Sort by grouping keys and ibor_date to ensure cumulative calculations are correct
    df = df.sort_values(by=group_keys + ['ibor_date'])

    # Calculate cumulative sums for quantity, local, and book within each group
    df['RunningQty'] = df.groupby(group_keys)['quantity'].cumsum()
    df['RunningLocal'] = df.groupby(group_keys)['local'].cumsum()
    df['RunningBook'] = df.groupby(group_keys)['book'].cumsum()

    # Update each Journals entry with calculated running balances
    for _, row in df.iterrows():
        entry = row['entry']  # Access the original Journals object
        entry.running_balances = (row['RunningQty'], row['RunningLocal'], row['RunningBook'])


def filter_and_sort_data(df, sort_fields):
    # Filter and sort DataFrame based on given fields
    df.sort_values(by=sort_fields, inplace=True)
    return df


def calculate_subtotals(df, group_fields, subtotal_fields):
    # Group by given fields and calculate subtotal for given fields
    subtotals = df.groupby(group_fields)[subtotal_fields].sum().reset_index()
    return subtotals


def insert_subtotals(df, group_by_cols, subtotal_cols):
    # Create an empty DataFrame to hold the data with subtotals
    data_with_subtotals = pd.DataFrame(columns=df.columns)

    # Group the DataFrame by the specified columns
    grouped = df.groupby(group_by_cols, as_index=False)

    for _, group in grouped:
        # Calculate the subtotal for the group
        subtotal_row = group[subtotal_cols].sum()

        # Since subtotal_row is a Series, we create a new DataFrame with one row
        # and reindex it to match the original DataFrame's columns
        subtotal_df = pd.DataFrame([subtotal_row], columns=subtotal_cols)

        # Set the non-subtotal columns to some identifier, e.g., 'Subtotal'
        for col in df.columns.difference(subtotal_cols):
            subtotal_df[col] = 'Subtotal' if col == 'Portfolio' else None

        # Append the group and its subtotal row to the data_with_subtotals DataFrame
        data_with_subtotals = pd.concat([data_with_subtotals, group, subtotal_df], ignore_index=True)

    return data_with_subtotals

import pandas as pd
import mark_to_market  # Ensure you import the required module

import central_processing_hub
from bookkeeping import BookkeepingSpace
from utilities import flatten_nested_tuples



import central_processing_hub
from bookkeeping import BookkeepingSpace

import central_processing_hub
from bookkeeping import BookkeepingSpace
from utilities import flatten_nested_tuples



def prepare_main_reporting_data(main_space):
    # Access the already-built main_space and prepare the data for reporting
    main_asset_liability_data = main_space.combined_assets_liabilities()
    all_main_bookkeeping_accounts = main_space.get_combined_space()

    return main_space, main_asset_liability_data, all_main_bookkeeping_accounts

def prepare_gl_reporting_data(combined_journals, period_cutoff):
    # Create a new instance of BookkeepingSpace specifically for GL reporting
    gl_space = bookkeeping.BookkeepingSpace()

    # Build the GL-specific bookkeeping repository using the combined journals
    gl_space.build_general_ledger_from_journals(combined_journals, period_cutoff)

    # Generate the necessary data for GL reporting
    gl_asset_liability_data = gl_space.combined_assets_liabilities()
    gl_report_data = gl_space.get_combined_space()

    return gl_space, gl_asset_liability_data, gl_report_data

def value_positions(bookkeeping_space_df, bookkeeping_space, date, mark_type):
    # Define the required columns
    required_columns = ['Portfolio', 'Investment', 'Lot_ID', 'Tax_Date', 'Ls', 'Location',
                        'Financial_Account', 'Quantity', 'Local', 'Book', 'Notional', 'Oface']

    # Check for missing columns
    missing_columns = [col for col in required_columns if col not in bookkeeping_space_df.columns]
    if missing_columns:
        raise KeyError(f"DataFrame is missing required columns: {missing_columns}")

    # Perform mark calculations
    f1, f2, mark_date = mark_to_market.get_data_and_format(date)
    marked_records = mark_to_market.calculate_marks(bookkeeping_space_df, bookkeeping_space, f1, f2, date)

    # Flatten the marked records
    flattened_records = []
    for rk, data in marked_records:
        record = list(rk) + list(data)
        flattened_records.append(record)

    # Define the final columns for the DataFrame
    final_columns = ['portfolio', 'investment', 'lotid', 'tax_date', 'ls', 'location', 'financial_account',
                     'quantity', 'local', 'book', 'mkt_val_local', 'mkt_val_book',
                     'pgain_local', 'pgain_book', 'totgain_book', 'fx_gain', 'investment_type']

    # Ensure the length of records matches the final columns
    for record in flattened_records:
        if len(record) != len(final_columns):
            raise ValueError("Mismatch between the number of columns in records and final_columns")

    # Create a DataFrame from the flattened records
    df_valued = pd.DataFrame(flattened_records, columns=final_columns)

    return df_valued


# Example of the mark_to_market.calculate_marks function
def calculate_marks(bookkeeping_space_df, bookkeeping_space, f1, f2, date):
    marked_records = []
    for _, subspace in bookkeeping_space_df.iterrows():
        pricing_factor_str = subspace.get_attribute_field("AIF", "Pricing_Factor") if subspace else None
        try:
            pricing_factor = float(pricing_factor_str) if pricing_factor_str is not None else 1.0
        except ValueError:
            pricing_factor = 1.0  # or handle the error as necessary

        # Example calculation logic (replace with actual logic)
        rk = (subspace['portfolio'], subspace['investment'], subspace['lotid'])
        data = (subspace['tax_date'], subspace['ls'], subspace['location'], subspace['financial_account'],
                subspace['quantity'], subspace['local'], subspace['book'], subspace['notional'], subspace['oface'],
                pricing_factor, pricing_factor * subspace['local'], pricing_factor * subspace['book'],
                0, 0, 0, 0, 'example_type')  # Replace with actual data structure

        marked_records.append((rk, data))

    # Return the calculated marks
    return marked_records


# Updated calculate_marks function to handle NoneType
def subtotal_data(df, subtotal_columns, group_by_cols):
    # Convert the DataFrame to a dictionary for calculations

    # Apply the insert_subtotals function to add subtotals
    df_with_subtotals = insert_subtotals(df, group_by_cols, subtotal_columns)

    return df_with_subtotals

# # Usage example
# group_by_columns = ['Investment', 'Financial Account']  # Columns to group by and insert subtotals after
# subtotal_columns = ['Quantity', 'Local

def calculate_grand_totals(df, total_columns):
    """
    Calculate the grand totals for specified columns in a DataFrame.

    Parameters:
    - df: pandas DataFrame containing the data.
    - total_columns: list of column names to sum for the grand total.

    Returns:
    - grand_totals: DataFrame containing the grand totals.
    """
    # Sum the specified columns to get the grand totals
    totals = df[total_columns].sum().to_frame().T
    # Set the non-total columns to a placeholder or descriptive string
    for col in df.columns.difference(total_columns):
        totals[col] = "Grand Total" if col == 'Portfolio' else ""
    return totals

import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import pandas as pd
from pandas.plotting import table

def df_to_pdf(df, pdf_filename, font_size=10, figsize=(11.7, 8.3)):
    """
    Save a DataFrame to a PDF file with custom font size and figure size.

    Parameters:
    - df: The pandas DataFrame to save as PDF.
    - pdf_filename: The filename of the PDF to create.
    - font_size: Font size for the table content.
    - figsize: Size of the figure (width, height in inches). Default is A4 size.
    """
    # Configure matplotlib font size
    plt.rcParams.update({'font.size': font_size})

    # Create a new PDF with PdfPages
    with PdfPages(pdf_filename) as pdf:
        # Initialize a figure
        fig, ax = plt.subplots(figsize=figsize)
        # Hide axes
        ax.axis('off')

        # Draw the table
        table(ax, df, loc='center')

        # Save the current figure to the PDF
        pdf.savefig(fig, bbox_inches='tight')

        # Close the figure
        plt.close(fig)

# # Example usage:
# df = pd.DataFrame({'A': [1, 2, 3], 'B': [4, 5, 6]})
# df_to_pdf(df, 'output.pdf', font_size=12, figsize=(8, 6))


def combine_and_drop_columns(df, columns_to_combine, new_column_name):
    """
    Sums specified numeric columns, adds the result as a new column, and drops the original columns.

    Parameters:
    - df: The DataFrame to process.
    - columns_to_combine: A list of column names to sum.
    - new_column_name: The name of the new column that will store the sum.

    Returns:
    - The DataFrame with the new column added and original columns dropped.
    """
    # Check that all columns exist in the DataFrame
    for col in columns_to_combine:
        if col not in df.columns:
            raise KeyError(f"Column {col} does not exist in the DataFrame.")

    # Create the new column as the sum of the columns to combine
    df[new_column_name] = df[columns_to_combine].sum(axis=1)

    # Drop the original columns
    df.drop(columns=columns_to_combine, inplace=True)

    return df

def apply_standard_formatting(excel_file_path, sheet_name = "Sheet1"):
    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook[sheet_name]    # Freeze the first row
    sheet.freeze_panes = sheet['A2']

    # Set column widths based on the maximum length of the content
    for column in sheet.columns:
        max_length = max((len(str(cell.value)) for cell in column[1:] if cell.value), default=0) + 1
        adjusted_width = (max_length + 2) * 1.2
        column_letter = get_column_letter(column[0].column)
        sheet.column_dimensions[column_letter].width = adjusted_width



    # Apply number formatting (assuming 'format_numbers' is a predefined function)
    # format_numbers(sheet)

    # Apply auto-sizing (assuming 'autosize_columns' is a predefined function)
    # autosize_columns(sheet)
    # Assuming 'Subtotal' is in a specific column, e.g., first column
    subtotal_column_index = 1  # Adjust as per your file

    # Iterate through rows and apply bold font to subtotal rows
    for row in sheet.iter_rows(min_row=2):  # Assuming the first row is headers
        if row[subtotal_column_index - 1].value == 'Subtotal':
            for cell in row:
                cell.font = Font(bold=True)

    # Apply zebra striping (assuming 'fill_blue_row' is a predefined function)
    for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if i % 2 == 0:  # Apply to every even row
            for cell in row:
                cell.fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

    num_format = '#,##0.00'
    num_format = '#,##0.00'
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, int) or isinstance(cell.value, float):
                cell.number_format = num_format
    # Save the changes to the workbook
    workbook.save(excel_file_path)
# Function to replace '1970-01-01' with 'avgcost'
# Function to replace '1970-01-01 00:00:00' dates with 'avgcost'
def replace_1970_with_avgcost(df):
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].apply(lambda x: 'avgcost' if x == pd.Timestamp('1970-01-01 00:00:00') else x)
    return df
# Function to replace '1970-01-01 00:00:00' dates with 'avgcost'



def apply_standard_formatting_optimized(excel_file_path, sheet_name = "Sheet1"):
    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook[sheet_name]    # Freeze the first row
    sheet.freeze_panes = sheet['A2']

    # Set column widths based on the maximum length of the content
    for column in sheet.columns:
        max_length = max((len(str(cell.value)) for cell in column[1:] if cell.value), default=0) + 2
        adjusted_width = (max_length + 2) * 1.2
        column_letter = get_column_letter(column[0].column)
        sheet.column_dimensions[column_letter].width = adjusted_width



    # Apply number formatting (assuming 'format_numbers' is a predefined function)
    # format_numbers(sheet)

    # Apply auto-sizing (assuming 'autosize_columns' is a predefined function)
    # autosize_columns(sheet)
    # Assuming 'Subtotal' is in a specific column, e.g., first column
    subtotal_column_index = 1  # Adjust as per your file

    # Iterate through rows and apply bold font to subtotal rows
    for row in sheet.iter_rows(min_row=2):  # Assuming the first row is headers
        if row[subtotal_column_index - 1].value == 'Subtotal':
            for cell in row:
                cell.font = Font(bold=True)

    # Apply zebra striping (assuming 'fill_blue_row' is a predefined function)
    for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if i % 2 == 0:  # Apply to every even row
            for cell in row:
                cell.fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

    num_format = '#,##0.00'
    num_format = '#,##0.00'
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, int) or isinstance(cell.value, float):
                cell.number_format = num_format
    # Save the changes to the workbook
    workbook.save(excel_file_path)


import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def add_title(excel_path, sheet_name, report_title):
    # Load the workbook and select the sheet
    workbook = load_workbook(excel_path)

    sheet = workbook[sheet_name]

    # Insert a new row at the top for the title
    sheet.insert_rows(1)

    # Insert a new row at the top for the title
    sheet.insert_rows(1)

    # Merge cells from A1 to L1 for the title
    sheet.merge_cells('A1:L1')

    # Write the title into the merged cell
    title_cell = sheet['A1']
    title_cell.value = f"{report_title}"

    title_cell.value = f"{report_title}"
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = f"{report_title}"
    title_font_size = 16.5
    title_cell.font = Font(bold=True, size=title_font_size)

    # Center the title
    title_cell.alignment = Alignment(horizontal='center')

    # Row number of your title
    title_row_number = 1

    # Get the current height of the title row
    current_height = sheet.row_dimensions[title_row_number].height

    # Set the new height to be twice the current height
    # If the current height is None (auto-height), you can assign a specific value
    new_height = current_height * 1.5 if current_height is not None else 30  # 30 is an example value

    # Apply the new height to the title row
    sheet.row_dimensions[title_row_number].height = new_height
    # Apply any other formatting you want for the title
    # ...

    # Save the workbook
    workbook.save(excel_path)

def format_excel_sheet(sheet):
    # Excel formatting logic
    pass


def save_to_excel(df, filename, sheet_name):
    # Save DataFrame to Excel with formatting
    df.to_excel(filename, index=False)
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheet_name]
    format_excel_sheet(sheet)
    workbook.save(filename)


def map_external_data(df, mapping_dict, account_column='Financial Account'):
    # Map external data to DataFrame
    df['MappedGroup'] = df[account_column].map(mapping_dict)
    return df
import pandas as pd
from datetime import datetime
import pandas as pd
from datetime import datetime

import pandas as pd

PERIOD_END_DATE = '12/30/2022'

def flatten_records_for_tax_lots(bookkeeping_records, primary_accounts, sub_ledger):
    grouped_entries = {}

    # Group bookkeeping records by key fields
    for record in bookkeeping_records:
        group_key = (record[0], record[1], record[2], record[3], record[4], record[5])  # Portfolio, Investment, Lot_Id, Tax_Date, Ls, Location
        if group_key not in grouped_entries:
            grouped_entries[group_key] = []
        grouped_entries[group_key].append(record)

    flattened_records = []

    # Process each group
    for group_key, entries in grouped_entries.items():
        primary_records = [record for record in entries if record[6] in primary_accounts]
        secondary_records = [record for record in entries if record[6] not in primary_accounts]

        if not primary_records:
            print(f"No primary records found for group: {group_key}")
            continue

        for primary_row in primary_records:
            primary_account = primary_row[6]
            linked_accounts = primary_accounts.get(primary_account, {}).get('SecondaryAccounts', {})

            # Initialize secondary data dictionaries, including placeholders for calculated columns
            secondary_data = {
                'MktValLocal': 0.0,
                'MktValBook': 0.0,
                'PriceGL': 0.0,  # Placeholder for calculated PriceGL
                'FXGL': 0.0      # Placeholder for calculated FXGL
            }

            # Process secondary records to update MktValLocal and MktValBook
            for secondary_row in secondary_records:
                secondary_account = secondary_row[6]
                if secondary_account == 'MarketVal':
                #    print(f"Found MarketVal record: Local={secondary_row[8]}, Book={secondary_row[9]}")
                    secondary_data['MktValLocal'] += secondary_row[8]  # Local value
                    secondary_data['MktValBook'] += secondary_row[9]  # Book value
                else:
                    print(f"Secondary account {secondary_account} does not match MarketVal")


            # Calculate PriceGL and FXGL from MarketValLocal and MarketValBook
            # These calculations should be adapted based on your specific logic
            # Example calculations:
            secondary_data['PriceGL'] = secondary_data['MktValLocal'] - primary_row[8]  # Subtract cost or another baseline
            secondary_data['FXGL'] = secondary_data['MktValBook'] - primary_row[9]  # Subtract book value or another baseline

            # Calculate the additional columns and create the primary record
            primary_record = {
                'Portfolio': primary_row[0],
                'Investment': primary_row[1],
                'Lot_ID': primary_row[2],
                'Tax_Date': primary_row[3],
                'Ls': primary_row[4],
                'Location': primary_row[5],
                'Financial_Account': primary_account,
                'Quantity': primary_row[7],
                'Local': primary_row[8],
                'Book': primary_row[9],
                'PriceGL': secondary_data['PriceGL'],  # Include the calculated PriceGL
                'FXGL': secondary_data['FXGL'],        # Include the calculated FXGL
                'MktValLocal': secondary_data['MktValLocal'],
                'MktValBook': secondary_data['MktValBook']
            }

            flattened_records.append(primary_record)

    return flattened_records




def position_report(bookkeeping_list, portfolio, ledger_space, edate, fname, price_data, fx_data):
    if not fname.endswith('.xlsx'):
        fname += '.xlsx'

    # Ensure output directory
    output_path = "BASE_PATH/reports"
    os.makedirs(output_path, exist_ok=True)

    # Construct the full filename with portfolio prepended
    full_filename = f"{portfolio}_{fname}"
    output_filepath = os.path.join(output_path, full_filename)

    # Handle case where bookkeeping_list is empty
    if not bookkeeping_list:
        print(f"No positions found as of {edate}. Generating a report with initial value as 0.")
        placeholder_df = pd.DataFrame(
            [{
                "Investment": "Initial Value", "Ls": "-", "Location": "-", "Quantity": 0,
                "Local_Cost": 0, "Book_Cost": 0, "MktVal_Local": 0, "MktVal_Book": 0,
                "PriceGLLocal": 0, "PriceGLBook": 0, "FXGLBook": 0
            }]
        )
        placeholder_df.to_excel(output_filepath, index=False, sheet_name='Position Report')
        print(f"Position report saved as {output_filepath}.")
        return

    # Process entries
    bookkeeping_space_list = []
    for entry in bookkeeping_list:
        portfolio, investment, lotid, tax_date, ls, location, financial_account, quantity, local, book, notional, oface = entry
        bookkeeping_space_list.append([portfolio, investment, ls, location, financial_account, quantity, local, book])

    df = pd.DataFrame(bookkeeping_space_list, columns=[
        'Portfolio', 'Investment', 'Ls', 'Location', 'Financial_Account', 'Quantity', 'Local_Cost', 'Book_Cost'
    ])

    df_aggregated = df.groupby(['Investment', 'Ls', 'Location']).agg({
        'Quantity': 'sum',
        'Local_Cost': 'sum',
        'Book_Cost': 'sum'
    }).reset_index()

    valuation_data = []
    for _, row in df_aggregated.iterrows():
        investment = row['Investment']
        quantity = row['Quantity']
        local_cost = row['Local_Cost']
        book_cost = row['Book_Cost']
        price = utilities.get_price(investment, edate, price_data)
        currency = ledger_space.get_attribute_field(investment, 'AIF', 'Currency')
        fx_rate = utilities.get_fx_rate(currency, edate, fx_data)
        if price is None or fx_rate is None:
            print(f"Price or FX rate not found for {investment} on {edate}")
            continue
        pricing_factor = ledger_space.get_attribute_field(investment, 'AIF', 'Pricing_Factor') or 1
        mktval_local = quantity * float(price) * float(pricing_factor)
        mktval_book = mktval_local * fx_rate
        price_gl_local = mktval_local - local_cost
        price_gl_book = price_gl_local * fx_rate
        fx_gl_book = mktval_book - book_cost - price_gl_book
        valuation_data.append({
            'Investment': investment,
            'Ls': row['Ls'],
            'Location': row['Location'],
            'Quantity': quantity,
            'Local_Cost': local_cost,
            'Book_Cost': book_cost,
            'MktVal_Local': mktval_local,
            'MktVal_Book': mktval_book,
            'PriceGLLocal': price_gl_local,
            'PriceGLBook': price_gl_book,
            'FXGLBook': fx_gl_book
        })

    df_valuations = pd.DataFrame(valuation_data)
    df_valuations['Portfolio'] = 'Subtotal'
    df_valuations['Financial_Account'] = ''

    df_final = df_valuations.sort_values(by=['Investment', 'Ls', 'Location']).reset_index(drop=True)

    grand_totals = df_final.agg({
        'Book_Cost': 'sum',
        'MktVal_Book': 'sum',
        'PriceGLBook': 'sum',
        'FXGLBook': 'sum'
    }).to_dict()
    grand_totals.update({
        'Portfolio': 'Grand Totals',
        'Investment': '',
        'Ls': '',
        'Location': '',
        'Financial_Account': ''
    })

    df_with_totals = pd.concat([df_final, pd.DataFrame([grand_totals])], ignore_index=True)

    columns = [col for col in df_with_totals.columns if col not in ['Portfolio', 'Location', 'Financial_Account']]
    columns += ['Portfolio', 'Location', 'Financial_Account']
    df_with_totals = df_with_totals[columns]

    # Save to Excel
    df_with_totals.to_excel(output_filepath, index=False, sheet_name='Position Report')
    print(f"Position report saved as {output_filepath}.")

def tax_lot_appraisal(bookkeeping_list, portfolio,  ledger_space, stat_repo, sdate, edate, fname, price_data, fx_data):
    if not fname.endswith('.xlsx'):
        fname += '.xlsx'

    # Check if bookkeeping_list is empty
    if not bookkeeping_list:
        print("Error: No positions found for the portfolio. Skipping appraisal.")
        return  # Terminate early and allow subsequent processes to run

    bookkeeping_space_list = []

    # Process each entry in bookkeeping_list to calculate market values
    for entry in bookkeeping_list:
        try:
            portfolio, investment, lotid, tax_date, ls, location, financial_account, quantity, local, book, notional, oface = entry

            # Fetch the price for the investment for the target date
            price = utilities.get_price(investment, edate, price_data)

            # Fetch the FX rate for the target date using fx_data
            currency = ledger_space.get_attribute_field(investment, 'AIF', 'Currency')
            fx_rate = utilities.get_fx_rate(currency, edate, fx_data)

            # Add processing logic (e.g., calculate market values) as needed
            # bookkeeping_space_list.append(processed_entry)

        except Exception as e:
            print(f"Error processing entry {entry}: {e}")
            continue  # Skip this entry and process the next one

        # Fetch the price for the investment for the target date
        price = utilities.get_price(investment, edate, price_data)

        # Fetch the FX rate for the target date using fx_data
        currency = ledger_space.get_information_field(investment, 'AIF', 'Currency')
        fx_rate = utilities.get_fx_rate(currency, edate, fx_data)

        # Ensure we have both a valid price and fx_rate, otherwise skip the entry
        if price is None:
            print(f"Price not found for {investment} on {edate}")
            continue

        if fx_rate is None:
            print(f"FX rate not found for {currency} on {edate}")
            continue

        # Fetch the pricing factor from bookkeeping (this can be specific to certain securities)
        pricing_factor = ledger_space.get_information_field(investment, 'AIF', 'Pricing_Factor')

        # Calculate market values based on the price and FX rate
        mktval_local = quantity * float(price) * float(pricing_factor) - notional
        mktval_book = mktval_local * float(fx_rate)

        # Calculate unrealized Price GL and FX GL
        price_gl_local = mktval_local - local  # Price GL in local currency
        price_gl_book = price_gl_local * float(fx_rate)  # Price GL in book currency
        fx_gl_book = mktval_book - book - price_gl_book  # FX GL in book currency

        # Construct a row for the bookkeeping space list
        booksp_row = [
            portfolio, investment, lotid, tax_date, ls, location, financial_account,
            quantity, local, book, mktval_local, mktval_book,
            price_gl_local, price_gl_book, fx_gl_book
        ]

        # Append the row to the bookkeeping space list
        bookkeeping_space_list.append(booksp_row)

    # Convert the bookkeeping space list into a DataFrame for further processing
    df = pd.DataFrame(bookkeeping_space_list, columns=[
        'Portfolio', 'Investment', 'Lot_ID', 'Tax_Date', 'Ls', 'Location', 'Financial_Account', 'Quantity',
        'Local_Cost', 'Book_Cost', 'MktVal_Local', 'MktVal_Book', 'PriceGLLocal', 'PriceGLBook', 'FXGLBook'
    ])

    # Add 'IsCurrency' column to help with sorting if needed
    df['IsCurrency'] = df['Investment'].apply(lambda inv: ledger_space.get_investment_attribute('AIF', inv, 'IsCurrency'))

    # Group by 'Investment', 'Ls', and 'Location' and calculate subtotals
    subtotal_df = df.groupby(['Investment', 'Ls', 'Location']).agg({
        'Quantity': 'sum',
        'Local_Cost': 'sum',
        'Book_Cost': 'sum',
        'MktVal_Local': 'sum',
        'MktVal_Book': 'sum',
        'PriceGLLocal': 'sum',
        'PriceGLBook': 'sum',
        'FXGLBook': 'sum'
    }).reset_index()

    # Label the subtotals for clarity
    subtotal_df['Portfolio'] = 'Subtotal'
    subtotal_df['Lot_ID'] = ''
    subtotal_df['Tax_Date'] = ''
    subtotal_df['Financial_Account'] = ''

    # Concatenate the original and subtotal DataFrames
    df_with_subtotals = pd.concat([df, subtotal_df], ignore_index=True)

    # Sort by 'Investment', 'Ls', and 'Location' to keep related entries together
    df_with_subtotals = df_with_subtotals.sort_values(by=['Investment', 'Ls', 'Location']).reset_index(drop=True)

    # Calculate grand totals
    grand_totals = df_with_subtotals.agg({
        'Book_Cost': 'sum',
        'MktVal_Book': 'sum',
        'PriceGLBook': 'sum',
        'FXGLBook': 'sum',
    }).to_dict()

    grand_totals.update({
        'Portfolio': 'Grand Totals',
        'Investment': '',
        'Lot_ID': '',
        'Tax_Date': '',
        'Ls': '',
        'Location': '',
        'Financial_Account': ''
    })

    df_with_subtotals = df_with_subtotals.sort_values(by=['Investment', 'Ls', 'Location']).reset_index(drop=True)

    # Calculate grand totals
    grand_totals = df_with_subtotals.agg({
        'Book_Cost': 'sum',
        'MktVal_Book': 'sum',
        'PriceGLBook': 'sum',
        'FXGLBook': 'sum',
    }).to_dict()

    grand_totals.update({
        'Portfolio': 'Grand Totals',
        'Investment': '',
        'Lot_ID': '',
        'Tax_Date': '',
        'Ls': '',
        'Location': '',
        'Financial_Account': ''
    })

    # Append grand totals to the DataFrame
    df_with_totals = pd.concat([df_with_subtotals, pd.DataFrame([grand_totals])], ignore_index=True)

    # Rearrange columns
    columns_order = [col for col in df_with_totals.columns if col not in ['Portfolio', 'Location', 'Financial_Account']]
    columns_order += ['Portfolio', 'Location', 'Financial_Account']
    df_with_totals = df_with_totals[columns_order]

    # Define file paths
    outputpath = "BASE_PATH/reports"
    output_filepath = f"{outputpath}/{portfolio}_{fname}"

    # Save the DataFrame to an Excel file
    utilities.save_report(df_with_totals, f"{fname}", portfolio, outputpath)

    # Apply formatting to the saved file
   # apply_standard_formatting(output_filepath)


def journals_style_reports(journal_entries, sdate, edate, fname):
    import pandas as pd

    # Create the DataFrame from the list of entries
    df = pd.DataFrame([[
        je.portfolio,
        je.ibor_date,
        je.investment,
        je.lotid,
        je.tax_date,
        je.ls,
        je.location,
        je.financial_account,
        je.quantity,
        je.local,
        je.book,

    ] for je in journal_entries], columns=[
        "Portfolio", "IBOR_Date", "Investment", "Lot_ID", "Tax_Date", "LS", "Location", "Financial_Account", "Quantity", "Local",
        "Book"
    ])
    df = filter_records(df, 'Financial_Account', ['MktVal', 'MktValRE'," UnrealGLAsset", "UnrealGLRE"], operator="OR", mode="exclude")

    total_columns = ['Book']  # Replace with actual column names
    group_by_columns = ['Investment', 'Tax_Date']  # Replace
    df.sort_values(group_by_columns, inplace=True)
    #Must calc grand totals and save
    grand_totals = calculate_grand_totals(df, total_columns)

    # Fetch and map groupings to the DataFrame created from journal entries
    df_merged = fetch_and_map_groupings(df, "C:/Users/hjmne/PycharmProjects/chest/refdata/investment_master.csv",
                            "C:/Users/hjmne/PycharmProjects/chest/refdata/chart_of_accounts.csv")

#    filter_records(df, 'SummaryReport', ['Skip'], operator="OR", mode="exclude")
    import pandas as pd


    df_with_subtotals = insert_subtotals(df_merged, group_by_columns, total_columns)

    # Append the Grand Totals row to the DataFrame with subtotals
    df_final_report = pd.concat([df_with_subtotals, grand_totals], ignore_index=True)


    # Save the DataFrame with subtotals and Grand Totals to an Excel file
    fnamechoice = "C:/Users/hjmne/PycharmProjects/chest/reports/JournalsByInvestmentFinancialAccount"+fname+".xlsx"
    df_final_report.to_excel(fnamechoice, index=False)



    report_title = "Tax Lot Appraisal" +fname  # Replace with your report title
    # Assuming 'Portfolio' is the column name where portfolio names are stored.
    # This filters the DataFrame to exclude any rows where 'Portfolio' is 'Subtotal' or 'Grand Totals'
    # and then gets the unique values in the 'Portfolio' column.
    unique_portfolios = df_final_report[
        ~df_final_report['Portfolio'].isin(['Subtotal', 'Grand Totals'])
    ]['Portfolio'].unique()
    # Check if there's at least one portfolio name and use the first one; otherwise, use a default.
    portfolio_name = unique_portfolios[0] if len(unique_portfolios) > 0 else "Unknown Portfolio"
    sdate1 = sdate.strftime('%Y-%m-%d')
    edate1 = edate.strftime('%Y-%m-%d')
    period_start_date = sdate1  # Assuming 'date' is a datetime object
    period_end_date = edate1  # Assuming 'date' is a datetime object
    # Combine title and date range
    full_title = f"{report_title} for {portfolio_name} ({period_start_date} - {period_end_date})"
    add_title(fnamechoice, "Sheet1",full_title)


    # Now apply standard formatting to the saved Excel file
    apply_standard_formatting(fnamechoice, "Sheet1")

def prior_period_adjustments(journal_entries, bookkeeping_space, sdate, edate, fname):
    import pandas as pd

    # Create the DataFrame from the list of entries
    df = pd.DataFrame([[
        je.portfolio,
        je.ibor_date,
        je.transaction,
        je.investment,
        je.lotid,
        je.tax_date,
        je.ls,
        je.location,
        je.financial_account,
        je.quantity,
        je.local,
        je.book,
        je.tranid,

    ] for je in journal_entries], columns=[
        "Portfolio", "IBOR Date", "Transaction", "Investment", "Lot ID", "Tax Date", "LS", "Location", "Financial Account", "Quantity", "Local",
        "Book", "Tran ID"
    ])
#    df = filter_records(df, 'Financial Account', ['MktVal', 'MktValRE'], operator="OR", mode="exclude")
    #Must calc grand totals and save



    # Fetch and map groupings to the DataFrame created from journal entries
    df_merged = fetch_and_map_groupings(df, "C:/Users/hjmne/PycharmProjects/chest/refdata/investment_master.csv",
                            "C:/Users/hjmne/PycharmProjects/chest/refdata/chart_of_accounts.csv")


    total_columns = ['Book']  # Replace with actual column names
    group_by_columns = ['IBOR Date', 'Transaction', 'BS_Group']  # Replace
    df_merged.sort_values(group_by_columns, inplace=True)

    df_merged = filter_records(df_merged, 'Transaction', ['PriorPeriodAdjustment'], operator="OR", mode="include")

    # ADD THIS FUTURE DIRECTION # Assuming 'events.xlsx' is your Excel file containing events data
    # events_df = pd.read_excel('events.xlsx')
    #
    # # Now you have 'events_df' as your DataFrame containing events data
    #
    # # Then you can proceed to merge it with your 'df_merged' DataFrame as shown in the previous response:
    # merged_df = pd.merge(df_merged, events_df, left_on='Tran ID', right_on='tranid', how='inner')

    import pandas as pd
    # Always do grand_totals before sub-totaling
    grand_totals = calculate_grand_totals(df_merged, total_columns)


    df_with_subtotals = insert_subtotals(df_merged, group_by_columns, total_columns)

    # Append the Grand Totals row to the DataFrame with subtotals
    df_final_report = pd.concat([df_with_subtotals, grand_totals], ignore_index=True)


    # Save the DataFrame with subtotals and Grand Totals to an Excel file
    fnamechoice = ("C:/Users/hjmne/PycharmProjects/chest/reports/PriorPeriodAdjustments"+fname+".xlsx")
    df_final_report.to_excel(fnamechoice, index=False)

    report_title = "Prior Period Adjustments" + fname  # Replace with your report title
    # Assuming 'Portfolio' is the column name where portfolio names are stored.
    # This filters the DataFrame to exclude any rows where 'Portfolio' is 'Subtotal' or 'Grand Totals'
    # and then gets the unique values in the 'Portfolio' column.
    unique_portfolios = df_final_report[
        ~df_final_report['Portfolio'].isin(['Subtotal', 'Grand Totals'])
    ]['Portfolio'].unique()
    # Check if there's at least one portfolio name and use the first one; otherwise, use a default.
    portfolio_name = unique_portfolios[0] if len(unique_portfolios) > 0 else "Unknown Portfolio"
    sdate1 = sdate.strftime('%Y-%m-%d')
    edate1 = edate.strftime('%Y-%m-%d')
    period_start_date = sdate1  # Assuming 'date' is a datetime object
    period_end_date = edate1  # Assuming 'date' is a datetime object
    # Combine title and date range
    full_title = f"{report_title} for {portfolio_name} ({period_start_date} - {period_end_date})"
    add_title(fnamechoice, "Sheet1", full_title)

    # Now apply standard formatting to the saved Excel file
    apply_standard_formatting(fnamechoice, "Sheet1")

def realized_gains_losses(journal_entries, bookkeeping_space, sdate, edate, fname, portfolio):
    import pandas as pd
    # Update running balances first
    # update_running_balances(journal_entries, ['portfolio', 'investment', 'ls', 'location', 'financial_account'])

    # Create the DataFrame from the list of entries
    df = pd.DataFrame([[
        je.portfolio,
        je.ibor_date,
        je.investment,
        je.lotid,
        je.tax_date,
        je.ls,
        je.location,
        je.financial_account,
        je.quantity,
        je.local,
        je.book,
        je.tranid,
        je.transaction,
     ] for je in journal_entries], columns=[
        "Portfolio", "IBOR Date", "Investment", "Lot ID", "Tax Date", "LS", "Location", "Financial Account", "Quantity", "Local",
        "Book", "Tran ID", "Transaction"
    ])

    import pandas as pd

    # Assuming filtered_df is your DataFrame
    # Example: filtered_df = pd.DataFrame({'Cost Local': [...], 'PriceGainLocal': [...], ...})




    closing_transactions = ['SellLong', 'CoverShort']  # List of closing transaction types
    df_filtered = df[df['Transaction'].isin(closing_transactions)]

    # Replace 'Quantity', 'Local', and 'Book' with the actual column names you need to sum up
    numeric_columns = ['Quantity', 'Local', 'Book']  # Update this with the actual columns you want to aggregate

    # Specify the financial accounts to be included in the pivot table
    included_accounts = ['Cost', 'PriceGainInvestment', 'FXGainInvestment']  # Update as needed

    # Filter the DataFrame for these accounts
    df_filtered = df_filtered[df_filtered['Financial Account'].isin(included_accounts)]
 #   df_filtered['Tax Date'] = pd.to_datetime(df['Tax Date'], errors='coerce')

    # 1) Change all signs on the numbers (for specific columns, specify them)
    df_filtered = df_filtered.apply(lambda x: -x if x.dtype == 'float64' else x)

    # Convert to numeric and fill NaN values if necessary
    numeric_cols = ['Cost Local', 'PriceGainLocal', 'Cost Book', 'PriceGainBook', 'FxGain']


    # Create the pivot table
    pivot_df = df_filtered.pivot_table(
        index=['Portfolio',  'IBOR Date', 'Investment', 'Lot ID','Tax Date','LS', 'Tran ID'],  # Index columns
        columns='Financial Account',  # Columns to pivot
        values=['Book', 'Local', 'Quantity'], # Values to aggregate
        aggfunc='sum',  # Aggregation function
        fill_value=0  # Fill value for missing data
    )


    # Flatten the MultiIndex in columns
    pivot_df.columns = [' '.join(col).strip() for col in pivot_df.columns.values]
    pivot_df = pivot_df.drop(['Local FXGainInvestment', 'Quantity FXGainInvestment', 'Quantity PriceGainInvestment'],
                             axis=1)

    # Reset index to turn the indices into columns again
    pivot_df.reset_index(inplace=True)

    # This order is essential and it is hard to assess why
    pivot_df.columns = [
        'Portfolio', 'IBOR Date', 'Investment', 'Lot ID', 'Tax Date',  'LS', 'Tran ID',
        'Book Cost', 'FX Gain Book', 'Price Gain Book', 'Local Cost',
        'Price Gain Local', 'Quantity'
        # Add other new column names as necessary
    ]
    pivot_df.reset_index(inplace=True)
    # Operate on the pivot_df DataFrame
#    pivot_df.sort_values(group_by_columns, inplace=True)
#    pivot_df[numeric_cols] = pivot_df[numeric_cols].apply(pd.to_numeric, errors='coerce').fillna(0)

    pivot_df['Market Value Local'] = pivot_df['Local Cost'] + pivot_df['Price Gain Local']
    pivot_df['Market Value Book'] = pivot_df['Book Cost'] + pivot_df['Price Gain Book'] + pivot_df['FX Gain Book']
    pivot_df.reset_index(inplace=True)
    #
    group_by_columns = ['IBOR Date', 'Tran ID']
    # Example column order
    new_order = ['Portfolio', 'IBOR Date', 'Tax Date', 'Lot ID', 'Investment',  'LS', 'Tran ID',
                'Quantity', 'Local Cost', 'Market Value Local', 'Price Gain Local', 'Book Cost', 'Market Value Book',
                'Price Gain Book', 'FX Gain Book']
    # Create new columns
 
    # Output the DataFrame to an Excel file to check
    pivot_df.to_excel("C:/Users/hjmne/PycharmProjects/chest/reports/output.xlsx", index=False)  # Change the path as needed
    # Calculate grand totals on the pivot_df DataFrame

    # Replace these with the actual names of the columns that contain your data
    total_columns = ['Book Cost',  'Market Value Book', 'Price Gain Book', 'FX Gain Book']

    # Calculating grand totals for specific columns
    grand_totals = calculate_grand_totals(pivot_df, total_columns)

    # Assuming 'group_by_columns' is defined and contains the columns to group by
    # Assuming 'subtotal_columns' contains all the columns for which you want subtotals
    subtotal_columns = ['Quantity', 'Local Cost', 'Book Cost', 'Price Gain Local', 'Price Gain Book', 'FX Gain Book',
                        'Market Value Local', 'Market Value Book']  # Add all required columns

    # Group by and calculate subtotals
    subtotals = pivot_df.groupby(group_by_columns)[subtotal_columns].sum().reset_index()
    subtotals['Type'] = 'Subtotal'

    # Empty DataFrame to hold data with subtotals
    df_with_subtotals = pd.DataFrame(columns=pivot_df.columns)

    for _, group_data in pivot_df.groupby(group_by_columns):
        # Insert the group data
        df_with_subtotals = pd.concat([df_with_subtotals, group_data], ignore_index=True)

        # Calculate and insert subtotal for the group
        subtotal_row = group_data[subtotal_columns].sum().to_frame().T
        # Add additional columns to match the structure of df_with_subtotals
        for col in df_with_subtotals.columns:
            if col not in subtotal_columns:
                subtotal_row[col] = 'Subtotal' if col == 'Type' else None
        df_with_subtotals = pd.concat([df_with_subtotals, subtotal_row], ignore_index=True)

    # Insert subtotals into the pivot_df DataFrame
    # df_with_subtotals = insert_subtotals(pivot_df, group_by_columns, total_columns)

    # Reassign the DataFrame columns to this new order
    df_with_subtotals = df_with_subtotals[new_order]
    df_with_subtotals.reset_index(inplace=True)

    grand_totals['Type'] = 'Grand Total'
    df_final_report = pd.concat([df_with_subtotals, grand_totals], ignore_index=True)

    if new_order:
        df_final_report = df_final_report[new_order]

    df_final_report.reset_index(inplace=True, drop=True)

    # Save the DataFrame with subtotals and Grand Totals to an Excel file
    fnamechoice = "C:/Users/hjmne/PycharmProjects/chest/reports/RealizedGainsLosses"+portfolio+".xlsx"
    if new_order:
        df_final_report = df_final_report[new_order]

    df_final_report.reset_index(inplace=True, drop=True)
    df_final_report.to_excel(fnamechoice, index=False)

    report_title = "Realized Gains/Losses " # Replace with your report title
    # Assuming 'Portfolio' is the column name where portfolio names are stored.
    # This filters the DataFrame to exclude any rows where 'Portfolio' is 'Subtotal' or 'Grand Totals'
    # and then gets the unique values in the 'Portfolio' column.
    unique_portfolios = df_final_report[
        ~df_final_report['Portfolio'].isin(['Subtotal', 'Grand Totals'])
    ]['Portfolio'].unique()
    # Check if there's at least one portfolio name and use the first one; otherwise, use a default.
    portfolio_name = unique_portfolios[0] if len(unique_portfolios) > 0 else "Unknown Portfolio"
    sdate1 = sdate.strftime('%Y-%m-%d')
    edate1 = edate.strftime('%Y-%m-%d')
    period_start_date = sdate1  # Assuming 'date' is a datetime object
    period_end_date = edate1  # Assuming 'date' is a datetime object
    # Combine title and date range
    full_title = f"{report_title} for {portfolio_name} ({period_start_date} - {period_end_date})"
    add_title(fnamechoice, "Sheet1",full_title)


    # Now apply standard formatting to the saved Excel file
    apply_standard_formatting(fnamechoice, "Sheet1")


def total_income_earned(journal_entries, bookkeeping_space, sdate, edate, fname):
    import pandas as pd
    # Update running balances first
    # update_running_balances(journal_entries, ['portfolio', 'investment', 'ls', 'location', 'financial_account'])
    coa_master = pd.read_csv("C:/Users/hjmne/PycharmProjects/chest/refdata/chart_of_accounts.csv")
    investment_master = pd.read_csv("C:/Users/hjmne/PycharmProjects/chest/refdata/investment_master.csv")

    for entry in journal_entries:
        if not hasattr(entry, 'financial_account'):
            print("Missing financial_account:", entry)

    from collections import namedtuple

    # Define a named tuple for Journals

    import pandas as pd

    # Convert namedtuples to dict within DataFrame
    def convert_to_dict(entry):
        try:
            return entry._asdict()
        except AttributeError:
            return entry  # or modify as per your requirements

    import pandas as pd
    from collections import namedtuple

    Journals = namedtuple('Journals',
                          ['portfolio', 'investment', 'ibor_date', 'ls', 'tranid', 'quantity', 'local', 'book',
                           'location', 'financial_account'])

    def normalize_journal_entries(journal_entries):
        data = [entry.to_dict() for entry in journal_entries]
        return pd.DataFrame(data)

    df = normalize_journal_entries(journal_entries)


    df = filter_records(df, 'financial_account', ['MktVal', 'MktValRE'], operator="OR", mode="exclude")
    import pandas as pd

    # Assuming df is your DataFrame
    # And assuming sdate and ibor_date are columns in your DataFrame

    df['tax_date'] = df.apply(lambda row: row['settledate'] if row['transaction'] == "Settlement" else row['ibor_date'],
                              axis=1)

    je_data = pd.merge(df, coa_master[['SystemName', 'SummaryReport']], left_on='financial_account', right_on='SystemName',
                       how='left').drop(
        'SystemName', axis=1)

    je_data['tax date'] = je_data['tax date'].astype(str)
    je_data = replace_column_name(je_data, 'tax date', 'tax date/lot num')
    df_filtered = filter_records(je_data, 'SummaryReport', ['RealizedIncome','Income'], operator="OR", mode="include")

    # Replace 'Quantity', 'Local', and 'Book' with the actual column names you need to sum up
    numeric_columns = ['quantity', 'local', 'book']  # Update this with the actual columns you want to aggregate

    # 1) Change all signs on the numbers (for specific columns, specify them)
    df_filtered = df_filtered.apply(lambda x: -x if x.dtype == 'float64' else x)

    group_by_columns = ['financial_account']
    # Create new columns

    # Output the DataFrame to an Excel file to check
    df_filtered.to_excel("C:/Users/hjmne/PycharmProjects/chest/reports/output.xlsx", index=False)  # Change the path as needed
    # Calculate grand totals on the pivot_df DataFrame

    # Replace these with the actual names of the columns that contain your data
    total_columns = ['book']

    # Calculating grand totals for specific columns
    grand_totals = calculate_grand_totals(df_filtered, total_columns)

    # Assuming 'group_by_columns' is defined and contains the columns to group by
    # Assuming 'subtotal_columns' contains all the columns for which you want subtotals
    subtotal_columns = ['book']  # Add all required columns

    # Group by and calculate subtotals
    subtotals = df_filtered.groupby(group_by_columns)[subtotal_columns].sum().reset_index()
    subtotals['Type'] = 'Subtotal'

    # Empty DataFrame to hold data with subtotals
    df_with_subtotals = pd.DataFrame(columns=df_filtered.columns)

    for _, group_data in df_filtered.groupby(group_by_columns):
        # Insert the group data
        df_with_subtotals = pd.concat([df_with_subtotals, group_data], ignore_index=True)

        # Calculate and insert subtotal for the group
        subtotal_row = group_data[subtotal_columns].sum().to_frame().T
        # Add additional columns to match the structure of df_with_subtotals
        for col in df_with_subtotals.columns:
            if col not in subtotal_columns:
                subtotal_row[col] = 'Subtotal' if col == 'Type' else None
        df_with_subtotals = pd.concat([df_with_subtotals, subtotal_row], ignore_index=True)

    # Insert subtotals into the pivot_df DataFrame
    # df_with_subtotals = insert_subtotals(pivot_df, group_by_columns, total_columns)

    grand_totals['Type'] = 'Grand Total'
    df_final_report = pd.concat([df_with_subtotals, grand_totals], ignore_index=True)

    df_final_report.reset_index(inplace=True, drop=True)

    # Save the DataFrame with subtotals and Grand Totals to an Excel file
    fnamechoice = "C:/Users/hjmne/PycharmProjects/chest/reports/PeriodIncomeEarned"+fname+".xlsx"


    df_final_report.reset_index(inplace=True, drop=True)
    df_final_report.to_excel(fnamechoice, index=False)


    report_title = "Period Income Earned"+ fname  # Replace with your report title
    # Assuming 'Portfolio' is the column name where portfolio names are stored.
    # This filters the DataFrame to exclude any rows where 'Portfolio' is 'Subtotal' or 'Grand Totals'
    # and then gets the unique values in the 'Portfolio' column.
    unique_portfolios = df_final_report[
        ~df_final_report['portfolio'].isin(['Subtotal', 'Grand Totals'])
    ]['portfolio'].unique()
    # Check if there's at least one portfolio name and use the first one; otherwise, use a default.
    portfolio_name = unique_portfolios[0] if len(unique_portfolios) > 0 else "Unknown Portfolio"
    sdate1 = sdate.strftime('%Y-%m-%d')
    edate1 = edate.strftime('%Y-%m-%d')
    period_start_date = sdate1  # Assuming 'date' is a datetime object
    period_end_date = edate1  # Assuming 'date' is a datetime object
    # Combine title and date range
    full_title = f"{report_title} for {portfolio_name} ({period_start_date} - {period_end_date})"
    add_title(fnamechoice, "Sheet1", full_title)

    # Now apply standard formatting to the saved Excel file
    apply_standard_formatting(fnamechoice, "Sheet1")


import pandas as pd


import pandas as pd

# def journals_by_tranid(journal_entries, bookkeeping_space, sdate, edate, fname, specific_investment):
#     # Assume update_running_balances and other required functions are defined elsewhere
#
#     # Create DataFrame from journal entries
#     df = pd.DataFrame([[
#         je.portfolio, je.ibor_date, je.transaction, je.investment, je.tax_date, je.ls, je.location,
#         je.financial_account, je.quantity, je.local, je.book, je.tranid, *je.running_balances
#     ] for je in journal_entries], columns=[
#         "Portfolio", "IBOR Date", "Transaction", "Investment", "Tax Date", "LS", "Location",
#         "Financial Account", "Quantity", "Local", "Book", "Tran ID", "RunningQty", "RunningLocal", "RunningBook"
#     ])
#
#     # Filter directly for specific investment and associated transaction IDs to optimize performance
#     df_specific_investment = df[df['Investment'] == specific_investment]
#     related_tranids = df_specific_investment['Tran ID'].unique()
#     df = df[df['Tran ID'].isin(related_tranids)]
#
#     # Fetch and map groupings to the DataFrame created from journal entries
#     df_merged = fetch_and_map_groupings(df, "C:/Users/hjmne/PycharmProjects/chest/refdata/investment_master.csv",
#                                 "C:/Users/hjmne/PycharmProjects/chest/refdata/chart_of_accounts.csv")
#
#     grand_totals = calculate_grand_totals(df_merged, ['Book'])
#     df_with_subtotals = insert_subtotals(df_merged, ['IBOR Date', 'Tran ID'], ['Book'])
#     df_final_report = pd.concat([df_with_subtotals, grand_totals], ignore_index=True)
#
#     # Save the DataFrame to an Excel file
#     fnamechoice = f"reports/JournalsBalancedByTranID{fname}.csv"
#     df_final_report.to_csv(fnamechoice, index=False)
#     import reportlab
#     # import pandas as pd
#     # from reportlab.lib.pagesizes import letter
#     # from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
#     # from reportlab.lib import colors
#     #
#     # def generate_pdf_report(df, filename):
#     #     # Create a PDF document with the specified filename
#     #     pdf = SimpleDocTemplate(filename, pagesize=letter)
#     #     elements = []
#     #
#     #     # Ensure DataFrame columns are strings if they're not already
#     #     df = df.astype(str)
#     #
#     #     # Convert DataFrame to a list of lists (including header)
#     #     data = [df.columns.tolist()] + df.values.tolist()
#     #
#     #     # Create a table to add to the elements list
#     #     table = Table(data)
#     #
#     #     # Add some style to the table
#     #     table.setStyle(TableStyle([
#     #         ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
#     #         ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
#     #         ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
#     #         ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
#     #         ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
#     #         ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
#     #         ('GRID', (0, 0), (-1, -1), 1, colors.black)
#     #     ]))
#     #
#     #     elements.append(table)
#     #
#     #     # Build the PDF
#     #     pdf.build(elements)
#     #
#     # # Example usage
#     # df_final_report = pd.DataFrame({
#     #     # Assuming df_final_report is already created and populated
#     # })
#     #
#     # filename = "Path_to_save_your_PDF_report/report.pdf"
#     # generate_pdf_report(df_final_report, filename)
#     #
#     # # Add title and apply formatting
#     # report_title = f"Journals By TransactionID and Date {fname}"
#     # unique_portfolios = df_final_report[~df_final_report['Portfolio'].isin(['Subtotal', 'Grand Totals'])]['Portfolio'].unique()
#     # portfolio_name = unique_portfolios[0] if len(unique_portfolios) > 0 else "Unknown Portfolio"
#     # period_start_date, period_end_date = sdate.strftime('%Y-%m-%d'), edate.strftime('%Y-%m-%d')
#     # full_title = f"{report_title} for {portfolio_name} ({period_start_date} - {period_end_date})"
#     # add_title(fnamechoice, "Sheet1", full_title)
#  #   apply_standard_formatting(fnamechoice, "Sheet1")
#
#     return df_final_report  # Optionally return the final DataFrame for further use or inspection
# def journals_by_tranid(journal_entries, bookkeeping_space, sdate, edate, fname, specific_investment):
#     # Update running balances first
#     update_running_balances(journal_entries, ['ibor_date', 'tranid', 'portfolio'])
#
#     # Create the DataFrame from the list of entries
#     df = pd.DataFrame([[
#         je.portfolio,
#         je.ibor_date,
#         je.transaction,
#         je.investment,
#         je.tax_date,
#         je.ls,
#         je.location,
#         je.financial_account,
#         je.quantity,
#         je.local,
#         je.book,
#         je.tranid,
#         je.running_balances[0],
#         je.running_balances[1],
#         je.running_balances[2]
#     ] for je in journal_entries], columns=[
#         "Portfolio", "IBOR Date", "Transaction", "Investment", "Tax Date", "LS", "Location", "Financial Account",
#         "Quantity", "Local",
#         "Book", "Tran ID", "RunningQty", "RunningLocal", "RunningBook"
#     ])
#
#     # Filter for the specific investment
#     df_filtered_for_investment = df[df['Investment'] == specific_investment]
#
#     # Find related transaction IDs
#     related_tranids = df_filtered_for_investment['Tran ID'].unique()
#
#     # Filter the original dataset by these transaction IDs to include all related journal entries
#     df_filtered_by_tranids = df[df['Tran ID'].isin(related_tranids)]
#
#     # Fetch and map groupings to the DataFrame created from journal entries
#     df_merged = fetch_and_map_groupings(df, "C:/Users/hjmne/PycharmProjects/chest/refdata/investment_master.csv",
#                                 "C:/Users/hjmne/PycharmProjects/chest/refdata/chart_of_accounts.csv")
#
#     total_columns = ['Book']  # Specify actual column names for totals
#     group_by_columns = ['IBOR Date', 'Tran ID']  # Specify columns to group by
#
#     df_merged.sort_values(group_by_columns, inplace=True)
#
#     grand_totals = calculate_grand_totals(df_merged, total_columns)
#     # df_with_subtotals = insert_subtotals(df_merged, group_by_columns, total_columns)
#
#     # Append the Grand Totals row to the DataFrame with subtotals
#     df_final_report = pd.concat([df_merged, grand_totals], ignore_index=True)
#
#     # Save the DataFrame with subtotals and Grand Totals to an Excel file
#     fnamechoice = f"C:/Users/hjmne/PycharmProjects/chest/reports/JournalsBalancedByTranID{fname}.xlsx"
#     df_final_report.to_excel(fnamechoice, index=False)
#
#     report_title = f"Journals By TransactionID and Date for {specific_investment}"
#     unique_portfolios = df_final_report[~df_final_report['Portfolio'].isin(['Subtotal', 'Grand Totals'])][
#         'Portfolio'].unique()
#     portfolio_name = unique_portfolios[0] if len(unique_portfolios) > 0 else "Unknown Portfolio"
#     period_start_date = sdate.strftime('%Y-%m-%d')
#     period_end_date = edate.strftime('%Y-%m-%d')
#     full_title = f"{report_title} ({period_start_date} - {period_end_date})"
#     add_title(fnamechoice, "Sheet1", full_title)
#
#     # Apply standard formatting to the saved Excel file
#     apply_standard_formatting(fnamechoice, "Sheet1")
#

def generate_category_reports(journal_entries, sdate, edate, save_path="BASE_PATH/reports"):
    """
    Generate reports for predefined categories within a given date range.

    Parameters:
    - journal_entries (list of objects): List of journal entry records with attributes.
    - sdate (str): Start date in 'YYYY-MM-DD' format.
    - edate (str): End date in 'YYYY-MM-DD' format.
    - save_path (str): Path to save the reports.

    Returns:
    - dict: A dictionary of DataFrames for each category.
    """
    # Ensure the save_path directory exists
    os.makedirs(save_path, exist_ok=True)

    # Define the mapping for the five categories
    categories = {
        "Cost Basis": ["Cost", "Receivable", "Payable", "AccruedInterestReceivable", "AccruedInterestPayable",
                       "SoldAccruedReceivable", "PurchasedAccruedPayable", "PurchasedInterest", "SoldInterest"],
        "Unrealized": ["UnrealPriceGL", "UnrealFXGL"],
        "Realized": ["PriceGainStatOffset", "FXGainStatOffset", "PriceGainInvestment", "FXGainInvestment"],
        "Income": ["DividendReceipt", "DividendExpense", "UnearnedIncome", "InterestIncome", "InterestReceipt",
                   "InterestExpense", 'FXGainTradeSettle', 'FXGainCurrency', 'OptionIncome'],
        "Capital": ["ContributedCost"]
    }

#     'Cost': (0, 1, 2),
#     'UnrealPriceGL': (0, 3, 4),
#     'UnrealFXGL': (0, 'skip', 5),
#     'PriceGainStatOffset': (0, 3, 4),
#     'FXGainStatOffset': (0, 'skip', 5),
#     'PriceGainInvestment': (0, 6, 7),
#     'FXGainInvestment': (0, 6, 7),
#     'DividendReceipt': (0, 8, 9),
#     'DividendExpense': (0, 8, 9),
#     'UnearnedIncome': (0, 8, 9),
#     'ContributedCost': (10, 11, 12),
#     'Receivable': (0, 1, 2),
#     'Payable': (0, 1, 2),
#     'AccruedInterestReceivable': (0, 1, 2),
#     'AccruedInterestPayable': (0, 1, 2),
#     'SoldAccruedReceivable': (0, 1, 2),
#     'PurchasedAccruedPayable': (0, 1, 2),
#     'PurchasedInterest': (0, 1, 2),
#     'PurchasedInterestExpense': (0, 8, 9),
#     'SoldInterest': (0, 1, 2),
#     'SoldInterestIncome': (0, 8, 9),
#     'DividendsReceivable': (0, 1, 2),
#     'DividendsPayable': (0, 1, 2),
#     'FXGainTradeSettle': (0, 8, 9),
#     'FXGainCurrency': (0, 8, 9),
#     'OptionIncome': (0, 8, 9),
#     'InterestIncome': (0, 8, 9),
#     'InterestReceipt': (0, 8, 9),
#     'InterestExpense': (0, 8, 9),
#     'SpotFxReceivable': (0, 1, 2),
#     'SpotFxPayable': (0, 1, 2)
#
# }


# Define input columns and output columns
    input_columns = [
        'portfolio', 'transaction', 'investment', 'ibor_date', 'tradedate',
        'settledate', 'lotid', 'tranid', 'tax_date', 'ls', 'location',
        'financial_account', 'quantity', 'local', 'book', 'notional', 'oface'
    ]
    output_columns = [
        'ibor_date', 'portfolio', 'investment', 'transaction', 'financial_account',
        'quantity', 'local', 'book', 'notional', 'oface', 'tranid', 'lotid',
        'tax_date', 'tradedate', 'settledate', 'ls', 'location'
    ]

    # Convert journal entries to DataFrame using object attributes
    data = [[getattr(je, col) for col in input_columns] for je in journal_entries]
    df = pd.DataFrame(data, columns=input_columns)

    # Reorder columns based on output_columns
    df = df[output_columns]

    # Convert ibor_date to datetime for filtering
    df['ibor_date'] = pd.to_datetime(df['ibor_date'])

    # Filter by start and end date
    sdate = pd.to_datetime(sdate)
    edate = pd.to_datetime(edate)
    df = df[(df['ibor_date'] >= sdate) & (df['ibor_date'] <= edate)]

    # Exclude records where quantity, local, and book are all zero or NaN
    df = df[~((df[['quantity', 'local', 'book']].fillna(0) == 0).all(axis=1))]

    # Generate and save reports for each category
    reports = {}
    for category, accounts in categories.items():
        filtered_df = df[df['financial_account'].isin(accounts)]
        filtered_df.sort_values(by=['ibor_date', 'tranid'], inplace=True)

        # Add totals row
        total_row = filtered_df[['quantity', 'local', 'book']].sum().to_frame().T
        total_row['portfolio'] = 'Total'
        filtered_df = pd.concat([filtered_df, total_row], ignore_index=True)

        # Save report to the specified directory
        filename = os.path.join(save_path, f"{category.replace(' ', '_')}_Report.xlsx")
        filtered_df.to_excel(filename, index=False, engine='openpyxl')

        reports[category] = filtered_df

    return reports

import pandas as pd
import os

import pandas as pd
import os

def prepare_data(journal_entries, investment_master_path, sdate, edate):
    """
    Prepare data by joining journal entries with the investment master file.

    Parameters:
    - journal_entries (list of objects): List of journal entry records with attributes.
    - investment_master_path (str): Path to the investment master file (CSV or Excel).
    - sdate (str): Start date in 'YYYY-MM-DD' format.
    - edate (str): End date in 'YYYY-MM-DD' format.

    Returns:
    - DataFrame: Merged DataFrame with journal entries and investment metadata.
    """
    # Convert journal entries to DataFrame
    input_columns = [
        'portfolio', 'transaction', 'investment', 'ibor_date', 'tradedate',
        'settledate', 'lotid', 'tranid', 'tax_date', 'ls', 'location',
        'financial_account', 'quantity', 'local', 'book', 'notional', 'oface'
    ]
    data = [[getattr(je, col) for col in input_columns] for je in journal_entries]
    df = pd.DataFrame(data, columns=input_columns)

    # Filter by date range
    df['ibor_date'] = pd.to_datetime(df['ibor_date'])
    sdate = pd.to_datetime(sdate)
    edate = pd.to_datetime(edate)
    df = df[(df['ibor_date'] >= sdate) & (df['ibor_date'] <= edate)]


    # Load investment master and join
    if investment_master_path.endswith('.csv'):
        investment_master = pd.read_csv(investment_master_path)
    else:
        investment_master = pd.read_excel(investment_master_path)

    # Merge investment metadata
    df = df.merge(investment_master[['Investment', 'Asset_Class']],
                  left_on='investment', right_on='Investment', how='left')

    return df
import os
import pandas as pd

def generate_nav_reports(df, save_path):
    """
    Generate NAV reports with sorting for each view.

    Parameters:
    - df (DataFrame): Prepared data with journal entries and investment metadata.
    - save_path (str): Directory to save the reports.

    Returns:
    - dict: A dictionary of DataFrames for each report.
    """
    os.makedirs(save_path, exist_ok=True)

    # Define the mapping for categories
    categories = {
        "Cost Basis": ["Cost", "Receivable", "Payable", "AccruedInterestReceivable", "AccruedInterestPayable",
                       "InterestReceivable", "InterestPayable", "DividendsPayable", "DividendsReceivable",
                       "SpotFXPayable", "SpotFXReceivable", "ForwardFXPayable", "ForwardFXReceivable",
                       "SoldAccruedReceivable", "PurchasedAccruedPayable", "PurchasedInterest", "SoldInterest"],
        "Unrealized": ["UnrealPriceGL", "UnrealFXGL"],
        "Realized": ["PriceGainStatOffset", "FXGainStatOffset", "PriceGainInvestment", "FXGainInvestment"],
        "Income": ["DividendReceipt", "DividendExpense", "UnearnedIncome", "InterestIncome", "InterestReceipt",
                   "InterestExpense", 'FXGainTradeSettle', 'FXGainCurrency', 'OptionIncome'],
        "Capital": ["ContributedCost"],
        "OffBalanceSheet": ["MarketVal"]
    }

    # Add Category column based on financial_account
    def map_category(account):
        for category, accounts in categories.items():
            if account in accounts:
                return category
        return "Unknown"



    df['Category'] = df['financial_account'].apply(map_category)
    portfolio = df.iloc[1, 0]

    # View 1: Summary by Category
    summary_df = df.groupby('Category')['book'].sum().reset_index()
    summary_df = summary_df.sort_values(by='Category')  # Sort by Category
    utilities.save_report(summary_df, "NAVSummary.xlsx", portfolio, save_path)
  #  summary_filename = os.path.join(save_path, "NAVSummary.xlsx")
  #  summary_df.to_excel(summary_filename, index=False, engine='openpyxl')

    # View 2: Breakdown by Category and Asset Type
    category_asset_type_df = df.groupby(['Category', 'Asset_Class'])['book'].sum().reset_index()
    category_asset_type_df = category_asset_type_df.sort_values(by=['Category', 'Asset_Class']) # Sort by Category and Asset_Class
    utilities.save_report(category_asset_type_df, "NAVAssetType.xlsx", portfolio, save_path)
    # category_asset_type_filename = os.path.join(save_path, "NAVAssetType.xlsx")
    # category_asset_type_df.to_excel(category_asset_type_filename, index=False, engine='openpyxl')

    # View 3: Detailed View sorted by Category, Asset_Class, Investment, and IBOR_Date
    view_3_columns = [
        'Category', 'Asset_Class', 'investment', 'tranid', 'ibor_date', 'quantity', 'local', 'book', 'notional',
        'financial_account', 'tradedate', 'settledate', 'tax_date', 'location'
    ]
    # Ensure all specified columns exist in the DataFrame before reordering
    existing_columns = [col for col in view_3_columns if col in df.columns]
    category_asset_investment_je_df = df[existing_columns]
    category_asset_investment_je_df = category_asset_investment_je_df.sort_values(
        by=['Category', 'Asset_Class', 'investment', 'ibor_date']  # Sort by specified hierarchy
    )

    # Save View 3
    utilities.save_report(category_asset_investment_je_df, "NAVDetail.xlsx", portfolio, save_path)

    return {
        "Summary": summary_df,
        "By Category and Asset Type": category_asset_type_df,
        "By Category, Asset Type, and Investment (Detailed)": category_asset_investment_je_df
    }

def generate_pivot_report(df, save_path):
    """
    Generate a pivot table for the third NAV view and save it as a report.

    Parameters:
    - df (DataFrame): Prepared data with journal entries and investment metadata.
    - save_path (str): Directory to save the pivot report.

    Returns:
    - DataFrame: The pivot table DataFrame.
    """
    pivot_df = df.pivot_table(
        values='book',
        index=['financial_account', 'Asset_Class', 'investment'],
        aggfunc='sum'
    ).reset_index()

    pivot_filename = os.path.join(save_path, "Pivot_Category_Asset_Investment.xlsx")
    pivot_df.to_excel(pivot_filename, index=False, engine='openpyxl')

    return pivot_df

def generate_all_nav_reports(df, save_path):
    generate_nav_reports(df, save_path)
   # generate_pivot_report(df, save_path)


def ChgInValue(journal_entries, sdate, edate, save_path="BASE_PATH/reports"):
    """
    Generate a consolidated report with totals for each category within a given date range.

    Parameters:
    - journal_entries (list of objects): List of journal entry records with attributes.
    - sdate (str): Start date in 'YYYY-MM-DD' format.
    - edate (str): End date in 'YYYY-MM-DD' format.
    - save_path (str): Path to save the consolidated report.

    Returns:
    - DataFrame: A DataFrame containing the consolidated report.
    """
    # Ensure the save_path directory exists
    os.makedirs(save_path, exist_ok=True)

    # Define the mapping for the five categories
    categories = {
        "Cost Basis": ["Cost", "Receivable", "Payable", "AccruedInterestReceivable", "AccruedInterestPayable",
                       "SoldAccruedReceivable", "PurchasedAccruedPayable"],
        "Unrealized": ["UnrealPriceGL", "UnrealFXGL"],
        "Realized": ["PriceGainStatOffset", "FXGainStatOffset", "PriceGainInvestment", "FXGainInvestment"],
        "Income": ["DividendReceipt", "DividendExpense", "UnearnedIncome", "InterestIncome", "InterestReceipt",
                   "InterestExpense", 'FXGainTradeSettle', 'FXGainCurrency'],
        "Capital": ["ContributedCost"]
    }

    # Convert journal entries to DataFrame using object attributes
    input_columns = [
        'portfolio', 'transaction', 'investment', 'ibor_date', 'tradedate',
        'settledate', 'lotid', 'tranid', 'tax_date', 'ls', 'location',
        'financial_account', 'quantity', 'local', 'book', 'notional', 'oface'
    ]
    data = [[getattr(je, col) for col in input_columns] for je in journal_entries]
    df = pd.DataFrame(data, columns=input_columns)

    # Convert ibor_date to datetime for filtering
    df['ibor_date'] = pd.to_datetime(df['ibor_date'])

    # Filter by start and end date
    sdate = pd.to_datetime(sdate)
    edate = pd.to_datetime(edate)
    df = df[(df['ibor_date'] >= sdate) & (df['ibor_date'] <= edate)]

    # Exclude records where book is zero or NaN
    df = df[df['book'].notna() & (df['book'] != 0)]

    # Initialize the consolidated report
    consolidated_data = []

    # Process each category and calculate totals
    for category, accounts in categories.items():
        filtered_df = df[df['financial_account'].isin(accounts)]
        category_total = filtered_df['book'].sum()
        consolidated_data.append({"Category": category, "Book Total": category_total})

    # Add grand total
    grand_total = sum(item["Book Total"] for item in consolidated_data)
    consolidated_data.append({"Category": "Grand Total", "Book Total": grand_total})

    # Create a DataFrame for the consolidated report
    consolidated_df = pd.DataFrame(consolidated_data)

    # Save the consolidated report to an Excel file
    filename = os.path.join(save_path, "Consolidated_Report.xlsx")
    consolidated_df.to_excel(filename, index=False, engine='openpyxl')

    return consolidated_df


def journals_by_tranid(journal_entries,  sdate, edate, fname):
    # Extract data from journal_entries, ensuring it includes all necessary fields
    data = [
        [
            je.portfolio,
            je.transaction,  # Include transaction
            je.investment,
            je.ibor_date,
            je.tradedate,
            je.settledate,
            je.lotid,  # Include lotid
            je.tranid,
            je.tax_date,
            je.ls,
            je.location,
            je.financial_account,
            je.quantity,
            je.local,
            je.book,
            je.notional,  # Include notional
            je.oface  # Include oface
        ]
        for je in journal_entries
    ]

    # Ensure columns match the data fields
    columns = [
        'portfolio',
        'transaction',  # Include transaction
        'investment',
        "ibor_date",
        "tradedate",
        "settledate",
        'lotid',  # Include lotid
        'tranid',
        'tax_date',
        'ls',
        'location',
        'financial_account',
        "quantity",
        'local',
        'book',
        'notional',  # Include notional
        'oface' # Include oface
    ]

    # Create the DataFrame from the list of entries
    df = pd.DataFrame(data, columns=columns)

    df = filter_records(df, 'transaction', ['Valuation'], operator="OR", mode="exclude")



    # Sort the DataFrame by group_by_columns
    group_by_columns = ['ibor_date', 'tranid']  # Replace with your actual column names
    df.sort_values(group_by_columns, inplace=True)
    #
    # # Fetch and map groupings to the DataFrame created from journal entries
    # df_merged = fetch_and_map_groupings(df, "C:/Users/hjmne/PycharmProjects/chest/refdata/investment_master.csv",
    #                                     "C:/Users/hjmne/PycharmProjects/chest/refdata/chart_of_accounts.csv")

    # Calculate grand totals before sub-totaling
    total_columns = ['book']  # Replace with actual column names
    grand_totals = calculate_grand_totals(df, total_columns)

    # Insert subtotals
    df_with_subtotals = insert_subtotals(df, group_by_columns, total_columns)

    # Append the Grand Totals row to the DataFrame with subtotals
    df_final_report = pd.concat([df_with_subtotals, grand_totals], ignore_index=True)
    import datetime

    # Save the DataFrame with subtotals and Grand Totals to an Excel file
    fnamechoice = "C:/Users/hjmne/PycharmProjects/chest/reports/JEs.xlsx"
    df_final_report.to_excel(fnamechoice, index=False)

    # Generate the report title
    report_title = "Journals By TransactionID and Date"  # Replace with your report title

    # Assuming 'Portfolio' is the column name where portfolio names are stored.
    unique_portfolios = df_final_report[
        ~df_final_report['portfolio'].isin(['Subtotal', 'Grand Totals'])
    ]['portfolio'].unique()
    portfolio_name = unique_portfolios[0] if len(unique_portfolios) > 0 else "Unknown Portfolio"
    #
    # if isinstance(sdate, datetime):
    #     sdate1 = sdate.strftime('%Y-%m-%d')
    # else:
    #     # Handle the case where sdate is not a datetime object
    #     sdate1 = str(sdate)  # or some other fallback logic
    #
    # if isinstance(sdate, datetime):
    #     edate1 = edate.strftime('%Y-%m-%d')
    # else:
    #     # Handle the case where sdate is not a datetime object
    #     edate1 = str(edate)  # or some other fallback logic

    # Assuming 'sdate' and 'edate' are your start and end date variables
    # sdate1 = sdate.strftime('%Y-%m-%d')
    # edate1 = edate.strftime('%Y-%m-%d')

    # Combine title and date range
    full_title = "Journals"
    add_title(fnamechoice, "Sheet1", full_title)

    # Apply standard formatting to the saved Excel file
    apply_standard_formatting(fnamechoice, "Sheet1")

def merge_journals_with_groupings(journal_entries, sdate, edate, view, portfolio):
    # Step 1: Extract and structure journal entries
    data = [
        [
            je.portfolio, je.transaction, je.investment, je.ibor_date, je.tradedate,
            je.settledate, je.lotid, je.tranid, je.tax_date, je.ls, je.location,
            je.financial_account, je.quantity, je.local, je.book, je.notional, je.oface
        ]
        for je in journal_entries
    ]

    columns = [
        'Portfolio', 'Transaction', 'Investment', 'IBOR Date', 'Trade Date',
        'Settle Date', 'LotID', 'TranID', 'Tax Date', 'LS', 'Location',
        'Financial Account', 'Quantity', 'Local', 'Book', 'Notional', 'Oface'
    ]

    df = pd.DataFrame(data, columns=columns)

    # Step 2: Filter records (excluding 'Valuation' transactions)
    df = filter_records(df, 'Transaction', ['Valuation'], operator="OR", mode="exclude")

    # Step 2.1: Remove grand total rows
    total_keywords = ['total', 'grand total', 'subtotal', 'summary']
    df = df[~df.apply(lambda row: any(str(row[col]).strip().lower() in total_keywords
                                      for col in ['Investment', 'Transaction', 'Financial Account']), axis=1)]

    # Step 3: Sort by IBOR Date & TranID
    df.sort_values(['IBOR Date', 'TranID'], inplace=True)

    # Step 4: Fetch investment master & chart of accounts data
    investment_master_df = pd.read_csv("C:/BASE_PATH/refdata/investment_master.csv")
    coa_df = pd.read_csv("C:/BASE_PATH/refdata/chart_of_accounts.csv")

    # Standardize column names
    investment_master_df = standardize_columns(investment_master_df)
    coa_df = standardize_columns(coa_df)

    # Select necessary columns from investment master
    investment_groupings = investment_master_df[['Ticker', 'Asset_Class', 'Currency', 'Country', 'Sector', 'Industry', 'Analyst']]

    # Select necessary columns from Chart of Accounts
    coa_groupings = coa_df[['System_Name', 'System_Type', 'BS_Group_Name', 'BS_Group', 'Performance_Category']]

    # Merge Journals with Investment Master Data
    df = df.merge(investment_groupings, left_on='Investment', right_on='Ticker', how='left')

    # Merge the result with Chart of Accounts Data
    df = df.merge(coa_groupings, left_on='Financial Account', right_on='System_Name', how='left')


    fname = f"C:/BASE_PATH/reports/{portfolio}_Journals{view}.xlsx"



    # Step 5: Save merged DataFrame to Excel
    if not fname.endswith('.xlsx'):
        fname += '.xlsx'

    print(f"Saving merged report to: {fname}")
    df.to_excel(fname, index=False)

    # Apply formatting
    apply_standard_formatting(fname, "Sheet1")

    return df

def journals_by_tranid_no_subtotals(journal_entries, sdate, edate, fname, portfolio):
    # Ensure the file path has the .xlsx extension if not provided
    if not fname.endswith('.xlsx'):
        fname += '.xlsx'  # Append .xlsx extension if missing

    # Extract data from journal_entries, ensuring it includes all necessary fields
    data = [
        [
            je.portfolio,
            je.transaction,  # Include transaction
            je.investment,
            je.ibor_date,
            je.tradedate,
            je.settledate,
            je.lotid,  # Include lotid
            je.tranid,
            je.tax_date,
            je.ls,
            je.location,
            je.financial_account,
            je.quantity,
            je.local,
            je.book,
            je.notional,  # Include notional
            je.oface  # Include oface
        ]
        for je in journal_entries
    ]

    # Ensure columns match the data fields
    columns = [
        'portfolio',
        'transaction',  # Include transaction
        'investment',
        'ibor_date',
        'tradedate',
        'settledate',
        'lotid',  # Include lotid
        'tranid',
        'tax_date',
        'ls',
        'location',
        'financial_account',
        'quantity',
        'local',
        'book',
        'notional',  # Include notional
        'oface'  # Include oface
    ]

    # Create the DataFrame from the list of entries
    df = pd.DataFrame(data, columns=columns)

    # Filter records (excluding 'Valuation' transactions)
    df = filter_records(df, 'transaction', ['Valuation'], operator="OR", mode="exclude")

    # No group_by or subtotals. Just sorting by 'tranid' and 'ibor_date' as example (optional)
    df.sort_values(['ibor_date', 'tranid'], inplace=True)

    # Save the DataFrame to an Excel file
    print(f"Saving report to: {fname}")
    df.to_excel(fname, index=False)

    # Add title to the sheet
    full_title = "Journals"
    add_title(fname, "Sheet1", full_title)

    # Apply standard formatting to the saved Excel file
    apply_standard_formatting(fname, "Sheet1")



def journals_assets_liabilities(journal_entries,  sdate, edate, fname):
    # Extract data from journal_entries, ensuring it includes all necessary fields
    update_running_balances(journal_entries, ['portfolio', 'investment', 'ls', 'location', 'financial_account'])
    data = [
        [
            je.portfolio,
            je.transaction,  # Include transaction
            je.investment,
            je.ibor_date,
            je.tradedate,
            je.settledate,
            je.lotid,  # Include lotid
            je.tranid,
            je.tax_date,
            je.ls,
            je.location,
            je.financial_account,
            je.quantity,
            je.local,
            je.book,
            je.notional,  # Include notional
            je.oface,  # Include oface
            je.entry_type,
            je.running_balances[0],
            je.running_balances[1],
            je.running_balances[2]
        ]
        for je in journal_entries
    ]

    # Ensure columns match the data fields
    columns = [
        'portfolio',
        'transaction',  # Include transaction
        'investment',
        "ibor_date",
        "tradedate",
        "settledate",
        'lotid',  # Include lotid
        'tranid',
        'tax_date',
        'ls',
        'location',
        'financial_account',
        "quantity",
        'local',
        'book',
        'notional',  # Include notional
        'oface',  # Include oface
        'entry_type',
        "RunningQty", "RunningLocal", "RunningBook"
    ]

    # Create the DataFrame from the list of entries
    df = pd.DataFrame(data, columns=columns)

    df = filter_records(df, 'transaction', ['Valuation'], operator="OR", mode="exclude")
    df = filter_records(df, 'entry_type', ['Asset/Liability'], operator="OR", mode="include")

    # Sort the DataFrame by group_by_columns
    group_by_columns = ['investment', 'ls', 'location', 'financial_account']  # Replace with your actual column names
    df.sort_values(group_by_columns, inplace=True)
    #
    # # Fetch and map groupings to the DataFrame created from journal entries
    # df_merged = fetch_and_map_groupings(df, "C:/Users/hjmne/PycharmProjects/chest/refdata/investment_master.csv",
    #                                     "C:/Users/hjmne/PycharmProjects/chest/refdata/chart_of_accounts.csv")

    # Calculate grand totals before sub-totaling
    total_columns = ['quantity', 'local', 'book']  # Replace with actual column names
    grand_totals = calculate_grand_totals(df, total_columns)

    # Insert subtotals
    df_with_subtotals = insert_subtotals(df, group_by_columns, total_columns)

    # Append the Grand Totals row to the DataFrame with subtotals
    df_final_report = pd.concat([df_with_subtotals, grand_totals], ignore_index=True)
    import datetime

    # Save the DataFrame with subtotals and Grand Totals to an Excel file
    fnamechoice = "C:/Users/hjmne/PycharmProjects/chest/reports/Journals.xlsx"
    df_final_report.to_excel(fnamechoice, index=False)

    # Generate the report title
    report_title = "Journals Asset Liability Accounts"  # Replace with your report title

    # Assuming 'Portfolio' is the column name where portfolio names are stored.
    unique_portfolios = df_final_report[
        ~df_final_report['portfolio'].isin(['Subtotal', 'Grand Totals'])
    ]['portfolio'].unique()
    portfolio_name = unique_portfolios[0] if len(unique_portfolios) > 0 else "Unknown Portfolio"
    #
    # if isinstance(sdate, datetime):
    #     sdate1 = sdate.strftime('%Y-%m-%d')
    # else:
    #     # Handle the case where sdate is not a datetime object
    #     sdate1 = str(sdate)  # or some other fallback logic
    #
    # if isinstance(sdate, datetime):
    #     edate1 = edate.strftime('%Y-%m-%d')
    # else:
    #     # Handle the case where sdate is not a datetime object
    #     edate1 = str(edate)  # or some other fallback logic

    # Assuming 'sdate' and 'edate' are your start and end date variables
    # sdate1 = sdate.strftime('%Y-%m-%d')
    # edate1 = edate.strftime('%Y-%m-%d')

    # # Combine title and date range
    # full_title = "Journals"
    # add_title(fnamechoice, "Sheet1", full_title)

    # Apply standard formatting to the saved Excel file
    apply_standard_formatting(fnamechoice, "Sheet1")


import pandas as pd
import logging
from pandas.tseries.offsets import BDay
def generate_investment_grouped_report(space_manager, journal_entries, ledger_choice, start_period, end_period, fund,
                                       output_file, derive_mktval, fx_data):
    start_period = pd.to_datetime(start_period)
    end_period = pd.to_datetime(end_period)
    start_period_adjusted = start_period - BDay(1)
    start_period_adjusted = start_period_adjusted.replace(hour=23, minute=59, second=59)

    # Convert the journal entries list to a DataFrame
    combined_journal_entries_df = pd.DataFrame([je.__dict__ for je in journal_entries])

    if 'portfolio' in combined_journal_entries_df.columns:
        combined_journal_entries_df = combined_journal_entries_df.drop(columns=['portfolio'])

    combined_journal_entries_df = clean_dates(combined_journal_entries_df, ['ibor_date', 'tradedate', 'settledate', 'tax_date'])
    combined_journal_entries_df = combined_journal_entries_df.sort_values(by=['investment', 'lotid', 'ibor_date', 'financial_account'])

    # Filter out MarketVal financial account
    combined_journal_entries_df = combined_journal_entries_df[
        (combined_journal_entries_df['financial_account'] != 'MarketVal') |
        (
                (combined_journal_entries_df['local'] == 0) &
                (combined_journal_entries_df['book'] == 0)
        )
    ]

    combined_journal_entries_df.to_csv('journal_entries_report.csv', index=False)

    try:
        logging.info("DataFrame info:")
        logging.info(combined_journal_entries_df.info())
        logging.info("First few rows of the DataFrame:")
        logging.info(combined_journal_entries_df.head())

        # Combine DataFrames
        combined_df = combined_journal_entries_df
        combined_df = combined_df.sort_values(by=['investment', 'ibor_date', 'financial_account'])

        rows = []

        # Process each group by investment
        for investment, group in combined_df.groupby(['investment']):
            subtotal_row = {
                'investment': investment,
                'quantity': group['quantity'].sum(),
                'local': group['local'].sum(),
                'book': group['book'].sum(),
                'unrealgllocal': group['unrealgllocal'].sum(),
                'unrealglbook': group['unrealglbook'].sum(),
                'unrealfxbook': group['unrealfxbook'].sum(),
                'realizedlocal': group['realizedlocal'].sum(),
                'realizedbook': group['realizedbook'].sum(),
                'incomelocal': group['incomelocal'].sum(),
                'incomebook': group['incomebook'].sum(),
                'capitalshares': group['capitalshares'].sum(),
                'capitallocal': group['capitallocal'].sum(),
                'capitalbook': group['capitalbook'].sum(),
                'tranid': 'Subtotal',
                'entry_type': 'Subtotal'
            }
            rows.append(subtotal_row)

            # Append individual rows in the group
            for _, entry in group.iterrows():
                rows.append(entry.to_dict())

        pivot_table_df = pd.DataFrame(rows)
        pivot_table_df = pivot_table_df.sort_values(by=['investment', 'ibor_date', 'transaction'])  # Final sort

        # Add Summary Rows
        book_sum = pivot_table_df['book'].sum()
        unrealglbook_sum = pivot_table_df['unrealglbook'].sum()
        unrealfxbook_sum = pivot_table_df['unrealfxbook'].sum()
        realizedbook_sum = -pivot_table_df['realizedbook'].sum()
        incomebook_sum = -pivot_table_df['incomebook'].sum()
        capitalbook_sum = -pivot_table_df['capitalbook'].sum()
        mkt_val_book = book_sum + unrealglbook_sum + unrealfxbook_sum
        comprehensive_sum = capitalbook_sum + realizedbook_sum + incomebook_sum + unrealfxbook_sum + unrealglbook_sum

        summary_data = pd.DataFrame([{
            'portfolio': 'Summary Book Value',
            'investment': '',
            'transaction': '',
            'lotid': '',
            'ibor_date': '',
            'ls': '',
            'location': '',
            'financial_account': '',
            'quantity': '',
            'local': '',
            'book': book_sum,
            'unrealgllocal': '',
            'unrealglbook': unrealglbook_sum,
            'unrealfxbook': unrealfxbook_sum,
            'realizedlocal': '',
            'realizedbook': realizedbook_sum,
            'incomelocal': '',
            'incomebook': incomebook_sum,
            'capitalshares': '',
            'capitallocal': '',
            'capitalbook': capitalbook_sum,
            'tranid': '',
            'entry_type': ''
        }, {
            'portfolio': 'Summary MarketVal Book Derived from Book Cost & Unrealized',
            'investment': '',
            'transaction': '',
            'lotid': '',
            'ibor_date': '',
            'ls': '',
            'location': '',
            'financial_account': '',
            'quantity': '',
            'local': '',
            'book': mkt_val_book,
            'unrealgllocal': '',
            'unrealglbook': '',
            'unrealfxbook': '',
            'realizedlocal': '',
            'realizedbook': '',
            'incomelocal': '',
            'incomebook': '',
            'capitalshares': '',
            'capitallocal': '',
            'capitalbook': '',
            'tranid': '',
            'entry_type': 'Summary MktValBook Derived from Capital & Net Earnings'
        }, {
            'portfolio': 'Summary',
            'investment': '',
            'transaction': '',
            'lotid': '',
            'ibor_date': '',
            'ls': '',
            'location': '',
            'financial_account': '',
            'quantity': '',
            'local': '',
            'book': comprehensive_sum,
            'unrealgllocal': '',
            'unrealglbook': '',
            'unrealfxbook': '',
            'realizedlocal': '',
            'realizedbook': '',
            'incomelocal': '',
            'incomebook': '',
            'capitalshares': '',
            'capitallocal': '',
            'capitalbook': '',
            'tranid': '',
            'entry_type': 'Comprehensive Sum'
        }])

        pivot_table_df = pd.concat([pivot_table_df, summary_data], ignore_index=True)

        logging.info("Saving pivot table to Excel file...")
        pivot_table_df.to_excel(output_file, index=False, engine='openpyxl')

        wb = load_workbook(output_file)
        ws = wb.active

        # Define the blue fill pattern
        blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

        # Apply blue fill to rows where 'tranid' column equals 'Subtotal'
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            if row[13].value == 'Subtotal':  # Assuming 'tranid' is the 14th column (index 13)
                for cell in row:
                    cell.fill = blue_fill

            # Apply number formatting and font sizes
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
                cell.font = Font(size=11)  # Set normal font size

        # Apply bold font for summary rows
        summary_rows = ['Summary Book Value', 'Summary MarketVal Book Derived from Book Cost & Unrealized', 'Comprehensive Sum']
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            if row[0].value in summary_rows:
                for cell in row:
                    cell.font = Font(bold=True)

        # Freeze the top row for column titles
        ws.freeze_panes = 'A2'

        # Auto-size the columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(output_file)
    except Exception as e:
        logging.error(f"An error occurred: {e}")
    finally:
        if 'pivot_table_df' in locals():
            logging.info(f"Pivot table saved with formatting. Number of records in pivot table: {len(pivot_table_df)}")
        else:
            logging.info("Pivot table was not created due to an error.")


def journals_by_sequence_number(journal_entries, sdate, edate, fname):
    """
    Generate a report of journal entries sorted by sequence number.

    Args:
    - journal_entries (list): List of Journals objects to process.
    - sdate (datetime): Start date for the report.
    - edate (datetime): End date for the report.
    - fname (str): File name to save the report to.
    """
    # Extract data from journal_entries, ensuring it includes all necessary fields
    data = [
        [
            je.sequence_number,  # Include sequence number for sorting
            je.portfolio,
            je.transaction,
            je.investment,
            je.ibor_date,
            je.tradedate,
            je.settledate,
            je.lotid,
            je.tranid,
            je.tax_date,
            je.ls,
            je.location,
            je.financial_account,
            je.quantity,
            je.local,
            je.book,
            je.notional,
            je.oface
        ]
        for je in journal_entries
    ]

    # Ensure columns match the data fields
    columns = [
        'sequence_number',  # Include sequence number for sorting
        'portfolio',
        'transaction',
        'investment',
        "ibor_date",
        "tradedate",
        "settledate",
        'lotid',
        'tranid',
        'tax_date',
        'ls',
        'location',
        'financial_account',
        "quantity",
        'local',
        'book',
        'notional',
        'oface'
    ]

    # Create the DataFrame from the list of entries
    df = pd.DataFrame(data, columns=columns)

    # Sort the DataFrame by sequence_number
    df.sort_values('sequence_number', inplace=True)

    # Save the sorted DataFrame to an Excel file
    fnamechoice = "C:/Users/hjmne/PycharmProjects/chest/reports/Journals_By_Sequence_Number.xlsx"
    df.to_excel(fnamechoice, index=False)

    # Generate the report title
    report_title = "Journals By Sequence Number"

    # Adding title to the Excel file
    add_title(fnamechoice, "Sheet1", report_title)

    # Apply standard formatting to the saved Excel file
    apply_standard_formatting(fnamechoice, "Sheet1")


import openpyxl
from openpyxl.styles import PatternFill, Font

# Make sure to include the insert_subtotals and apply_standard_formatting functions here
def test_combine():
    import pandas as pd

    try:
        df1 = pd.read_csv('C:/Users/hjmne/PycharmProjects/chest/reports/CostBasisBalanceSheet.xlsx', encoding='utf-8')
    except UnicodeDecodeError:
        df1 = pd.read_csv('C:/Users/hjmne/PycharmProjects/chest/reports/CostBasisBalanceSheet.xlsx', encoding='latin1')  # Try a different encoding

    try:
        df2 = pd.read_csv('C:/Users/hjmne/PycharmProjects/chest/reports/PeriodIncomeEarned.xlsx', encoding='utf-8')
    except UnicodeDecodeError:
        df2 = pd.read_csv('C:/Users/hjmne/PycharmProjects/chest/reports/PeriodIncomeEarned.xlsx', encoding='latin1')  # Try a different encoding


    # Optional: Add a delimiter row for visual separation
    delimiter_row = pd.DataFrame({col: '' for col in df1.columns}, index=[0])

    # Concatenate the DataFrames with the delimiter row in between
    combined_df = pd.concat([df1, delimiter_row, df2], ignore_index=True)

    # Save the combined DataFrame to a new file or proceed with further processing
    combined_df.to_csv('C:/Users/hjmne/PycharmProjects/chest/reports/tworeports.xlsx', index=False)

import pandas as pd


def valuation_style_reports(bookkeeping_space, sub_ledger, edate, fname):
    # Step 1: Create the DataFrame from the list of entries
    # Assuming sub_ledger_list is the list shown above
    bookkeeping_space_list = []
    for entry in bookkeeping_space:
        portfolio, investment, lotid, tax_date, ls, location, financial_account, quantity, local, book, notional, oface = entry
        booksp_row = [portfolio, investment, lotid, tax_date, ls, location, financial_account, quantity, local, book, notional, oface]
        bookkeeping_space_list.append(booksp_row)


    df = pd.DataFrame(bookkeeping_space_list, columns=[
        'Portfolio', 'Investment', 'Lot_Id', 'Tax_Date', 'Ls', 'Location',
        'Financial_Account', 'Quantity', 'Local', 'Book', 'Notional', 'Oface'
    ])

    #Step 2 - value the positions
    # Convert the DataFrame to a dictionary and pass it to the mark_calculations function
    bookkeeping_records = df.set_index([
        'Portfolio', 'Investment', 'Lot_Id', 'Tax_Date', 'Ls', 'Location', 'Financial_Account']).T.to_dict('list')
    df = filter_records(df, 'Financial Account', ['MktVal', 'MktValRE'], operator="OR", mode="exclude")
    # Assuming these are the paths to your data files
    price_data = "C:/Users/hjmne/PycharmProjects/chest/refdata/price_master.csv"
    fx_data = "C:/Users/hjmne/PycharmProjects/chest/refdata/fx_master.csv"

    mark_type = "Valuation"
    df_valued = value_positions(bookkeeping_records,sub_ledger, price_data, fx_data, edate, mark_type)


    #Step 3 - merge with COA/Inv Master Columns

    df_merged = fetch_and_map_groupings(df_valued, "C:/Users/hjmne/PycharmProjects/chest/refdata/investment_master.csv",
                                        "C:/Users/hjmne/PycharmProjects/chest/refdata/chart_of_accounts.csv")

    # Step 4 - important calc grand totals first. What you are grouping by will dictate subtotal_columns vs total_cols
    group_by_columns = ['Investment', 'Financial Account', 'LS']
    total_columns = ['Book',  'Mkt Val Book', 'PGain Book','TotGain Book', 'Fx Gain']
    subtotal_columns = ['Quantity', 'Local', 'Book', 'Mkt Val Local', 'Mkt Val Book', 'PGain Local', 'PGain Book',
                        'TotGain Book', 'Fx Gain']

    #Always do grand_totals before sub-totaling
    grand_totals = calculate_grand_totals(df_merged, total_columns)

    df_subtotals = subtotal_data(df_merged, subtotal_columns, group_by_columns)

    df_subtotals = df_subtotals[df_subtotals['Quantity'] != 0]
    df_subtotals.sort_values(by=['Investment', 'Financial Account', 'LS'], inplace=True)
    df_subtotals = insert_subtotals(df_subtotals, ['Investment', 'Financial Account', 'LS'], subtotal_columns)

    # Append the Grand Totals row to the DataFrame with subtotals
    df_final_report = pd.concat([df_subtotals, grand_totals], ignore_index=True)

    # Step 5: Save the DataFrame to an Excel file
    fnamechoice = f"C:/Users/hjmne/PycharmProjects/chest/reports/TaxLotHoldings"+fname+".xlsx"
    df_final_report.to_excel(fnamechoice, index=False)

    report_title = "Tax Lot Holdings"+fname  # Replace with your report title
    # Assuming 'Portfolio' is the column name where portfolio names are stored.
    # This filters the DataFrame to exclude any rows where 'Portfolio' is 'Subtotal' or 'Grand Totals'
    # and then gets the unique values in the 'Portfolio' column.
    unique_portfolios = df_final_report[
        ~df_final_report['Portfolio'].isin(['Subtotal', 'Grand Totals'])
    ]['Portfolio'].unique()
    # Check if there's at least one portfolio name and use the first one; otherwise, use a default.
    portfolio_name = unique_portfolios[0] if len(unique_portfolios) > 0 else "Unknown Portfolio"
    edate1 = edate.strftime('%Y-%m-%d')
    period_end_date = edate1  # Assuming 'date' is a datetime object
    # Combine title and date range
    full_title = f"{report_title} for {portfolio_name} ({period_end_date})"
    add_title(fnamechoice, "Sheet1", full_title)

  # Now apply standard formatting to the saved Excel file
    apply_standard_formatting(fnamechoice, "Sheet1")


def cost_basis_balance_sheet(portfolio, bookkeeping_space, edate, fname):
    bookkeeping_space_list = []
    for entry in bookkeeping_space:
        # Extract required fields
        portfolio = entry[0]  # Portfolio
        investment = entry[1]  # Investment
        ls = entry[4]  # LS
        financial_account = entry[6]  # Financial Account
        quantity = entry[7]  # Quantity
        local = entry[8]  # Local
        book = entry[9]  # Book

        # Construct row and add to the list
        booksp_row = [portfolio, investment, ls, financial_account, quantity, local, book]
        bookkeeping_space_list.append(booksp_row)

    # Create DataFrame with correct column names
    df = pd.DataFrame(bookkeeping_space_list, columns=[
        'Portfolio', 'Investment', 'LS', 'Financial Account', 'Quantity', 'Local', 'Book'
    ])

    # **Step 1: Remove Grand Totals & Subtotals Before Any Processing**
    total_keywords = ['total', 'grand total', 'subtotal', 'summary']
    df = df[~df.apply(lambda row: any(str(row[col]).strip().lower() in total_keywords
                                      for col in ['Portfolio', 'Investment', 'Financial Account']), axis=1)]

    # **Step 2: If DataFrame is empty, create an empty report with correct headers**
    if df.empty:
        print("⚠ WARNING: No data available. Creating an empty report with column headers.")

        # restore header columns
        df = pd.DataFrame(bookkeeping_space_list, columns=[
            'Portfolio', 'Investment', 'LS', 'Financial Account', 'Quantity', 'Local', 'Book'
        ])

        # Save empty report (headers only, no data)
        fnamechoice = f"C:/Users/hjmne/PycharmProjects/chest/reports/{portfolio}_{fname}.xlsx"
        df.to_excel(fnamechoice, index=False)

        print(f"✅ Empty report with headers saved: {fnamechoice}")

        # Apply formatting
        # add_title(fnamechoice, "Sheet1", f"{portfolio} - No Data Available")
        # apply_standard_formatting(fnamechoice, "Sheet1")

        return df  # ✅ Return an empty DataFrame with headers

    # **Step 3: Remove Unrealized Gains/Losses**
    df = df[~df['Financial Account'].isin(['UnrealPriceGL', 'UnrealFXGL'])]

    # **Step 4: Merge with COA & Investment Master**
    df_merged = fetch_and_map_groupings(df, "C:/Users/hjmne/PycharmProjects/chest/refdata/investment_master.csv",
                                        "C:/Users/hjmne/PycharmProjects/chest/refdata/chart_of_accounts.csv")

    # **Step 5: Select required columns and sort**
    df_primary_structure = df_merged[['Portfolio', 'Investment', 'BS_Group', 'BS_Group_Name', 'Asset_Class',
                                      'LS', 'Financial Account', 'Quantity', 'Local', 'Book']]
    sorted_df = df_primary_structure.sort_values(by=['BS_Group', 'Asset_Class', 'Financial Account'],
                                                 ascending=[True, True, True])

    # **Step 6: Generate Subtotals**
    group_by_columns = ['BS_Group']
    subtotal_columns = ['Book']
    df_subtotals = subtotal_data(sorted_df, subtotal_columns, group_by_columns)

    # **Step 7: Sort & Insert Subtotals**
    df_subtotals.sort_values(by=['BS_Group', 'Asset_Class', 'Financial Account', 'LS'], inplace=True)
    df_subtotals = insert_subtotals(df_subtotals, ['BS_Group', 'Asset_Class', 'Financial Account', 'LS'],
                                    subtotal_columns)

    # **Step 8: Remove Subtotals & Grand Totals Again After Calculations**
    df_subtotals = df_subtotals[~df_subtotals.apply(lambda row: any(str(row[col]).strip().lower() in total_keywords
                                                                    for col in
                                                                    ['Portfolio', 'Investment', 'Financial Account']),
                                                    axis=1)]

    # **Step 9: If Subtotals Are Empty, Still Create Report With Headers**
    if df_subtotals.empty:
        print(
            "⚠ WARNING: Subtotals DataFrame is empty after removing totals. Creating a 0-record report with column headers.")

        # Create empty DataFrame with proper structure
        empty_df = pd.DataFrame(columns=df_primary_structure.columns)
        fnamechoice = f"C:/Users/hjmne/PycharmProjects/chest/reports/{portfolio}_{fname}.xlsx"
        empty_df.to_excel(fnamechoice, index=False)

        print(f"✅ Empty report with headers saved: {fnamechoice}")

        # Apply formatting
        add_title(fnamechoice, "Sheet1", f"{portfolio} - No Data Available")
        apply_standard_formatting(fnamechoice, "Sheet1")

        return empty_df  # ✅ Return empty DataFrame with correct headers

    # **Step 10: Save the Final Report**
    fnamechoice = f"C:/Users/hjmne/PycharmProjects/chest/reports/{portfolio}_{fname}.xlsx"
    df_subtotals.to_excel(fnamechoice, index=False)

    # **Step 11: Generate Title for Report**
    # unique_portfolios = df_subtotals['Portfolio'].unique()
    # portfolio_name = unique_portfolios[0] if len(unique_portfolios) > 0 else "Unknown Portfolio"
    # edate1 = edate.strftime('%Y-%m-%d')
    # full_title = f"{portfolio_name}_{fname}"
    #
    # # **Step 12: Apply Formatting**
    # add_title(fnamechoice, "Sheet1", full_title)
    # apply_standard_formatting(fnamechoice, "Sheet1")

    return df_subtotals  # ✅ Return the final processed DataFrame

    # # **Step 11: Generate Title for Report**
    # unique_portfolios = df_subtotals['Portfolio'].unique()
    # portfolio_name = unique_portfolios[0] if len(unique_portfolios) > 0 else "Unknown Portfolio"
    # edate1 = edate.strftime('%Y-%m-%d')
    # full_title = f"{portfolio_name}_{fname}"
    #
    # # **Step 12: Apply Formatting**
    # add_title(fnamechoice, "Sheet1", full_title)
    # apply_standard_formatting(fnamechoice, "Sheet1")

    return df_subtotals  # ✅ Return the final processed DataFrame

# def position_report_by_sector(bookkeeping_space, sub_ledger, edate, kd, fname, price_data, fx_data):
#     # Step 1: Create the DataFrame from the list of entries
#     bookkeeping_space_list = []
#     for entry in bookkeeping_space:
#         portfolio, investment, lotid, tax_date, ls, location, financial_account, quantity, local, book, notional, oface = entry
#         booksp_row = [portfolio, investment, lotid, tax_date, ls, location, financial_account, quantity, local, book, None, None]
#         bookkeeping_space_list.append(booksp_row)
#
#     df = pd.DataFrame(bookkeeping_space_list, columns=[
#         'Portfolio', 'Investment', 'Lot_ID', 'Tax_Date', 'Ls', 'Location',
#         'Financial_Account', 'Quantity', 'Local', 'Book', 'Notional', 'Oface'
#     ])
#     df = filter_records(df, 'Financial_Account', ['MktVal', 'MktValRE'], operator="OR", mode="exclude")
#
#     # Step 2 - value the positions
#     # Convert the DataFrame to a dictionary and pass it to the mark_calculations function
#     mark_type = "Valuation"
#     df_valued = value_positions(df, sub_ledger,  edate, mark_type)
#
#     # Replace the existing merge operations with the call to fetch_and_map_groupings
#     df_valued = fetch_and_map_groupings(df,
#                                         "C:/Users/hjmne/PycharmProjects/chest/refdata/investment_master.csv",
#                                         "C:/Users/hjmne/PycharmProjects/chest/refdata/chart_of_accounts.csv")
#
#     # Fill NaN values in 'Sector' and 'Asset_Class' columns
#     df_valued['Sector'] = df_valued['Sector'].fillna('Undefined')
#     df_valued['Asset_Class'] = df_valued['Asset_Class'].fillna('Undefined')
#
#     # Standardize columns and key values before groupby and agg operations
#     df_valued = standardize_columns(df_valued)
#     df_valued = standardize_key_values(df_valued,
#                                        ['Portfolio', 'Asset_Class', 'Sector', 'Investment', 'Ls', 'Financial_Account'])
#
#     # Group by the remaining columns excluding 'Tax_Lot' and sum the numeric columns
#     group_by_columns = ['Portfolio', 'Asset_Class', 'Sector', 'Investment', 'Ls', 'Financial_Account']
#     aggregate_columns = ['Quantity', 'Local', 'Book', 'Mkt_Val_Local', 'Mkt_Val_Book', 'Pgain_Local', 'Pgain_Book',
#                          'Totgain_Book', 'Fx_Gain']
#
#     # Perform groupby and agg operations
#     df_merged = df_valued.groupby(group_by_columns)[aggregate_columns].sum().reset_index()
#
#
#     # Step 4 - important calc grand totals first. What you are grouping by will dictate subtotal_columns vs total_cols
#     group_by_columns = ['Asset_Class', 'Sector']
#     total_columns = ['Book', 'Mkt_Val_Book', 'PGain_Book', 'TotGain_Book', 'Fx_Gain']
#     subtotal_columns = ['Book', 'Mkt_Val_Book', 'PGain_Book', 'TotGain_Book', 'Fx_Gain']
#
#     # Always do grand_totals before sub-totaling
#     grand_totals = calculate_grand_totals(df_merged, total_columns)
#
#     # Step 5 - Subtotals
#     df_subtotals = subtotal_data(df_merged, subtotal_columns, group_by_columns)
#     df_subtotals = df_subtotals[df_subtotals['Quantity'] != 0]
#     df_subtotals.sort_values(by=['Asset_Class', 'Sector', 'Investment', 'LS'], inplace=True)
#     df_subtotals = insert_subtotals(df_subtotals, ['Asset_Class', 'Sector'], subtotal_columns)
#
#     # Step 6 - Add % of portfolio
#     df_subtotals = calculate_percentage_of_portfolio(df_subtotals, grand_totals, 'Book')
#     df_subtotals = move_column_in_dataframe(df_subtotals, 'Percent_Of_Portfolio', 3)
#
#     # Step 7 - Reclassify 'payable' and 'receivable' under 'Cash & Equivalents'
#     cash_equiv_accounts = df_subtotals[(df_subtotals['Financial_Account'] == 'Payable') |
#                                        (df_subtotals['Financial_Account'] == 'Receivable') |
#                                        (df_subtotals['Financial_Account'] == 'InterestReceivable') |
#                                        (df_subtotals['Financial_Account'] == 'AccruedInterestReceivable') |
#                                        (df_subtotals['Financial_Account'] == 'DividendsReceivable') |
#                                        (df_subtotals['Financial_Account'] == 'InterestPayable') |
#                                        (df_subtotals['Financial_Account'] == 'DividendsPayable') |
#                                        (df_subtotals['Financial_Account'] == 'AccruedInterestPayable') |
#                                        (df_subtotals['Financial_Account'] == 'ExpensesPayable')]
#     cash_equiv_accounts['LS'] = 'n'  # Change LS designation to 'n'
#
#     # Step 8 - Filter out 'payable' and 'receivable' from the original subtotals DataFrame
#     df_subtotals = df_subtotals[~((df_subtotals['Financial_Account'] == 'Payable') |
#                                   (df_subtotals['Financial_Account'] == 'Receivable') |
#                                   (df_subtotals['Financial_Account'] == 'InterestReceivable') |
#                                   (df_subtotals['Financial_Account'] == 'AccruedInterestReceivable') |
#                                   (df_subtotals['Financial_Account'] == 'DividendsReceivable') |
#                                   (df_subtotals['Financial_Account'] == 'InterestPayable') |
#                                   (df_subtotals['Financial_Account'] == 'DividendsPayable') |
#                                   (df_subtotals['Financial_Account'] == 'AccruedInterestPayable') |
#                                   (df_subtotals['Financial_Account'] == 'ExpensesPayable'))]
#
#     # Step 9 - Concatenate 'cash_equiv_accounts' with the original 'Cash & Equivalents' subtotals
#     cash_equiv_df = pd.concat([df_subtotals[df_subtotals['LS'] == 'n'], cash_equiv_accounts])
#
#     # Step 10 - Break out positions by three categories: Cash & Equivalents, Long Positions, and Short Positions
#     long_positions_df = df_subtotals[df_subtotals['LS'] == 'l']
#     short_positions_df = df_subtotals[df_subtotals['LS'] == 's']
#
#     # Step 11 - Create Category Headers
#     cash_equiv_header = pd.DataFrame({'Portfolio': ['Cash & Equivalents']})
#     long_positions_header = pd.DataFrame({'Portfolio': ['Long Positions']})
#     short_positions_header = pd.DataFrame({'Portfolio': ['Short Positions']})
#
#     # Step 12 - Concatenate DataFrames
#     cash_equiv_final = pd.concat([cash_equiv_header, cash_equiv_df], ignore_index=True)
#     long_positions_final = pd.concat([long_positions_header, long_positions_df], ignore_index=True)
#     short_positions_final = pd.concat([short_positions_header, short_positions_df], ignore_index=True)
#
#     # Step 13 - Concatenate all DataFrames
#     df_final_report = pd.concat([cash_equiv_final, long_positions_final, short_positions_final, grand_totals], ignore_index=True)
#
#     # Step 14 - Save the DataFrame to an Excel file
#     fnamechoice = f"C:/Users/hjmne/PycharmProjects/chest/reports/PositionReportBySector-"+fname+".xlsx"
#     df_final_report.to_excel(fnamechoice, index=False)
#
#     report_title = ("Positions Report by Industry Sector" + fname+'-KnowledgeDate-'+str(kd)) # Replace with your report title
#     # Assuming 'Portfolio' is the column name where portfolio names are stored.
#     # This filters the DataFrame to exclude any rows where 'Portfolio' is 'Subtotal' or 'Grand Totals'
#     # and then gets the unique values in the 'Portfolio' column.
#     unique_portfolios = df_final_report[
#         ~df_final_report['Portfolio'].isin(['Subtotal', 'Grand Totals'])
#     ]['Portfolio'].unique()
#     # Check if there's at least one portfolio name and use the first one; otherwise, use a default.
#     portfolio_name = unique_portfolios[0] if len(unique_portfolios) > 0 else "Unknown Portfolio"
#     edate1 = edate.strftime('%Y-%m-%d')
#     period_end_date = edate1  # Assuming 'date' is a datetime object
#     # Combine title and date rangem
#     full_title = f"{report_title} as of {period_end_date}"
#     add_title(fnamechoice, "Sheet1", full_title)
#
#     # Now apply standard formatting to the saved Excel file
#    apply_standard_formatting(fnamechoice, "Sheet1")
import pandas as pd
import os


def top_bottom_by_analyst(portfolio_name, report_file, master_file):
    top_output_file = f"C:/Users/hjmne/pycharmprojects/chest/reports/{portfolio_name}_Analyst_Top_Ten.xlsx"
    bottom_output_file = f"C:/Users/hjmne/pycharmprojects/chest/reports/{portfolio_name}_Analyst_Bottom_Ten.xlsx"

    # Check if files exist
    if not os.path.exists(report_file):
        raise FileNotFoundError(f"Report file not found: {report_file}")

    if not os.path.exists(master_file):
        raise FileNotFoundError(f"Master file not found: {master_file}")

    # Load the investment report (Excel file)
    report_df = pd.read_excel(report_file)

    # Load the investment master data (CSV file)
    master_df = pd.read_csv(master_file)

    # Standardize column names for merging
    report_df.rename(columns={'investment': 'Investment'}, inplace=True)

    # Merge data on the "Investment" column (assumed to be the common key)
    merged_df = report_df.merge(master_df, on="Investment", how="left")

    # Sort by Book_To_Date_Percent (descending order)
    sorted_df = merged_df.sort_values(by="Book_To_Date_Percent", ascending=False)

    # Extract top 10 best and bottom 10 worst performing investments
    top_10 = sorted_df.head(10)
    bottom_10 = sorted_df.tail(10)

    # Reset index for readability
    top_10.reset_index(drop=True, inplace=True)
    bottom_10.reset_index(drop=True, inplace=True)

    # Add Analyst column
    top_10['Analyst'] = merged_df['Analyst']
    bottom_10['Analyst'] = merged_df['Analyst']

    # Save top and bottom reports to separate Excel files
    top_10.to_excel(top_output_file, index=False)
    bottom_10.to_excel(bottom_output_file, index=False)

    return top_10, bottom_10


def calculate_unrealized_gains(bookkeeping_space, portfolio_key, financial_account_key):
    unrealized_gains = []

    for entry in bookkeeping_space:
        portfolio = entry[portfolio_key]  # Portfolio
        investment = entry['investment']  # Investment
        tx_date = entry['tx_date']  # Transaction Date (ibor_date)
        lot_id = entry['lotid']  # Lot ID
        location = entry['location']  # Location
        financial_account = entry[financial_account_key]  # Financial Account
        quantity = entry['quantity']  # Quantity
        book_value = entry['book_value']  # Book Value
        market_value = entry['market_value']  # Market Value (already present)

        unrealized_gain_loss = (market_value - book_value) * quantity

        unrealized_gains.append(
            [portfolio, investment, tx_date, lot_id, location, financial_account, quantity, book_value, market_value,
             unrealized_gain_loss])

    # Create DataFrame for Unrealized Gains/Losses
    unrealized_gains_df = pd.DataFrame(unrealized_gains, columns=[
        'Portfolio', 'Investment', 'Transaction Date', 'Lot ID', 'Location', 'Financial Account', 'Quantity',
        'Book Value', 'Market Value', 'Unrealized Gain/Loss'
    ])

    # Group by Financial Account and then by other keys like Portfolio, Investment, etc.
    grouped_unrealized_gains_df = unrealized_gains_df.groupby(
        ['Portfolio', 'Investment', 'Transaction Date', 'Lot ID', 'Location', 'Financial Account']
    ).agg({
        'Quantity': 'sum',
        'Book Value': 'sum',
        'Market Value': 'sum',
        'Unrealized Gain/Loss': 'sum'
    }).reset_index()

    return unrealized_gains_df, grouped_unrealized_gains_df



def position_report_by_sector(bookkeeping_space, portfolio_name, sub_ledger,
                              edate, fname, price_data, fx_data, calendar):
    if not fname.endswith('.xlsx'):
        fname += '.xlsx'

    # Ensure output directory
    output_path = "BASE_PATH/reports"
    os.makedirs(output_path, exist_ok=True)

    # Ensure 'edate' is a proper date object
    if not hasattr(edate, 'month'):
        raise AttributeError("'edate' must be a datetime object with 'month', 'day', and 'year' attributes.")

    # Construct the full filename
    output_filepath = os.path.join(output_path, f"{portfolio_name}_{calendar}PositionsBySector.xlsx")

    # Step 1: Create the DataFrame from the list of entries
    bookkeeping_space_list = []
    for entry in bookkeeping_space:
        portfolio, investment, lotid, tax_date, ls, location, financial_account, quantity, local, book, notional, oface = entry
        bookkeeping_space_list.append([portfolio, investment, lotid, tax_date, ls, location, financial_account, quantity, local, book])

    df = pd.DataFrame(bookkeeping_space_list, columns=[
        'Portfolio', 'Investment', 'Lot_ID', 'Tax_Date', 'Ls', 'Location',
        'Financial Account', 'Quantity', 'Local_Cost', 'Book_Cost'
    ])

    # Step 2: Perform valuation logic
    valuation_data = []
    for _, row in df.iterrows():
        investment = row['Investment']
        quantity = row['Quantity']
        local_cost = row['Local_Cost']
        book_cost = row['Book_Cost']

        price = utilities.get_price(investment, edate, price_data)
        currency = sub_ledger.get_information_field(investment, 'AIF', 'Currency')
        fx_rate = utilities.get_fx_rate(currency, edate, fx_data)

        if price is None or fx_rate is None:
            print(f"Price or FX rate not found for {investment} on {edate}")
            continue

        pricing_factor = sub_ledger.get_information_field(investment, 'AIF', 'Pricing_Factor') or 1
        mktval_local = quantity * float(price) * float(pricing_factor)
        mktval_book = mktval_local * fx_rate
        price_gl_local = mktval_local - local_cost
        price_gl_book = price_gl_local * fx_rate
        fx_gl_book = mktval_book - book_cost - price_gl_book

        valuation_data.append({
            'Portfolio': row['Portfolio'],
            'Investment': investment,
            'Lot_ID': row['Lot_ID'],
            'Tax_Date': row['Tax_Date'],
            'Ls': row['Ls'],
            'Location': row['Location'],
            'Financial Account': row['Financial Account'],
            'Quantity': quantity,
            'Local_Cost': local_cost,
            'Book_Cost': book_cost,
            'MktVal_Local': mktval_local,
            'MktVal_Book': mktval_book,
            'PriceGLLocal': price_gl_local,
            'PriceGLBook': price_gl_book,
            'FXGLBook': fx_gl_book
        })

    df_valuations = pd.DataFrame(valuation_data)

    # Step 3: Merge with investment master to include Full_Name
    investment_master_path = "C:/Users/hjmne/PycharmProjects/chest/refdata/investment_master.csv"
    investment_master = pd.read_csv(investment_master_path, usecols=['Investment', 'Full_Name'])
    df_valuations = df_valuations.merge(investment_master, on='Investment', how='left')

    # Step 4: Group by Sector and asset class
    df_valuations = fetch_and_map_groupings(df_valuations,
                                            "C:/Users/hjmne/PycharmProjects/chest/refdata/investment_master.csv",
                                            "C:/Users/hjmne/PycharmProjects/chest/refdata/chart_of_accounts.csv")

    df_valuations['Sector'] = df_valuations['Sector'].fillna('Undefined')

    # Step 5: Aggregate data by Sector and asset class
    group_by_columns = ['Portfolio', 'Sector', 'Investment', 'Full_Name', 'Ls', 'Financial Account']
    aggregate_columns = ['Quantity', 'Local_Cost', 'Book_Cost', 'MktVal_Local', 'MktVal_Book', 'PriceGLLocal', 'PriceGLBook', 'FXGLBook']

    df_grouped = df_valuations.groupby(group_by_columns)[aggregate_columns].sum().reset_index()

    # Step 6: Add grand totals
    grand_totals = df_grouped[aggregate_columns].sum().to_dict()
    grand_totals.update({col: '' for col in group_by_columns})
    grand_totals['Portfolio'] = 'Grand Totals'
    df_with_totals = pd.concat([df_grouped, pd.DataFrame([grand_totals])], ignore_index=True)

    # Step 7: Save to Excel
    df_with_totals.to_excel(output_filepath, index=False, sheet_name="PositionsBySector")

    print(f"Position report by Sector saved as {output_filepath}.")


def position_report_by_industry(bookkeeping_space, portfolio_name, sub_ledger, edate, fname, price_data, fx_data):
    if not fname.endswith('.xlsx'):
        fname += '.xlsx'

    # Ensure output directory
    output_path = "BASE_PATH/reports"
    os.makedirs(output_path, exist_ok=True)

    # Ensure 'edate' is a proper date object
    if not hasattr(edate, 'month'):
        raise AttributeError("'edate' must be a datetime object with 'month', 'day', and 'year' attributes.")

    # Construct the full filename
    output_filepath = os.path.join(output_path, f"{portfolio_name}_PositionsByIndustry.xlsx")

    # Step 1: Create the DataFrame from the list of entries
    bookkeeping_space_list = []
    for entry in bookkeeping_space:
        portfolio, investment, lotid, tax_date, ls, location, financial_account, quantity, local, book, notional, oface = entry
        bookkeeping_space_list.append([portfolio, investment, lotid, tax_date, ls, location, financial_account, quantity, local, book])

    df = pd.DataFrame(bookkeeping_space_list, columns=[
        'Portfolio', 'Investment', 'Lot_ID', 'Tax_Date', 'Ls', 'Location',
        'Financial Account', 'Quantity', 'Local_Cost', 'Book_Cost'
    ])

    # Step 2: Perform valuation logic
    valuation_data = []
    for _, row in df.iterrows():
        investment = row['Investment']
        quantity = row['Quantity']
        local_cost = row['Local_Cost']
        book_cost = row['Book_Cost']

        price = utilities.get_price(investment, edate, price_data)
        currency = sub_ledger.get_information_field(investment, 'AIF', 'Currency')
        fx_rate = utilities.get_fx_rate(currency, edate, fx_data)

        if price is None or fx_rate is None:
            print(f"Price or FX rate not found for {investment} on {edate}")
            continue

        pricing_factor = sub_ledger.get_information_field(investment, 'AIF', 'Pricing_Factor') or 1
        mktval_local = quantity * float(price) * float(pricing_factor)
        mktval_book = mktval_local * fx_rate
        price_gl_local = mktval_local - local_cost
        price_gl_book = price_gl_local * fx_rate
        fx_gl_book = mktval_book - book_cost - price_gl_book

        valuation_data.append({
            'Portfolio': row['Portfolio'],
            'Investment': investment,
            'Lot_ID': row['Lot_ID'],
            'Tax_Date': row['Tax_Date'],
            'Ls': row['Ls'],
            'Location': row['Location'],
            'Financial Account': row['Financial Account'],
            'Quantity': quantity,
            'Local_Cost': local_cost,
            'Book_Cost': book_cost,
            'MktVal_Local': mktval_local,
            'MktVal_Book': mktval_book,
            'PriceGLLocal': price_gl_local,
            'PriceGLBook': price_gl_book,
            'FXGLBook': fx_gl_book
        })

    df_valuations = pd.DataFrame(valuation_data)

    # Step 3: Merge with investment master to include Full_Name
    investment_master_path = "C:/Users/hjmne/PycharmProjects/chest/refdata/investment_master.csv"
    investment_master = pd.read_csv(investment_master_path, usecols=['Investment', 'Full_Name'])
    df_valuations = df_valuations.merge(investment_master, on='Investment', how='left')

    # Step 4: Group by industry and asset class
    df_valuations = fetch_and_map_groupings(df_valuations,
                                            "C:/Users/hjmne/PycharmProjects/chest/refdata/investment_master.csv",
                                            "C:/Users/hjmne/PycharmProjects/chest/refdata/chart_of_accounts.csv")

    df_valuations['Industry'] = df_valuations['Industry'].fillna('Undefined')

    # Step 5: Aggregate data by industry and asset class
    group_by_columns = ['Portfolio', 'Industry', 'Investment', 'Full_Name', 'Ls', 'Financial Account']
    aggregate_columns = ['Quantity', 'Local_Cost', 'Book_Cost', 'MktVal_Local', 'MktVal_Book', 'PriceGLLocal', 'PriceGLBook', 'FXGLBook']

    df_grouped = df_valuations.groupby(group_by_columns)[aggregate_columns].sum().reset_index()

    # Step 6: Add grand totals
    grand_totals = df_grouped[aggregate_columns].sum().to_dict()
    grand_totals.update({col: '' for col in group_by_columns})
    grand_totals['Portfolio'] = 'Grand Totals'
    df_with_totals = pd.concat([df_grouped, pd.DataFrame([grand_totals])], ignore_index=True)

    # Step 7: Save to Excel
    df_with_totals.to_excel(output_filepath, index=False, sheet_name="PositionsByIndustry")

    print(f"Position report by industry saved as {output_filepath}.")

def position_report_by_analyst(bookkeeping_space, portfolio_name, sub_ledger, edate, fname, price_data, fx_data):
    if not fname.endswith('.xlsx'):
        fname += '.xlsx'

    # Ensure output directory
    output_path = "BASE_PATH/reports"
    os.makedirs(output_path, exist_ok=True)

    # Ensure 'edate' is a proper date object
    if not hasattr(edate, 'month'):
        raise AttributeError("'edate' must be a datetime object with 'month', 'day', and 'year' attributes.")

    # Construct the full filename
    output_filepath = os.path.join(output_path, f"{portfolio_name}_PositionsByAnalyst.xlsx")

    # Step 1: Create the DataFrame from the list of entries
    bookkeeping_space_list = []
    for entry in bookkeeping_space:
        portfolio, investment, lotid, tax_date, ls, location, financial_account, quantity, local, book, notional, oface = entry
        bookkeeping_space_list.append([portfolio, investment, lotid, tax_date, ls, location, financial_account, quantity, local, book])

    df = pd.DataFrame(bookkeeping_space_list, columns=[
        'Portfolio', 'Investment', 'Lot_ID', 'Tax_Date', 'Ls', 'Location',
        'Financial Account', 'Quantity', 'Local_Cost', 'Book_Cost'
    ])

    # Step 2: Perform valuation logic
    valuation_data = []
    for _, row in df.iterrows():
        investment = row['Investment']
        quantity = row['Quantity']
        local_cost = row['Local_Cost']
        book_cost = row['Book_Cost']

        price = utilities.get_price(investment, edate, price_data)
        currency = sub_ledger.get_information_field(investment, 'AIF', 'Currency')
        fx_rate = utilities.get_fx_rate(currency, edate, fx_data)

        if price is None or fx_rate is None:
            print(f"Price or FX rate not found for {investment} on {edate}")
            continue

        pricing_factor = sub_ledger.get_information_field(investment, 'AIF', 'Pricing_Factor') or 1
        mktval_local = quantity * float(price) * float(pricing_factor)
        mktval_book = mktval_local * fx_rate
        price_gl_local = mktval_local - local_cost
        price_gl_book = price_gl_local * fx_rate
        fx_gl_book = mktval_book - book_cost - price_gl_book

        valuation_data.append({
            'Portfolio': row['Portfolio'],
            'Investment': investment,
            'Lot_ID': row['Lot_ID'],
            'Tax_Date': row['Tax_Date'],
            'Ls': row['Ls'],
            'Location': row['Location'],
            'Financial Account': row['Financial Account'],
            'Quantity': quantity,
            'Local_Cost': local_cost,
            'Book_Cost': book_cost,
            'MktVal_Local': mktval_local,
            'MktVal_Book': mktval_book,
            'PriceGLLocal': price_gl_local,
            'PriceGLBook': price_gl_book,
            'FXGLBook': fx_gl_book
        })

    df_valuations = pd.DataFrame(valuation_data)

    # Step 3: Merge with investment master to include Full_Name
    investment_master_path = "C:/Users/hjmne/PycharmProjects/chest/refdata/investment_master.csv"
    investment_master = pd.read_csv(investment_master_path, usecols=['Investment', 'Full_Name'])
    df_valuations = df_valuations.merge(investment_master, on='Investment', how='left')

    # Step 4: Group by Analyst and asset class
    df_valuations = fetch_and_map_groupings(df_valuations,
                                            "C:/Users/hjmne/PycharmProjects/chest/refdata/investment_master.csv",
                                            "C:/Users/hjmne/PycharmProjects/chest/refdata/chart_of_accounts.csv")

    df_valuations['Analyst'] = df_valuations['Analyst'].fillna('Undefined')

    # Step 5: Aggregate data by Analyst and asset class
    group_by_columns = ['Portfolio', 'Analyst', 'Investment', 'Full_Name', 'Ls', 'Financial Account']
    aggregate_columns = ['Quantity', 'Local_Cost', 'Book_Cost', 'MktVal_Local', 'MktVal_Book', 'PriceGLLocal', 'PriceGLBook', 'FXGLBook']

    df_grouped = df_valuations.groupby(group_by_columns)[aggregate_columns].sum().reset_index()

    # Step 6: Add grand totals
    grand_totals = df_grouped[aggregate_columns].sum().to_dict()
    grand_totals.update({col: '' for col in group_by_columns})
    grand_totals['Portfolio'] = 'Grand Totals'
    df_with_totals = pd.concat([df_grouped, pd.DataFrame([grand_totals])], ignore_index=True)

    # Step 7: Save to Excel
    df_with_totals.to_excel(output_filepath, index=False, sheet_name="PositionsByAnalyst")

    print(f"Position report by Analyst saved as {output_filepath}.")

def generic_positions_report(bookkeeping_space, portfolio_name, sub_ledger, edate, fname, price_data, fx_data):
    if not fname.endswith('.xlsx'):
        fname += '.xlsx'

    # Ensure output directory
    output_path = "BASE_PATH/reports"
    os.makedirs(output_path, exist_ok=True)

    # Ensure 'edate' is a proper date object
    if not hasattr(edate, 'month'):
        raise AttributeError("'edate' must be a datetime object with 'month', 'day', and 'year' attributes.")

    # Construct the full filename
    output_filepath = os.path.join(output_path, f"{portfolio_name}_GenericPositions.xlsx")

    # Step 1: Create the DataFrame from the list of entries
    bookkeeping_space_list = []
    for entry in bookkeeping_space:
        portfolio, investment, lotid, tax_date, ls, location, financial_account, quantity, local, book, notional, oface = entry
        bookkeeping_space_list.append([portfolio, investment, lotid, tax_date, ls, location, financial_account, quantity, local, book])

    df = pd.DataFrame(bookkeeping_space_list, columns=[
        'Portfolio', 'Investment', 'Lot_ID', 'Tax_Date', 'Ls', 'Location',
        'Financial Account', 'Quantity', 'Local_Cost', 'Book_Cost'
    ])

    # Step 2: Perform valuation logic
    valuation_data = []
    for _, row in df.iterrows():
        investment = row['Investment']
        quantity = row['Quantity']
        local_cost = row['Local_Cost']
        book_cost = row['Book_Cost']

        price = utilities.get_price(investment, edate, price_data)
        currency = sub_ledger.get_information_field(investment, 'AIF', 'Currency')
        fx_rate = utilities.get_fx_rate(currency, edate, fx_data)

        if price is None or fx_rate is None:
            print(f"Price or FX rate not found for {investment} on {edate}")
            continue

        pricing_factor = sub_ledger.get_information_field(investment, 'AIF', 'Pricing_Factor') or 1
        mktval_local = quantity * float(price) * float(pricing_factor)
        mktval_book = mktval_local * fx_rate
        price_gl_local = mktval_local - local_cost
        price_gl_book = price_gl_local * fx_rate
        fx_gl_book = mktval_book - book_cost - price_gl_book

        valuation_data.append({
            'Portfolio': row['Portfolio'],
            'Investment': investment,
            'Lot_ID': row['Lot_ID'],
            'Tax_Date': row['Tax_Date'],
            'Ls': row['Ls'],
            'Location': row['Location'],
            'Financial Account': row['Financial Account'],
            'Quantity': quantity,
            'Local_Cost': local_cost,
            'Book_Cost': book_cost,
            'MktVal_Local': mktval_local,
            'MktVal_Book': mktval_book,
            'PriceGLLocal': price_gl_local,
            'PriceGLBook': price_gl_book,
            'FXGLBook': fx_gl_book
        })

    df_valuations = pd.DataFrame(valuation_data)

    # Step 3: Merge with investment master to include all tags
    investment_master_path = "C:/Users/hjmne/PycharmProjects/chest/refdata/investment_master.csv"
    investment_master = pd.read_csv(investment_master_path)
    df_valuations = df_valuations.merge(investment_master, on='Investment', how='left')

    # Step 4: Save to Excel
    df_valuations.to_excel(output_filepath, index=False, sheet_name="GenericPositions")

    print(f"Generic positions report saved as {output_filepath}.")


def position_report_by_asset_class(bookkeeping_space, edate, fname):
    # Step 1: Create the DataFrame from the list of entries
    bookkeeping_space_list = []
    for entry in bookkeeping_space:
        portfolio, investment, tax_lot_num, ls, location, financial_account, quantity, local, book = entry
        booksp_row = [portfolio, investment, tax_lot_num, ls, location, financial_account, quantity, local, book]
        bookkeeping_space_list.append(booksp_row)

    df = pd.DataFrame(bookkeeping_space_list, columns=[
        'Portfolio', 'Investment', 'Tax Lot', 'LS', 'Location',
        'Financial Account', 'Quantity', 'Local', 'Book'
    ])
    df = filter_records(df, 'Financial Account', ['MktVal', 'MktValRE'], operator="OR", mode="exclude")

    # Step 2 - value the positions
    # Convert the DataFrame to a dictionary and pass it to the mark_calculations function
    bookkeeping_records = df.set_index([
        'Portfolio', 'Investment', 'Tax Lot', 'LS', 'Location', 'Financial Account']).T.to_dict('list')

    # Assuming these are the paths to your data files
    price_data = "C:/Users/hjmne/PycharmProjects/chest/refdata/price_master.csv"
    fx_data = "C:/Users/hjmne/PycharmProjects/chest/refdata/fx_master.csv"
    mark_type = "Valuation"
    df_valued = value_positions(bookkeeping_records,bookkeeping_space, price_data, fx_data, edate, mark_type)

    # Step 3 - merge with COA/Inv Master Columns

    df_valued = df_valued.drop(columns=['Tax Lot', 'Location', 'Financial Account'])

    # df_merged = fetch_and_map_groupings(df_valued, "C:/Users/hjmne/PycharmProjects/chest/refdata/investment_master.csv",
    #                                     "C:/Users/hjmne/PycharmProjects/chest/refdata/chart_of_accounts.csv")

    df_valued = fetch_from_inv_master_and_merge(df_valued,
                    "C:/Users/hjmne/PycharmProjects/chest/refdata/investment_master.csv", 'asset_class')

    df_valued['sector'] = df_valued['asset_class'].fillna('undefined')
    # Group by the remaining columns excluding 'Tax Lot' and sum the numeric columns
    group_by_columns = ['Portfolio', 'asset_class', 'Investment', 'LS']
    aggregate_columns = ['Quantity', 'Local', 'Book', 'Mkt Val Local', 'Mkt Val Book', 'PGain Local', 'PGain Book',
                         'TotGain Book', 'Fx Gain']

    df_merged = df_valued.groupby(group_by_columns)[aggregate_columns].sum().reset_index()
    # Step 4 - important calc grand totals first. What you are grouping by will dictate subtotal_columns vs total_cols
    group_by_columns = ['asset_class']
    total_columns = ['Book', 'Mkt Val Book', 'PGain Book', 'TotGain Book', 'Fx Gain']
    subtotal_columns = ['Book', 'Mkt Val Book', 'PGain Book','TotGain Book', 'Fx Gain']

    # Always do grand_totals before sub-totaling
    grand_totals = calculate_grand_totals(df_merged, total_columns)

    # Step 5 - Subtotals
    df_subtotals = subtotal_data(df_merged, subtotal_columns, group_by_columns)

    df_subtotals = df_subtotals[df_subtotals['Quantity'] != 0]
    df_subtotals.sort_values(by=['asset_class', 'Investment', 'LS'], inplace=True)
    df_subtotals = insert_subtotals(df_subtotals, ['asset_class'], subtotal_columns)

    # Append the Grand Totals row to the DataFrame with subtotals
    df_final_report = pd.concat([df_subtotals, grand_totals], ignore_index=True)

    # Step 6 - Add % of portfolio
    df_final_report = calculate_percentage_of_portfolio(df_final_report, grand_totals, 'Book')
    df_final_report = move_column_in_dataframe(df_final_report, 'Percent of Portfolio', 3)
      # # Step 8: Save the DataFrame to an Excel file
    fnamechoice = f"C:/Users/hjmne/PycharmProjects/chest/reports/PositionReportByAssetClass"+fname+".xlsx"
    df_final_report.to_excel(fnamechoice, index=False)

    report_title = "Positions Report by asset_class"+fname  # Replace with your report title
    # Assuming 'Portfolio' is the column name where portfolio names are stored.
    # This filters the DataFrame to exclude any rows where 'Portfolio' is 'Subtotal' or 'Grand Totals'
    # and then gets the unique values in the 'Portfolio' column.
    unique_portfolios = df_final_report[
        ~df_final_report['Portfolio'].isin(['Subtotal', 'Grand Totals'])
    ]['Portfolio'].unique()
    # Check if there's at least one portfolio name and use the first one; otherwise, use a default.
    portfolio_name = unique_portfolios[0] if len(unique_portfolios) > 0 else "Unknown Portfolio"
    portfolio_name = unique_portfolios[0] if len(unique_portfolios) > 0 else "Unknown Portfolio"
    edate1 = edate.strftime('%Y-%m-%d')
    period_end_date = edate1  # Assuming 'date' is a datetime object
    # Combine title and date range
    full_title = f"{report_title} for {portfolio_name} ({period_end_date})"
    add_title(fnamechoice, "Sheet1", full_title)

    # Now apply standard formatting to the saved Excel file
    apply_standard_formatting(fnamechoice, "Sheet1")
    # remove columns not wanted
#    drop_columns_by_name(df_final_report, "asset_class")

    # output to pdf
    df_to_pdf(df_final_report, 'df_output.pdf', font_size=12, figsize=(8, 6))


def valuation(sub_ledger):
    # Convert sub_ledger to a list of booksp_rows
    sub_ledger_list = []
    print("test")
    for key, (quantity, local, book) in sub_ledger:
        portfolio, investment, tax_lot_num, ls, location, financial_account = key
        booksp_row = [portfolio, investment, tax_lot_num, ls, location, financial_account, quantity, local, book]
        sub_ledger_list.append(booksp_row)

    return

def objects_to_dicts(obj_list):
    return [vars(obj) for obj in obj_list]

def dump_fund_structures_journals(journal_entries, filename):
    # Load the Chart of Accounts
    coa_df = pd.read_csv("C:/Users/hjmne/PycharmProjects/chest/refdata/chart_of_accounts.csv")
    print(coa_df.columns)
    chart_of_accounts = dict(zip(coa_df['SystemName'], coa_df['Group1']))

    # Create the DataFrame from the list of entries
    df = pd.DataFrame([[
        je.portfolio,
        je.ibor_date,
        je.investment,
        je.tax_date,
        je.ls,
        je.location,
        je.financial_account,
        je.quantity,
        je.local,
        je.book,
        je.tranid,
        je.transaction,
        je.feeder,
        je.tradedate,
        je.settledate,
        je.kdbegin

    ] for je in journal_entries], columns=[
        "Entity", "IBOR Date", "Investment", "Tax Date", "LS", "Master", "Financial Account", "Quantity", "Local",
        "Book", "Tran ID", "Transaction", "Feeder Specific", "Trade Date", "SettleDate", "KnowledgeDate"
    ])
    # Prepare the mapping from SystemName to BSGroup
    account_to_bucket = dict(zip(coa_df['SystemName'], coa_df['BS_Group']))

    # Map the Financial Account in df to its corresponding BS_Group
    df['BS_Group'] = df['Financial Account'].map(account_to_bucket)

    # Map the Financial Account to its Economic Bucket and add as a new column

    # Create sorted entries and subtotals
    #df.sort_values(by=['BS_Group', 'Entity', 'Financial Account', 'IBOR Date'], inplace=True)
    df[['Quantity', 'Local', 'Book']] = df[['Quantity', 'Local', 'Book']].apply(pd.to_numeric, errors='coerce')
    #subtotals = df.groupby(['Investment', 'LS', 'Financial Account'])[['Quantity', 'Local', 'Book']].sum().reset_index()

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

    def df_to_pdf_table(dataframe, filename):
        doc = SimpleDocTemplate(filename, pagesize=letter)
        elements = []

        # Convert DataFrame to a list of lists for the table data
        data = [dataframe.columns.tolist()] + dataframe.values.tolist()

        # Create a Table object
        t = Table(data)

        # Add style to the table (optional)
        t.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                               ('GRID', (0, 0), (-1, -1), 1, colors.black)]))

        elements.append(t)
        doc.build(elements)

    df_to_pdf_table(df, "output_table.pdf")
    # # Save the DataFrame to an Excel file
    # fnamechoice = "C:/Users/hjmne/PycharmProjects/chest/repdata/raw/fundstructuresdump.xlsx"
    # df.to_excel(fnamechoice)



from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

def add_subtotal(df):
    numeric_columns = ['Quantity', 'Local', 'Book', 'Mkt Val Local', 'Mkt Val Book', 'PGain Local', 'PGain Book', 'TotGain Book', 'Fx Gain']
    # add or remove columns to suit your actual dataframe

    out = pd.DataFrame()
    for key, values in df.groupby(['Portfolio', 'Investment', 'LS', 'Location', 'Financial Account']):
        # use .loc to select the numeric columns you want to sum up
        subtotal = values.loc[:, numeric_columns].sum()
        subtotal_line = pd.Series([f'Subtotal', f'{key[1]}', f'{key[2]}', f'{key[3]}', f'{key[4]}'],
                                  index=['Portfolio', 'Investment', 'LS', 'Location', 'Financial Account'])
        subtotal = pd.concat([subtotal, subtotal_line])
        values = pd.concat([values, subtotal.T.to_frame().T])
        out = pd.concat([out, values])
    return out

def add_subtotal_new(df):
    subtotal_columns = ['Primary Quantity', 'Primary Local', 'Primary Book', 'MktValLocal', 'MktValBook']
    subtotal_df = df.groupby(['Investment', 'LS', 'Location'])[subtotal_columns].sum().reset_index()
    subtotal_df['Portfolio'] = 'Subtotal'
    return pd.concat([df, subtotal_df], ignore_index=True)

def create_accounting_reports(journal_entries, fname, bookkeeping_space, asset_liability_list, date):

    mark_type = "Valuation"
    # Create the DataFrame from the list of entries
    df = pd.DataFrame([[
        je.portfolio,
        je.ibor_date,
        je.investment,
        je.tax_date,
        je.ls,
        je.location,
        je.financial_account,
        je.quantity,
        je.local,
        je.book,
        je.tranid,
        je.transaction
    ] for je in journal_entries], columns=[
        "Portfolio", "IBOR Date", "Investment", "Tax Date/Lot Num", "LS", "Location", "Financial Account", "Quantity", "Local",
        "Book", "Tran ID", "Transaction"
    ])

    # Filter the DataFrame to include only rows where IBOR Date is greater than or equal to period_start
#    df = df[df["IBOR Date"] >= period_start]


    # Define the types you want to exclude
    # excluded_types = ['MktVal', 'MktValRE']
    #
    # # Filter df1 to exclude rows where 'Financial Account' matches the excluded types
    # df = df[~df['Financial Account'].isin(excluded_types)]

    fnamechoice ="C:/Users/hjmne/PycharmProjects/chest/reports/" + fname + ".xlsx"

    print(df['Investment'].dtype)
    print(df['Financial Account'].dtype)
    print(df['IBOR Date'].dtype)
    # Create sorted entries and subtotals
 #   df.sort_values(by=['Investment', 'Financial Account', 'IBOR Date'], inplace=True)
 #    df[['Quantity', 'Local', 'Book']] = df[['Quantity', 'Local', 'Book']].apply(pd.to_numeric, errors='coerce')
 #    subtotals = df.groupby(['Investment', 'LS', 'Financial Account'])[['Quantity', 'Local', 'Book']].sum().reset_index()
 #
 #    # Replace 'Portfolio' with 'Subtotal'
 #    subtotals['Portfolio'] = 'Subtotal'

    # Create Grand Totals row based on the original dataframe
    df_copy = df.copy()  # Make a copy before appending subtotals
    grand_totals = df_copy.sum(numeric_only=True)
    grand_totals = grand_totals.to_frame().transpose()
    grand_totals['Portfolio'] = 'Grand Totals'

    # Append subtotals and grand totals to the original dataframe
#    df = pd.concat([df, subtotals, grand_totals], ignore_index=True)

    # Save the DataFrame to an Excel file
    df.to_excel(fnamechoice)

    # Open the workbook with openpyxl to add Excel-specific formatting
    workbook = openpyxl.load_workbook(fnamechoice)
    sheet = workbook.active
    sheet.title = "AccountingJournals"

    # Adjust column widths (starting from row 2 to avoid title)
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column[1:]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.1
        sheet.column_dimensions[column_letter].width = adjusted_width

        sheet.freeze_panes = sheet['A2']
        format_numbers(sheet)
        for column in sheet.columns:
            max_length = max((len(str(cell.value)) for cell in column[1:] if cell.value), default=0) + 2
            adjusted_width = (max_length + 2) * 1.2
            column_letter = get_column_letter(column[0].column)
            sheet.column_dimensions[column_letter].width = adjusted_width

        for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            if i % 2 == 0:  # If the row number is even
                fill_blue_row(row)


    workbook.save(fnamechoice)

    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    # Your list of bookkeeping entries
    bookkeeping_space_list = []
    for entry in asset_liability_list:
        portfolio, investment, lotid, tax_lot_num, ls, location, financial_account, quantity, local, book, notional, oface = entry
        booksp_row = [portfolio, investment, lotid, tax_lot_num, ls, location, financial_account, quantity, local, book, None, None]
        bookkeeping_space_list.append(booksp_row)

    bookkeeping_space_df = pd.DataFrame(bookkeeping_space_list, columns=[
        'Portfolio', 'Investment', 'Lot ID', 'Tax Lot', 'LS', 'Location',
        'Financial Account', 'Quantity', 'Local', 'Book', 'Notional', 'OFace'])
    print(bookkeeping_space_df.head())

    # Convert Quantity, Local, and Book columns to numeric type
    df[['Quantity', 'Local', 'Book']] = df[['Quantity', 'Local', 'Book']].apply(pd.to_numeric, errors='coerce')

    # Filter out records with zero or near-zero Quantity, Local, and Book values
    df = df.query('Quantity != 0 and Local != 0 and Book != 0')

    # Filter to keep only records with financial account = cost, payable, or receivable
    df = df[df['Financial Account'].isin(['cost', 'payable', 'receivable'])]

    # Sort by LS, Investment, and Financial Account
    df.sort_values(by=['Investment', 'Financial Account', 'LS'], inplace=True)

    # Assuming 'df' is your DataFrame

    # Set the values of Quantity, Local, Mkt Val Local, and PGain Local to 0 where Quantity == 0
    df.loc[df['Quantity'] == 0, ['Quantity', 'Local', 'Mkt Val Local', 'PGain Local']] = 0

    # Set the values of Quantity, Local, Mkt Val Local, and PGain Local to 'NA' where Quantity != 0
    # df.loc[df['Quantity'] != 0, ['Quantity', 'Local', 'Mkt Val Local', 'PGain Local']] = 'NA'

    # Convert the DataFrame to a dictionary and pass it to the mark_calculations function
    bookkeeping_records = bookkeeping_space_df.set_index([
        'Portfolio', 'Investment', 'Tax Lot', 'LS', 'Location', 'Financial Account']).T.to_dict('list')

    # Assuming price_data and fx_data are file paths to your .csv files.
    price_data = "C:/Users/hjmne/PycharmProjects/chest/refdata/price_master.csv"
    fx_data = "C:/Users/hjmne/PycharmProjects/chest/refdata/fx_master.csv"

 #   f1, f2, mark_date = mark_to_market.get_data_and_format(date)
    mark_type = "Valuation"
    marked_records = value_positions(bookkeeping_records,bookkeeping_space, price_data, fx_data, date, mark_type)
    asset_liability_accounts_list = bookkeeping_space.get_all_asset_liability_bookkeeping_info()


#    flattened_records = [(rk + data[:-1]) for rk, data in marked_records]
    # Define your DataFrame
    df = pd.DataFrame.from_records(marked_records, columns=[
        'Portfolio', 'Investment', 'Tax Lot', 'LS', 'Location', 'Financial Account',
        'Quantity', 'Local', 'Book', 'Mkt Val Local', 'Mkt Val Book',
        'PGain Local', 'PGain Book', 'TotGain Book', 'Fx Gain'])

    subtotals_df = add_subtotal(df)  # Adjust this with your actual subtotals function

    # Compute grand total on the original data
    numeric_columns = ['Quantity', 'Local', 'Book', 'Mkt Val Local', 'Mkt Val Book',
                       'PGain Local', 'PGain Book', 'TotGain Book', 'Fx Gain']
    grand_total = df[numeric_columns].sum()
    grand_total_line = pd.Series(
        ['Grand Total'] + [''] * (len(df.columns) - len(numeric_columns) - 1) + grand_total.tolist(),
        index=df.columns)

    # Create a new sheet for the report
    wb = load_workbook(fnamechoice)
    ws_summary = wb.create_sheet('Valuation Summary')



    # Write the DataFrame to the "Summary" sheet
    for r in dataframe_to_rows(subtotals_df, index=False, header=True):
        ws_summary.append(r)


    # Add styles to the "Summary" sheet
    for i, row in enumerate(ws_summary.iter_rows(min_row=2), start=2):
        if i % 2 == 0:  # If the row number is even
            fill_blue_row(row)
        if ws_summary.cell(row=i, column=1).value.startswith(
                'Subtotal'):  # if the first cell in the row starts with "Subtotal"
            underline_row(ws_summary, row)
            bold_row(ws_summary, row)
    dark_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    dark_green_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
    dark_blue_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")

    # Style the green row for both sheets
# make sure these three lines are executed in order

    ws_summary.append([''] * len(df.columns))  # Add a blank row
    last_blank_row_summary = ws_summary.max_row  # Identify this new row's number

    for cell in ws_summary[last_blank_row_summary]:
        cell.fill = dark_blue_fill  # Color each cell in the row dark blue

    ws_summary.append(list(grand_total_line))
    last_row_summary = ws_summary.max_row
    from openpyxl.styles import Font

    bold_font = Font(bold=True)
    # # For the "Summary" sheet
    # ws_summary.append(list(grand_total_line))
    # last_row_summary = ws_summary.max_row

    for cell in ws_summary[last_row_summary]:
        cell.font = bold_font

    # Add the grand total to the "Summary" sheet
    # ws_summary.append(list(grand_total_line))
    # last_row_summary = ws_summary.max_row
    # fill_light_green_row(ws_summary[last_row_summary])



    # After appending grand_total_line to ws_detail
    columns_to_exclude = ['Investment', 'Tax Lot', 'LS', 'Location', 'Financial Account', 'Quantity', 'Local', 'Mkt Val Local', 'PGain Local']


    # Similarly, after appending grand_total_line to ws_summary
    for col_name in columns_to_exclude:
        col_num = df.columns.get_loc(col_name) + 1  # +1 since Excel is 1-indexed
        ws_summary.cell(row=last_row_summary, column=col_num).fill = dark_blue_fill


    ws_summary.freeze_panes = ws_summary['A2']
    format_numbers(ws_summary)
    #autosize_columns(ws_summary)
    for column in sheet.columns:
        max_length = max((len(str(cell.value)) for cell in column[1:] if cell.value), default=0) + 2
        adjusted_width = (max_length + 2) * 1.2
        column_letter = get_column_letter(column[0].column)
        sheet.column_dimensions[column_letter].width = adjusted_width

import pandas as pd
import numpy as np

def flatten_records(journal_entries, account_links):
    primary_accounts = set(account_links.keys())
    grouped_entries = {}

    # Group journal entries by key fields
    for je in journal_entries:
        group_key = (je.portfolio, je.investment, je.lotid, je.ibor_date, je.ls, je.location)
        if group_key not in grouped_entries:
            grouped_entries[group_key] = []
        grouped_entries[group_key].append(je)

    flattened_records = []

    # Process each group
    for group_key, entries in grouped_entries.items():
        primary_records = [je for je in entries if je.financial_account in primary_accounts]
        if not primary_records:
            print(f"No primary records found for group: {group_key}")
            continue

        for primary_row in primary_records:
            primary_account = primary_row.financial_account
            linked_accounts = account_links.get(primary_account, [])

            # Initialize secondary data dictionaries
            secondary_data = {
                'PriceGainInvestment Local': 0.0,
                'PriceGainInvestment Book': 0.0,
                'FXGainInvestment Book': 0.0
            }

            # Process secondary records
            for secondary_row in entries:
                secondary_account = secondary_row.financial_account
                if secondary_account in linked_accounts:
                    if secondary_account == 'PriceGainInvestment':
                        secondary_data['PriceGainInvestment Local'] += secondary_row.local
                        secondary_data['PriceGainInvestment Book'] += secondary_row.book
                    elif secondary_account == 'FXGainInvestment':
                        secondary_data['FXGainInvestment Book'] += secondary_row.book

            # Calculate the additional columns
            price_gain_local = secondary_data['PriceGainInvestment Local']
            price_gain_book = secondary_data['PriceGainInvestment Book']
            fx_gain_investment = secondary_data['FXGainInvestment Book']

            proceeds_local = primary_row.local + price_gain_local + fx_gain_investment
            proceeds_book = primary_row.book + price_gain_book + fx_gain_investment

            primary_record = {
                'Portfolio': primary_row.portfolio,
                'Tran ID': primary_row.tranid,
                'Transaction': primary_row.transaction,
                'Investment': primary_row.investment,
                'Lot ID': primary_row.lotid,
                'Close Date': primary_row.ibor_date,
                'Open Date': primary_row.tax_date,
                'LS': primary_row.ls,
                'Location': primary_row.location,
                'Account': primary_account,
                'Quantity': -primary_row.quantity,  # Reverse sign
                'Cost Local': -primary_row.local,  # Reverse sign
                'Cost Book': -primary_row.book,  # Reverse sign
                'Proceeds Local': -proceeds_local,  # Reverse sign
                'Proceeds Book': -proceeds_book,  # Reverse sign
                'PriceGainInvestment Local': -price_gain_local,  # Reverse sign
                'PriceGainInvestment Book': -price_gain_book,  # Reverse sign
                'FXGainInvestment Book': -fx_gain_investment,  # Reverse sign
            }

            flattened_record = {**primary_record}
            flattened_records.append(flattened_record)

    return flattened_records

def add_subtotals_and_grand_totals(df):
    subtotals = []

    for tran_id, group in df.groupby('Tran ID'):
        subtotal = group.sum(numeric_only=True)
        subtotal['Portfolio'] = 'Subtotal'
        subtotal['Tran ID'] = tran_id
        subtotal['Transaction'] = ''
        subtotal['Investment'] = ''
        subtotal['Lot ID'] = ''
        subtotal['Close Date'] = ''
        subtotal['Open Date'] = ''
        subtotal['LS'] = ''
        subtotal['Location'] = ''
        subtotal['Account'] = ''
        subtotals.append(subtotal)

    subtotals_df = pd.DataFrame(subtotals)
    df_with_subtotals = pd.concat([df, subtotals_df]).sort_values(by=['Tran ID', 'Portfolio'])
    df_with_subtotals.reset_index(drop=True, inplace=True)

    # Move subtotals below the details
    df_final = []
    for tran_id, group in df_with_subtotals.groupby('Tran ID'):
        group = group.reset_index(drop=True)
        subtotal_row = group[group['Portfolio'] == 'Subtotal']
        details_rows = group[group['Portfolio'] != 'Subtotal']
        df_final.append(details_rows)
        df_final.append(subtotal_row)

    final_df = pd.concat(df_final).reset_index(drop=True)

    # Calculate grand totals for specific book columns
    grand_totals = final_df.sum(numeric_only=True)
    grand_totals['Portfolio'] = 'Grand Totals'
    grand_totals['Tran ID'] = ''
    grand_totals['Transaction'] = ''
    grand_totals['Investment'] = ''
    grand_totals['Lot ID'] = ''
    grand_totals['Close Date'] = ''
    grand_totals['Open Date'] = ''
    grand_totals['LS'] = ''
    grand_totals['Location'] = ''
    grand_totals['Account'] = ''

    # Only keep grand totals for specified book columns
    book_columns = ['Cost Book', 'Proceeds Book', 'PriceGainInvestment Book', 'FXGainInvestment Book']
    grand_totals = grand_totals[book_columns]
    grand_totals['Portfolio'] = 'Grand Totals'

    final_df = pd.concat([final_df, pd.DataFrame([grand_totals])], ignore_index=True)

    return final_df

def add_title_and_formatting(writer, sheet_name, report_title, df):
    worksheet = writer.sheets[sheet_name]

    # Add title
    worksheet.merge_range('A1:N1', report_title, writer.book.add_format({'bold': True, 'font_size': 14}))
    worksheet.set_row(0, 30)  # Adjust row height for title

    # Freeze panes
    worksheet.freeze_panes(1, 0)

    # Apply formatting
    format1 = writer.book.add_format({'num_format': '#,##0.00'})  # Formatting numbers
    format2 = writer.book.add_format({'bold': True, 'num_format': '#,##0.00'})  # Formatting subtotals and grand totals
    format_blue = writer.book.add_format({'bg_color': '#B0E0E6'})  # Light blue background

    # Apply column formats
    for col_num, col_name in enumerate(df.columns):
        column_len = max(df[col_name].astype(str).str.len().max(), len(col_name)) + 2  # Adjust column width
        worksheet.set_column(col_num, col_num, column_len, format1 if df[col_name].dtype in [np.float64, np.int64] else None)

    # Apply row formats
    for row_num, row in df.iterrows():
        row_format = format_blue if row_num % 2 == 0 else None
        worksheet.set_row(row_num + 1, None, row_format)
        if 'Subtotal' in row['Portfolio'] or 'Grand Totals' in row['Portfolio']:
            worksheet.set_row(row_num + 1, None, format2)

def gl_test(journal_entries, fname):
    account_links = {
        'Cost': ['PriceGainInvestment', 'FXGainInvestment']
    }

    # Filter for 'SellLong' transactions
    filtered_entries = [je for je in journal_entries if je.transaction == 'SellLong']

    flattened_records = flatten_records(filtered_entries, account_links)

    # Convert the flattened records into a DataFrame
    flattened_df = pd.DataFrame(flattened_records)

    # Add subtotals and grand totals by Tran ID
    flattened_df = add_subtotals_and_grand_totals(flattened_df)

    # Write the DataFrame to an Excel file with formatting
    fnamechoice = f"C:/Users/hjmne/PycharmProjects/chest/reports/{fname}.xlsx"
    with pd.ExcelWriter(fnamechoice, engine='xlsxwriter') as writer:
        flattened_df.to_excel(writer, index=False, startrow=1, sheet_name='Realized GL Report')

        # Add title and apply formatting
        report_title = "Positions Report by asset_class" + fname  # Replace with your report title
        unique_portfolios = flattened_df[~flattened_df['Portfolio'].isin(['Subtotal', 'Grand Totals'])]['Portfolio'].unique()
        portfolio_name = unique_portfolios[0] if len(unique_portfolios) > 0 else "Unknown Portfolio"
        period_end_date = pd.to_datetime('today').strftime('%Y-%m-%d')  # Assuming you want the current date
        full_title = f"{report_title} for {portfolio_name} ({period_end_date})"
        add_title_and_formatting(writer, "Realized GL Report", full_title, flattened_df)

# You can now call gl_test with your actual journal_entries and file name
# gl_test(journal_entries, 'realized_gl_report')

def create_raw_accounting_reports(journal_entries, usefordiff, fname, period_start, period_cutoff, sub_ledger, account_to_bucket):
    # Create the DataFrame from the list of entries
    coa_df = pd.read_csv("C:/Users/hjmne/PycharmProjects/chest/refdata/chart_of_accounts.csv")
    print(coa_df.columns)
    chart_of_accounts = dict(zip(coa_df['SystemName'], coa_df['Group1']))
    df = pd.DataFrame52([[
        je.portfolio,
        je.ibor_date,
        je.investment,
        je.tax_date,
        je.ls,
        je.location,
        je.financial_account,
        je.quantity,
        je.local,
        je.book,
        je.tranid,
        je.transaction,
        je.feeder
    ] for je in journal_entries], columns=[
        "Portfolio", "IBOR Date", "Investment", "Tax Date", "LS", "Location", "Financial Account", "Quantity", "Local",
        "Book", "Tran ID", "Transaction", "Feeder"
    ])

    # Ensure that account_to_bucket is a dictionary mapping 'SystemName' to 'BS_Group'.
    account_to_bucket = dict(zip(coa_df['SystemName'], coa_df['BS_Group']))

    # Map external data to DataFrame
    df = map_external_data(df, 'Financial Account', account_to_bucket, 'BS_Group')

    # Compute the 'BSGroup' column
    df['BS_Group'] = df['Financial Account'].map(account_to_bucket)

    # Sort the DataFrame by 'BSGroup'
    df = df.sort_values(by='BS_Group')
    # Group by 'BSGroup' and sum the 'Book' column for the original DataFrame
    bsgroup_subtotals = df.groupby("BS_Group")["Book"].sum().reset_index()

    # Rename the columns in the subtotal DataFrame
    bsgroup_subtotals.columns = ["BS_Group", "Book Subtotal"]

    # Merge the subtotals back to the original DataFrame
    df = pd.merge(df, bsgroup_subtotals, on="BS_Group", how="left")

    # Save the sorted DataFrame with subtotals to an Excel file
    fnamechoice = "C:/Users/hjmne/PycharmProjects/chest/repdata/" + fname + ".xlsx"
    df.to_excel(fnamechoice, index=False)

def split_columns_into_slices(input_file, output_file):
    # Load the Excel file
    df = pd.read_excel(input_file)

    slices = [2, 3, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2,
              0, 3, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2]

    num_slices = len(slices)

    # Function to split a value into slices based on percentages
    def split_value_into_slices(value):
        total_slices = sum(slices)
        slices_percentage = [s / total_slices for s in slices]
        return [value * sp for sp in slices_percentage]

    # Create a new DataFrame to store the sliced data
    sliced_data = []

    # Iterate through the records
    for index, row in df.iterrows():
        if pd.isna(row['Investment']):
            continue  # Skip blank rows

        investment = row['Investment']
        ls = row['LS']
        financial_account = row['Financial Account']
        quantity_diff = row['Quantity_diff']
        local_diff = row['Local_diff']
        book_diff = row['Book_diff']

        entity_counter = 1  # Reset entity counter for each record

        quantity_slices = split_value_into_slices(quantity_diff)
        local_slices = split_value_into_slices(local_diff)
        book_slices = split_value_into_slices(book_diff)

        for i in range(num_slices):
            sliced_row = {
                'Entity': f'Entity{entity_counter}',
                'Investment': investment,
                'LS': ls,
                'Financial Account': financial_account,
                'Quantity_diff': quantity_slices[i],
                'Local_diff': local_slices[i],
                'Book_diff': book_slices[i],
                'Allocation %': slices[i]
            }
            sliced_data.append(sliced_row)
            entity_counter += 1

    # Create a new DataFrame for the sliced data
    sliced_df = pd.DataFrame(sliced_data)

    # Save the sliced data to a new Excel file
    sliced_df.to_excel(output_file, index=False)

from openpyxl.styles import NamedStyle, PatternFill
from openpyxl import load_workbook


def diff_two_excel_files(f1, f2):
    import os
    # Paths for convenience
    # path_prefix ="C:/Users/hjmne/PycharmProjects/chest/repdata/"

    # Read the Excel files into DataFrames

    df2 = pd.read_excel("C:/Users/hjmne/PycharmProjects/chest/reports/"+f2,
                                   sheet_name='AccountingJournals')

    df1 = pd.read_excel("C:/Users/hjmne/PycharmProjects/chest/reports/"+f1,
                        sheet_name='AccountingJournals')

    output_file = ("C:/Users/hjmne/PycharmProjects/chest/reports/AccountingComparison.xlsx")

    # Explicitly drop columns containing "Running Balance" in their name
    columns_to_exclude = [col for col in df1.columns if "Running" in col or "Tran" in col or "Unnamed" in col]
    df1.drop(columns=columns_to_exclude, inplace=True)
    df2.drop(columns=columns_to_exclude, inplace=True)

    print(df1.index)
    print(df2.index)

    print(df1.index.is_unique)
    print(df2.index.is_unique)


    # Define aggregation keys
    aggregation_keys = ['Investment', 'LS', 'Location', 'Financial Account']

    # Columns to aggregate (Exclude datetime or non-numeric columns)
    numeric_cols = df1.select_dtypes(include=['number']).columns.tolist()
    columns_to_aggregate = list(set(numeric_cols) - set(aggregation_keys))

    # Aggregate based on the keys for both dataframes
    df1_aggregated = df1.groupby(aggregation_keys)[columns_to_aggregate].sum().reset_index()
    df2_aggregated = df2.groupby(aggregation_keys)[columns_to_aggregate].sum().reset_index()

    # Merge the aggregated dataframes using outer join to keep all records
    merged_df = pd.merge(df1_aggregated, df2_aggregated, on=aggregation_keys, how='outer',
                         suffixes=('_file1', '_file2'))

    # Replace NaN with 0 to handle missing records
    merged_df.fillna(0, inplace=True)

    # Get the column names for differences and calculate them
    diff_columns = ['Quantity', 'Local', 'Book']
    for col_name in diff_columns:
        merged_df[f'{col_name}_diff'] = merged_df[f'{col_name}_file2'] - merged_df[f'{col_name}_file1']


    # Save the aggregated differences to an Excel file
    with pd.ExcelWriter(output_file) as writer:
        merged_df.to_excel(writer, sheet_name='Aggregated Differences', index=False)

    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    apply_standard_formatting(output_file, "Aggregated Differences")
    # Save the aggregated differences to an Excel file
   # output_file = path_prefix + f1 + f2 + ".xlsx"
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        merged_df.to_excel(writer, sheet_name='Aggregated Differences', index=False)

    # Load the workbook again using openpyxl to apply styles
    wb = load_workbook(output_file)
    ws = wb['Aggregated Differences']

    # Define a style with a number format that includes commas
    from openpyxl.styles import NamedStyle
    comma_format = NamedStyle(name="comma_format", number_format="#,##0")

    # Apply the number format style to the numeric columns
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.style = comma_format

    # Apply the fill for every second row in the worksheet
    fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    for idx, row in enumerate(ws.iter_rows(), start=1):
        if idx % 2 == 0:  # If the row number is even
            for cell in row:
                cell.fill = fill
    # Freeze the panes
    ws.freeze_panes = "A2"

    # Identify columns for Quantity_diff, Local_diff, and Book_diff
    columns = {cell.value: cell.column for cell in next(ws.iter_rows())}
    qty_diff_col = columns['Quantity_diff']
    local_diff_col = columns['Local_diff']
    book_diff_col = columns['Book_diff']

    # Identify the next available row
    next_row = ws.max_row + 1

    # Add the title "Net Adjustments"
    ws[f"A{next_row}"] = ""
    next_row += 1

    # Identify the column that contains the account type
    account_type_col = columns['Financial Account']  # Adjust this column title based on your actual data
    apply_standard_formatting(output_file, "Aggregated Differences")
    # Make sure 'Financial Account' column doesn't have leading/trailing spaces
    df1['Financial Account'] = df1['Financial Account'].str.strip()
    df2['Financial Account'] = df2['Financial Account'].str.strip()

    desired_types = ['Cost', 'Receivable', 'Payable', 'UnrealGLAsset', 'DividendsReceivable', 'DividendsPayable','SpotFxPayable','SpotFxReceivable',
                     'ExpensesPayable', 'InterestReceivable', 'InterestPayable','AccruedInterestPayable','AccruedInterestReceivable']

    net_change_accumulator = 0  # This will store the accumulated value

    # Copy non-blank records
    for row in ws.iter_rows(min_row=2):
        qty_diff = row[qty_diff_col - 1].value
        local_diff = row[local_diff_col - 1].value
        book_diff = row[book_diff_col - 1].value

        account_type = row[account_type_col - 1].value
        if account_type in desired_types:
            if book_diff is not None:
                net_change_accumulator += book_diff
            else:
                # Handle the case when book_diff is None
                # For example, log an error message, or add a default value.
                net_change_accumulator += 0  #

        if not (qty_diff == 0 and local_diff == 0 and book_diff == 0):
            for idx, cell in enumerate(row):
                new_cell = ws.cell(row=next_row, column=idx + 1, value=cell.value)
                new_cell.fill = green_fill
            next_row += 1
    # Define the light purple fill
    purple_fill = PatternFill(start_color="D8BFD8", end_color="D8BFD8", fill_type="solid")

    # After adding "Net Adjustments" title, add the accumulated value
    ws[f"A{next_row}"] = "Net Adjustments in Assets/Liabilities"
    ws[f"M{next_row}"] = net_change_accumulator
    ws[f"M{next_row}"].style = comma_format

    # Fill the row with purple
    for cell in ws[f"A{next_row}:M{next_row}"]:
        for c in cell:
            c.fill = green_fill
    next_row += 1

    # For df1
    book_sum_df1 = df1[df1['Financial Account'].isin(desired_types)]['Book'].sum()
    # For df2
    book_sum_df2 = df2[df2['Financial Account'].isin(desired_types)]['Book'].sum()

    # Calculate the difference
    net_book_difference = book_sum_df2 - book_sum_df1

    # After the Net Adjustments calculations, before saving:

    # Add the aggregate values and their difference
    next_row += 1  # Maybe add a blank row for separation
    ws[f"A{next_row}"] = "Net Market Value in File1"
    ws[f"M{next_row}"] = book_sum_df1
    ws[f"M{next_row}"].style = comma_format

    # Fill the row with purple
    for cell in ws[f"A{next_row}:M{next_row}"]:
        for c in cell:
            c.fill = purple_fill
    next_row += 1

    ws[f"A{next_row}"] = "Net Market Value in File2"
    ws[f"M{next_row}"] = book_sum_df2
    ws[f"M{next_row}"].style = comma_format

    # Fill the row with purple
    for cell in ws[f"A{next_row}:M{next_row}"]:
        for c in cell:
            c.fill = purple_fill
    next_row += 1

    ws[f"A{next_row}"] = "Net Market Value Difference"
    ws[f"M{next_row}"] = net_book_difference
    ws[f"M{next_row}"].style = comma_format

    # Fill the row with purple
    for cell in ws[f"A{next_row}:M{next_row}"]:
        for c in cell:
            c.fill = purple_fill
    next_row += 1
    # Save the changes
    wb.save(output_file)

def reporttest():
    import pandas as pd

    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

    # Read data from Excel file using pandas
    file_path = "C:/Users/hjmne/PycharmProjects/chest/repdata/summary_report.xlsx"
    df = pd.read_excel(file_path)

    # Convert the DataFrame to a list of lists
    data = [df.columns.tolist()] + df.values.tolist()

    # Create a new PDF with the title 'Investment Report'
    doc = SimpleDocTemplate("Investment_Report.pdf", pagesize=landscape(letter))

    # Create a table with the data
    table = Table(data, repeatRows=1)
    table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]
    table.setStyle(TableStyle(table_style))

    # Add the table to the PDF
    doc.build([table])

    print("PDF generated: Investment_Report.pdf")

import pandas as pd



def position_dump(bookkeeping_space, edate, fname):
    # Step 1: Create the DataFrame from the list of entries
    # Assuming sub_ledger_list is the list shown above
    bookkeeping_space_list = []
    for entry in bookkeeping_space:
        portfolio, investment, tax_lot_num, ls, location, financial_account, quantity, local, book = entry
        booksp_row = [portfolio, investment, tax_lot_num, ls, location, financial_account, quantity, local, book]
        bookkeeping_space_list.append(booksp_row)


    df = pd.DataFrame(bookkeeping_space_list, columns=[
        'Portfolio', 'Investment', 'Tax Lot', 'LS', 'Location',
        'Financial Account', 'Quantity', 'Local', 'Book'
    ])

    # #Step 2 - value the positions
    # # Convert the DataFrame to a dictionary and pass it to the mark_calculations function
    # bookkeeping_records = df.set_index([
    #     'Portfolio', 'Investment', 'Tax Lot', 'LS', 'Location', 'Financial Account']).T.to_dict('list')
    # df = filter_records(df, 'Financial Account', ['MktVal', 'MktValRE'], operator="OR", mode="exclude")
    # # Assuming these are the paths to your data files
    # price_data_path = "C:/Users/hjmne/PycharmProjects/chest/refdata/price_master.csv"
    # fx_data_path = "C:/Users/hjmne/PycharmProjects/chest/refdata/fx_master.csv"

    # df_valued = value_positions(bookkeeping_records,bookkeeping_space, price_data, fx_data, date, mark_type)
    #
    # #Step 3 - merge with COA/Inv Master Columns
    #
    # df_merged = fetch_and_map_groupings(df_valued, "C:/Users/hjmne/PycharmProjects/chest/refdata/investment_master.csv",
    #                                     "C:/Users/hjmne/PycharmProjects/chest/refdata/chart_of_accounts.csv")
    #
    # # Step 4 - important calc grand totals first. What you are grouping by will dictate subtotal_columns vs total_cols
    # group_by_columns = ['Investment', 'Financial Account', 'LS']
    # total_columns = ['Book',  'Mkt Val Book', 'PGain Book','TotGain Book', 'Fx Gain']
    # subtotal_columns = ['Quantity', 'Local', 'Book', 'Mkt Val Local', 'Mkt Val Book', 'PGain Local', 'PGain Book',
    #                     'TotGain Book', 'Fx Gain']
    #
    # #Always do grand_totals before sub-totaling
    # grand_totals = calculate_grand_totals(df_merged, total_columns)
    #
    # df_subtotals = subtotal_data(df_merged, subtotal_columns, group_by_columns)
    #
    # df_subtotals = df_subtotals[df_subtotals['Quantity'] != 0]
    # df_subtotals.sort_values(by=['Investment', 'Financial Account', 'LS'], inplace=True)
    # df_subtotals = insert_subtotals(df_subtotals, ['Investment', 'Financial Account', 'LS'], subtotal_columns)
    #
    # # Append the Grand Totals row to the DataFrame with subtotals
    # df_final_report = pd.concat([df_subtotals, grand_totals], ignore_index=True)

    # Step 5: Save the DataFrame to an Excel file
    fnamechoice = f"C:/Users/hjmne/PycharmProjects/chest/reports/TaxLotHoldings.xlsx"
    df.to_excel(fnamechoice, index=False)

    report_title = "Tax Lot Holdings"  # Replace with your report title
    # Assuming 'Portfolio' is the column name where portfolio names are stored.
    # This filters the DataFrame to exclude any rows where 'Portfolio' is 'Subtotal' or 'Grand Totals'
    # and then gets the unique values in the 'Portfolio' column.
  #   unique_portfolios = df_final_report[
  #       ~df_final_report['Portfolio'].isin(['Subtotal', 'Grand Totals'])
  #   ]['Portfolio'].unique()
  #   # Check if there's at least one portfolio name and use the first one; otherwise, use a default.
  #   portfolio_name = unique_portfolios[0] if len(unique_portfolios) > 0 else "Unknown Portfolio"
  #   edate1 = edate.strftime('%Y-%m-%d')
  #   period_end_date = edate1  # Assuming 'date' is a datetime object
  #   # Combine title and date range
  #   full_title = f"{report_title} for {portfolio_name} ({period_end_date})"
  #   add_title(fnamechoice, "Sheet1", full_title)
  #
  # # Now apply standard formatting to the saved Excel file
  #   apply_standard_formatting(fnamechoice, "Sheet1"

import os
import pickle


# Function to concatenate portfolio_name with sub_ledger accounting object and store them
def concatenate_and_store_ledger(space_manager, events_sheet, portfolio_name, process_start_date,
                                 current_period_start, current_period_cutoff, current_period_knowledge,
                                 sub_ledger, general_ledger, tdate_fx, scheduler, stat_repo, price_data, fx_data, smf,
                                 mark_daily, aggregate_marks, include_marks):

    # Directory to store the concatenated ledgers
    accounting_dir = "BASE_PATH/accounting_spaces"
    os.makedirs(accounting_dir, exist_ok=True)

    # Concatenate portfolio name with sub_ledger
    ledger_filename = f"{portfolio_name}_sub_ledger.pkl"
    ledger_filepath = os.path.join(accounting_dir, ledger_filename)

    # Save the concatenated object
    with open(ledger_filepath, 'wb') as ledger_file:
        pickle.dump(sub_ledger, ledger_file)

    print(f"Sub ledger for portfolio '{portfolio_name}' successfully stored at: {ledger_filepath}")

# Example usage
# concatenate_and_store_ledger(space_manager, events_sheet, "PortfolioABC", process_start_date,
#                              current_period_start, current_period_cutoff, current_period_knowledge,
#                              sub_ledger, general_ledger, tdate_fx, scheduler, stat_repo, price_data, fx_data, smf,
#                              mark_daily=True, aggregate_marks=True, include_marks=False)
def store_master_query_space(portfolio, journal_entries):
    """
    Stores Journal Entries in a Master Query Space for a given portfolio.
    Must be called manually after process_events().
    """
    if not journal_entries:
        print(f"⚠ WARNING: No Journal Entries to store for {portfolio}. Skipping.")
        return

    # ✅ Convert to DataFrame
    je_df = pd.DataFrame(journal_entries)

    # ✅ Store in the global dictionary (exact return order)


    print(f"✅ Master Query Space stored for {portfolio} ({len(je_df)} entries)")


def get_master_query_journals(portfolio):
    """
    Retrieves Journal Entries for a given portfolio in original order.
    Returns an empty DataFrame if none exist.
    """
    return MASTER_QUERY_SPACES.get(portfolio, pd.DataFrame())
